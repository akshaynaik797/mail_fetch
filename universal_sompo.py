import pandas as pd
from email.mime.text import MIMEText
import smtplib
import time
import imaplib
import sys
import email
import camelot
import PyPDF2
import csv
import xlsxwriter
from xlrd import open_workbook
import xlwt
import os
import glob
import os.path
import xlrd
from os import listdir
from os import path
import pdftotext
from os.path import isfile, join
from html.parser import HTMLParser
import pdfkit
import pandas as pd
import html2text
import re
import subprocess

from decode_error import check_subject, read_from_delete
from make_log import log_exceptions

try:

    fg = []
    subprocess.run(["python", "updation.py", "1", "max", "9", "X"])

    repeat = []
    # sys.argv = ['universal_sompo.py', 'mediclaim@inamdarhospital.org', 'Mediclaim@2019', '10-Aug-2020', '10-Aug-2020', 'imap.gmail.com', 'inamdar', '6230']
    # sys.argv = ['universal_sompo.py', 'Tpappg@maxhealthcare.com', 'May@2020', '10-Aug-2020', '10-Aug-2020', 'outlook.office365.com', 'Max', '10245']
    def read_email_from_gmail():
        SMTP_SERVER = str(sys.argv[5])
        mail = imaplib.IMAP4_SSL(SMTP_SERVER)
        e_id = str(sys.argv[1])
        pswd = str(sys.argv[2])
        srt = str(sys.argv[3])
        stp = str(sys.argv[4])
        mail.login(user=e_id, password=pswd)
        mail.select("inbox", readonly=True)
        ###############################################<
        mail_uid = str(sys.argv[7])
        if mail_uid == -1:
            type, data = mail.search(None,
                                     '(FROM "donotreply@universalsompo.co.in" SUBJECT "Paid for Patient Name-" since ' + srt + ' before ' + stp + ')')
            ids = data[0]
            id_list = ids.split()
        else:
            ids = mail_uid  # data is a list.
            # accept id from outside and put in id_list akshay var name = id

            id_list = []  # ids is a space separated string
            id_list.append(ids)
        ###############################################>
        # type, data = mail.search(None,
        #                          '(FROM "donotreply@universalsompo.co.in" SUBJECT "Paid for Patient Name-" since ' + srt + ' before ' + stp + ')')
        # ids = data[0]  # data is a list.
        # id_list = ids.split()  # ids is a space separated string
        # # print(id_list)
        for i in range(0, len(id_list)):
            latest_email_id = id_list[i]  # get the latest
            result, data = mail.fetch(latest_email_id,
                                      "(RFC822)")  # fetch the email body (RFC822)             for the given ID

            ##################################################ak
            try:
                # raw_email = data[0][1].decode('utf-8')
                try:
                    raw_email = data[0][1].decode('utf-8')
                except UnicodeDecodeError:
                    try:
                        raw_email = data[0][1].decode('ISO-8859-1')
                    except UnicodeDecodeError:
                        try:
                            raw_email = data[0][1].decode('ascii')
                        except UnicodeDecodeError:
                            pass
                email_message = email.message_from_string(raw_email)
                subject = email_message['Subject']
                result, sys.argv[8] = check_subject(subject, sys.argv[8], mail)
                if result == 'Changed':
                    # raise Exception('subject not matched')
                    raise Exception('subject not matched', )
            except:
                try:
                    log_exceptions(syssubject=sys.argv[8], subject=subject, error='subject not matched')
                except:
                    pass
                if result != 'OK':
                    data = {'server': SMTP_SERVER,
                            'hospmail': e_id,
                            'pass': pswd,
                            'subject': sys.argv[8]}
                    try:
                        data = read_from_delete(data)
                        if data == None:
                            raise Exception("Not found")
                    except:
                        log_exceptions(msg='not found in deleted', subject=sys.argv[8])
            ##################################################akend

            raw_email = data[0][1].decode('utf-8')
            email_message = email.message_from_string(raw_email)
            if email_message['Subject'] not in fg:
                for part in email_message.walk():
                    if part.get_content_maintype() == 'multipart':
                        # print part.as_string()
                        continue
                    if part.get('Content-Disposition') is None:
                        # print part.as_string()
                        continue
                    fileName = part.get_filename()
                    detach_dir = (os.getcwd() + '/universal_sompo/attachments_' + str(sys.argv[6]))
                    if bool(fileName):
                        filePath = os.path.join(detach_dir, fileName)
                        if not os.path.isfile(filePath):
                            from reportlab.pdfgen import canvas
                            c = canvas.Canvas(fileName + '.pdf')
                            # print (fileName)
                            fp = open(filePath, 'wb')
                            fp.write(part.get_payload(decode=True))
                            fp.close()
            else:
                repeat.append(email_message['Subject'])

            fg.append(email_message['Subject'])


    mypath = os.getcwd() + '/universal_sompo'
    if not path.exists(mypath):
        os.mkdir(mypath)
    if not path.exists(mypath + '/attachments_' + str(sys.argv[6])):
        os.mkdir(mypath + '/attachments_' + str(sys.argv[6]))
    mypath = os.getcwd() + '/universal_sompo/attachments_' + str(sys.argv[6]) + '/'

    for filename in os.listdir(mypath):
        file_path = os.path.join(mypath, filename)
        if os.path.isfile(file_path) or os.path.islink(file_path):
            os.unlink(file_path)
    read_email_from_gmail()
    import sys

    for filename in os.listdir(mypath):
        file_path = os.path.join(mypath, filename)
        # print(filename)
        if filename.find('.PDF') == -1:
            # print(file_path)
            os.remove(file_path)
    onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
    if path.exists(r'universal_sompo/universal_sompo' + str(sys.argv[6]) + '.xlsx'):
        os.remove(r'universal_sompo/universal_sompo' + str(sys.argv[6]) + '.xlsx')
    import openpyxl

    po = []
    wbkName = 'universal_sompo/universal_sompo' + str(sys.argv[6]) + '.xlsx'
    wbk = openpyxl.Workbook()
    wbk.create_sheet('1')
    s1 = wbk.worksheets[0]
    s2 = wbk.worksheets[1]
    wq = 0
    for t in range(0, len(onlyfiles)):
        # try:
        tables = camelot.read_pdf(mypath + onlyfiles[t], pages='all')
        tables.export('universal_sompo/foo1.xls', f='excel')
        loc = ("universal_sompo/foo1.xls")
        wb = xlrd.open_workbook(loc)
        with open(mypath + onlyfiles[t], "rb") as f:
            pdf = pdftotext.PDF(f)

        with open('universal_sompo/output.txt', 'w') as f:
            f.write(" ".join(pdf))
        with open('universal_sompo/output.txt', 'r') as myfile:
            f = myfile.read()
        sh1 = ['sr no', 'CCN', 'IP NO', 'Patient Name', 'doa', 'dod', 'diagnosis', 'Beneficiary Name', 'Acc No.',
               'Bank name', 'IFSC code', 'UTR No.', 'NEFT Date', 'BilledAmount', 'SettledAmount', 'TDS', 'NetPayable',
               'DiscountAmt', 'COPay', 'deduction', 'Cashless Authorized']
        for i in range(0, len(sh1)):
            s1.cell(row=1, column=i + 1).value = sh1[i]
        sh2 = ['sr no', 'CCN', 'category', 'deduction', 'reason']
        for i in range(0, len(sh2)):
            s2.cell(row=1, column=i + 1).value = sh2[i]
        hg = []

        regex = r'(?<=Claim Registration Number:) *\d+'
        result = re.search(regex, f)
        if result:
            hg.append(result.group().strip())
        else:
            w = f.find('Claim No:') + 10
            g = f[w:]
            u = g.find('\n') + w
            hg.append(f[w:u])

        w = f.find('Patient IP NO:') + 14
        g = f[w:]
        u = g.find('Claimed Amount:') + w
        hg.append(f[w:u])

        w = f.find('Patient Name:') + 13
        g = f[w:]
        u = g.find('Approved Amount') + w
        hg.append(f[w:u])

        w = f.find('Date of Admission:') + 18
        g = f[w:]
        u = g.find('Co Pay Amount:') + w
        hg.append(f[w:u])

        w = f.find('Date of Discharge:') + 18
        g = f[w:]
        u = g.find('TDS Deducted:') + w
        hg.append(f[w:u])

        w = f.find('Ailment:') + 10
        g = f[w:]
        u = g.find('Amount not') + w
        hg.append(f[w:u])

        w = f.find('Beneficiary Name:') + 17
        g = f[w:]
        u = g.find('NEFT Date:') + w
        hg.append(f[w:u])

        w = f.find('Beneficiary Acc No:') + 19
        g = f[w:]
        u = g.find('UTR No:') + w
        hg.append(f[w:u])

        w = f.find('Bank Name:') + 10
        g = f[w:]
        u = g.find('\n') + w
        hg.append(f[w:u])

        w = f.find('IFSC Code:') + 10
        g = f[w:]
        u = g.find('\n') + w
        hg.append(f[w:u])

        w = f.find('UTR No:') + 7
        g = f[w:]
        u = g.find('\n') + w
        hg.append(f[w:u])

        w = f.find('NEFT Date:') + 10
        g = f[w:]
        u = g.find('\n') + w
        hg.append(f[w:u])

        w = f.find('Claimed Amount:') + 14
        g = f[w:]
        u = g.find('\n') + w
        hg.append(f[w:u])

        w = f.find('Approved Amount:') + 15
        g = f[w:]
        u = g.find('\n') + w
        hg.append(f[w:u])

        w = f.find('TDS Deducted:') + 12
        g = f[w:]
        u = g.find('\n') + w
        hg.append(f[w:u])

        regex = r'(?<=Paid Amount after) *\d+'
        result = re.search(regex, f)
        if result:
            hg.append(result.group().strip())
        else:
            w = f.find('Paid Amount after TDS') + 22
            g = f[w:]
            u = g.find('\n') + w
            hg.append(f[w:u])

        w = f.find('Discount Amount:') + 15
        g = f[w:]
        u = g.find('\n') + w
        hg.append(f[w:u])

        w = f.find('Co Pay Amount:') + 13
        g = f[w:]
        u = g.find('\n') + w
        hg.append(f[w:u])

        w = f.find('Amount not paid*:') + 16
        g = f[w:]
        u = g.find('\n') + w
        hg.append(f[w:u])

        w = f.find('Cashless Authorized Amount') + 26
        g = f[w:]
        u = g.find('\n') + w
        hg.append(f[w:u])

        hg = [sub.replace('  ', '') for sub in hg]
        hg = [sub.replace(':', '') for sub in hg]

        # print(hg)

        for i in range(0, len(hg)):
            s1.cell(row=t + 2, column=1).value = t + 1
            s1.cell(row=t + 2, column=i + 2).value = hg[i]

        regex = r'(?<=Reason for Deduction)\r?\n[ \S\n]+(?=In case of any variance)'
        regex2 = r'(?P<category>[\S ]+[^\d](?=\d+.0{2}))(?P<deduction>\d+.0{2})(?P<reason>[ \S]+)'
        result = re.search(regex, f)
        if result:
            raw = result.group().strip()
            s2_data = [match.groupdict() for match in re.compile(regex2).finditer(raw)]

        for i in s2_data:
            row_num = s2.max_row
            s2.cell(row=row_num + 1, column=1).value = row_num
            s2.cell(row=row_num + 1, column=2).value = hg[0]
            s2.cell(row=row_num + 1, column=3).value = i['category'].strip()
            s2.cell(row=row_num + 1, column=4).value = i['deduction'].strip()
            s2.cell(row=row_num + 1, column=5).value = i['reason'].strip()

        # w = f.find('Reason for Deduction') + 21
        # g = f[w:]
        # u = g.find('In case') + w
        # temp = f[w:u]
        # temp = temp.split('\n')
        # temp.pop()
        # so1 = []
        # so2 = []
        # so3 = []
        # for k in temp:
        #     w = k.find('.')
        #     g = k[:w]
        #     h = k[w:]
        #     km = " "
        #     u = g.rindex(km)
        #     u1 = h.find(' ') + w
        #     so1.append(k[:u])
        #     so2.append(k[u:u1])
        #     so3.append(k[u1:])
        # so1 = [sub.replace('  ', '') for sub in so1]
        # so2 = [sub.replace('  ', '') for sub in so2]
        # so3 = [sub.replace('  ', '') for sub in so3]
        # for i in range(0, len(so1)):
        #     wq += 1
        #     row_num = s2.max_row
        #     s2.cell(row=row_num + 1, column=1).value = wq
        #     s2.cell(row=row_num + 1, column=2).value = hg[0]
        #     s2.cell(row=row_num + 1, column=3).value = so1[i]
        #     s2.cell(row=row_num + 1, column=4).value = so2[i]
        #     s2.cell(row=row_num + 1, column=5).value = so3[i]
        # os.rename(os.getcwd() + '/universal_sompo/attachments_' + str(sys.argv[6]) + '/' + onlyfiles[t],
        #           os.getcwd() + '/universal_sompo/attachments_' + str(sys.argv[6]) + '/' + hg[0] + '.pdf')
        # except Exception as e:
        #     s1.cell(row=t + 2, column=1).value = 'error'
        #     os.rename(os.getcwd() + '/universal_sompo/attachments_' + str(sys.argv[6]) + '/' + onlyfiles[t],
        #               os.getcwd() + '/universal_sompo/attachments_' + str(sys.argv[6]) + '/' + hg[0] + '.pdf')

    # print(po)
    print("Done")
    wbk.save(wbkName)
    wbk.close
    wbkName = 'count/count.xlsx'
    wbk = openpyxl.load_workbook(wbkName)
    s1 = wbk.worksheets[0]

    row_ = s1.max_row + 1
    s1.cell(row=row_, column=1).value = 'universal sompo'
    s1.cell(row=row_, column=2).value = str(sys.argv[6])
    s1.cell(row=row_, column=3).value = len(fg)
    s1.cell(row=row_, column=4).value = len(onlyfiles)
    s1.cell(row=row_, column=5).value = len(repeat)
    wbk.save(wbkName)
    wbk.close
    s2_data
    pass
    subprocess.run(["python", "updation.py", "1", "max", "9", " "])
except:
    log_exceptions()