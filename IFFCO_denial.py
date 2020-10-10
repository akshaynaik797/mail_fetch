import os
import sys, re
import struct, time
import subprocess
from datetime import date
import datetime
import openpyxl
import pdftotext
import time
import requests

now = datetime.datetime.now()
'''
wbkName = 'log file.xlsx'
wbk= openpyxl.load_workbook(wbkName)
s1=wbk.worksheets[0]
s2=wbk.worksheets[1]
row_count_1 = s2.max_row
s2.cell(row_count_1+1, column=1).value=sys.argv[2]
s2.cell(row_count_1+1, column=2).value=sys.argv[3]
s2.cell(row_count_1+1, column=3).value=sys.argv[4]
s2.cell(row_count_1+1, column=5).value=now
s2.cell(row_count_1+1, column=7).value=sys.argv[5]
s2.cell(row_count_1+1, column=8).value=sys.argv[6]
'''

subprocess.run(["python", "updation.py", "1", "max1", "1", sys.argv[2]])
subprocess.run(["python", "updation.py", "1", "max", "2", sys.argv[3]])
subprocess.run(["python", "updation.py", "1", "max", "3", sys.argv[4]])
subprocess.run(["python", "updation.py", "1", "max", "5", str(now)])
subprocess.run(["python", "updation.py", "1", "max", "7", sys.argv[5]])
subprocess.run(["python", "updation.py", "1", "max", "8", sys.argv[6]])

with open(sys.argv[1], "rb") as f:
    pdf = pdftotext.PDF(f)

with open('IFFCO/output.txt', 'w') as f:
    f.write(" ".join(pdf))
with open('IFFCO/output.txt', 'r') as myfile:
    f = myfile.read()
try:
    badchars = ('/', ',', ':', '-')
    datadict = {}
    regexdict = {'preid': [r"(?<=Cashless Authorization No).*(?=\.)"],
                 'memid': [r"(?<=MemberID:).*"],
                 'pname': [r"(?<=of).*(?=, *MemberID)"],
                 'polno': [r"(?<=Policy).*(?=of)"],
                 }

    for i in regexdict:
        for j in regexdict[i]:
            data = re.compile(j).search(f)
            if data is not None:
                temp = data.group().strip()
                for k in badchars:
                    temp = temp.replace(k, "")
                datadict[i] = temp.strip()
                break
            datadict[i] = ""
    a = 1

    subprocess.run(["python", "updation.py", "1", "max", "9", 'Yes'])
    subprocess.run(["python", "updation.py", "1", "max", "10", 'NA'])

    try:
        subprocess.run(
            ["python", "test_api.py", datadict["preid"], "", datadict["polno"], '', 'Denail', sys.argv[6], sys.argv[1],
             '', datadict["memid"], datadict["pname"]])
        '''wbk= openpyxl.load_workbook(wbkName)
        s2=wbk.worksheets[1]
        s2.cell(row_count_1+1, column=11).value='YES'
        '''
        subprocess.run(["python", "updation.py", "1", "max", "11", 'Yes'])
    except Exception as e:
        # s2.cell(row_count_1+1, column=11).value='NO'
        subprocess.run(["python", "updation.py", "1", "max", "11", 'No'])
except Exception as e:
    # s2.cell(row_count_1+1, column=9).value='No'
    # s2.cell(row_count_1+1, column=11).value='NO'
    subprocess.run(["python", "updation.py", "1", "max", "9", 'Yes'])
    subprocess.run(["python", "updation.py", "1", "max", "11", 'No'])
now = datetime.datetime.now()
# s2.cell(row_count_1+1, column=6).value=now
# wbk.save(wbkName)
subprocess.run(["python", "updation.py", "1", "max", "6", str(now)])
