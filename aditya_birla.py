import re

import pandas as pd
from email.mime.text import MIMEText
import smtplib
import time
import openpyxl
import imaplib
import sys
import email
import camelot
import PyPDF2
import csv
import xlsxwriter
# from xlrd import open_workbook
# import xlwt
import os
import glob
import os.path
import xlrd
from os import listdir
from os import path
import pdftotext
from os.path import isfile, join
from html.parser import HTMLParser
import pdfkit
import pandas as pd
import html2text
import subprocess

from decode_error import check_subject, read_from_delete
from make_log import log_exceptions

try:
    subprocess.run(["python", "updation.py", "1", "max", "9", "X"])
    fg = []
    path_wkhtmltopdf = r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe'
    config = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf)
    # config = pdfkit.configuration(wkhtmltopdf='/usr/bin/wkhtmltopdf')

    def read_email_from_gmail():
        SMTP_SERVER = str(sys.argv[5])
        mail = imaplib.IMAP4_SSL(SMTP_SERVER)
        e_id = str(sys.argv[1])
        pswd = str(sys.argv[2])
        srt = str(sys.argv[3])
        stp = str(sys.argv[4])

        mail_uid = str(sys.argv[7])

        mail.login(user=e_id, password=pswd)
        mail.select("inbox", readonly=True)
    ###############################################<
        if mail_uid == -1:
            type, data = mail.search(None, '(SUBJECT "Claims settlement-NEFT" since ' + srt + ' before ' + stp + ')')
            ids = data[0]
            id_list = ids.split()
        else:
            ids = mail_uid  # data is a list.
            # accept id from outside and put in id_list akshay var name = id

            id_list = []  # ids is a space separated string
            id_list.append(ids)
    ###############################################>
        # print(id_list)
        for i in range(0, len(id_list)):
            latest_email_id = id_list[i]  # get the latest
            result, data = mail.fetch(latest_email_id, "(RFC822)")
            ##################################################ak
            try:
                # raw_email = data[0][1].decode('utf-8')
                try:
                    raw_email = data[0][1].decode('utf-8')
                except UnicodeDecodeError:
                    try:
                        raw_email = data[0][1].decode('ISO-8859-1')
                    except UnicodeDecodeError:
                        try:
                            raw_email = data[0][1].decode('ascii')
                        except UnicodeDecodeError:
                            pass
                email_message = email.message_from_string(raw_email)
                subject = email_message['Subject']
                result, sys.argv[8] = check_subject(subject, sys.argv[8], mail)
                if result == 'Changed':
                    # raise Exception('subject not matched')
                    raise Exception('subject not matched', )
            except:
                try:
                    log_exceptions(syssubject=sys.argv[8], subject=subject, error='subject not matched')
                except:
                    pass
                if result != 'OK':
                    data = {'server': SMTP_SERVER,
                            'hospmail': e_id,
                            'pass': pswd,
                            'subject': sys.argv[8]}
                    try:
                        data = read_from_delete(data)
                        if data == None:
                            raise Exception("Not found")
                    except:
                        log_exceptions(msg='not found in deleted', subject=sys.argv[8])
            ##################################################akend
            raw_email = data[0][1].decode('utf-8')
            email_message = email.message_from_string(raw_email)
            # if path.exists(r'/home/shivam/Desktop/vnu_scripts/Paramount/email.html'):
            # os.remove(r'email.html')
            # Body details
            # if email_message['Subject'] not in fg:

            for part in email_message.walk():
                # print(part.get_content_type())
                if part.get_content_type() == "text/html":
                    # print('hi')
                    body = part.get_payload(decode=True)
                    file_name = "aditya_birla/email.html"
                    output_file = open(file_name, 'w')
                    output_file.write("Body: %s" % (body.decode('utf-8')))
                    output_file.close()
                    with open('aditya_birla/email.html', 'r') as myfile:
                        f = myfile.read()
                    x1 = f.find('Claim registration number:') + 36
                    g = f[x1:]
                    x2 = g.find('<') + x1
                    b = f[x1:x2]

                    pdfkit.from_file('aditya_birla/email.html',
                                     'aditya_birla/attachments_' + str(sys.argv[6]) + '/' + b + '.pdf',
                                     configuration=config)
                else:
                    continue
            fg.append(email_message['Subject'])


    mypath = os.getcwd() + '/aditya_birla'
    if not path.exists(mypath):
        os.mkdir(mypath)
    if not path.exists(mypath + '/attachments_' + str(sys.argv[6])):
        os.mkdir(mypath + '/attachments_' + str(sys.argv[6]))
    mypath = os.getcwd() + '/aditya_birla/attachments_' + str(sys.argv[6]) + '/'
    for filename in os.listdir(mypath):
        file_path = os.path.join(mypath, filename)
        if os.path.isfile(file_path) or os.path.islink(file_path):
            os.unlink(file_path)
    read_email_from_gmail()
    onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
    if path.exists(r'aditya_birla/aditya' + str(sys.argv[6]) + '.xlsx'):
        os.remove(r'aditya_birla/aditya' + str(sys.argv[6]) + '.xlsx')
    po = []
    wbkName = 'aditya_birla/aditya' + str(sys.argv[6]) + '.xlsx'
    wbk = openpyxl.Workbook()
    wbk.create_sheet('1')
    wbk.create_sheet('Sheet3')
    s1 = wbk.worksheets[0]
    s2 = wbk.worksheets[1]
    s3 = wbk.worksheets[2]
    for t in range(0, len(onlyfiles)):
        tables = camelot.read_pdf(mypath+onlyfiles[t],pages = 'all', line_scale=100)
        tables.export('aditya_birla/foo1.xls', f='excel')
        loc = ("aditya_birla/foo1.xls")
        with open(mypath + onlyfiles[t], "rb") as f:
            pdf = pdftotext.PDF(f)

        with open('aditya_birla/output.txt', 'w') as f:
            f.write(" ".join(pdf))
        with open('aditya_birla/output.txt', 'r') as myfile:
            f = myfile.read()
        f = f.replace('\n', '$$ ')
        # print(f)

        wb = xlrd.open_workbook(loc)
        sh1 = ['Sum Insured', 'Claimed amount (Rs.)', 'AL Amount(in case of cashless)', 'Approved amount (Rs.)',
               'Deduction amount (Rs.)', 'Hospital Amount', 'TDS', 'Discount Amount', 'Reason for deduction',
               'Amount Utilised', 'CoPay']
        for i in range(0, len(sh1)):
            s2.cell(row=1, column=i + 3).value = sh1[i]
        sh2 = ['Policy Number', 'Member id', 'Patient Name', 'Policy Holder', 'Payee Bank name', 'Payee account number',
               'Amount of transfer', 'UTR number', 'Diagnosis', 'doa', 'dod', 'Deduction amount', 'discount',
               'transaction date']
        for i in range(0, len(sh2)):
            s1.cell(row=1, column=i + 3).value = sh2[i]
        sh3 = ['Sr', 'CCN', 'Deduction amount', 'Deduction reason']
        for i in range(0, len(sh3)):
            s3.cell(row=1, column=i + 1).value = sh3[i]
        hg = []
        w = f.find('Policy Number') + 15
        g = f[w:]
        u = g.find('$$') + w
        hg.append(f[w:u])

        w1 = f.find('Member id') + 11
        g = f[w1:]
        u1 = g.find('$$') + w1
        hg.append(f[w1:u1])

        w2 = f.find('Patient Name') + 14
        g = f[w2:]
        u2 = g.find('$$') + w2
        hg.append(f[w2:u2])

        w3 = f.find('Policy Holder') + 14
        g = f[w3:]
        u3 = g.find('$$') + w3
        hg.append(f[w3:u3])

        w4 = f.find('Payee Bank name') + 17
        g = f[w4:]
        u4 = g.find('$$') + w4
        hg.append(f[w4:u4])

        w5 = f.find('Payee account number') + 22
        g = f[w5:]
        u5 = g.find('$$') + w5
        hg.append(f[w5:u5])

        w6 = f.find('Amount of transfer') + 25
        g = f[w6:]
        u6 = g.find('$$') + w6
        hg.append(f[w6:u6])

        w7 = f.find('UTR number') + 12
        g = f[w7:]
        u7 = g.find('$$') + w7
        hg.append(f[w7:u7])

        w8 = f.find('Ailment Name') + 14
        g = f[w8:]
        u8 = g.find('Please note') + w8
        hg.append(f[w8:u8])

        w8 = f.find('Date of Admission:') + 19
        g = f[w8:]
        u8 = g.find('$$') + w8
        hg.append(f[w8:u8])

        w8 = f.find('Date of Discharge:') + 19
        g = f[w8:]
        u8 = g.find('$$') + w8
        hg.append(f[w8:u8])

        w8 = f.find('Deduction amount') + 23
        g = f[w8:]
        u8 = g.find('$$') + w8
        hg.append(f[w8:u8])

        w8 = f.find('Date of transfer :') + 19
        g = f[w8:]
        u8 = g.find('$$') + w8
        hg.append(f[w8:u8])

        w9 = f.find('Claim registration number') + 26
        g = f[w9:]
        u9 = g.find('$$') + w9
        ccn = (f[w9:u9])
        hg = [sub.replace('  ', '') for sub in hg]
        hg = [sub.replace('$$ ', '') for sub in hg]
        # print(hg)
        transaction_date = hg[12]
        for i in range(0, len(hg)):
            s1.cell(row=t + 2, column=i + 3).value = hg[i]

        for wd in wbk.worksheets[:2]:
            wd.cell(row=1, column=1).value = 'Sr. No.'
            wd.cell(row=1, column=2).value = 'CCN'
            wd.cell(row=t + 2, column=1).value = t + 1
            wd.cell(row=t + 2, column=2).value = ccn

        # wb = xlrd.open_workbook('foo1.xls', on_demand=True)
        sheetno = len(wb.sheet_names())
        sheet_with_details = 0
        for i in range(sheetno):
            worksheet = wb.sheet_by_index(i)
            a = worksheet.cell_value(1, 1)
            if 'Details' in a:
                sheet_contains_details = i
                break

        sheet_2 = wb.sheet_by_index(sheet_contains_details)
        sheet_2.cell_value(0, 0)
        b = []

        for i in range(2, sheet_2.nrows):
            b.append(sheet_2.cell_value(i, 2))

    ############################################################<
    # code block returns list of amounts = clean[] in Approved amount (Rs.) column

        for i, elem in enumerate(b):
            if 'MOU' in elem or 'mou' in elem:
                text = b[i]

        templist, clean = [], []
        temp = text.split('\t')
        for i in temp:
            if '/-' in i:
                templist.append(i)
        for i in templist:
            num = i.split('/')[0]
            if 'rs' in num:
                num = num.strip(',rs.')
                num = num.replace('s', '')
            clean.append(int(num))
        rd = text
    ###########################################################>
        rd = rd.replace('Rs', 'rs')
        rd = rd.replace('RS', 'rs')
        rd = rd.replace('\t', ' ')

        reason = rd.split(',rs')
        deduct = []
        deduct_res = []
        for i in reason:
            w8 = i.find('rs') + 2
            g = i[w8:]
            u8 = g.find('/-') + w8
            # deduct.append(i[w8:u8])
            w8 = i.find('/-') + 2
            deduct_res.append(i[w8:])
            if '' in deduct_res:
                deduct_res.remove('')
        # print(deduct,deduct_res,reason)
    #############################################################<
        # send clean array to deducts coloumn
        deduct.extend(clean)
    #############################################################>

    #######################################################<
    #fixing s2 sheet

        with open(sys.argv[0].strip('.py')+'/output.txt') as f:
            txt = f.read()
        x = re.search(r"Copay +\d+", txt)
        x1 = x.group()
        x2 = re.search(r"\d+", x1)
        cp = x2.group()

        temp = b
        temp[6] = temp[11]
        temp[7] = temp[5]
        temp[8] = temp[9]
        temp[9] = temp[13]
        temp[5] = temp[12]
        temp[10] = cp

    ########################################################>
        for i in range(0, len(b)):
            # s2 is foo1.xls -> 1
            s2.cell(row=t + 2, column=i + 3).value = temp[i]
    #########################################################<
    # fixing s1 sheet
        hg[11] = temp[4]
        hg[12] = clean[0]
        hg.append(transaction_date)
        for i in range(0, len(hg)):
            s1.cell(row=t + 2, column=i + 3).value = hg[i]
    #########################################################>
        for i in range(0, len(deduct)):
            if (deduct_res[i].find('MOU Discount') != -1 or deduct_res[i].find('MOU discount') != -1):
                discount = deduct[i]
                s1.cell(row=t + 2, column=14).value = float(hg[-2]) - float(discount)
                s1.cell(row=t + 2, column=15).value = discount
            if (deduct[i] != ''):
                #########################################################<
                # start srno of s3 with 2
                # row = s3.max_row
                #########################################################>
                row = s3.max_row + 1
                s3.cell(row=row, column=1).value = row + 1
                s3.cell(row=row, column=2).value = ccn
                s3.cell(row=row, column=3).value = deduct[i]
                s3.cell(row=row, column=4).value = deduct_res[i]
        s1.cell(row=t + 2, column=16).value = hg[-1]

        # try:
        #     for i in range(2, sheet_2.nrows):
        #         b.append(sheet_2.cell_value(i, 2))
        #     rd = b[8]
        #     rd = rd.replace('Rs', 'rs')
        #     rd = rd.replace('RS', 'rs')
        #     rd = rd.replace('\t', ' ')
        #
        #     reason = rd.split(',rs')
        #     deduct = []
        #     deduct_res = []
        #     for i in reason:
        #         w8 = i.find('rs') + 2
        #         g = i[w8:]
        #         u8 = g.find('/-') + w8
        #         deduct.append(i[w8:u8])
        #         w8 = i.find('/-') + 2
        #         deduct_res.append(i[w8:])
        #     # print(deduct,deduct_res,reason)
        #     for i in range(0, len(b)):
        #         s2.cell(row=t + 2, column=i + 3).value = b[i]
        #     for i in range(0, len(deduct)):
        #         if (deduct_res[i].find('MOU Discount') != -1 or deduct_res[i].find('MOU discount') != -1):
        #             discount = deduct[i]
        #             s1.cell(row=t + 2, column=14).value = float(hg[-2]) - float(discount)
        #             s1.cell(row=t + 2, column=15).value = discount
        #         if (deduct[i] != ''):
        #             row = s3.max_row + 1
        #             s3.cell(row=row, column=1).value = row + 1
        #             s3.cell(row=row, column=2).value = ccn
        #             s3.cell(row=row, column=3).value = deduct[i]
        #             s3.cell(row=row, column=4).value = deduct_res[i]
        #     s1.cell(row=t + 2, column=16).value = hg[-1]
        # except:
        #     print("Done")
        #     wbk.save(wbkName)
        #     wbk.close()
        #     Name = 'count/count.xlsx'
        #     wbk = openpyxl.Workbook()
        #     s1 = wbk.worksheets[0]
        #     s1.cell(row=1, column=1).value = 'insurance id'
        #     s1.cell(row=1, column=2).value = 'mail count'
        #     s1.cell(row=1, column=3).value = 'attachments count'
        #     s1.cell(row=2, column=1).value = 'aditya birla'
        #     s1.cell(row=2, column=2).value = len(fg)
        #     s1.cell(row=2, column=3).value = len(onlyfiles)
        #     wbk.save(Name)
        #     wbk.close()
        #     print("hi")


    print("Done")
    wbk.save(wbkName)
    wbk.close()
    Name = 'count/count.xlsx'
    wbk = openpyxl.Workbook()
    s1 = wbk.worksheets[0]
    s1.cell(row=1, column=1).value = 'insurance id'
    s1.cell(row=1, column=2).value = 'mail count'
    s1.cell(row=1, column=3).value = 'attachments count'
    s1.cell(row=2, column=1).value = 'aditya birla'
    s1.cell(row=2, column=2).value = len(fg)
    s1.cell(row=2, column=3).value = len(onlyfiles)
    wbk.save(Name)
    wbk.close()
    subprocess.run(["python", "updation.py", "1", "max", "9", " "])
except:
    log_exceptions()