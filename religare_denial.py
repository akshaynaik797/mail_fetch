import os
import sys
import struct, time
import subprocess
from datetime import date
import datetime
import openpyxl
import pdftotext
import time
import requests
now = datetime.datetime.now()

subprocess.run(["python", "updation.py","1","max1","1",sys.argv[2]])
subprocess.run(["python", "updation.py","1","max","2",sys.argv[3]])
subprocess.run(["python", "updation.py","1","max","3",sys.argv[4]])
subprocess.run(["python", "updation.py","1","max","5",str(now)])
subprocess.run(["python", "updation.py","1","max","7",sys.argv[5]])
subprocess.run(["python", "updation.py","1","max","8",sys.argv[6]])

with open(sys.argv[1], "rb") as f:
        pdf = pdftotext.PDF(f)


# with open(r"C:\Users\91798\Desktop\varun sir\trial -live\religare\attachments_pdf_denial\RejectionLetter_20818.pdf", "rb") as f:
#         pdf = pdftotext.PDF(f)


with open('religare/output1.txt', 'w') as f:
        f.write(" ".join(pdf))     
with open('religare/output1.txt', 'r') as myfile:
        f = myfile.read()

try:
        hg=[]
        if f.find('AL No')!= -1:
                w=f.find('AL No')+len('AL No')
                g=f[w:]
                u=g.find('\n')+w
                g=f[w:u]
                hg.append(g.strip())

        elif f.find('Intimation No :')!= -1:
                w=f.find('Intimation No :')+len('Intimation No :')
                g=f[w:]
                u=g.find('\n')+w
                g=f[w:u]
                hg.append(g.strip())

        # if g.find('-'):
        #       u=g.find('-')+w
        # else:
        #       u=g.find('-')+w

        w=f.find('Pre-Auth for')+len('Pre-Auth for')
        g=f[w:]  
        u=g.find('\n')+w
        g=f[w:u]
        g=g.replace('.','')
        hg.append(g.strip())

        w=f.find('Member ID')+len('Member ID')
        g=f[w:]
        u=g.find('Provisional')+w
        g=f[w:u]
        hg.append(g.strip())

        w=f.find('Policy Number')+len('Policy Number')
        g=f[w:]
        u=g.find('\n')+w
        g=f[w:u]
        hg.append(g.strip())

        hg=[sub.replace(':','') for sub in hg]  
        hg=[sub.replace('  ','') for sub in hg]
        hg=[sub.replace('Rs.','') for sub in hg]

        subprocess.run(["python", "updation.py","1","max","9",'Yes'])
        subprocess.run(["python", "updation.py","1","max","10",'NA'])
        
        try:
                subprocess.run(["python", "test_api.py",hg[0],'',hg[3],'','Denial',sys.argv[6],sys.argv[1],'',hg[2],hg[1]])
                '''wbk= openpyxl.load_workbook(wbkName)
                s2=wbk.worksheets[1]
                s2.cell(row_count_1+1, column=11).value='YES'
                '''
                subprocess.run(["python", "updation.py","1","max","11",'Yes'])  
        except Exception as e:
                #s2.cell(row_count_1+1, column=11).value='NO'
                subprocess.run(["python", "updation.py","1","max","11",'No'])
except Exception as e:
        #s2.cell(row_count_1+1, column=9).value='No'
        #s2.cell(row_count_1+1, column=11).value='NO'
        subprocess.run(["python", "updation.py","1","max","9",'Yes'])
        subprocess.run(["python", "updation.py","1","max","11",'No'])
now = datetime.datetime.now()
#s2.cell(row_count_1+1, column=6).value=now
#wbk.save(wbkName)
subprocess.run(["python", "updation.py","1","max","6",str(now)])

        
