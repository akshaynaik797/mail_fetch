import re
import pandas as pd
from email.mime.text import MIMEText
import smtplib
import time
import imaplib
import email
import camelot
import sys
import PyPDF2
import csv
import xlsxwriter
from xlrd import open_workbook
import xlwt
import os
import glob
import os.path
import xlrd
from os import listdir
from os import path
from os.path import isfile, join
from html.parser import HTMLParser
import pdfkit
import pandas as pd
import html2text
import subprocess

from decode_error import check_subject, read_from_delete
from make_log import log_exceptions
if path.exists(r'Paramount/Paramount'+str(sys.argv[6])+'.xlsx'):
	os.remove(r'Paramount/Paramount'+str(sys.argv[6])+'.xlsx')
import openpyxl
po=[]
repeat=[]
fg=[]

try:
	subprocess.run(["python", "updation.py", "1", "max", "9", "X"])

	def read_email_from_gmail():
		b=0
		SMTP_SERVER = str(sys.argv[5])
		mail = imaplib.IMAP4_SSL(SMTP_SERVER)
		e_id =str(sys.argv[1])
		pswd=str(sys.argv[2])
		srt =str(sys.argv[3])
		stp=str(sys.argv[4])
		if srt[-4:]<='2020':
			srt='1-jan-2020'
		mail.login(user =e_id, password =pswd)
		mail.select("inbox", readonly=True)
		###############################################<
		mail_uid = str(sys.argv[7])
		if mail_uid == -1:
			type, data = mail.search(None, '(SUBJECT "Neft payment advice generated" since ' + srt + ' before ' + stp + ')')
			ids = data[0]
			id_list = ids.split()
		else:
			ids = mail_uid  # data is a list.
			# accept id from outside and put in id_list akshay var name = id

			id_list = []  # ids is a space separated string
			id_list.append(ids)
		###############################################>
		# type, data = mail.search(None,'(SUBJECT "Neft payment advice generated" since '+srt+' before '+stp+')' )
		# ids = data[0] # data is a list.
		# id_list = ids.split() # ids is a space separated string
		#print(id_list)
		for i in range (0,len(id_list)):
			latest_email_id = id_list[i] # get the latest
			result, data = mail.fetch(latest_email_id, "(RFC822)")
			##################################################ak
			try:
				# raw_email = data[0][1].decode('utf-8')
				try:
					raw_email = data[0][1].decode('utf-8')
				except UnicodeDecodeError:
					try:
						raw_email = data[0][1].decode('ISO-8859-1')
					except UnicodeDecodeError:
						try:
							raw_email = data[0][1].decode('ascii')
						except UnicodeDecodeError:
							pass
				email_message = email.message_from_string(raw_email)
				subject = email_message['Subject']
				result, sys.argv[8] = check_subject(subject, sys.argv[8], mail)
				if result == 'Changed':
					# raise Exception('subject not matched')
					raise Exception('subject not matched', )
			except:
				try:
					log_exceptions(syssubject=sys.argv[8], subject=subject, error='subject not matched')
				except:
					pass
				if result != 'OK':
					data = {'server': SMTP_SERVER,
							'hospmail': e_id,
							'pass': pswd,
							'subject': sys.argv[8]}
					try:
						data = read_from_delete(data)
						if data == None:
							raise Exception("Not found")
					except:
						log_exceptions(msg='not found in deleted', subject=sys.argv[8])
			##################################################akend
			raw_email = data[0][1].decode('utf-8')
			email_message = email.message_from_string(raw_email)
			#if path.exists(r'/home/shivam/Desktop/vnu_scripts/Paramount/email.html'):
				#os.remove(r'email.html')
			# Body details
			if email_message['Subject'] not in fg:
				b+=1
				for part in email_message.walk():
					if part.get_content_type() == "text/html":
						#print('hi')
						body = part.get_payload(decode=True)
						file_name = "Paramount/email.html"
						output_file = open(file_name, 'w')
						output_file.write("Body: %s" %(body.decode('utf-8')))
						output_file.close()
						file_name = 'Paramount/attachments_mail_'+str(sys.argv[6])+'/output'+'%s'%(b)+'.html'
						output_file = open(file_name, 'w')
						output_file.write("Body: %s" %(body.decode('utf-8')))
						output_file.close()
						pass
						#pdfkit.from_file('Paramount/email.html', 'Paramount/attachments_pdf/'+str(i)+'.pdf')

					else:
						continue
			else:
				repeat.append(email_message['Subject'])

			for t, df in enumerate(pd.read_html(r'Paramount/email.html')):
				writer = pd.ExcelWriter('Paramount/attachments_'+str(sys.argv[6])+'/output'+'%s'%(b)+'.xlsx',engine='xlsxwriter')

				if(t==1):
					df_1=df
				if(t==2):
					df_2=df
				if(t==3):
					df_3=df
				if(t==4):
					df_4=df
				if(t==5):
					df_5=df
				if(t==6):
					df_6=df
				if(t==7):
					df_7=df
			df_1.to_excel(writer, '1')
			df_2.to_excel(writer, '2')
			df_3.to_excel(writer, '3')
			df_4.to_excel(writer, '4')
			df_5.to_excel(writer, '5')
			df_6.to_excel(writer, '6')
			df_7.to_excel(writer, '7')
			writer.save()
			fg.append(email_message['Subject'])
					#df.to_csv('attachments/'+'%s'%i+'myfile_%s.csv'% t)
	sheet_list=[]
	mypath=os.getcwd()+'/Paramount'
	if not path.exists(mypath):
		os.mkdir(mypath)
	if not path.exists(mypath+'/attachments_'+str(sys.argv[6])):
		os.mkdir(mypath+'/attachments_'+str(sys.argv[6]))
	if not path.exists(mypath+'/attachments_mail_'+str(sys.argv[6])):
		os.mkdir(mypath+'/attachments_mail_'+str(sys.argv[6]))

	mypath = os.getcwd()+'/Paramount/attachments_mail_'+str(sys.argv[6])+'/'
	for filename in os.listdir(mypath):
		file_path = os.path.join(mypath, filename)
		if os.path.isfile(file_path) or os.path.islink(file_path):
			os.unlink(file_path)

	mypath = os.getcwd()+'/Paramount/attachments_'+str(sys.argv[6])+'/'
	for filename in os.listdir(mypath):
		file_path = os.path.join(mypath, filename)
		if os.path.isfile(file_path) or os.path.islink(file_path):
			os.unlink(file_path)

	read_email_from_gmail()
	wbkName = 'Paramount/Paramount'+str(sys.argv[6])+'.xlsx'
	wbk = openpyxl.Workbook()
	wbk.create_sheet('1')
	#wbk.create_sheet('Sheet3')
	s1=wbk.worksheets[0]
	s3=wbk.worksheets[1]

	#CCN='20439541'
	onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
	#print(onlyfiles)
	for t in range(0,len(onlyfiles)):
		#print(t+1)
		try:
			def select_sheet(wks,sheet_name):


				if not sheet_name in wks.sheetnames:
					wks.create_sheet(sheet_name)
					#print(sheet_name)

				wks.save('Paramount/Paramount'+str(sys.argv[6])+'.xlsx')




			def select_column(wks,s,ro):
				sheet = wks.worksheets[s]

				max_col = sheet.max_column
				s=[]
				for i in range(1, max_col + 1):
					cell_obj = sheet.cell(row = 1, column = i)
					s.append(cell_obj.value)
				#print(s)

			loc = (r'Paramount/attachments_'+str(sys.argv[6])+'/output'+'%s'%(t+1)+'.xlsx')
			wb = xlrd.open_workbook(loc)
			sh1=['Member PHS ID','Name of the Patient','Name of Insurance co.','Policy No','Employee No','Group Name']
			sh2=['Payment To','Insurer UTR No.','Date of Payment to Hospital']
			sh3=['Date of Admission','Date of Discharge','Amount Claimed','Claim Amt Settled','Amt Paid to Hospital','Amt Paid to Member']

			sheet_4 = wb.sheet_by_index(1)
			sheet_4.cell_value(0, 0)
			jf=[]
			kl=[]
			for i in range(2,sheet_4.nrows):
				if  sheet_4.cell_value(i,1) in sh1:
					jf.append(sheet_4.cell_value(i,1))
					kl.append(sheet_4.cell_value(i,3))
			#print(jf,kl)

			sheet_5 = wb.sheet_by_index(2)
			sheet_5.cell_value(0, 0)
			rt=[]
			ty=[]
			jndf=[]
			for i in range(2,sheet_5.nrows):
				jndf.append(sheet_5.cell_value(i,1))
				if  sheet_5.cell_value(i,1) in sh2:
					rt.append(sheet_5.cell_value(i,1))
					ty.append(sheet_5.cell_value(i,3))
			#print(jndf)
			#print(rt,ty)

			sheet_6 = wb.sheet_by_index(3)
			sheet_6.cell_value(0, 0)
			q=[]
			w=[]
			for i in range(2,sheet_6.nrows):
				if  sheet_6.cell_value(i,1) in sh3:
					q.append(sheet_6.cell_value(i,1))
					w.append(sheet_6.cell_value(i,3))
				if  sheet_6.cell_value(i,1) =='CCN No.':
					ccn=sheet_6.cell_value(i,3)
			#print(q,w)
			for wd in wbk.worksheets[:2]:
				wd.cell(row=1, column=1).value = 'Sr. No.'
				wd.cell(row=1, column=2).value = 'CCN'
				wd.cell(row=t+2, column=1).value = t+1
				wd.cell(row=t+2, column=2).value = ccn
			for i in range (0,len(jf)):
				s1.cell(row=1, column=i+3).value = jf[i]
				s1.cell(row=t+2, column=i+3).value = kl[i]
			for x in range (0,len(rt)):
				s1.cell(row=1, column=x+i+3).value = rt[x]
				s1.cell(row=t+2, column=x+i+3).value = ty[x]

			w = [pl.replace('Rs.', ' ') for pl in w]
			w = [pl.replace('/-', ' ') for pl in w]
			#print(w)
			for e in range (0,len(q)):
				s1.cell(row=1, column=e+x+i+3).value = q[e]
				s1.cell(row=t+2, column=e+x+i+3).value = w[e]
			s1.cell(row=1, column=16).value = rt[-1]
			s1.cell(row=t+2, column=16).value = ty[-1]
			s1.cell(row=1, column=17).value = jf[-1]
			s1.cell(row=t+2, column=17).value = kl[-1]
			s=[]
			d=[]
			sheet_3 = wb.sheet_by_index(5)
			sheet_3.cell_value(0, 0)


			for i in range(2,9):
				s.append(sheet_3.cell_value(0,i))
				d.append(sheet_3.cell_value(1, i))
			#print(s,d)

			for i in range (0,len(s)):
				s3.cell(row=1, column=i+3).value = s[i]
				s3.cell(row=t+2, column=i+3).value = d[i]


			s_v=[]

			sheet_2 = wb.sheet_by_index(6)
			sheet_2.cell_value(0, 0)
			b=[]
			p=[]
			np=[]
			r=[]
			for i in range(1,sheet_2.nrows):
				s_v.append(sheet_2.cell_value(i,3))
				b.append(sheet_2.cell_value(i, 2))
				p.append(sheet_2.cell_value(i,4))
				np.append(sheet_2.cell_value(i,5))
				r.append(sheet_2.cell_value(i, 6))
			for i in range (len(s_v)):
				select_sheet(wbk,s_v[i])
			ro=[]
			ro.append(sheet_2.row_values(0))
			ro=ro[0][2:]
			#print(b)
			xls = xlrd.open_workbook(r'Paramount/Paramount'+str(sys.argv[6])+'.xlsx', on_demand=True)
			sheet_list=xls.sheet_names()
			sheet_list = [x.replace('sheet','').replace('sheet1','').replace('sheet2','') for x in sheet_list]
			#print(sheet_list)
			for i in range (len(s_v)):
				sheet_name=s_v[i]
				for y in range (len(sheet_list)):
					if sheet_name in sheet_list[y]:
						for k in range (0,len(ro)):
							sheet = wbk.worksheets[y]
							sheet.cell(row=1, column=1).value = 'Sr. No.'
							sheet.cell(row=1, column=2).value = 'CCN'
							sheet.cell(row=1, column=k+3).value = ro[k]
						#print(sheet)
						row_num=sheet.max_row
						sheet.cell(row=row_num+1, column=1).value = t+1
						sheet.cell(row=row_num+1, column=2).value = ccn
						sheet.cell(row=row_num+1, column=3).value = b[i]
						sheet.cell(row=row_num+1, column=4).value = s_v[i]
						sheet.cell(row=row_num+1, column=5).value = p[i]
						sheet.cell(row=row_num+1, column=6).value = np[i]
						sheet.cell(row=row_num+1, column=7).value = r[i]

			os.rename(os.getcwd()+'/Paramount/attachments_mail_'+str(sys.argv[6])+'/output'+'%s'%(t+1)+'.html',os.getcwd()+'/Paramount/attachments_mail_'+str(sys.argv[6])+'/'+ccn+'.html')
		except Exception as e:
			log_exceptions()
			s1.cell(row=t+2, column=1).value = 'error'
			os.rename(os.getcwd()+'/Paramount/attachments_mail_'+str(sys.argv[6])+'/output'+'%s'%(t+1)+'.html',os.getcwd()+'/Paramount/attachments_mail_'+str(sys.argv[6])+'/'+ccn+'.html')

	for y in range (len(sheet_list)):
		sheet = wbk.worksheets[y]
		sheet.title = 'sheet'+str(y+2)
	print("Done")
	wbk.save(wbkName)
	wbk.close()

	##########################################################ak
	try:
		wb = openpyxl.load_workbook(wbkName)
		b = wb.worksheets[0]
		c, r = b.max_row, b.max_column
		movecells = ((3, 4), (4, 5), (5, 6), (6, 7), (17, 3), (7, 17))

		isdigit = re.compile(r'\d')
		for i in range(2, b.max_row+1):
			if isdigit.match(b.cell(i,3).value):
				temprow = []
				for j in movecells:
					temprow.append(b.cell(i, j[0]).value)
				if len(movecells) == len(temprow):
					for j, data in zip(movecells, temprow):
						b.cell(i, j[1]).value = data

		f = b.cell(2,4).value
		b.cell(2,4).value = f+f
		wb.save(wbkName)
		wb.close()
	except:
		log_exceptions()
	##########################################################akend

	wbkName = 'count/count.xlsx'
	wbk= openpyxl.load_workbook(wbkName)
	s1=wbk.worksheets[0]

	row_=s1.max_row+1
	s1.cell(row=row_, column=1).value = 'Paramount'
	s1.cell(row=row_, column=2).value = str(sys.argv[6])
	s1.cell(row=row_, column=3).value = len(fg)
	s1.cell(row=row_, column=4).value = len(onlyfiles)
	s1.cell(row=row_, column=5).value = len(repeat)
	wbk.save(wbkName)
	pass
	subprocess.run(["python", "updation.py", "1", "max", "9", " "])
except:
	log_exceptions()
	pass
