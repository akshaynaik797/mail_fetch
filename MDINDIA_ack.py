import os
import sys
import struct, time
import subprocess
from datetime import date
import datetime
import openpyxl
import pdftotext
import time
import requests
now = datetime.datetime.now()

subprocess.run(["python", "updation.py","1","max1","1",sys.argv[2]])
subprocess.run(["python", "updation.py","1","max","2",sys.argv[3]])
subprocess.run(["python", "updation.py","1","max","3",sys.argv[4]])
subprocess.run(["python", "updation.py","1","max","5",str(now)])
subprocess.run(["python", "updation.py","1","max","7",sys.argv[5]])
subprocess.run(["python", "updation.py","1","max","8",sys.argv[6]])

with open(sys.argv[1], "rb") as f:
        pdf = pdftotext.PDF(f)

with open('mdindia/output.txt', 'w') as f:
        f.write(" ".join(pdf))     
with open('mdindia/output.txt', 'r') as myfile:
        f = myfile.read()
    #print("d")

try:            
        hg=[]
        #Added by Ashish.This field available in  output.txt file
        # Policy Number : OG/18/1000/8403/00000000.

        # MDID Number :MDI5-0030853492.

        # CCN Number :MDI4377912.

    #Added by Ashish
        # w=f.find('Claim No')+9  
        if f.find('Claim No ')!=-1:
                        temp='Claim No '
                        temp1='\n'
    
        elif f.find('CCN Number :')!=-1:
                        temp='CCN Number :'
                        temp1='.'

                                
        w=f.find(temp)+len(temp)
        g=f[w:]
        u=g.find(temp1)+w
        hg.append(f[w:u])
        
        if f.find('MDID Number :')!=-1:
                w=f.find('MDID Number :')+len("MDI ID Number :")
                            
                g=f[w:]
                u=g.find('\n')+w
                hg.append(f[w:u])
        
        else:
                hg.append('')

        if f.find('Policy Number :')!=-1:
                w=f.find('Policy Number :')+len("Policy Number :")
                
                g=f[w:]
                u=g.find('\n')+w
                hg.append(f[w:u].strip('.'))
    
        else:
                hg.append('')
        if f.find('facility for ')!=-1:
                        temp='facility for ' 
                        temp1=','
        elif f.find('Patient :')!=-1:
                        temp='Patient :'
                        temp1='\n'
                                
        w=f.find(temp)+len(temp)
        g=f[w:]
        u=g.find(temp1)+w
        try:
                hg.append(f[w:u].split('-')[-1])
        except:
                hg.append(f[w:u])

        hg=[sub.replace(':','') for sub in hg]  
        hg=[sub.replace('  ','') for sub in hg]
        hg=[sub.replace('.','') for sub in hg]
        hg=[sub.replace('Rs.','') for sub in hg]
        
        subprocess.run(["python", "updation.py","1","max","9",'Yes'])
        subprocess.run(["python", "updation.py","1","max","10",'NA'])
        
        try:
                subprocess.run(["python", "test_api.py",hg[0],'',hg[2],'','Acknowledgement',sys.argv[6],sys.argv[1],'',hg[1],hg[3]])
                '''wbk= openpyxl.load_workbook(wbkName)
                s2=wbk.worksheets[1]
                s2.cell(row_count_1+1, column=11).value='YES'
                '''
                subprocess.run(["python", "updation.py","1","max","11",'Yes'])  
        except Exception as e:
                #s2.cell(row_count_1+1, column=11).value='NO'
                subprocess.run(["python", "updation.py","1","max","11",'No'])
except Exception as e:
        #s2.cell(row_count_1+1, column=9).value='No'
        #s2.cell(row_count_1+1, column=11).value='NO'
        subprocess.run(["python", "updation.py","1","max","9",'Yes'])
        subprocess.run(["python", "updation.py","1","max","11",'No'])
now = datetime.datetime.now()
#s2.cell(row_count_1+1, column=6).value=now
#wbk.save(wbkName)
subprocess.run(["python", "updation.py","1","max","6",str(now)])
