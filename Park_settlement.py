import re
from email.mime.text import MIMEText
import smtplib
import time
import imaplib
import email
import camelot
import PyPDF2
import sys
import xlsxwriter
from xlrd import open_workbook
from openpyxl import load_workbook
# from tabula import read_pdf
import xlwt
import json
import os
import os.path
import xlrd
from os import listdir
from os import path
from os.path import isfile, join
import pdftotext
import tabula
import pandas as pd
# from tabula import read_pdf
from decode_error import check_subject, read_from_delete
from make_log import log_exceptions

try:
    if path.exists(r'park/park' + str(sys.argv[6]) + '.xlsx'):
        os.remove(r'park/park' + str(sys.argv[6]) + '.xlsx')
    import openpyxl

    wq = 0
    eu = []
    fg = []
    repeat = []
    from openpyxl.styles import Color, PatternFill, Font, Border

    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')


    def read_email_from_gmail():
        SMTP_SERVER = str(sys.argv[5])
        mail = imaplib.IMAP4_SSL(SMTP_SERVER)
        e_id = str(sys.argv[1])
        pswd = str(sys.argv[2])
        srt = str(sys.argv[3])
        stp = str(sys.argv[4])
        mail.login(user=e_id, password=pswd)
        mail.select("inbox", readonly=True)
        ###############################################<
        mail_uid = str(sys.argv[7])
        if mail_uid == -1:
            type, data = mail.search(None,
                                     '(SUBJECT "Payment Advice for claim ref no" since ' + srt + ' before ' + stp + ')')
            ids = data[0]
            id_list = ids.split()
        else:
            ids = mail_uid  # data is a list.
            # accept id from outside and put in id_list akshay var name = id

            id_list = []  # ids is a space separated string
            id_list.append(ids)
        ###############################################>
        # type, data = mail.search(None,
        #                          '(SUBJECT "Payment against Claim Reference Number:" since ' + srt + ' before ' + stp + ')')
        # ids = data[0]  # data is a list.
        # id_list = ids.split()  # ids is a space separated string
        # # print(id_list)
        for i in range(0, len(id_list)):
            latest_email_id = id_list[i]  # get the latest
            result, data = mail.fetch(latest_email_id,
                                      "(RFC822)")  # fetch the email body (RFC822)             for the given ID

            ##################################################ak
            try:
                try:
                    raw_email = data[0][1].decode('utf-8')
                except UnicodeDecodeError:
                    try:
                        raw_email = data[0][1].decode('ISO-8859-1')
                    except UnicodeDecodeError:
                        try:
                            raw_email = data[0][1].decode('ascii')
                        except UnicodeDecodeError:
                            pass
                email_message = email.message_from_string(raw_email)
                subject = email_message['Subject']
                result, sys.argv[8] = check_subject(subject, sys.argv[8], mail)
                if result == 'Changed':
                    # raise Exception('subject not matched')
                    raise Exception('subject not matched', )
            except:
                try:
                    log_exceptions(syssubject=sys.argv[8], subject=subject, error='subject not matched')
                except:
                    pass
                if result != 'OK':
                    data = {'server': SMTP_SERVER,
                            'hospmail': e_id,
                            'pass': pswd,
                            'subject': sys.argv[8]}
                    try:
                        data = read_from_delete(data)
                        if data == None:
                            raise Exception("Not found")
                    except:
                        log_exceptions(msg='not found in deleted', subject=sys.argv[8])
            ##################################################akend

            raw_email = data[0][1].decode('utf-8')
            email_message = email.message_from_string(raw_email)
            if email_message['Subject'] not in fg:
                for part in email_message.walk():

                    if part.get_content_maintype() == 'multipart':
                        # print part.as_string()
                        continue
                    if part.get('Content-Disposition') is None:
                        # print part.as_string()
                        continue
                    fileName = str(id_list[i]) + '.pdf'  # part.get_filename()
                    detach_dir = (os.getcwd() + '/park/attachments_' + str(sys.argv[6]) + '/')
                    if bool(fileName):
                        # print (email_message['Subject'])
                        filePath = os.path.join(detach_dir, fileName)
                        if not os.path.isfile(filePath):
                            from reportlab.pdfgen import canvas
                            c = canvas.Canvas(fileName)
                            # print (fileName)
                            fp = open(filePath, 'wb')
                            fp.write(part.get_payload(decode=True))
                            fp.close()
            else:
                repeat.append(email_message['Subject'])
            fg.append(email_message['Subject'])


    mypath = os.getcwd() + '/park'
    # if not path.exists(mypath):
    #     os.mkdir(mypath)
    # if not path.exists(mypath + '/attachments_' + str(sys.argv[6])):
    #     os.mkdir(mypath + '/attachments_' + str(sys.argv[6]))
    mypath = os.getcwd() + '/park/attachments_settlement/'
    # for filename in os.listdir(mypath):
    #     file_path = os.path.join(mypath, filename)
    #     if os.path.isfile(file_path) or os.path.islink(file_path):
    #         os.unlink(file_path)
    # read_email_from_gmail()

    wbkName = 'park/park' + str(sys.argv[6]) + '.xlsx'
    wbk = openpyxl.Workbook()
    wbk.create_sheet('Sheet1')
    s1 = wbk.worksheets[0]
    s2 = wbk.worksheets[1]

    onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
    for t in range(0, len(onlyfiles)):
        sh1 = ['S.no', 'Claim No.', 'Policy No.', 'Bank Name', 'Proposer Name', 'Name of Patient',
               'Instrument/ NEFT No', 'Instrument/ NEFT Date', 'Date of admission', 'Date of Discharge', 'bill amt',
               'net amt', 'co-pay', 'deductible', 'Employee ID', 'Employee Name', 'hospital discount', 'Al no.',
               'total deduction']
        sh2 = ['S.no', 'Claim No.', 'Description', 'Rate per day/Quantity', 'No.of Day/Visits/Quantity', 'Bill Amount',
               'Admissible Amount', 'Deducted Amount', 'Deduction Reason']
        for i in range(0, len(sh2)):
            s2.cell(row=1, column=i + 1).value = sh2[i]
        for i in range(0, len(sh1)):
            s1.cell(row=1, column=i + 1).value = sh1[i]

        print(onlyfiles[t])
        with open(mypath + onlyfiles[t], "rb") as f:
            pdf = pdftotext.PDF(f)

        with open('park/output.txt', 'w') as f:
            f.write(" ".join(pdf))
        with open('park/output.txt', 'r') as myfile:
            f = myfile.read()

        if os.path.exists(mypath + onlyfiles[t]):
            os.remove(mypath + onlyfiles[t])


        badchars = ('/',)
        regexdict = {'policy_no': [r'(?<=policy no).*'],
                     'pname': [r'(?<=Patient Name).*(?=and)'],
                     'amount': [r'(?<=Rs.)[\d ]+'],
                     'utr_no': [r'(?<=UTR No.).*(?=dated)'],
                     'date': [r'(?<=dated).*(?=of)']}
        datadict = {}
        for i in regexdict:
            for j in regexdict[i]:
                data = re.compile(j).search(f)
                if data is not None:
                    temp = data.group().strip()
                    for k in badchars:
                        temp = temp.replace(k, "")
                    datadict[i] = temp.strip()
                    break
                datadict[i] = ""

        hg = []
        w = f.find('Policy No.') + 10
        g = f[w:]
        u = g.find('\n') + w
        hg.append(datadict['policy_no'])

        w1 = f.find('Bank Name') + 10
        g = f[w1:]
        u1 = g.find('Successful') + w1
        hg.append('')

        w2 = f.find('Proposer Name') + 14
        g = f[w2:]
        u2 = g.find('Policy No.') + w2
        u3 = g.find('\n') + w2
        u4 = g.find('Claimed') + w2
        hg.append('')

        w2 = f.find('Name of Patient') + 16
        g = f[w2:]
        u2 = g.find('\n') + w2
        u3 = g.find('Date of admission') + w2
        hg.append(datadict['pname'])

        w3 = f.find('Instrument/ NEFT No') + 19
        g = f[w3:]
        u3 = g.find('\n') + w3
        hg.append(datadict['utr_no'])

        w4 = g.find('Instrument/ NEFT') + w3 + 17
        g = f[w4:]
        u4 = g.find('\n') + w4
        hg.append('')

        w5 = f.find('Date of admission') + 18
        g = f[w5:]
        u5 = g.find('Date of Discharge') + w5
        hg.append('')

        w6 = f.find('Date of Discharge') + 18
        g = f[w6:]
        u6 = g.find('\n') + w6
        hg.append(datadict['date'])

        w6 = f.find('Bill Amount') + 18
        g = f[w6:]
        u6 = g.find('Instrument') + w6
        hg.append('')

        w6 = f.find('Amount Paid') + 18
        g = f[w6:]
        u6 = g.find('Bank') + w6
        hg.append(datadict['amount'])

        if f.find('Co pay') != -1:
            w6 = f.find('Co pay') + 18
            g = f[w6:]
            u6 = g.find('Deductible') + w6
            hg.append('')
        else:
            hg.append(' ')
        w6 = f.find('Deductible') + 18
        g = f[w6:]
        u6 = g.find('\n') + w6
        hg.append('')

        w2 = f.find('Employee ID') + 14
        g = f[w2:]
        u3 = g.find('Employee Name') + w2
        hg.append('')

        w2 = f.find('Employee Name') + 14
        g = f[w2:]
        u3 = g.find('Name of Proposer') + w2
        hg.append('')

        w2 = f.find('Hospital Discount') + 25
        g = f[w2:]
        u3 = g.find('AL approved') + w2
        hg.append('')

        w9 = f.find('AL No.') + 6
        g = f[w9:]
        u9 = g.find('\n') + w9 + 3
        ccn = (f[w9:u9])

        w9 = f.find('Claim No') + 15
        g = f[w9:]
        u9 = g.find('-') + w9 + 3
        hg.append('')
        ccn = ccn.replace('\n', '')
        ccn = ccn.replace('.', '')
        ccn = ccn.replace(' ', '')
        if ccn == None:
            ccn = hg[-1]
        hg = [sub.replace('  ', '') for sub in hg]
        hg = [sub.replace(':', '') for sub in hg]
        hg = [sub.replace('\n', ' ') for sub in hg]
        s1.cell(row=t + 2, column=1).value = t + 1
        s1.cell(row=t + 2, column=2).value = ''#ccn
        for i in range(0, len(hg)):
            s1.cell(row=t + 2, column=i + 3).value = hg[i]
    print("Done")
    wbk.save(wbkName)
    wbk.close
    wbkName = 'count/count.xlsx'
    wbk = openpyxl.load_workbook(wbkName)
    s1 = wbk.worksheets[0]

    row_ = s1.max_row + 1
    s1.cell(row=row_, column=1).value = 'park'
    s1.cell(row=row_, column=2).value = str(sys.argv[6])
    s1.cell(row=row_, column=3).value = len(fg)
    s1.cell(row=row_, column=4).value = len(onlyfiles)
    s1.cell(row=row_, column=5).value = len(repeat)
    wbk.save(wbkName)
except:
    log_exceptions()
    pass