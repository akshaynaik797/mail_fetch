import os
import sys
import struct, time
import subprocess
from datetime import date
import datetime
import openpyxl
import pdftotext
import time
import requests

from patient_name_fun import pname_fun

now = datetime.datetime.now()

# Added by Ashish
subprocess.run(["python", "updation.py", "1", "max1", "1", sys.argv[2]])
subprocess.run(["python", "updation.py", "1", "max", "2", sys.argv[3]])
subprocess.run(["python", "updation.py", "1", "max", "3", sys.argv[4]])
subprocess.run(["python", "updation.py", "1", "max", "5", str(now)])
subprocess.run(["python", "updation.py", "1", "max", "7", sys.argv[5]])
subprocess.run(["python", "updation.py", "1", "max", "8", sys.argv[6]])
# 

with open(sys.argv[1], "rb") as f:
    pdf = pdftotext.PDF(f)

with open('fgh/output1.txt', 'w') as f:
    f.write(" ".join(pdf))
with open('fgh/output1.txt', 'r') as myfile:
    f = myfile.read()

try:
    hg = []

    w = f.find('Policy Number') + 20
    k = f[w:]
    u = k.find("\n") + w
    g = f[w:u]
    hg.append(g)

    w = f.find('Claim Number') + 20
    k = f[w:]
    u = k.find("\n") + w
    g = f[w:u]
    hg.append(g)

    w = f.find('Member ID Number') + 20
    k = f[w:]
    u = k.find("\n") + w
    g = f[w:u]
    hg.append(g)

    w = f.find('Patient Name') + 40
    k = f[w:]
    u = k.find("\n") + w
    g = f[w:u]
    hg.append(g)

    # Added by Ashish
    subprocess.run(["python", "updation.py", "1", "max", "9", 'Yes'])
    subprocess.run(["python", "updation.py", "1", "max", "10", 'NA'])
    regex_list = [r"(?<=Patient Name).*"]
    pname = pname_fun(f, regex_list)
    #
    try:
        subprocess.run(
            ["python", "test_api.py", hg[1], '', hg[0], '', 'Acknowledgement', sys.argv[6], sys.argv[1], '', hg[3], pname])

        '''wbk= openpyxl.load_workbook(wbkName)
        s2=wbk.worksheets[1]
        s2.cell(row_count_1+1, column=11).value='YES'
        '''
        subprocess.run(["python", "updation.py", "1", "max", "11", 'Yes'])
    except Exception as e:
        # s2.cell(row_count_1+1, column=11).value='NO'
        subprocess.run(["python", "updation.py", "1", "max", "11", 'No'])
except Exception as e:
    # s2.cell(row_count_1+1, column=9).value='No'
    # s2.cell(row_count_1+1, column=11).value='NO'
    subprocess.run(["python", "updation.py", "1", "max", "9", 'Yes'])
    subprocess.run(["python", "updation.py", "1", "max", "11", 'No'])
now = datetime.datetime.now()
# s2.cell(row_count_1+1, column=6).value=now
# wbk.save(wbkName)
subprocess.run(["python", "updation.py", "1", "max", "6", str(now)])
