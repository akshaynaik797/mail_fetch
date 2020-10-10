import sqlite3
import pandas as pd
from email.mime.text import MIMEText
import smtplib
import time
import imaplib
import email
import camelot
import PyPDF2
import sys
import csv
import xlsxwriter
from xlrd import open_workbook
import xlwt
import os
import glob
import os.path
import xlrd
from os import listdir
from os import path
from os.path import isfile, join
from html.parser import HTMLParser
import pdfkit
import pandas as pd
import html2text
import pdftotext
import openpyxl
import re
import PyPDF2
import pdftotext, openpyxl
from openpyxl import load_workbook
import os, sys, shutil
import re
import pdftotext
import os, sys, shutil
import re
import pdftotext
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
#from xvfbwrapper import Xvfb
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.proxy import Proxy, ProxyType
import time
import subprocess
from decode_error import read_from_delete, check_subject
from make_log import log_exceptions
# sys.argv = ['bajaj.py', 'mediclaim@inamdarhospital.org', 'Mediclaim@2019', '03-Aug-2020', '03-Aug-2020', 'imap.gmail.com', 'inamdar', '24507']

try:

    fg, repeat = [], []
    subprocess.run(["python", "updation.py", "1", "max", "9", "X"])

    chromeOptions = webdriver.ChromeOptions()
    chromeOptions.add_argument("--headless")

    attachment_path = r"C:\Users\Administrator\Downloads/"

    def read_email_from_gmail():
        SMTP_SERVER = str(sys.argv[5])
        mail = imaplib.IMAP4_SSL(SMTP_SERVER)
        # mail.login(user = 'Mediclaim@therisingmedicare.com', password = 'cef@2018')

        e_id = str(sys.argv[1])
        pswd = str(sys.argv[2])
        srt = str(sys.argv[3])
        stp = str(sys.argv[4])
        mail.login(user=e_id, password=pswd)
        mail.select("inbox", readonly=True)
        ###############################################<
        mail_uid = str(sys.argv[7])
        if mail_uid == -1:
            type, data = mail.search(None, '(SUBJECT "Advice from Standard Chartered Bank" since ' + srt + ' before ' + stp + ')')
            ids = data[0]
            id_list = ids.split()
        else:
            ids = mail_uid  # data is a list.
            # accept id from outside and put in id_list akshay var name = id

            id_list = []  # ids is a space separated string
            id_list.append(ids)
        ###############################################>
        # type, data = mail.search(None, '(SUBJECT "payment processed for claim no" since ' + srt + ' before ' + stp + ')')
        # ids = data[0]  # data is a list.
        # id_list = ids.split()  # ids is a space separated string
        # # print(id_list)
        # b = 0
        for i in range(0, len(id_list)):
            latest_email_id = id_list[i]  # get the latest
            result, data = mail.fetch(latest_email_id,
                                      "(RFC822)")  # fetch the email body (RFC822)             for the given ID

            ##################################################ak
            try:
                # raw_email = data[0][1].decode('utf-8')
                try:
                    raw_email = data[0][1].decode('utf-8')
                except UnicodeDecodeError:
                    try:
                        raw_email = data[0][1].decode('ISO-8859-1')
                    except UnicodeDecodeError:
                        try:
                            raw_email = data[0][1].decode('ascii')
                        except UnicodeDecodeError:
                            pass
                email_message = email.message_from_string(raw_email)
                subject = email_message['Subject']
                result, sys.argv[8] = check_subject(subject, sys.argv[8], mail)
                if result == 'Changed':
                    # raise Exception('subject not matched')
                    raise Exception('subject not matched', )
            except:
                # log_exceptions(syssubject=sys.argv[8], subject=subject, error='subject not matched')
                if result != 'OK':
                    data = {'server': SMTP_SERVER,
                            'hospmail': e_id,
                            'pass': pswd,
                            'subject': sys.argv[8]}
                    try:
                        data = read_from_delete(data)
                        if data == None:
                            raise Exception("Not found")
                    except:
                        log_exceptions(msg='not found in deleted', subject=sys.argv[8])
            ##################################################akend
            raw_email = data[0][1].decode('utf-8')
            email_message = email.message_from_string(raw_email)
            if email_message['Subject'] not in fg:
                for part in email_message.walk():

                    if part.get_content_maintype() == 'multipart':
                        # print part.as_string()
                        continue
                    if part.get('Content-Disposition') is None:
                        # print part.as_string()
                        continue
                    fileName = part.get_filename()

                    detach_dir = (os.getcwd() + '/bajaj/attachments_pdf_' + str(sys.argv[6]))
                    if bool(fileName):
                        # print (email_message['Subject'])
                        filePath = os.path.join(detach_dir, fileName)
                        if not os.path.isfile(filePath):
                            from reportlab.pdfgen import canvas
                            c = canvas.Canvas(fileName)
                            # print (fileName)
                            fp = open(filePath, 'wb')
                            fp.write(part.get_payload(decode=True))
                            fp.close()
                            #code for claim no
            else:
                repeat.append(email_message['Subject'])
            fg.append(email_message['Subject'])


    mypath = os.getcwd() + '/bajaj'

    if not path.exists(mypath):
        os.mkdir(mypath)
    if not path.exists(mypath + '/attachments_' + str(sys.argv[6])):
        os.mkdir(mypath + '/attachments_' + str(sys.argv[6]))
    if not path.exists(mypath + '/attachments_pdf_' + str(sys.argv[6])):
        os.mkdir(mypath + '/attachments_pdf_' + str(sys.argv[6]))

    mypath = os.getcwd() + '/bajaj/attachments_' + str(sys.argv[6]) + '/'
    for filename in os.listdir(mypath):
        file_path = os.path.join(mypath, filename)
        if os.path.isfile(file_path) or os.path.islink(file_path):
            os.unlink(file_path)

    pdfpath = os.getcwd() + '/bajaj/attachments_pdf_' + str(sys.argv[6]) + '/'
    pdffiles = [f for f in listdir(pdfpath) if isfile(join(pdfpath, f))]

    for filename in os.listdir(pdfpath):
        file_path = os.path.join(pdfpath, filename)
        if os.path.isfile(file_path) or os.path.islink(file_path):
            os.unlink(file_path)
    read_email_from_gmail()

    onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
    t = 0
    yu = []
    # print(onlyfiles)
    for i in onlyfiles:
        if i[-1] == '#':
            yu.append(t)
        t += 1
    # print(yu)
    t = 0
    for i in yu:
        onlyfiles.pop(i - t)
        t = t + 1
    pdfpath = os.getcwd() + '/bajaj/attachments_pdf_' + str(sys.argv[6]) + '/'
    pdffiles = [f for f in listdir(pdfpath) if isfile(join(pdfpath, f))]
    pdffiles.sort()
    # print(pdffiles)


    with sqlite3.connect("database1.db") as con:
        cur = con.cursor()
        q = f'select id, password from bajaj_credentials where hospital="{sys.argv[6]}";'
        print(q)
        cur.execute(q)
        r = cur.fetchone()

        if r:
            hosp_mail = r[0]
            hosp_pass = r[1]
        else:
            hosp_mail = ''
            hosp_pass = ''

    for t in range(0, len(pdffiles)):
        with open(pdfpath + pdffiles[t], "rb") as f:
            pdf = pdftotext.PDF(f)
        data = "\n\n".join(pdf)
        with open('temppdf.txt', "w") as f:
            f.write(data)
        with open('temppdf.txt', "r") as f:
            data = f.readlines()
        with open('temppdf.txt', "r") as f:
            f = f.read()

        regex = r'(?<=UTR Reference).*'
        x = re.search(regex, f)
        if x:
            utr = x.group().strip()
        else:
            utr = ''

        regex = r'\d+ +\d{2}/\d{2}/.+'
        list1 = re.findall(regex, f)
        list1 = [re.split(r' {2,}', i) for i in list1]
        regex = r'\d{4} +\d{4}.*'
        list2 = re.findall(regex, f)
        list2 = [re.split(r' {2,}', i) for i in list2]

        table = []
        if len(list1) == len(list2):
            for i, j in zip(list1, list2):
                datadict = {}
                datadict['appr_no'] = i[0] + j[0]
                datadict['date'] = i[1] + j[1]
                if j[2].isdigit() is True:
                    datadict['name'] = i[2]
                else:
                    datadict['patientname'] = i[2] + ' ' + j[2]
                if len(j) > 3:
                    datadict['claimno'] = i[3] + j[3]
                else:
                    datadict['claimno'] = i[3] + j[2]
                datadict['amt'] = i[-1].replace(',', '')
                datadict['tds'] = i[-2]
                datadict['utr_ref'] = utr
                table.append(datadict)
            claimno_list = [i['claimno'] for i in table]
            for i in claimno_list:
                try:
                    claimno = i
                    driver = webdriver.Chrome(r'C:\Users\Administrator\Downloads\chromedriver_win32\chromedriver.exe', options=chromeOptions)
                    driver.get("https://hcm.bajajallianz.com/BagicHCM/hlogin.jsp")
                    driver.find_element_by_id("j_username").click()
                    driver.find_element_by_id("j_username").send_keys(hosp_mail)
                    driver.find_element_by_id("j_password").send_keys(hosp_pass)
                    driver.find_element_by_id("Login").click()

                    driver.find_element_by_link_text("Payment Details").click()

                    driver.find_element_by_id("p_search_criteria.stringval3").send_keys(i)
                    driver.find_element_by_id("payment").click()

                    driver.find_element_by_xpath("/html/body/table[1]/tbody/tr/td/form/div/div/div[3]/fieldset[2]/div/table/tbody/tr[2]/td[42]/img").click()
                    time.sleep(20)

                    # driver.find_element_by_id("p_search_criteria.stringval3").click()
                    driver.quit()
                except Exception as e:
                    print(e)
                    log_exceptions(claimno=claimno)
                    if driver in locals():
                        driver.quit()
                if path.exists(attachment_path+'claimCoveringLetter.pdf'):
                    os.replace(attachment_path+'claimCoveringLetter.pdf', 'bajaj/'+i+'.pdf')

                    filepath = 'bajaj/'+i+'.pdf'
                    with open(filepath, "rb") as f:
                        pdf = pdftotext.PDF(f)
                    data = "\n\n".join(pdf)
                    with open('bajaj/temppdf.txt', "w") as f:
                        f.write(data)
                    with open('bajaj/temppdf.txt', "r") as f:
                        f = f.read()
                    datadict = dict()

                    params = (
                        ('patientname', r'(?<=Name Of The Patient)[ \S]+'),
                        ('idcardno', r'(?<=ID Card No)[ \S]+'),
                        ('claim_id', r'(?<=Claim ID)[ \S]+'),
                        ('claim_no', r'(?<=Claim Number)[ \S]+'),
                        ('doa', r'(?<=DOA:) ?\S+'),
                        ('dod', r'(?<=DOD:) ?\S+'),
                        ('appr_no', r'(?<=Approval Number)[ \S]+'),
                        ('utr_no', r'(?<=UTR No)[ \S]+'),
                        ('bill_amount', r'\d+(?=\s+Paid Amount)'),
                        ('paid_amount', r'\d+(?=\s+Disallowed Amount)'),
                        ('disallowed_amount', r'\d+(?=\s+TDS Amount)'),
                        ('tds_amount', r'\d+(?=\s+Hospital Service Tax No)'),
                    )

                    for i in params:
                        regex = i[1]
                        x = re.search(regex, f)
                        keyname = i[0]
                        if x:
                            datadict[keyname] = x.group().strip()
                        else:
                            datadict[keyname] = ''

                    regex = r'\w+ ?Charges[\s\S]+(?=\n[\s\S]+Payment Details)'
                    x = re.search(regex, f)
                    # keyname = i[0]
                    if x:
                        data = x.group().split('\n')
                        data = [re.split(r' {2,}', i) for i in data]
                        for i, j in enumerate(data):
                            if len(j) > 2:
                                for x, y in enumerate(data[i + 1:]):
                                    if len(y) <= 2:
                                        j[-1] = j[-1] + ' ' + y[-1]
                                    else:
                                        break
                        clean = []
                        for i, j in enumerate(data):
                            if len(j) > 2:
                                clean.append(j)
                    else:
                        data = ''

                    tempdata = table[claimno_list.index(claimno)]

                    mylist = [i[0] for i in params]
                    mylist.append('date')
                    mydata = [datadict[i[0]] for i in params]
                    mydata.append(tempdata['date'].replace('/','-'))
                    wbk = openpyxl.Workbook()
                    wbk.create_sheet('1')
                    s1 = wbk.worksheets[0]
                    s2 = wbk.worksheets[1]
                    rowno = s1.max_row+1
                    for i, j in enumerate(mylist):
                        s1.cell(row=1, column=i+1).value = j
                        s1.cell(row=rowno, column=i+1).value = mydata[i]

                    mylist = ['Sr no','Particular', 'Bill Amount', 'Disallowed Amount', 'Approved Amount', 'Disallowance Reason']
                    for i, j in enumerate(mylist):
                        s2.cell(row=1, column=i+1).value = j
                    for i, j in enumerate(clean):
                        for x, y in enumerate(j):
                            s2.cell(row=i+2, column=1).value = i+1
                            s2.cell(row=i+2, column=x + 2).value = y

                    wbname = 'bajaj'+sys.argv[6]+'.xlsx'
                    wbk.save('bajaj/'+wbname)
                    wbk.close()

        filelist = []
        fileext = ['pdf']

        pathname = os.path.dirname(sys.argv[0])
        fullp = os.path.abspath(pathname) + '/'
        folder_name = 'bajaj'
        for i in fileext:
            for root, dirs, files in os.walk(fullp+folder_name):
                for file in files:
                    if file.endswith(i):
                        filelist.append(os.path.join(root, file))


        for myfile in filelist:
            try:
                dest = os.path.dirname(myfile)+'/'+datadict['claim_no']+'.pdf'
            except KeyError as e:
                dest = os.path.dirname(myfile)+'/'+datadict['claimno']+'.pdf'
            try:
                shutil.copy(myfile, dest)
            except shutil.SameFileError:
                pass
    subprocess.run(["python", "updation.py", "1", "max", "9", " "])
except:
    log_exceptions()
    pass