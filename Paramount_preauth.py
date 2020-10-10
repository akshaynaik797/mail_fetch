import os
import re
import sys
import struct, time
import subprocess
from datetime import date
import datetime
import openpyxl
import pdftotext
import time
import requests

from patient_name_fun import pname_fun

now = datetime.datetime.now()
'''
wbkName = 'log file.xlsx'
wbk= openpyxl.load_workbook(wbkName)
s1=wbk.worksheets[0]
s2=wbk.worksheets[1]
row_count_1 = s2.max_row
s2.cell(row_count_1+1, column=1).value=sys.argv[2]
s2.cell(row_count_1+1, column=2).value=sys.argv[3]
s2.cell(row_count_1+1, column=3).value=sys.argv[4]
s2.cell(row_count_1+1, column=5).value=now
s2.cell(row_count_1+1, column=7).value=sys.argv[5]
s2.cell(row_count_1+1, column=8).value=sys.argv[6]
'''

subprocess.run(["python", "updation.py","1","max1","1",sys.argv[2]])
subprocess.run(["python", "updation.py","1","max","2",sys.argv[3]])
subprocess.run(["python", "updation.py","1","max","3",sys.argv[4]])
subprocess.run(["python", "updation.py","1","max","5",str(now)])
subprocess.run(["python", "updation.py","1","max","7",sys.argv[5]])
subprocess.run(["python", "updation.py","1","max","8",sys.argv[6]])

with open(sys.argv[1], "rb") as f:
	pdf = pdftotext.PDF(f)

with open('Paramount/output.txt', 'w') as f:
	f.write(" ".join(pdf))     
with open('Paramount/output.txt', 'r') as myfile:
 	f = myfile.read()
try:
	hg=[]
	w=f.find('Claim no')+9
	g=f[w:]
	u=g.find('has')+w
	hg.append(f[w:u])

	status='Approved'
	w=f.find('TOTAL AL ISSUED')+15
	g=f[w:]
	u=g.find('\n')+w
	hg.append(f[w:u])

	regex = r"(?<=Policy No :).*"
	result = re.search(regex, f)
	if result is not None:
		x = result
		x1 = x.group()
		hg.append(x1)
	elif f.find('POLICY NO') != -1:
		w=f.find('POLICY NO')+9
		g=f[w:]
		u=g.find('\n')+w
		hg.append(f[w:u])
	else:
		hg.append('')

	w=f.find('PHS ID')+7
	g=f[w:]
	u=g.find('\n')+w
	hg.append(f[w:u])
	#hg.append('')#member id
	#hg=[sub.replace('-','') for sub in hg]
	hg=[sub.replace(':','') for sub in hg]
	hg=[sub.replace('  ','') for sub in hg]
	hg=[sub.replace('Rs','') for sub in hg]
	hg=[sub.replace('/','') for sub in hg]
	hg=[sub.replace('-','') for sub in hg]
	#s2.cell(row_count_1+1, column=9).value='Yes'
	#s2.cell(row_count_1+1, column=10).value='NA'
	#wbk.save(wbkName)
	
	subprocess.run(["python", "updation.py","1","max","9",'Yes'])
	subprocess.run(["python", "updation.py","1","max","10",'NA'])
	
	try:
		regex_list = [r"(?<=PATIENT NAME & RELATION).*", r"(?<=PATIENT NAME AND).*"]
		pname = pname_fun(f, regex_list)

		subprocess.run(["python", "test_api.py",hg[0],hg[1],hg[2],'Pa',status,sys.argv[6],sys.argv[1],'',hg[3], pname])
		'''wbk= openpyxl.load_workbook(wbkName)
		s2=wbk.worksheets[1]
		s2.cell(row_count_1+1, column=11).value='YES'
		'''
		subprocess.run(["python", "updation.py","1","max","11",'Yes'])	
	except Exception as e:
		#s2.cell(row_count_1+1, column=11).value='NO'
		subprocess.run(["python", "updation.py","1","max","11",'No'])
except Exception as e:
	#s2.cell(row_count_1+1, column=9).value='No'
	#s2.cell(row_count_1+1, column=11).value='NO'
	subprocess.run(["python", "updation.py","1","max","9",'Yes'])
	subprocess.run(["python", "updation.py","1","max","11",'No'])
now = datetime.datetime.now()
#s2.cell(row_count_1+1, column=6).value=now
#wbk.save(wbkName)
subprocess.run(["python", "updation.py","1","max","6",str(now)])
