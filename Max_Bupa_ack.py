import os, re
import sys
import struct, time
import subprocess
from datetime import date
import datetime
import openpyxl
import pdftotext
import time
import requests
from make_log import log_exceptions
now = datetime.datetime.now()
subprocess.run(["python", "updation.py","1","max1","1",sys.argv[2]])
subprocess.run(["python", "updation.py","1","max","2",sys.argv[3]])
subprocess.run(["python", "updation.py","1","max","3",sys.argv[4]])
subprocess.run(["python", "updation.py","1","max","5",str(now)])
subprocess.run(["python", "updation.py","1","max","7",sys.argv[5]])
subprocess.run(["python", "updation.py","1","max","8",sys.argv[6]])

with open(sys.argv[1], "rb") as f:
    pdf = pdftotext.PDF(f)

with open('Max_Bupa/output.txt', 'w') as f:
    f.write(" ".join(pdf))
with open('Max_Bupa/output.txt', 'r') as myfile:
    f = myfile.read()

try:		
    hg=[]


    if f.find('Customer ID') != -1:
        x = re.search(r"(?<=Customer ID :) *\d+", f)
        if x != None:
            x1 = x.group().strip()
            hg.append(x1)
    elif f.find('Customer ID') != -1:
        if f.find('Customer ID') == -1:
                # len function used(Ashish)
                w=f.find('request number')+len('request number ')
        else:
                w=f.find('Customer ID')+len('Customer ID')
        k=f[w:]
        if f.find('Customer Id') == -1:
                # find should search after w position(Ashish)
                u=w+k.find(".")
        else:
                u=k.find("\n")+w
        g=f[w:u]
        hg.append(g)


    if f.find("Patient's Name") != -1:
        x = re.search(r"(?<=Patient's Name).*", f)
        if x != None:
            x1 = x.group().strip()
            hg.append(x1)
        else:
            hg.append('')
    else:
        # if f.find("Patient's Name") == -1:
        #     # len function used(Ashish)
        #     w=f.find("respect of")+len('respect of ')
        # else:
        #     w=f.find("Patient's Name")+len("Patient's Name")
        #     k=f[w:]
        # if(k.find("Patient's Name")) == -1:
        #     u=w+k.find('we')
        # else:
        #     u=k.find("\n")+w
        # g=f[w:u]
        hg.append("")


    if f.find('request ID') != -1:
        x = re.search(r"(?<=cashless request ID) *\d+", f)
        if x != None:
            x1 = x.group().strip()
            hg.append(x1)
    else:
        hg.append('')


    # if f.find('request ID') == -1:
    # 		print('No request Id is present')
    # else:
    # 	w=f.find('request ID')+len('request ID')
    # 	k=f[w:]
    # 	u=k.find("\n")+w
    # 	g=f[w:u]
    # 	hg.append(g.trim())


    subprocess.run(["python", "updation.py","1","max","9",'Yes'])
    subprocess.run(["python", "updation.py","1","max","10",'NA'])

    try:
        subprocess.run(["python", "test_api.py",hg[0],'','','','Acknowledgement',sys.argv[6],sys.argv[1],'','',hg[1]])
        '''wbk= openpyxl.load_workbook(wbkName)
        s2=wbk.worksheets[1]
        s2.cell(row_count_1+1, column=11).value='YES'
        '''
        subprocess.run(["python", "updation.py","1","max","11",'Yes'])
    except Exception as e:
        log_exceptions()
        #s2.cell(row_count_1+1, column=11).value='NO'
        subprocess.run(["python", "updation.py","1","max","11",'No'])
except Exception as e:
    log_exceptions()
    #s2.cell(row_count_1+1, column=9).value='No'
    #s2.cell(row_count_1+1, column=11).value='NO'
    subprocess.run(["python", "updation.py","1","max","9",'Yes'])
    subprocess.run(["python", "updation.py","1","max","11",'No'])
now = datetime.datetime.now()
#s2.cell(row_count_1+1, column=6).value=now
#wbk.save(wbkName)
subprocess.run(["python", "updation.py","1","max","6",str(now)])
pass
