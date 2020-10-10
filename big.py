from email.mime.text import MIMEText
import smtplib
import time
import imaplib
import email
import camelot
import sys
import pdftotext
import xlsxwriter
from xlrd import open_workbook
import xlwt
import os
import os.path
import xlrd
from os import listdir
from os import path
from os.path import isfile, join
import pdfkit
import subprocess

from openpyxl.styles import Color, PatternFill, Font, Border

from decode_error import read_from_delete, check_subject
from make_log import log_exceptions

try:
    subprocess.run(["python", "updation.py", "1", "max", "9", "X"])
    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')
    fg = []
    eu = []
    path_wkhtmltopdf = r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe'
    config = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf)
    # config = pdfkit.configuration(wkhtmltopdf='/usr/bin/wkhtmltopdf')



    def read_email_from_gmail():
        SMTP_SERVER = str(sys.argv[5])
        mail = imaplib.IMAP4_SSL(SMTP_SERVER)
        e_id = str(sys.argv[1])
        pswd = str(sys.argv[2])
        srt = str(sys.argv[3])
        stp = str(sys.argv[4])
        mail.login(user=e_id, password=pswd)
        mail.select("inbox", readonly=True)
        ###############################################<
        mail_uid = str(sys.argv[7])
        if mail_uid == -1:
            type, data = mail.search(None,
                                     '(FROM "claims.payment@starhealth.biz" SUBJECT "Intimation No:" since ' + srt + ' before ' + stp + ')')
            ids = data[0]
            id_list = ids.split()
        else:
            ids = mail_uid  # data is a list.
            # accept id from outside and put in id_list akshay var name = id

            id_list = []  # ids is a space separated string
            id_list.append(ids)
        ###############################################>
        # type, data = mail.search(None,
        #                          '(FROM "claims.payment@starhealth.biz" SUBJECT "Intimation No:" since ' + srt + ' before ' + stp + ')')
        # ids = data[0]  # data is a list.
        # id_list = ids.split()  # ids is a space separated string
        # # print(id_list)
        for i in range(0, len(id_list)):
            latest_email_id = id_list[i]  # get the latest
            result, data = mail.fetch(latest_email_id,
                                      "(RFC822)")  # fetch the email body (RFC822)             for the given ID

            ##################################################ak
            try:
                # raw_email = data[0][1].decode('utf-8')
                try:
                    raw_email = data[0][1].decode('utf-8')
                except UnicodeDecodeError:
                    try:
                        raw_email = data[0][1].decode('ISO-8859-1')
                    except UnicodeDecodeError:
                        try:
                            raw_email = data[0][1].decode('ascii')
                        except UnicodeDecodeError:
                            pass
                email_message = email.message_from_string(raw_email)
                subject = email_message['Subject']
                result, sys.argv[8] = check_subject(subject, sys.argv[8], mail)
                if result == 'Changed':
                    # raise Exception('subject not matched')
                    raise Exception('subject not matched', )
            except:
                try:
                    log_exceptions(syssubject=sys.argv[8], subject=subject, error='subject not matched')
                except:
                    pass
                if result != 'OK':
                    data = {'server': SMTP_SERVER,
                            'hospmail': e_id,
                            'pass': pswd,
                            'subject': sys.argv[8]}
                    try:
                        data = read_from_delete(data)
                        if data == None:
                            raise Exception("Not found")
                    except:
                        log_exceptions(msg='not found in deleted', subject=sys.argv[8])
            ##################################################akend
            raw_email = data[0][1].decode('utf-8')

            email_message = email.message_from_string(raw_email)
            if email_message['Subject'] not in fg:
                for part in email_message.walk():
                    if part.get_content_maintype() == 'multipart':
                        # print part.as_string()
                        continue
                    if part.get('Content-Disposition') is None:
                        # print part.as_string()
                        continue
                    # fileName = part.get_filename()
                    f = email_message['Subject']
                    x1 = f.find('Intimation No:') + 14
                    x2 = f.find('-')
                    fileName = f[x1:x2] + '.pdf'
                    fileName = fileName.replace(' ', '')
                    fileName = fileName.replace('/', '')
                    detach_dir = (os.getcwd() + '/star/attachments_2_' + str(sys.argv[6]))
                    if bool(fileName):
                        filePath = os.path.join(detach_dir, fileName)
                        if not os.path.isfile(filePath):
                            from reportlab.pdfgen import canvas
                            c = canvas.Canvas(fileName + '.pdf')
                            # print (fileName)
                            fp = open(filePath, 'wb')
                            fp.write(part.get_payload(decode=True))
                            fp.close()

            fg.append(email_message['Subject'])


    mypath = os.getcwd() + '/star'
    if not path.exists(mypath):
        os.mkdir(mypath)
    if not path.exists(mypath + '/attachments_2_' + str(sys.argv[6])):
        os.mkdir(mypath + '/attachments_2_' + str(sys.argv[6]))
    mypath = os.getcwd() + '/star/attachments_2_' + str(sys.argv[6]) + '/'
    for filename in os.listdir(mypath):
        file_path = os.path.join(mypath, filename)
        if os.path.isfile(file_path) or os.path.islink(file_path):
            os.unlink(file_path)
    read_email_from_gmail()
    wq = 0
    qe = 0
    onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
    if path.exists(r'star/star' + str(sys.argv[6]) + '.xlsx'):
        os.remove(r'star/star' + str(sys.argv[6]) + '.xlsx')
    import openpyxl

    po = []
    wbkName = 'star/star' + str(sys.argv[6]) + '.xlsx'
    wbk = openpyxl.Workbook()
    wbk.create_sheet('Sheet1')
    s1 = wbk.worksheets[0]
    s3 = wbk.worksheets[1]

    for t in range(0, len(onlyfiles)):
        try:
            for wd in wbk.worksheets:
                wd.cell(row=1, column=1).value = 'Sr. No.'
                wd.cell(row=1, column=2).value = 'INTIMATION NO'
            sh1 = ['Policy number', 'Diagnosis', 'DOA', 'DOD', 'Claimant Name', 'ICD Codes Desc', 'Total amount claimed',
                   'Hospitalisation payable amount', 'Pre hospitalisation payable amount',
                   'Post hospitalisation payable amount', 'Add on Benefit(Hospital Cash / Patient care)',
                   'Total Claim Payable Amount', 'deducted']
            sh2 = ['Nature of Expenditure', 'Amount Claimed', '	Approve d Amount', 'Disallowance Reasons / Remarks']
            for i in range(0, len(sh1)):
                s1.cell(row=1, column=i + 3).value = sh1[i]
            for i in range(0, len(sh2)):
                s3.cell(row=1, column=i + 3).value = sh2[i]


            def select_sheet(wks, sheet_name):

                if not sheet_name in wks.sheetnames:
                    wks.create_sheet(sheet_name)
                # print(sheet_name)

                wks.save('star.xlsx')


            def select_column(wks, s, ro):
                sheet = wks.worksheets[s]

                max_col = sheet.max_column
                s = []
                for i in range(1, max_col + 1):
                    cell_obj = sheet.cell(row=1, column=i)
                    s.append(cell_obj.value)


            # print(s)

            # print(CCN)

            tables = camelot.read_pdf(mypath + onlyfiles[t], pages='all', Line_scale=100)
            # print(tables)
            with open(mypath + onlyfiles[t], "rb") as f:
                pdf = pdftotext.PDF(f)

            with open('star/output.txt', 'w') as f:
                f.write(" ".join(pdf))
            with open('star/output.txt', 'r') as myfile:
                f = myfile.read()
            # print(data)

            gh = []
            sd = []
            w2 = f.find('Intimation No') + 21
            g = f[w2:]
            u2 = g.find('Bill') + w2 - 1
            c = f[w2:u2]
            cli = c.replace(' ', '')
            w = f.find('Policy No') + 9
            g = f[w:]
            u = u2 = g.find('\n') + w
            gh.append(f[w:u])
            # print(f[w:u])
            w1 = f.find('Diagnosis') + 19
            g = f[w1:]
            u1 = g.find(',') + w1
            gh.append(f[w1:u1])

            if gh[-1].find('\n') != -1:
                o = gh[-1]
                x = gh[-1].find('\n')

                gh[-1] = o[:x] + o[x + 33:]

            w = f.find('DOA') + 3
            g = f[w:]
            u = u2 = g.find('\n') + w
            gh.append(f[w:u])

            w1 = f.find('DOD') + 3
            g = f[w1:]
            u1 = g.find('\n') + w1
            gh.append(f[w1:u1])

            w = f.find('Claimant Name') + 13
            g = f[w:]
            u = u2 = g.find('Product Name') + w
            gh.append(f[w:u])

            w1 = f.find('ICD Codes Desc') + 14
            g = f[w1:]
            u1 = g.find(',') + w1
            gh.append(f[w1:u1])

            gh = [sub.replace('  ', '') for sub in gh]
            # print(gh)
            s1.cell(row=t + 2, column=1).value = t + 1
            s1.cell(row=t + 2, column=2).value = cli
            for i in range(0, len(gh)):
                s1.cell(row=t + 2, column=i + 3).value = gh[i]

            tables.export('star/foo1.xls', f='excel')
            loc = ("star/foo1.xls")
            wb = xlrd.open_workbook(loc)
            sheet_1 = wb.sheet_by_index(0)
            sheet_1.cell_value(0, 0)
            sheet_2 = wb.sheet_by_index(1)
            sheet_2.cell_value(0, 0)
            max_row = sheet_1.nrows
            hg = []
            b = []
            p = []
            np = []
            m = 0
            for i in range(3, max_row):
                if sheet_1.cell_value(i, 1) == 'Total':
                    m = 1
                    break
                hg.append(sheet_1.cell_value(i, 2))
                b.append(sheet_1.cell_value(i, 5))
                p.append(sheet_1.cell_value(i, 8))

                k = sheet_1.cell_value(i, 9)
                if (k[0:5] == 'Refer'):
                    u = k[11:]
                    # print(u)
                    sheet_4 = wb.sheet_by_index(3)
                    sheet_4.cell_value(0, 0)
                    for rw in range(0, sheet_4.nrows):
                        ty = sheet_4.cell_value(rw, 1)
                        # print(ty)
                        if (ty == u):
                            np.append(sheet_4.cell_value(rw, 2))
                else:
                    np.append(sheet_1.cell_value(i, 9))
            # Refer Note #1
            sheet_1.cell_value(0, 0)
            hg = [sub.replace('a.ii)', '') for sub in hg]
            max_row = sheet_2.nrows
            if (m == 0):
                for i in range(3, max_row):
                    if sheet_2.cell_value(i, 1) == 'Total':
                        break
                    hg.append(sheet_2.cell_value(i, 2))
                    b.append(sheet_2.cell_value(i, 5))
                    p.append(sheet_2.cell_value(i, 8))
                    k = sheet_2.cell_value(i, 9)
                    if (k[0:5] == 'Refer'):
                        u = k[11:]
                        # print(u)
                        sheet_4 = wb.sheet_by_index(3)
                        sheet_4.cell_value(0, 0)
                        for rw in range(0, sheet_4.nrows):
                            ty = sheet_4.cell_value(rw, 1)
                            # print(ty)
                            if (ty == u):
                                np.append(sheet_4.cell_value(rw, 2))
                    else:
                        np.append(sheet_2.cell_value(i, 9))

            # print(hg)
            for i in range(0, len(hg)):
                row_num = s3.max_row + 1
                wq += 1
                s3.cell(row=row_num, column=1).value = wq
                s3.cell(row=row_num, column=2).value = cli
                s3.cell(row=row_num, column=3).value = hg[i]
                s3.cell(row=row_num, column=4).value = b[i]
                s3.cell(row=row_num, column=5).value = p[i]
                s3.cell(row=row_num, column=6).value = np[i]
            a = 0
            for wd in wb.sheets():
                if wd.cell_value(1, 1) == 'Section':
                    break
                a += 1
            # print(a)
            sheet_2 = wb.sheet_by_index(a)
            r = []
            ro = []
            max_row = sheet_2.nrows
            for i in range(2, max_row):
                # r.append(sheet_2.cell_value(i,1))
                ro.append(sheet_2.cell_value(i, 2))
            for i in range(0, len(ro)):
                # s2.cell(row=row_num, column=3).value = r[i]
                s1.cell(row=t + 2, column=i + 9).value = ro[i]
            # print(ro)
            if (f.find('Total Deduction') != -1):
                w1 = f.find('Total Deduction') + 16
                g = f[w1:]
                u1 = g.find('\n') + w1
                ty = (f[w1:u1])
                # print(f[w1:u1])
                ty = ty.replace('  ', '')
                s1.cell(row=t + 2, column=15).value = ty
            if(f.find('Less: Hospital Discounts')!=-1):
                w1=f.find('Less: Hospital Discounts')+25
                g=f[w1:]
                u1=g.find('\n')+w1
                ty=(f[w1:u1])
                dis1=ty.replace('  ','')
            else:dis1=0
            if(f.find('Less: Network Hospital Discount')!=-1):
                w1=f.find('Less: Network Hospital Discount')+31
                g=f[w1:]
                u1=g.find('\n')+w1
                ty=(f[w1:u1])
                dis2=ty.replace('  ','')
            else:dis2=0
            s1.cell(row=t+2, column=16).value = float(dis1)+float(dis2)
        except Exception as e:
            log_exceptions()
            eu.append(t)
            print(onlyfiles[t], e)
    for t in range(0, len(onlyfiles)):
        if t in eu:
            s1.cell(row=t + 2, column=1).fill = redFill
            s1.cell(row=t + 2, column=1).value = 'error'
    print("Done")
    wbk.save(wbkName)
    wbk.close
    wbkName = 'count/count.xlsx'
    wbk = openpyxl.load_workbook(wbkName)
    wbk.create_sheet('count_star')
    s1 = wbk.worksheets[1]
    s1.cell(row=1, column=1).value = 'insurance id'
    s1.cell(row=1, column=2).value = 'mail count'
    s1.cell(row=1, column=3).value = 'attachments count'
    s1.cell(row=2, column=1).value = 'star_big'
    s1.cell(row=2, column=2).value = len(fg)
    s1.cell(row=2, column=3).value = len(onlyfiles)
    wbk.save(wbkName)
    wbk.close()
    pass
    subprocess.run(["python", "updation.py", "1", "max", "9", " "])
except:
    log_exceptions()