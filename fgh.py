from email.mime.text import MIMEText
import smtplib
import time
import imaplib
import email
import camelot
import PyPDF2
import xlsxwriter
from xlrd import open_workbook
from openpyxl import load_workbook
# from tabula import read_pdf
import xlwt
import os
import os.path
import sys
import xlrd
from os import listdir
from os import path
from os.path import isfile, join
import pdftotext
import tabula
import pandas as pd
# from tabula import read_pdf
import glob

from decode_error import check_subject, read_from_delete
from make_log import log_exceptions
import subprocess
try:

    fg = []
    repeat = []
    subprocess.run(["python", "updation.py", "1", "max", "9", "X"])


    def read_email_from_gmail():
        b = 0
        SMTP_SERVER = str(sys.argv[5])
        mail = imaplib.IMAP4_SSL(SMTP_SERVER)
        e_id = str(sys.argv[1])
        pswd = str(sys.argv[2])
        srt = str(sys.argv[3])
        stp = str(sys.argv[4])
        mail.login(user=e_id, password=pswd)
        mail.select("inbox", readonly=True)
        ###############################################<
        mail_uid = str(sys.argv[7])
        if mail_uid == -1:
            type, data = mail.search(None,
                                     '(SUBJECT "NEFT Claim Payout - Patient Name" since ' + srt + ' before ' + stp + ')')
            ids = data[0]
            id_list = ids.split()
        else:
            ids = mail_uid  # data is a list.
            # accept id from outside and put in id_list akshay var name = id

            id_list = []  # ids is a space separated string
            id_list.append(ids)
        ###############################################>
        # type, data = mail.search(None, '(SUBJECT "NEFT Claim Payout - Patient Name" since ' + srt + ' before ' + stp + ')')
        # ids = data[0]  # data is a list.
        # id_list = ids.split()  # ids is a space separated string
        # # print(id_list)
        for i in range(0, len(id_list)):
            latest_email_id = id_list[i]  # get the latest
            result, data = mail.fetch(latest_email_id,
                                      "(RFC822)")  # fetch the email body (RFC822)             for the given ID

            ##################################################ak
            try:
                # raw_email = data[0][1].decode('utf-8')
                try:
                    raw_email = data[0][1].decode('utf-8')
                except UnicodeDecodeError:
                    try:
                        raw_email = data[0][1].decode('ISO-8859-1')
                    except UnicodeDecodeError:
                        try:
                            raw_email = data[0][1].decode('ascii')
                        except UnicodeDecodeError:
                            pass
                email_message = email.message_from_string(raw_email)
                subject = email_message['Subject']
                result, sys.argv[8] = check_subject(subject, sys.argv[8], mail)
                if result == 'Changed':
                    # raise Exception('subject not matched')
                    raise Exception('subject not matched', )
            except:
                try:
                    log_exceptions(syssubject=sys.argv[8], subject=subject, error='subject not matched')
                except:
                    pass
                if result != 'OK':
                    data = {'server': SMTP_SERVER,
                            'hospmail': e_id,
                            'pass': pswd,
                            'subject': sys.argv[8]}
                    try:
                        data = read_from_delete(data)
                        if data == None:
                            raise Exception("Not found")
                    except:
                        log_exceptions(msg='not found in deleted', subject=sys.argv[8])
            ##################################################akend

            raw_email = data[0][1].decode('utf-8')
            email_message = email.message_from_string(raw_email)
            if email_message['Subject'] not in fg:
                for part in email_message.walk():

                    if part.get_content_maintype() == 'multipart':
                        # print part.as_string()
                        continue
                    if part.get('Content-Disposition') is None:
                        # print part.as_string()
                        continue
                    fileName = part.get_filename()

                    detach_dir = (os.getcwd() + '/FGH/attachments_' + str(sys.argv[6]))
                    if bool(fileName):
                        # print (email_message['Subject'])
                        filePath = os.path.join(detach_dir, fileName)
                        if not os.path.isfile(filePath):
                            from reportlab.pdfgen import canvas
                            c = canvas.Canvas(fileName)
                            # print (fileName)
                            fp = open(filePath, 'wb')
                            fp.write(part.get_payload(decode=True))
                            fp.close()
            else:
                repeat.append(email_message['Subject'])
            fg.append(email_message['Subject'])


    mypath = os.getcwd() + '/FGH'
    if not path.exists(mypath):
        os.mkdir(mypath)
    if not path.exists(mypath + '/attachments_' + str(sys.argv[6])):
        os.mkdir(mypath + '/attachments_' + str(sys.argv[6]))
    mypath = os.getcwd() + '/FGH/attachments_' + str(sys.argv[6]) + '/'
    for filename in os.listdir(mypath):
        file_path = os.path.join(mypath, filename)
        if os.path.isfile(file_path) or os.path.islink(file_path):
            os.unlink(file_path)
    read_email_from_gmail()
    # CCN='20439541'
    wq = 0
    onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
    valid = []
    '''
    for name in glob.glob(mypath + '/Claim*?'):
        valid.append(name)
    for t in range(0, len(onlyfiles)):
        if (mypath + onlyfiles[t] not in valid):
            os.remove(mypath + onlyfiles[t])
    onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
    if path.exists(r'FGH/FGH' + str(sys.argv[6]) + '.xlsx'):
        os.remove(r'FGH/FGH' + str(sys.argv[6]) + '.xlsx')
    '''
    import openpyxl

    po = []
    wbkName = 'FGH/FGH' + str(sys.argv[6]) + '.xlsx'
    wbk = openpyxl.Workbook()
    wbk.create_sheet('Sheet1')
    # wbk.create_sheet('Sheet3')
    s1 = wbk.worksheets[0]
    s2 = wbk.worksheets[1]
    for t in range(0, len(onlyfiles)):
        try:
            sh1 = ['Policy Number', 'Reference/UTR No.', 'Payment Date', 'ID Card Number', 'Discount Deduction',
                   'Co-Payment']
            for i in range(0, len(sh1)):
                s1.cell(row=1, column=i + 16).value = sh1[i]
            sh2 = ['Sr. No.', 'Claim Number', 'Biling Head', 'Bill No.', 'Claimed', 'Disallowed', 'Approved',
                   'Disallowed Reason']
            sh3 = ['Patient Name', 'Hospital Name', 'Date of Admission', 'Date of Discharge', 'Hospital Bill Number',
                   'Beneficiary Name', 'Payment Type', 'A/C No', 'Bank Name', 'Bill Amount', 'Disallowed Amount',
                   'Service Tax', 'Approved Amount', 'TDS']
            for i in range(0, len(sh2)):
                s2.cell(row=1, column=i + 1).value = sh2[i]
            tables = camelot.read_pdf(mypath + onlyfiles[t], pages='all', Line_scale=100)
            tables.export('FGH/foo1.xls', f='excel')
            loc = ("FGH/foo1.xls")
            wb = xlrd.open_workbook(loc)
            s = []
            d = []
            sheet_3 = wb.sheet_by_index(0)
            sheet_3.cell_value(0, 0)

            for i in range(1, sheet_3.nrows):
                s.append(sheet_3.cell_value(i, 1))
                d.append(sheet_3.cell_value(i, 2))
            ccn = d[5]
            s.pop(5)
            d.pop(5)
            if (len(s) == 13):
                for i in range(0, len(s)):
                    s1.cell(row=1, column=i + 3).value = s[i]
                    s1.cell(row=t + 2, column=i + 3).value = d[i]
            else:
                ry = 0
                for i in range(0, len(s)):

                    if s[i] in sh3:
                        # print(s[i],  d[i])
                        s1.cell(row=1, column=ry + 3).value = s[i]
                        s1.cell(row=t + 2, column=ry + 3).value = d[i]
                        ry = ry + 1
            p = []
            r = []
            ro = []
            po = []
            e = []
            eo = []
            sheet_2 = wb.sheet_by_index(1)
            max_row = sheet_2.nrows
            for i in range(2, max_row):
                p.append(sheet_2.cell_value(i, 1))
                r.append(sheet_2.cell_value(i, 2))
                ro.append(sheet_2.cell_value(i, 3))
                po.append(sheet_2.cell_value(i, 4))
                e.append(sheet_2.cell_value(i, 5))
                eo.append(sheet_2.cell_value(i, 6))
            # print(p)
            for i in range(0, len(p)):
                if (p[i] == 'Discount Deduction'):
                    discount = po[i]
                    s1.cell(row=t + 2, column=20).value = discount
                # s1.cell(row=t+2, column=12).value =float(s1.cell(row=t+2, column=12).value)-float(discount)
                if (p[i] == 'Co-Payment'):
                    CoPayment = po[i]
                    s1.cell(row=t + 2, column=21).value = CoPayment
                # s1.cell(row=t+2, column=12).value =float(s1.cell(row=t+2, column=12)).value-float(CoPayment)
                row_num = s2.max_row
                wq += 1
                s2.cell(row=row_num + 1, column=1).value = wq
                s2.cell(row=row_num + 1, column=2).value = ccn
                s2.cell(row=row_num + 1, column=3).value = p[i]
                s2.cell(row=row_num + 1, column=4).value = r[i]
                s2.cell(row=row_num + 1, column=5).value = ro[i]
                s2.cell(row=row_num + 1, column=6).value = po[i]
                s2.cell(row=row_num + 1, column=7).value = e[i]
                s2.cell(row=row_num + 1, column=8).value = eo[i]
            with open(mypath + onlyfiles[t], "rb") as f:
                pdf = pdftotext.PDF(f)

            with open('FGH/output.txt', 'w') as f:
                f.write(" ".join(pdf))
            with open('FGH/output.txt', 'r') as myfile:
                f = myfile.read()

            hg = []
            w = f.find('Policy Number') + 13
            g = f[w:]
            u = g.find('Reference/UTR No.') + w
            hg.append(f[w:u])

            w1 = f.find('Reference/UTR No.') + 17
            g = f[w1:]
            u1 = g.find('\n') + w1
            hg.append(f[w1:u1])

            if f.find('Payment Date') != -1:
                w2 = f.find('Payment Date') + 13
                g = f[w2:]
                u2 = g.find('\n') + w2
                hg.append(f[w2:u2])
            else:
                w2 = f.find('Date') + 6
                g = f[w2:]
                u2 = g.find('\n') + w2
                hg.append(f[w2:u2])

            w9 = f.find('ID Card Number :') + 17
            g = f[w9:]
            u9 = g.find('\n') + w9
            hg.append(f[w9:u9])

            hg = [sub.replace('  ', '') for sub in hg]
            hg = [sub.replace(':', '') for sub in hg]
            hg = [sub.replace('\n', ' ') for sub in hg]
            s1.cell(row=1, column=1).value = 'Sr. No.'
            s1.cell(row=1, column=2).value = 'claim number'
            s1.cell(row=t + 2, column=1).value = t + 1
            s1.cell(row=t + 2, column=2).value = ccn
            for i in range(0, len(hg)):
                s1.cell(row=t + 2, column=i + 16).value = hg[i]
            os.rename(os.getcwd() + '/FGH/attachments_' + str(sys.argv[6]) + '/' + onlyfiles[t],
                      os.getcwd() + '/FGH/attachments_' + str(sys.argv[6]) + '/' + ccn + '.pdf')
        except Exception as e:
            print(e)
            s1.cell(row=t + 2, column=1).value = 'error'
            os.rename(os.getcwd() + '/FGH/attachments_' + str(sys.argv[6]) + '/' + onlyfiles[t],
                      os.getcwd() + '/FGH/attachments_' + str(sys.argv[6]) + '/' + ccn + '.pdf')

    # print(po)
    print("Done")
    wbk.save(wbkName)
    wbk.close
    wbkName = 'count/count.xlsx'
    wbk = openpyxl.load_workbook(wbkName)
    s1 = wbk.worksheets[0]

    row_ = s1.max_row + 1
    s1.cell(row=row_, column=1).value = 'fgh'
    s1.cell(row=row_, column=2).value = str(sys.argv[6])
    s1.cell(row=row_, column=3).value = len(fg)
    s1.cell(row=row_, column=4).value = len(onlyfiles)
    s1.cell(row=row_, column=5).value = len(repeat)
    wbk.save(wbkName)
    subprocess.run(["python", "updation.py", "1", "max", "9", " "])
except:
    log_exceptions()