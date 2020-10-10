import os
import sys, re
import struct, time
import subprocess
from datetime import date
import datetime
import openpyxl
import pdftotext
import time
import requests
now = datetime.datetime.now()
'''
wbkName = 'log file.xlsx'
wbk= openpyxl.load_workbook(wbkName)
s1=wbk.worksheets[0]
s2=wbk.worksheets[1]
row_count_1 = s2.max_row
s2.cell(row_count_1+1, column=1).value=sys.argv[2]
s2.cell(row_count_1+1, column=2).value=sys.argv[3]
s2.cell(row_count_1+1, column=3).value=sys.argv[4]
s2.cell(row_count_1+1, column=5).value=now
s2.cell(row_count_1+1, column=7).value=sys.argv[5]
s2.cell(row_count_1+1, column=8).value=sys.argv[6]
'''

subprocess.run(["python", "updation.py","1","max1","1",sys.argv[2]])
subprocess.run(["python", "updation.py","1","max","2",sys.argv[3]])
subprocess.run(["python", "updation.py","1","max","3",sys.argv[4]])
subprocess.run(["python", "updation.py","1","max","5",str(now)])
subprocess.run(["python", "updation.py","1","max","7",sys.argv[5]])
subprocess.run(["python", "updation.py","1","max","8",sys.argv[6]])

with open(sys.argv[1], "rb") as f:
	pdf = pdftotext.PDF(f)

with open('health_heritage/output.txt', 'w',encoding='utf-8') as f:
	f.write(" ".join(pdf))     
with open('health_heritage/output.txt', 'r',encoding='utf-8') as myfile:
 	f = myfile.read()
with open('health_heritage/output.txt', 'r',encoding='utf-8') as myfile:
 	templist = myfile.readlines()
try:		
	hg=[]

	if f.find('Claim No') != -1:
		x = re.search(r"Claim No +\S+", f)
		x1 = x.group()
		claimno = x1.split(':')[-1]
		hg.append(claimno)
	elif f.find('Claim No') != -1:
		w=f.find('Claim No')+7
		g=f[w:]
		u=g.find('Policy')+w
		hg.append(f[w:u])
	else:
		hg.append('')

	status='Information Awaiting'
	hg.append('')

	if f.find('Policy No.') != -1:
		for i, j in enumerate(templist):
			if 'Claim No :' in j:
				a = templist[i+2]
				x = re.search(r"\S+", a)
				polno = x.group()
				hg.append(polno)
				break
	elif f.find('Policy No.') != -1:
		w=f.find('Policy No.')+10
		g=f[w:]
		u=g.find('Agent')+w
		hg.append(f[w:u])

	hg.append('')
	# maake hg[4] of patient name

	if f.find("Patient's Name :") != -1:
		x = re.search(r"Patient's Name : +\S+", f)
		x1 = x.group()
		x2 = x1.split(":")[-1]
		hg.append(x2)	
	else:
		hg.append('')

	hg=[sub.replace(':','') for sub in hg]	
	hg=[sub.replace('  ','') for sub in hg]
	hg=[sub.replace('Rs.','') for sub in hg]
	#s2.cell(row_count_1+1, column=9).value='Yes'
	#s2.cell(row_count_1+1, column=10).value='NA'
	#wbk.save(wbkName)
	
	subprocess.run(["python", "updation.py","1","max","9",'Yes'])
	subprocess.run(["python", "updation.py","1","max","10",'NA'])
	
	try:
		subprocess.run(["python", "test_api.py",hg[0],hg[1],hg[2],'',status,sys.argv[6],sys.argv[1],'',hg[3],hg[4]])
		'''wbk= openpyxl.load_workbook(wbkName)
		s2=wbk.worksheets[1]
		s2.cell(row_count_1+1, column=11).value='YES'
		'''
		subprocess.run(["python", "updation.py","1","max","11",'Yes'])	
	except Exception as e:
		#s2.cell(row_count_1+1, column=11).value='NO'
		subprocess.run(["python", "updation.py","1","max","11",'No'])
except Exception as e:
	#s2.cell(row_count_1+1, column=9).value='No'
	#s2.cell(row_count_1+1, column=11).value='NO'
	subprocess.run(["python", "updation.py","1","max","9",'Yes'])
	subprocess.run(["python", "updation.py","1","max","11",'No'])
now = datetime.datetime.now()
#s2.cell(row_count_1+1, column=6).value=now
#wbk.save(wbkName)
subprocess.run(["python", "updation.py","1","max","6",str(now)])
