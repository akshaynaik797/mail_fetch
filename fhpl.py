import re
import pandas as pd
from email.mime.text import MIMEText
import smtplib
import time
import imaplib
import email
import camelot
import PyPDF2
import csv
import xlsxwriter
import sys
from xlrd import open_workbook
import xlwt
import os
import glob
import os.path
import xlrd
from os import listdir
from os import path
from os.path import isfile, join
from html.parser import HTMLParser
import pdfkit
import pandas as pd
import pdftotext
import html2text
from openpyxl.styles import Color, PatternFill, Font, Border
import openpyxl
import subprocess
from decode_error import check_subject, read_from_delete
from make_log import log_exceptions

try:

    # sys.argv = ['fhpl.py', 'mediclaim@inamdarhospital.org', 'Mediclaim@2019', '30-Jul-2020', '30-Jul-2020', 'imap.gmail.com', 'inamdar', '23659']
    claim = None
    subprocess.run(["python", "updation.py", "1", "max", "9", "X"])
    fg = []
    repeat = []

    if path.exists(r'fhpl/fhpl' + str(sys.argv[6]) + '.xlsx'):
        os.remove(r'fhpl/fhpl' + str(sys.argv[6]) + '.xlsx')

    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')
    wq = 0


    def read_email_from_gmail():
        global claim
        SMTP_SERVER = str(sys.argv[5])
        mail = imaplib.IMAP4_SSL(SMTP_SERVER)
        # mail.login(user = 'Mediclaim@therisingmedicare.com', password = 'cef@2018')

        e_id = str(sys.argv[1])
        pswd = str(sys.argv[2])
        srt = str(sys.argv[3])
        stp = str(sys.argv[4])
        mail.login(user=e_id, password=pswd)
        mail.select("inbox", readonly=True)
        ###############################################<
        mail_uid = str(sys.argv[7])
        if mail_uid == -1:
            type, data = mail.search(None,
                                     '(SUBJECT "Cashless Settlement Letter : Patient Name" since ' + srt + ' before ' + stp + ')')
            ids = data[0]
            id_list = ids.split()
        else:
            ids = mail_uid  # data is a list.
            # accept id from outside and put in id_list akshay var name = id

            id_list = []  # ids is a space separated string
            id_list.append(ids)
        ###############################################>
        # type, data = mail.search(None,
        #                          '(SUBJECT "Cashless Settlement Letter : Patient Name" since ' + srt + ' before ' + stp + ')')
        # ids = data[0]  # data is a list.
        # id_list = ids.split()  # ids is a space separated string
        # # print(id_list)
        b = 0
        for i in range(0, len(id_list)):
            latest_email_id = id_list[i]  # get the latest
            result, data = mail.fetch(latest_email_id, "(RFC822)")

            ##################################################ak
            try:
                # raw_email = data[0][1].decode('utf-8')
                try:
                    raw_email = data[0][1].decode('utf-8')
                except UnicodeDecodeError:
                    try:
                        raw_email = data[0][1].decode('ISO-8859-1')
                    except UnicodeDecodeError:
                        try:
                            raw_email = data[0][1].decode('ascii')
                        except UnicodeDecodeError:
                            pass
                email_message = email.message_from_string(raw_email)
                subject = email_message['Subject']
                result, sys.argv[8] = check_subject(subject, sys.argv[8], mail)
                if result == 'Changed':
                    # raise Exception('subject not matched')
                    raise Exception('subject not matched', )
            except:
                try:
                    log_exceptions(syssubject=sys.argv[8], subject=subject, error='subject not matched')
                except:
                    pass
                if result != 'OK':
                    data = {'server': SMTP_SERVER,
                            'hospmail': e_id,
                            'pass': pswd,
                            'subject': sys.argv[8]}
                    try:
                        data = read_from_delete(data)
                        if data == None:
                            raise Exception("Not found")
                    except:
                        log_exceptions(msg='not found in deleted', subject=sys.argv[8])
            ##################################################akend
            raw_email = data[0][1].decode('utf-8')
            email_message = email.message_from_string(raw_email)
            # if path.exists(r'/home/shivam/Desktop/vnu_scripts/Paramount/email.html'):
            # os.remove(r'email.html')
            # Body details
            if email_message['Subject'] not in fg:
                b += 1
                for part in email_message.walk():
                    # print(part.get_content_type())
                    if part.get_content_type() == "text/html":
                        # print('hi')
                        body = part.get_payload(decode=True)
                        file_name = 'fhpl/attachments_' + str(sys.argv[6]) + '/email' + str(b) + ".html"
                        output_file = open(file_name, 'w')
                        output_file.write("Body: %s" % (body.decode('utf-8')))
                        output_file.close()
            else:
                repeat.append(email_message['Subject'])
            if not 'Claim' in email_message['Subject']:
                fg.append(email_message['Subject'])
                claim = None
            else:
                fg.append(email_message['Subject'])
                claim = email_message['Subject']

    def read_alno(ccn, d):
        SMTP_SERVER = str(sys.argv[5])
        mail = imaplib.IMAP4_SSL(SMTP_SERVER)
        # mail.login(user = 'Mediclaim@therisingmedicare.com', password = 'cef@2018')

        e_id = str(sys.argv[1])
        pswd = str(sys.argv[2])
        mail.login(user=e_id, password=pswd)
        mail.select("inbox", readonly=True)
        d = d.replace(' ', '-')
        # print(d)
        type, data = mail.search(None, '(SUBJECT "Patient Name: ' + ccn + ', Uhid No:" since ' + d + ')')
        count = 0
        # print(data)
        if (data == [b'']):
            count = 1
            type, data = mail.search(None,
                                     '(SUBJECT "Cashless Initial Approval of Patient Name:' + ccn + '" since ' + d + ')')
        try:
            ids = data[0]  # data is a list.
            id_list = ids.split()
            latest_email_id = id_list[0]
            result, data = mail.fetch(latest_email_id, "(RFC822)")

            # raw_email = data[0][1].decode('utf-8')
            try:
                raw_email = data[0][1].decode('utf-8')
            except UnicodeDecodeError:
                try:
                    raw_email = data[0][1].decode('ISO-8859-1')
                except UnicodeDecodeError:
                    try:
                        raw_email = data[0][1].decode('ascii')
                    except UnicodeDecodeError:
                        pass

            email_message = email.message_from_string(raw_email)
            if (count == 0):
                v = email_message['Subject']
                x1 = v.find('PreauthID:') + 11
                return (v[x1:])
            if (count == 1):
                for part in email_message.walk():
                    # print(part.get_content_type())
                    if part.get_content_type() == "text/html":
                        # print('hi')
                        body = part.get_payload(decode=True)
                        v = body.decode('utf-8')
                        x1 = v.find('Claim Number:') + 13
                        x2 = v.find('(')
                        return (v[x1:x2])
        except IndexError as error:
            s1.cell(row=t + 2, column=14).fill = redFill
            s1.cell(row=t + 2, column=1).value = 'error'


    mypath = os.getcwd() + '/fhpl'
    if not path.exists(mypath):
        os.mkdir(mypath)
    if not path.exists(mypath + '/attachments_' + str(sys.argv[6])):
        os.mkdir(mypath + '/attachments_' + str(sys.argv[6]))
    mypath = os.getcwd() + '/fhpl/attachments_' + str(sys.argv[6]) + '/'
    for filename in os.listdir(mypath):
        file_path = os.path.join(mypath, filename)
        if os.path.isfile(file_path) or os.path.islink(file_path):
            os.unlink(file_path)
    read_email_from_gmail()
    onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
    wbkName = 'fhpl/fhpl' + str(sys.argv[6]) + '.xlsx'
    wbk = openpyxl.Workbook()
    wbk.create_sheet('1')
    wbk.create_sheet('Sheet3')
    s1 = wbk.worksheets[0]
    s2 = wbk.worksheets[1]
    s3 = wbk.worksheets[2]
    for t in range(0, len(onlyfiles)):

        try:
            sh1 = ['Sr No.', 'Claim ID', 'Patient Name', 'Policy No', 'Employee ID', 'Diagnosis', 'Card No.',
                   'Date of Admission', 'Date of Discharge', 'NEFT transaction number', 'transaction date',
                   'Corporate Name', 'insurance company', 'alno']
            sh2 = ['Sr No.', 'Claim ID', 'Claimed', 'Billed amount', 'Discount', 'Disallowed', 'Settled amount', 'Less TDS',
                   'Net Paid Amount']
            sh3 = ['Sr No.', 'Claim ID', 'category', 'Disallowance amount', 'Disallowance Reasons']

            for i in range(0, len(sh1)):
                s1.cell(row=1, column=i + 1).value = sh1[i]
            for i in range(0, len(sh2)):
                s2.cell(row=1, column=i + 1).value = sh2[i]
            for i in range(0, len(sh3)):
                s3.cell(row=1, column=i + 1).value = sh3[i]
            mypath = os.getcwd() + '/fhpl/attachments/'

            html = open('fhpl/attachments_' + str(sys.argv[6]) + '/email' + str(t + 1) + ".html")
            f = str(html.read())
            w = open("fhpl/out.txt", "w")
            w.write(html2text.html2text(f))
            html.close()
            w.close()
            with open('fhpl/out.txt', 'r') as myfile:
                f = myfile.read()

            if claim is None:
                gh = []
                w2 = f.find('Claim Of')
                g = f[w2:]
                w4 = g.find(':') + w2 + 2
                g = f[w4:]
                u2 = g.find('\n') + w4
                gh.append(f[w4:u2])

                w = f.find('Policy No')
                g = f[w:]
                w5 = g.find(':') + w + 2
                g = f[w5:]
                u3 = g.find('\n') + w5
                gh.append(f[w5:u3])

                w = f.find('Employee ID')
                g = f[w:]
                w5 = g.find(':') + w + 2
                g = f[w5:]
                u3 = g.find('\n') + w5
                gh.append(f[w5:u3])

                w3 = f.find('Diagnosis')
                g = f[w3:]
                w1 = g.find(':') + w3 + 2
                g = f[w1:]
                u1 = g.find('\n') + w1
                # print(w1,u1)
                gh.append(f[w1:u1])

                w3 = f.find('Card No.')
                g = f[w3:]
                w1 = g.find(':') + w3 + 2
                g = f[w1:]
                u1 = g.find('\n') + w1
                gh.append(f[w1:u1])

                w3 = f.find('Date of Admission')
                g = f[w3:]
                w1 = g.find(':') + w3 + 2
                g = f[w1:]
                u1 = g.find('\n') + w1
                gh.append(f[w1:u1])

                w3 = f.find('Date of Discharge')
                g = f[w3:]
                w1 = g.find(':') + w3 + 2
                g = f[w1:]
                u1 = g.find('\n') + w1
                gh.append(f[w1:u1])
                # print(w1,u1)

                w3 = f.find('number') + 7
                g = f[w3:]
                w1 = g.find('dated') + w3
                gh.append(f[w3:w1])
                g = f[w1:]
                u1 = g.find('.') + w1
                gh.append(f[w1 + 5:w1 + 19])

                w3 = f.find('Corporate Name')
                g = f[w3:]
                w1 = g.find(':') + w3 + 2
                g = f[w1:]
                u1 = g.find('\n') + w1
                gh.append(f[w1:u1])

                w3 = f.find('Good wishes from') + 17
                g = f[w3:]
                u1 = g.find('**') + w3
                gh.append(f[w3:u1])

                w3 = f.find('Claim ID')
                g = f[w3:]
                w1 = g.find(':') + w3 + 2
                g = f[w1:]
                u1 = g.find('\n') + w1
                ccn = f[w1:u1]
                gh = [sub.replace('\n', '') for sub in gh]
                # print(ccn)
                s1.cell(row=t + 2, column=1).value = t + 1
                s1.cell(row=t + 2, column=2).value = ccn
                gh = [sub.replace('**', '') for sub in gh]
                gh = [sub.replace('\n', '') for sub in gh]
                for i in range(0, len(gh)):
                    s1.cell(row=t + 2, column=i + 3).value = gh[i]
                # print(gh[4])
                subj = read_alno(gh[0], gh[5])
                if subj == None:
                    subj = ccn
                s1.cell(row=t + 2, column=2).value = ccn
                s1.cell(row=t + 2, column=14).value = subj
                # print(subj)
                gh = []
                w2 = f.find('Claimed')
                g = f[w2:]
                w4 = g.find('Rs.') + w2 + 4
                g = f[w4:]

                u2 = g.find('/-') + w4
                gh.append(f[w4:u2])

                w = f.find('Billed')
                g = f[w:]
                w5 = g.find(':') + w
                u3 = g.find('/-') + w
                gh.append(f[w5:u3])

                w = f.find('Discount')
                g = f[w:]
                w5 = g.find('Rs.') + w
                g = f[w5:]
                u3 = g.find('/-') + w5
                gh.append(f[w5:u3])

                w3 = f.find('Disallowed')
                g = f[w3:]
                w1 = g.find(':') + w3
                u1 = g.find('/-') + w3
                gh.append(f[w1:u1])

                w3 = f.find('Settled')
                g = f[w3:]
                w1 = g.find(':') + w3 + 2
                u1 = g.find('/-') + w3
                gh.append(f[w1:u1])

                w3 = f.find('Less TDS')
                g = f[w3:]
                w1 = g.find('Rs.') + w3
                u1 = g.find('/-') + w3
                gh.append(f[w1:u1])

                w3 = f.find('Net Paid Amount')
                g = f[w3:]
                w1 = g.find(':') + w3 + 2
                u1 = g.find('/-') + w3
                gh.append(f[w1:u1])
                s2.cell(row=t + 2, column=1).value = t + 1
                s2.cell(row=t + 2, column=2).value = ccn
                gh = [sub.replace('Rs.', '') for sub in gh]
                gh = [sub.replace(':', '') for sub in gh]
                gh = [sub.replace(' ', '') for sub in gh]
                for i in range(0, len(gh)):
                    s2.cell(row=t + 2, column=2).value = ccn
                    s2.cell(row=t + 2, column=i + 3).value = gh[i]
                # print(gh) Rs.

                w3 = f.find('Disallowance Reasons')
                g = f[w3:]
                w1 = g.find('Claimed') + w3
                g = f[w3:w1]
                # mylist = [item for item in g.split('\n')]
                mylist = [item for item in g.split('Rs.')]
                # mylist.remove('')
                # print(mylist)
                x = ''
                hj = 0
                for i in mylist:
                    x = i

                    if (hj == 0):
                        if x.find('_') != -1:
                            m = x.find('_')
                            l = x[m + 1:].find('_') + m
                            sd = x[m + 1:l]
                        # print(x)
                        hj = 1
                        continue

                    else:
                        row_num = s3.max_row + 1
                        if (x != ''):
                            wq += 1
                            m = i.find('\n')
                            i = i[:m]
                            m = i.find('.')
                            s = i[0:m + 3]
                            l = i[m + 3:]
                            s3.cell(row=row_num, column=1).value = wq
                            s3.cell(row=row_num, column=2).value = ccn
                            s3.cell(row=row_num, column=3).value = sd
                            s3.cell(row=row_num, column=4).value = s
                            s3.cell(row=row_num, column=5).value = l
                    if x.find('_') != -1:
                        m = x.find('_')
                        l = x[m + 1:].find('_') + m
                        sd = x[m + 1:l]
                    # print(x)
                ccn = ccn.replace(' ', '')
                os.rename(os.getcwd() + '/fhpl/attachments_' + str(sys.argv[6]) + '/email' + str(t + 1) + '.html',
                        os.getcwd() + '/fhpl/attachments_' + str(sys.argv[6]) + '/' + ccn + '.html')
            else:
                data = dict()
                if f.find('Employee ID') != -1:
                    regex = r'\S+(?= *\|\s*Claim ID)'
                    x = re.search(regex, f)
                    if x:
                        x1 = x.group().strip()
                        data['empid'] = x1
                    else:
                        data['empid'] = ''
                else:
                    data['empid'] = ''

                if f.find('Diagnosis') != -1:
                    regex = r'[ \w]+(?=\s*Date of Admission)'
                    x = re.search(regex, f)
                    if x:
                        x1 = x.group().strip()
                        data['diagnosis'] = x1
                    else:
                        data['diagnosis'] = ''
                else:
                    data['diagnosis'] = ''

                if f.find('Claim ID') != -1:
                    regex = r'\S+(?=\s*Hospital Name)'
                    x = re.search(regex, f)
                    if x:
                        x1 = x.group().strip()
                        data['claim_id'] = x1
                    else:
                        data['claim_id'] = ''
                else:
                    data['claim_id'] = ''

                if f.find('Claim Of') != -1:
                    regex = r'\S+ ?\S+(?=\s*---\|---)'
                    x = re.search(regex, f)
                    if x:
                        x1 = x.group().strip()
                        data['patient_name'] = x1
                    else:
                        data['patient_name'] = ''
                else:
                    data['patient_name'] = ''

                if f.find('Policy No.') != -1:
                    regex = r'\S+(?=\s*Card No.)'
                    x = re.search(regex, f)
                    if x:
                        x1 = x.group().strip()
                        data['policy_no'] = x1
                    else:
                        data['policy_no'] = ''
                else:
                    data['policy_no'] = ''

                if f.find('Card No.') != -1:
                    regex = r'\S+(?=\s*Payee Name)'
                    x = re.search(regex, f)
                    if x:
                        x1 = x.group().strip()
                        data['card_no'] = x1
                    else:
                        data['card_no'] = ''
                else:
                    data['card_no'] = ''

                if f.find('Date of Admission') != -1:
                    regex = r'( ?\S+){3}(?= *\|\s*Date of Discharge)'
                    x = re.search(regex, f)
                    if x:
                        x1 = x.group().strip()
                        data['doa'] = x1
                    else:
                        data['doa'] = ''
                else:
                    data['doa'] = ''

                if f.find('Date of Discharge') != -1:
                    regex = r'( ?\S+){3}(?= *\|?\s*Relation)'
                    x = re.search(regex, f)
                    if x:
                        x1 = x.group().strip()
                        data['dod'] = x1
                    else:
                        data['dod'] = ''
                else:
                    data['dod'] = ''

                if f.find('NEFT') != -1:
                    regex = r'\d+(?=\r?\n* *dated)'
                    x = re.search(regex, f)
                    if x:
                        x1 = x.group().strip()
                        data['neft'] = x1
                    else:
                        data['neft'] = ''
                else:
                    data['neft'] = ''

                if f.find('dated') != -1:
                    regex = r'(?<=dated)\s*\S+'
                    x = re.search(regex, f)
                    if x:
                        x1 = x.group().strip()
                        data['transaction_date'] = x1
                    else:
                        data['transaction_date'] = ''
                else:
                    data['transaction_date'] = ''

                if f.find('dated') != -1:
                    regex = r'(?<=dated)\s*\S+'
                    x = re.search(regex, f)
                    if x:
                        x1 = x.group().strip()
                        data['transaction_date'] = x1
                    else:
                        data['transaction_date'] = ''
                else:
                    data['transaction_date'] = ''

                if f.find('Corporate Name') != -1:
                    regex = r'(?<=Corporate Name)\s*\S+[ \S]+(?=\r?\n)'
                    x = re.search(regex, f)
                    if x:
                        x1 = x.group().strip()
                        data['corporate_name'] = x1
                    else:
                        data['corporate_name'] = ''
                else:
                    data['corporate_name'] = ''

                if f.find('instructions of') != -1:
                    regex = r'(?<=instructions of)[ \S]+(?=your)'
                    x = re.search(regex, f)
                    if x:
                        x1 = x.group().strip()
                        data['insurance_company'] = x1
                    else:
                        data['insurance_company'] = ''
                else:
                    data['insurance_company'] = ''

                data['alno'] = data['claim_id']

                if f.find('Claimed') != -1:
                    regex = r'\S+(?=/-\s*---\|---)'
                    x = re.search(regex, f)
                    if x:
                        x1 = x.group().strip()
                        data['claimed_amount'] = x1
                    else:
                        data['claimed_amount'] = ''
                else:
                    data['claimed_amount'] = ''

                if f.find('Billed') != -1:
                    regex = r'\S+(?=/-\s*Discount)'
                    x = re.search(regex, f)
                    if x:
                        x1 = x.group().strip()
                        data['billed_amount'] = x1
                    else:
                        data['billed_amount'] = ''
                else:
                    data['billed_amount'] = ''

                if f.find('Discount') != -1:
                    regex = r'\S+(?=/-\s*Disallowed)'
                    x = re.search(regex, f)
                    if x:
                        x1 = x.group().strip()
                        data['discount_amount'] = x1
                    else:
                        data['discount_amount'] = ''
                else:
                    data['discount_amount'] = ''

                if f.find('Disallowed') != -1:
                    regex = r'(?<=Rs.)\S+(?=/-\s*Settled)'
                    x = re.search(regex, f)
                    if x:
                        x1 = x.group().strip()
                        data['disallowed_amount'] = x1
                    else:
                        data['disallowed_amount'] = ''
                else:
                    data['disallowed_amount'] = ''

                if f.find('Settled') != -1:
                    regex = r'\S+(?=/-\s*Less TDS)'
                    x = re.search(regex, f)
                    if x:
                        x1 = x.group().strip()
                        data['settled_amount'] = x1
                    else:
                        data['settled_amount'] = ''
                else:
                    data['settled_amount'] = ''

                if f.find('Less TDS') != -1:
                    regex = r'\S+(?=/-\s*Net Paid Amount:)'
                    x = re.search(regex, f)
                    if x:
                        x1 = x.group().strip()
                        data['less_tds'] = x1
                    else:
                        data['less_tds'] = ''
                else:
                    data['less_tds'] = ''

                if f.find('Net Paid Amount') != -1:
                    regex = r'\S+(?=/-\s*Co-Payment Amount)'
                    x = re.search(regex, f)
                    if x:
                        x1 = x.group().strip()
                        data['net_paid_amount'] = x1
                    else:
                        data['net_paid_amount'] = ''
                else:
                    data['net_paid_amount'] = ''

                chars = ['\n', '|', '-']
                for key, value in data.items():
                    for i in chars:
                        value = value.replace(i, '')
                        data[key] = value.strip()

                # pprint(data)

                if f.find('Deduction Reason') != -1:
                    regex = r'(?<=Deduction Reason\*\*)[\s\S]+(?=\*\*\s*Claimed)'
                    x = re.search(regex, f)
                    if x:
                        x1 = x.group().strip().replace('---|---', "")

                regex = '(?<=Rs.)\d+'
                amounts = re.findall(regex, x1, re.MULTILINE)

                regex = '[ \S]+(?=\|)'
                catogories = re.findall(regex, x1, re.MULTILINE)

                regex = '(?<=:-)[ \w]+'
                reasons = re.findall(regex, x1, re.MULTILINE)
                if len(reasons) < len(amounts):
                    regex = '(?<=:-)[[\s\S][^,\n]*]*'
                    reasons = re.findall(regex, x1, re.MULTILINE)
                    reasons = [i.replace('\n', '') for i in reasons]



                table = []
                ccn = data['claim_id']
                for i, j in enumerate(amounts):
                    try:
                        table.append((catogories[i], j, reasons[i]))
                    except IndexError as e:
                        table.append((catogories[-1], j, reasons[i]))

                gh = [data['claim_id'], data['patient_name'], data['policy_no'], data['empid'], data['diagnosis'],
                      data['card_no'], data['doa'], data['dod'], data['neft'], data['transaction_date'],
                      data['corporate_name'], data['insurance_company'], data['alno'], ]
                s1.cell(row=t + 2, column=1).value = s1.max_row
                for i in range(0, len(gh)):
                    s1.cell(row=t + 2, column=i + 2).value = gh[i]
                gh = [data['claim_id'], data['claimed_amount'], data['billed_amount'], data['discount_amount'],
                      data['discount_amount'], data['settled_amount'], data['less_tds'],  data['net_paid_amount']]
                s2.cell(row=t + 2, column=1).value = s2.max_row
                for i in range(0, len(gh)):
                    s2.cell(row=t + 2, column=i + 2).value = gh[i]
                rowno = s3.max_row
                for i in table:
                    rowno = s3.max_row + 1
                    s3.cell(row=rowno, column=1).value = rowno-1
                    s3.cell(row=rowno, column=2).value = data['claim_id']
                    s3.cell(row=rowno, column=3).value = i[0]
                    s3.cell(row=rowno, column=4).value = i[1]
                    s3.cell(row=rowno, column=5).value = i[2]

        except Exception as e:
            print(e)
            s1.cell(row=t + 2, column=1).value = 'error'
            os.rename(os.getcwd() + '/fhpl/attachments_' + str(sys.argv[6]) + '/email' + str(t + 1) + '.html',
                      os.getcwd() + '/fhpl/attachments_' + str(sys.argv[6]) + '/' + ccn + '.html')

    print("Done")
    wbk.save(wbkName)
    wbk.close
    wbkName = 'count/count.xlsx'
    wbk = openpyxl.load_workbook(wbkName)
    s1 = wbk.worksheets[0]

    row_ = s1.max_row + 1
    s1.cell(row=row_, column=1).value = 'fhpl'
    s1.cell(row=row_, column=2).value = str(sys.argv[6])
    s1.cell(row=row_, column=3).value = len(fg)
    s1.cell(row=row_, column=4).value = len(onlyfiles)
    s1.cell(row=row_, column=5).value = len(repeat)
    wbk.save(wbkName)
    subprocess.run(["python", "updation.py", "1", "max", "9", " "])
except:
    log_exceptions()