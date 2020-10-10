import os
import sys
import struct, time
import subprocess
from datetime import date
import datetime
import openpyxl
import pdftotext
import time
import requests
now = datetime.datetime.now()

subprocess.run(["python", "updation.py","1","max1","1",sys.argv[2]])
subprocess.run(["python", "updation.py","1","max","2",sys.argv[3]])
subprocess.run(["python", "updation.py","1","max","3",sys.argv[4]])
subprocess.run(["python", "updation.py","1","max","5",str(now)])
subprocess.run(["python", "updation.py","1","max","7",sys.argv[5]])
subprocess.run(["python", "updation.py","1","max","8",sys.argv[6]])

with open(sys.argv[1], "rb") as f:
    pdf = pdftotext.PDF(f)

with open('health_insurance/output1.txt', 'w') as f:
    f.write(" ".join(pdf))  
with open('health_insurance/output1.txt', 'r') as myfile:
 	  f = myfile.read()  
           

try:
    hg=[]
    if f.find('Claim Number:') != -1:
      w=f.find('Claim Number:')+len('Claim Number:')
      k=f[w:]
      u=k.find("\n")+w
      g=f[w:u]
      if g.find('(') != -1:
          index=g.find('(')
          g=g[0:index]
      hg.append(g)
    else:
      print("Claim Number: not found in output1.text")
      hg.append('')

    if f.find ('Patient’s Member UHID' ) != -1 :
        w = f.find('Patient’s Member UHID') + len ('Patient’s Member UHID')
        k = f[w:]
        u = k.find ("\n" ) + w
        g = f[w :u]
        hg.append ( g )
    else :
        print ( "Patient’s Member UHID not found in output1.text" )
        hg.append('')

    if f.find('Patient Name') != -1:
      w=f.find('Patient Name')+len('Patient Name')
      k=f[w:]
      u=k.find("\n")+w
      g=f[w:u]
      if g.find('Age') != -1:
          index=g.find('Age')
          g=g[0:index]
      hg.append(g)
    else:
      print("Patient Name not found in output1.text")
      hg.append('')


    if f.find('Policy Number') != -1:
        w=f.find('Policy Number')+len('Policy Number')
        k=f[w:]
        u=k.find("\n")+w
        g=f[w:u]
        if g.find('Expected')!= -1:
          index=g.find('Expected')
          g=g[0:index]
          hg.append(g)
    else:
      print("Policy Number not found in output1.text")
      hg.append('')

    hg=[sub.replace(':','') for sub in hg]
    hg=[sub.replace('  ','') for sub in hg]
    hg=[sub.replace('Rs.','') for sub in hg]
    

    subprocess.run(["python", "updation.py","1","max","9",'Yes'])
    subprocess.run(["python", "updation.py","1","max","10",'NA'])
    try:  
      
      subprocess.run(["python", "test_api.py",hg[0],'',hg[3],'','Information Awaiting',sys.argv[6],sys.argv[1],'',hg[1],hg[2]])

      subprocess.run(["python", "updation.py","1","max","11",'Yes'])
    except Exception as e:
        #s2.cell(row_count_1+1, column=11).value='NO'
        subprocess.run(["python", "updation.py","1","max","11",'No'])
except Exception as e:
    #s2.cell(row_count_1+1, column=9).value='No'
    #s2.cell(row_count_1+1, column=11).value='NO'
    subprocess.run(["python", "updation.py","1","max","9",'Yes'])
    subprocess.run(["python", "updation.py","1","max","11",'No'])
now = datetime.datetime.now()
#s2.cell(row_count_1+1, column=6).value=now
#wbk.save(wbkName)
subprocess.run(["python", "updation.py","1","max","6",str(now)])
