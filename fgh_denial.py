import os
import sys
import struct, time
import subprocess
from datetime import date
import datetime
import openpyxl
import pdftotext
import time
import requests
now = datetime.datetime.now()

subprocess.run(["python", "updation.py","1","max1","1",sys.argv[2]])
subprocess.run(["python", "updation.py","1","max","2",sys.argv[3]])
subprocess.run(["python", "updation.py","1","max","3",sys.argv[4]])
subprocess.run(["python", "updation.py","1","max","5",str(now)])
subprocess.run(["python", "updation.py","1","max","7",sys.argv[5]])
subprocess.run(["python", "updation.py","1","max","8",sys.argv[6]])

with open(sys.argv[1], "rb") as f:
	pdf = pdftotext.PDF(f)

# For Dennial pdf is written to output1.txt and for ack pdf is written to output.txt
with open('fgh/output1.txt', 'w', encoding='utf-8') as f:
	f.write(" ".join(pdf))     
with open('fgh/output1.txt', 'r', encoding='utf-8') as myfile:
 	f = myfile.read()

# with open('fgh/output.txt', 'w', encoding='utf-8') as f:
# 	f.write(" ".join(pdf))     
# with open('fgh/output.txt', 'r', encoding='utf-8') as myfile:
#  	f = myfile.read()

try:

	hg=[]
	if f.find('Claim Number')!= -1:
		w = f.find('Claim Number') + 13
		k = f[w:]
		u = k.find("\n") + w
		g = f[w:u]
		hg.append(g)
	else:
		print("Claim number not found in output1.txt")
    
	if f.find('Hospitalization of')!= -1:
		w = f.find('Hospitalization of') + 18
		k = f[w:]
		u = k.find("\n") + w
		g = f[w:u]
		hg.append(g)
    		
	if f.find('Card No.')!= -1:
		w = f.find('Card No.') + 8
		k = f[w:]
		u = k.find("\n") + w
		g = f[w:u]
		hg.append(g)
	else:
		print("Card No not found in output1.txt")
		
	# print("uyt")
	if f.find('DL No :')!= -1:
		w = f.find('DL No :') + 5
		# print(w)
		k = f[w:]
		u = k.find("Date") + w
		g = f[w:u]
		hg.append(g)
	else:
		print("DL No not found in output1.txt")

	hg=[sub.replace(':','') for sub in hg]
	hg=[sub.replace('  ','') for sub in hg]
	hg=[sub.replace('Rs.','') for sub in hg]
	#s2.cell(row_count_1+1, column=9).value='Yes'
	#s2.cell(row_count_1+1, column=10).value='NA'
	#wbk.save(wbkName)
	
	subprocess.run(["python", "updation.py","1","max","9",'Yes'])
	subprocess.run(["python", "updation.py","1","max","10",'NA'])
	
	try:
		subprocess.run(["python", "test_api.py",hg[0],'',hg[2],'','Denial',sys.argv[6],sys.argv[1],'',hg[3],hg[1]])
		'''wbk= openpyxl.load_workbook(wbkName)
		s2=wbk.worksheets[1]
		s2.cell(row_count_1+1, column=11).value='YES'
		'''
		subprocess.run(["python", "updation.py","1","max","11",'Yes'])	
	except Exception as e:
		#s2.cell(row_count_1+1, column=11).value='NO'
		subprocess.run(["python", "updation.py","1","max","11",'No'])
except Exception as e:
	#s2.cell(row_count_1+1, column=9).value='No'
	#s2.cell(row_count_1+1, column=11).value='NO'
	subprocess.run(["python", "updation.py","1","max","9",'Yes'])
	subprocess.run(["python", "updation.py","1","max","11",'No'])
now = datetime.datetime.now()
#s2.cell(row_count_1+1, column=6).value=now
#wbk.save(wbkName)
subprocess.run(["python", "updation.py","1","max","6",str(now)])
