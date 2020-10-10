import os
import re
import sys
import struct, time
import subprocess
from datetime import date
import datetime
import openpyxl
import pdftotext
import time
import requests

from patient_name_fun import pname_fun

now = datetime.datetime.now()

subprocess.run(["python", "updation.py","1","max1","1",sys.argv[2]])
subprocess.run(["python", "updation.py","1","max","2",sys.argv[3]])
subprocess.run(["python", "updation.py","1","max","3",sys.argv[4]])
subprocess.run(["python", "updation.py","1","max","5",str(now)])
subprocess.run(["python", "updation.py","1","max","7",sys.argv[5]])
subprocess.run(["python", "updation.py","1","max","8",sys.argv[6]])

with open(sys.argv[1], "rb") as f:
	pdf = pdftotext.PDF(f)

with open('Medi_Assist/output.txt', 'w') as f:
	f.write(" ".join(pdf))     
with open('Medi_Assist/output.txt', 'r') as myfile:
 	f = myfile.read()

try:		
	hg=[]
	'''
	w=f.find('CCN')+4
	g=f[w:]
	u=g.find('\n')+w
	hg.append(f[w:u])
	if hg[0].find('MDI ID No')!=-1:
		temp_y=hg[0].find('MDI')
		g=hg[0]
		hg[0]=g[:temp_y]
	'''
	id = sys.argv[5]
	hg.append(id[38:])
	status='Information Awaiting'
	hg.append('')

	regex = r"(?<=Policy Number :).*(?=.)"
	result = re.search(regex, f)
	if result is not None:
		x = result
		x1 = x.group()
		hg.append(x1)
	elif f.find('Policy No') != -1:
		w = f.find('Policy No') + 10
		g = f[w:]
		u = g.find('\n') + w
		hg.append(f[w:u])
	else:
		hg.append('')

	regex = r"(?<=MDID Number :).*(?=.)"
	result = re.search(regex, f)
	if result is not None:
		x = result
		x1 = x.group()
		hg.append(x1)
	elif f.find('MDI ID No') != -1:
		w = f.find('MDI ID No') + 10
		g = f[w:]
		u = g.find('\n') + w
		hg.append(f[w:u])
	else:
		hg.append('')


	hg=[sub.replace(':','') for sub in hg]	
	hg=[sub.replace('  ','') for sub in hg]
	hg = [sub.replace('/', '') for sub in hg]
	hg = [sub.replace('-', '') for sub in hg]

	#s2.cell(row_count_1+1, column=9).value='Yes'
	#s2.cell(row_count_1+1, column=10).value='NA'
	#wbk.save(wbkName)
	
	subprocess.run(["python", "updation.py","1","max","9",'Yes'])
	subprocess.run(["python", "updation.py","1","max","10",'NA'])
	
	try:
		regex_list = [r"(?<=Patient).*"]
		pname = pname_fun(f, regex_list)
		subprocess.run(["python", "test_api.py",hg[0],hg[1],hg[2],'',status,sys.argv[6],sys.argv[1],'',hg[3], pname])
		'''wbk= openpyxl.load_workbook(wbkName)
		s2=wbk.worksheets[1]
		s2.cell(row_count_1+1, column=11).value='YES'
		'''
		subprocess.run(["python", "updation.py","1","max","11",'Yes'])	
	except Exception as e:
		#s2.cell(row_count_1+1, column=11).value='NO'
		subprocess.run(["python", "updation.py","1","max","11",'No'])
except Exception as e:
	#s2.cell(row_count_1+1, column=9).value='No'
	#s2.cell(row_count_1+1, column=11).value='NO'
	subprocess.run(["python", "updation.py","1","max","9",'Yes'])
	subprocess.run(["python", "updation.py","1","max","11",'No'])
now = datetime.datetime.now()
#s2.cell(row_count_1+1, column=6).value=now
#wbk.save(wbkName)
subprocess.run(["python", "updation.py","1","max","6",str(now)])
