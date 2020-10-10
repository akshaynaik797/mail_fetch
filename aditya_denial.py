import os
import sys
import struct, time
import subprocess
from datetime import date
import datetime
import openpyxl
import pdftotext
import time
import requests
from make_log import log_exceptions
from patient_name_fun import pname_fun

now = datetime.datetime.now()
'''
wbkName = 'log file.xlsx'
wbk= openpyxl.load_workbook(wbkName)
s1=wbk.worksheets[0]
s2=wbk.worksheets[1]
row_count_1 = s2.max_row
s2.cell(row_count_1+1, column=1).value=sys.argv[2]
s2.cell(row_count_1+1, column=2).value=sys.argv[3]
s2.cell(row_count_1+1, column=3).value=sys.argv[4]
s2.cell(row_count_1+1, column=5).value=now
s2.cell(row_count_1+1, column=7).value=sys.argv[5]
s2.cell(row_count_1+1, column=8).value=sys.argv[6]
'''

subprocess.run(["python", "updation.py","1","max1","1",sys.argv[2]])
subprocess.run(["python", "updation.py","1","max","2",sys.argv[3]])
subprocess.run(["python", "updation.py","1","max","3",sys.argv[4]])
subprocess.run(["python", "updation.py","1","max","5",str(now)])
subprocess.run(["python", "updation.py","1","max","7",sys.argv[5]])
subprocess.run(["python", "updation.py","1","max","8",sys.argv[6]])

#  by Ashish as 1) pdf file name is hard cord and 2)output1.txt file path is incorrect(hdfc/output1.txt)
with open(sys.argv[1], "rb") as f:
	pdf = pdftotext.PDF(f)

with open('aditya/output1.txt', 'w') as f:
	f.write(" ".join(pdf))     
with open('aditya/output1.txt', 'r') as myfile:
 	f = myfile.read()


# with open("13-18-0002835-01-00.pdf", "rb") as f:
# 	pdf = pdftotext.PDF(f)

# with open('hdfc/output1.txt', 'w') as f:
# 	f.write(" ".join(pdf))     
# with open('hdfc/output1.txt', 'r') as myfile:
#  	f = myfile.read()

try:
    hg=[]
    if f.find('Policy Number') != -1:
      w=f.find('Policy Number')+len('Policy Number')
      k=f[w:]
      u=k.find("\n")+w
      g=f[w:u]
      hg.append(g)
    else:
          print("Policy No not found in output1.text")

    if f.find('Claim number') != -1:
      w=f.find('Claim number')+len('Claim number')
      k=f[w:]
      u=k.find("\n")+w
      g=f[w:u]
      hg.append(g)
    else:
      print("Claim number not found in output1.text")

    if f.find('Patient Name') != -1:
      w=f.find('Patient Name')+len('Patient Name')
      k=f[w:]
      u=k.find("\n")+w
      g=f[w:u]
      hg.append(g)
    else:
      print("Patient Name not found in output1.text")

    if f.find('Claimed Amount') != -1:
      w=f.find('Claimed Amount')+len('Claimed Amount')
      k=f[w:]
      u=k.find("\n")+w
      g=f[w:u]
      hg.append(g.strip())
    else:
       print("Claimed Amount not found in output1.text")   

       

    hg=[sub.replace(':','') for sub in hg]	
    hg=[sub.replace('  ','') for sub in hg]
    hg=[sub.replace('Rs.','') for sub in hg]		
    hg=[sub.replace('-','') for sub in hg]
    regex_list = [r"(?<=Patient Name).*(?=Age)", r"(?<=Patient Name).*"]
    pname = pname_fun(f, regex_list)

    #s2.cell(row_count_1+1, column=9).value='Yes'
	#s2.cell(row_count_1+1, column=10).value='NA'
	#wbk.save(wbkName)
	
    subprocess.run(["python", "updation.py","1","max","9",'Yes'])
    subprocess.run(["python", "updation.py","1","max","10",'NA'])
    try:  
      # by Ashish sys.arv[1] has been added
      subprocess.run(["python", "test_api.py",hg[1],hg[3],hg[0],'','Denial',sys.argv[6],sys.argv[1],'','',hg[2]])

      # subprocess.run(["python", "test_api.py",hg[1],hg[3],hg[2],'','Denial',sys.argv[6],'',hg[0],'',hg[2]])

      subprocess.run(["python", "updation.py","1","max","11",'Yes'])
    except Exception as e:
		#s2.cell(row_count_1+1, column=11).value='NO'
        log_exceptions()
        subprocess.run(["python", "updation.py","1","max","11",'No'])
except Exception as e:
	#s2.cell(row_count_1+1, column=9).value='No'
	#s2.cell(row_count_1+1, column=11).value='NO'
    log_exceptions()
    subprocess.run(["python", "updation.py","1","max","9",'Yes'])
    subprocess.run(["python", "updation.py","1","max","11",'No'])
now = datetime.datetime.now()
#s2.cell(row_count_1+1, column=6).value=now
#wbk.save(wbkName)
subprocess.run(["python", "updation.py","1","max","6",str(now)])
