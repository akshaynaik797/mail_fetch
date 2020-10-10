import pandas as pd
from email.mime.text import MIMEText
import smtplib
import time
import imaplib
import email
import camelot
import PyPDF2
import csv
import xlsxwriter
from xlrd import open_workbook
import xlwt
import os
import glob
import sys
import os.path
import xlrd
from os import listdir
from os import path
from os.path import isfile, join
from html.parser import HTMLParser
import pdfkit
import pandas as pd
import pdftotext
import subprocess

from decode_error import check_subject, read_from_delete

fg = []
eu = []
ue = []
repeat = []
from openpyxl.styles import Color, PatternFill, Font, Border
from make_log import log_exceptions

try:
    subprocess.run(["python", "updation.py", "1", "max", "9", "X"])

    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')
    yFill = PatternFill(start_color='FFC7CE',
                        end_color='FFC7CE',
                        fill_type='solid')


    def read_email_from_gmail():
        SMTP_SERVER = str(sys.argv[5])
        mail = imaplib.IMAP4_SSL(SMTP_SERVER)
        # mail.login(user = 'Mediclaim@therisingmedicare.com', password = 'cef@2018')

        e_id = str(sys.argv[1])
        pswd = str(sys.argv[2])
        srt = str(sys.argv[3])
        stp = str(sys.argv[4])
        mail.login(user=e_id, password=pswd)
        mail.select("inbox", readonly=True)
        ###############################################<
        mail_uid = str(sys.argv[7])
        if mail_uid == -1:
            type, data = mail.search(None,
                                     '(SUBJECT "HDFC ERGO - Settlement of Claim" since ' + srt + ' before ' + stp + ')')
            ids = data[0]
            id_list = ids.split()
        else:
            ids = mail_uid  # data is a list.
            # accept id from outside and put in id_list akshay var name = id

            id_list = []  # ids is a space separated string
            id_list.append(ids)
        ###############################################>
        # type, data = mail.search(None, '(SUBJECT "HDFC ERGO - Settlement of Claim" since ' + srt + ' before ' + stp + ')')
        # ids = data[0]  # data is a list.
        # id_list = ids.split()  # ids is a space separated string
        # # print(id_list)
        for i in range(0, len(id_list)):
            latest_email_id = id_list[i]  # get the latest
            result, data = mail.fetch(latest_email_id, "(RFC822)")

            ##################################################ak
            try:
                # raw_email = data[0][1].decode('utf-8')
                try:
                    raw_email = data[0][1].decode('utf-8')
                except UnicodeDecodeError:
                    try:
                        raw_email = data[0][1].decode('ISO-8859-1')
                    except UnicodeDecodeError:
                        try:
                            raw_email = data[0][1].decode('ascii')
                        except UnicodeDecodeError:
                            pass
                email_message = email.message_from_string(raw_email)
                subject = email_message['Subject']
                result, sys.argv[8] = check_subject(subject, sys.argv[8], mail)
                if result == 'Changed':
                    # raise Exception('subject not matched')
                    raise Exception('subject not matched', )
            except:
                try:
                    log_exceptions(syssubject=sys.argv[8], subject=subject, error='subject not matched')
                except:
                    pass
                if result != 'OK':
                    data = {'server': SMTP_SERVER,
                            'hospmail': e_id,
                            'pass': pswd,
                            'subject': sys.argv[8]}
                    try:
                        data = read_from_delete(data)
                        if data == None:
                            raise Exception("Not found")
                    except:
                        log_exceptions(msg='not found in deleted', subject=sys.argv[8])
            ##################################################akend
            raw_email = data[0][1].decode('utf-8')
            email_message = email.message_from_string(raw_email)
            if email_message['Subject'] not in fg:
                for part in email_message.walk():

                    if part.get_content_maintype() == 'multipart':
                        # print part.as_string()
                        continue
                    if part.get('Content-Disposition') is None:
                        # print part.as_string()
                        continue
                    fileName = part.get_filename()
                    fileName = fileName.replace('rptSettlementLetterIndivisual_', '')
                    x1 = fileName.find('_')
                    fileName = fileName[:x1] + '.pdf'
                    detach_dir = (os.getcwd() + '/hdfc/attachments_' + str(sys.argv[6]))
                    if bool(fileName):
                        filePath = os.path.join(detach_dir, fileName)
                        if not os.path.isfile(filePath):
                            from reportlab.pdfgen import canvas
                            c = canvas.Canvas(fileName + '.pdf')
                            # print (fileName)
                            fp = open(filePath, 'wb')
                            fp.write(part.get_payload(decode=True))
                            fp.close()
            else:
                repeat.append(email_message['Subject'])
            fg.append(email_message['Subject'])


    wq = 0
    mypath = os.getcwd() + '/hdfc'
    if not path.exists(mypath):
        os.mkdir(mypath)
    if not path.exists(mypath + '/attachments_' + str(sys.argv[6])):
        os.mkdir(mypath + '/attachments_' + str(sys.argv[6]))
    mypath = os.getcwd() + '/hdfc/attachments_' + str(sys.argv[6]) + '/'
    for filename in os.listdir(mypath):
        file_path = os.path.join(mypath, filename)
        if os.path.isfile(file_path) or os.path.islink(file_path):
            os.unlink(file_path)
    read_email_from_gmail()
    print(len(fg))
    import os, zipfile

    dir_name = os.getcwd() + '/hdfc/attachments_' + str(sys.argv[6])
    extension = ".zip"

    os.chdir(dir_name)  # change directory from working dir to dir with files

    for item in os.listdir(dir_name):  # loop through items in dir
        if item.endswith(extension):  # check for ".zip" extension
            file_name = os.path.abspath(item)  # get full path of files
            zip_ref = zipfile.ZipFile(file_name)  # create zipfile object
            zip_ref.extractall(dir_name)  # extract file to dir
            zip_ref.close()  # close file
            os.remove(file_name)
    os.chdir('..')
    os.chdir('..')

    onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
    if path.exists(r'hdfc/hdfc' + str(sys.argv[6]) + '.xlsx'):
        os.remove(r'hdfc/hdfc' + str(sys.argv[6]) + '.xlsx')
    import openpyxl

    wbkName = 'hdfc/hdfc' + str(sys.argv[6]) + '.xlsx'
    wbk = openpyxl.Workbook()
    wbk.create_sheet('1')
    wbk.create_sheet('2')
    s1 = wbk.worksheets[0]
    s2 = wbk.worksheets[1]
    s3 = wbk.worksheets[2]

    for t in range(0, len(onlyfiles)):
        try:
            sh1 = ['Sr No.', 'CCN', 'HDFC ERGO ID', 'Patient Name', 'Policy No', 'Account No.', 'Bank name', 'Diagnosis',
                   'Settled amount', 'Main Member', 'UTR No', 'Transaction Date', 'doa', 'dod']
            sh2 = ['Sr No.', 'CCN', 'Service Type', 'Claimed  Amount', 'Deduction  Amount', 'Discount', 'Settled  Amount',
                   'Remarks']
            sh3 = ['Sr No.', 'CCN', 'Service Tax', 'Total with Service Tax', 'TDS', 'Co-Payment', 'Cheque Amount',
                   'total discont', 'deductible', 'settled amount']
            for i in range(0, len(sh1)):
                s1.cell(row=1, column=i + 1).value = sh1[i]
            for i in range(0, len(sh2)):
                s2.cell(row=1, column=i + 1).value = sh2[i]
            for i in range(0, len(sh3)):
                s3.cell(row=1, column=i + 1).value = sh3[i]

            with open(mypath + onlyfiles[t], "rb") as f:
                pdf = pdftotext.PDF(f)
            with open('hdfc/output.txt', 'w') as f:
                f.write(" ".join(pdf))
            with open('hdfc/output.txt', 'r') as myfile:
                f = myfile.read()

            gh = []

            w5 = f.find('HDFC ERGO ID : ') + 15
            g = f[w5:]
            u3 = g.find('\n') + w5
            gh.append(f[w5:u3])

            w4 = f.find('Patient Name :') + 14
            g = f[w4:]
            u2 = g.find('Main Member') + w4
            gh.append(f[w4:u2])

            w = f.find('policy number') + 13
            g = f[w:]
            u3 = g.find(',') + w
            gh.append(f[w:u3])

            w3 = f.find('Account No.') + 11
            g = f[w3:]
            u1 = g.find('with') + w3
            gh.append(f[w3:u1])

            u2 = g.find('and') + w3
            gh.append(f[u1 + 5:u2])

            w3 = f.find('Ailment :') + 9
            g = f[w3:]
            u1 = g.find('Hospitalization') + w3
            di = f[w3:u1]
            di = di.replace('HOSPITAL', '')
            gh.append(di)

            w = f.find('claim with CCN') + 14
            g = f[w:]
            u3 = g.find(',') + w
            ccn = f[w:u3]

            w3 = f.find('sum of') + 7
            g = f[w3:]
            u1 = g.find('(') + w3
            gh.append(f[w3:u1])

            w5 = f.find('Main Member') + 13
            g = f[w5:]
            u3 = g.find('\n') + w5
            gh.append(f[w5:u3])

            w5 = f.find('UTR') + 7
            g = f[w5:]
            u3 = g.find('and') + w5
            gh.append(f[w5:u3])

            w5 = f.find('Transaction Date') + 16
            g = f[w5:]
            u3 = g.find('\n') + w5
            gh.append(f[w5:u3])

            w5 = f.find('From :') + 7
            g = f[w5:]
            u3 = g.find('To') + w5
            gh.append(f[w5:u3])

            w5 = u3 + 4
            g = f[w5:]
            u3 = g.find('\n') + w5
            gh.append(f[w5:u3])
            gh = [sub.replace('Rs.', '') for sub in gh]
            gh = [sub.replace('  ', '') for sub in gh]
            gh = [sub.replace('.', '') for sub in gh]
            gh = [sub.replace('\n', '') for sub in gh]
            ccn = ccn.replace('  ', '')
            # print(gh,ccn)
            s1.cell(row=t + 2, column=1).value = t + 1
            s1.cell(row=t + 2, column=2).value = ccn

            for i in range(0, len(gh)):
                s1.cell(row=t + 2, column=i + 3).value = gh[i]

            tables = camelot.read_pdf(mypath + onlyfiles[t], pages='all', line_scale=10)
            tables.export('hdfc/foo1.xls', f='excel')
            loc = ("hdfc/foo1.xls")
            wb = xlrd.open_workbook(loc)
            sheet_1 = wb.sheet_by_index(0)
            sheet_1.cell_value(0, 0)
            sheet_2 = wb.sheet_by_index(1)
            sheet_2.cell_value(0, 0)
            max_row = sheet_1.nrows
            hg = []
            b = []
            p = []
            np = []
            r = []
            rt = []
            m = 0
            for i in range(2, max_row):

                hg.append(sheet_1.cell_value(i, 2))
                b.append(sheet_1.cell_value(i, 3))
                p.append(sheet_1.cell_value(i, 4))
                np.append(sheet_1.cell_value(i, 5))
                r.append(sheet_1.cell_value(i, 6))
                rt.append(sheet_1.cell_value(i, 7))
                if sheet_1.cell_value(i, 2) == 'Total':
                    m = 1
                    sett = sheet_1.cell_value(i, 3)
                    dis = sheet_1.cell_value(i, 5)
                    ded = sheet_1.cell_value(i, 4)
                    break
            hg = [sub.replace('a.ii)', '') for sub in hg]
            max_row = sheet_2.nrows
            max_col = sheet_2.ncols
            # print(max_col)
            if (m == 0 and max_col == 8):
                for i in range(1, max_row):
                    hg.append(sheet_2.cell_value(i, 2))
                    b.append(sheet_2.cell_value(i, 3))
                    p.append(sheet_2.cell_value(i, 4))
                    np.append(sheet_2.cell_value(i, 5))
                    r.append(sheet_2.cell_value(i, 6))
                    rt.append(sheet_2.cell_value(i, 7))
                    if sheet_2.cell_value(i, 2) == 'Total':
                        sett = sheet_2.cell_value(i, 3)
                        dis = sheet_2.cell_value(i, 5)
                        ded = sheet_2.cell_value(i, 4)
                        # print(dis)
                        break
            elif (m == 0 and max_col == 7):
                for i in range(1, max_row):
                    x = sheet_2.cell_value(i, 1)
                    if x[0] >= '0' and x[0] <= '9':
                        w3 = x.find('\n') + 1
                        x = x[w3:]
                    # print(x)
                    hg.append(x)
                    b.append(sheet_2.cell_value(i, 2))
                    p.append(sheet_2.cell_value(i, 3))
                    np.append(sheet_2.cell_value(i, 4))
                    r.append(sheet_2.cell_value(i, 5))
                    rt.append(sheet_2.cell_value(i, 6))
                    if sheet_2.cell_value(i, 1) == 'Total':
                        sett = sheet_2.cell_value(i, 2)
                        ded = sheet_2.cell_value(i, 3)
                        dis = sheet_2.cell_value(i, 4)
                        # print(dis)
                        break
            for i in range(0, len(hg)):
                row_num = s2.max_row + 1
                wq += 1
                s2.cell(row=row_num, column=1).value = wq
                s2.cell(row=row_num, column=2).value = ccn
                s2.cell(row=row_num, column=3).value = hg[i]
                s2.cell(row=row_num, column=4).value = b[i]
                s2.cell(row=row_num, column=5).value = p[i]
                s2.cell(row=row_num, column=6).value = np[i]
                s2.cell(row=row_num, column=7).value = r[i]
                s2.cell(row=row_num, column=8).value = rt[i]
            # print(hg,b,p,np,r,rt)
            gh = []
            if f.find('Service Tax') != -1:
                w5 = f.find('Service Tax') + 12
                g = f[w5:]
                u3 = g.find('\n') + w5
                gh.append(f[w5:u3])
            else:
                gh.append(' ')
            if f.find('Total with Service Tax') != -1:
                w4 = f.find('Total with Service Tax') + 23
                g = f[w4:]
                u2 = g.find('\n') + w4
                gh.append(f[w4:u2])
            else:
                gh.append(' ')

            if f.find('TDS') != -1:
                w = f.find('TDS') + 4
                g = f[w:]
                u3 = g.find('\n') + w
                gh.append(f[w:u3])
            else:
                gh.append(' ')
            if f.find('Co-Payment') != -1:
                w = f.find('Co-Payment') + 10
                g = f[w:]
                u3 = g.find('\n') + w
                gh.append(f[w:u3])
            else:
                gh.append(' ')
            if f.find('Cheque Amount') != -1:
                w = f.find('Cheque Amount') + 13
                g = f[w:]
                u3 = g.find('\n') + w
                gh.append(f[w:u3])
            else:
                gh.append(' ')

            gh = [sub.replace('  ', '') for sub in gh]

            for i in range(0, len(gh)):
                s3.cell(row=t + 2, column=i + 3).value = gh[i]
            s3.cell(row=t + 2, column=1).value = t + 1
            s3.cell(row=t + 2, column=2).value = ccn
            s3.cell(row=t + 2, column=8).value = dis
            s3.cell(row=t + 2, column=9).value = ded
            s3.cell(row=t + 2, column=10).value = sett
        except Exception as e:
            print(e)
            s1.cell(row=t + 2, column=1).value = 'error'
    print("Done")
    wbk.save(wbkName)
    wbk.close
    wbkName = 'count/count.xlsx'
    wbk = openpyxl.load_workbook(wbkName)
    s1 = wbk.worksheets[0]

    row_ = s1.max_row + 1
    s1.cell(row=row_, column=1).value = 'hdfc'
    s1.cell(row=row_, column=2).value = str(sys.argv[6])
    s1.cell(row=row_, column=3).value = len(fg)
    s1.cell(row=row_, column=4).value = len(onlyfiles)
    s1.cell(row=row_, column=5).value = len(repeat)
    subprocess.run(["python", "updation.py", "1", "max", "9", " "])
except:
    log_exceptions()