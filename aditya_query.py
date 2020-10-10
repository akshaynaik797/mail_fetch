import os
import sys
import struct, time
import subprocess
from datetime import date
import datetime
import openpyxl
import pdftotext
import time
import requests
now = datetime.datetime.now()
subprocess.run(["python", "updation.py","1","max1","1",sys.argv[2]])
subprocess.run(["python", "updation.py","1","max","2",sys.argv[3]])
subprocess.run(["python", "updation.py","1","max","3",sys.argv[4]])
subprocess.run(["python", "updation.py","1","max","5",str(now)])
subprocess.run(["python", "updation.py","1","max","7",sys.argv[5]])
subprocess.run(["python", "updation.py","1","max","8",sys.argv[6]])

with open(sys.argv[1], "rb") as f:
	pdf = pdftotext.PDF(f)

with open('aditya/output.txt', 'w') as f:
	f.write(" ".join(pdf))
with open('aditya/output.txt', 'r') as myfile:
 	f = myfile.read()
print(sys.argv[1])
if(sys.argv[1].find('_pdf_')==-1):
	try:

		#print(f)

			f=f.replace('\n',' ')
			f=f.replace('  ',' ')
			hg=[]
			w=f.find('Preauth number: ')+16
			g=f[w:]
			u=g.find(' ')+w
			#print(f)

			hg.append(f[w:u])

			hg.append("")
			status='Information Awaiting'
			
			hg.append("")


			w=f.find('Policy Number: ')+15
			g=f[w:]
			u=g.find(' ')+w
			hg.append(f[w:u])
			hg.append('')#member id
			hg=[sub.replace('  ','') for sub in hg]
			w1="The documents that need to be submitted are:"
			u1="We have communicated the details to him/her and would like your help in the submission of the same."
			w=f.find(w1)+len(w1)
			u=f.find(u1)
			text=f[w:u]
			text=text.replace('\n',' ')
			text=text.replace('    ',' ')

			hg.append(text)
			subprocess.run(["python", "updation.py","1","max","9",'Yes'])
			subprocess.run(["python", "updation.py","1","max","10",'NA'])
			try:
				subprocess.run(["python", "test_api.py",hg[0],hg[1],hg[3],'',status,sys.argv[6],sys.argv[1],'',hg[4],hg[5]])
				subprocess.run(["python", "updation.py","1","max","11",'Yes'])
			except Exception as e:
				subprocess.run(["python", "updation.py","1","max","11",'No'])
	except Exception as e:
		#s2.cell(row_count_1+1, column=9).value='No'
		#s2.cell(row_count_1+1, column=11).value='NO'
		subprocess.run(["python", "updation.py","1","max","9",'Yes'])
		subprocess.run(["python", "updation.py","1","max","11",'No'])
else:
	try:
		hg=[]
		w=f.find('Preauth Reference no')+21
		g=f[w:]
		u=g.find('\n')+w
		hg.append(f[w:u])

		status='Information Awaiting'
		hg.append(status)

		w=f.find('Policy Number')+14
		g=f[w:]
		u=g.find('\n')+w
		hg.append(f[w:u])
		hg.append('')#member id
		hg=[sub.replace(':','') for sub in hg]
		hg=[sub.replace('  ','') for sub in hg]
		hg=[sub.replace('Rs.','') for sub in hg]
		w1="documents we have found that we need the below mentioned details to process the claim."
		u1="To avail the cashless benefit the above"
		w=f.find(w1)+len(w1)
		u=f.find(u1)
		text=f[w:u]
		text=text.replace('\n',' ')
		text=text.replace('           ',' ')

		hg.append(text)

		subprocess.run(["python", "updation.py","1","max","9",'Yes'])
		subprocess.run(["python", "updation.py","1","max","10",'NA'])

		try:
			subprocess.run(["python", "test_api.py",hg[0],'',hg[2],'',status,sys.argv[6],sys.argv[1],'',hg[3],hg[4]])
			'''wbk= openpyxl.load_workbook(wbkName)
			s2=wbk.worksheets[1]
			s2.cell(row_count_1+1, column=11).value='YES'
			'''
			subprocess.run(["python", "updation.py","1","max","11",'Yes'])
		except Exception as e:
			#s2.cell(row_count_1+1, column=11).value='NO'
			subprocess.run(["python", "updation.py","1","max","11",'No'])

	except Exception as e:
		#s2.cell(row_count_1+1, column=9).value='No'
		#s2.cell(row_count_1+1, column=11).value='NO'
		subprocess.run(["python", "updation.py","1","max","9",'Yes'])
		subprocess.run(["python", "updation.py","1","max","11",'No'])



print(hg)

now = datetime.datetime.now()
#s2.cell(row_count_1+1, column=6).value=now
#wbk.save(wbkName)
subprocess.run(["python", "updation.py","1","max","6",str(now)])
