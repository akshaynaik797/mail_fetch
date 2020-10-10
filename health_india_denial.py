import os
import sys
import struct, time
import subprocess
from datetime import date
import datetime
import openpyxl
import pdftotext
import time
import requests
now = datetime.datetime.now()
'''
wbkName = 'log file.xlsx'
wbk= openpyxl.load_workbook(wbkName)
s1=wbk.worksheets[0]
s2=wbk.worksheets[1]
row_count_1 = s2.max_row
s2.cell(row_count_1+1, column=1).value=sys.argv[2]
s2.cell(row_count_1+1, column=2).value=sys.argv[3]
s2.cell(row_count_1+1, column=3).value=sys.argv[4]
s2.cell(row_count_1+1, column=5).value=now
s2.cell(row_count_1+1, column=7).value=sys.argv[5]
s2.cell(row_count_1+1, column=8).value=sys.argv[6]
'''

subprocess.run(["python", "updation.py","1","max1","1",sys.argv[2]])
subprocess.run(["python", "updation.py","1","max","2",sys.argv[3]])
subprocess.run(["python", "updation.py","1","max","3",sys.argv[4]])
subprocess.run(["python", "updation.py","1","max","5",str(now)])
subprocess.run(["python", "updation.py","1","max","7",sys.argv[5]])
subprocess.run(["python", "updation.py","1","max","8",sys.argv[6]])

# added by ashish. pdf file name was hardcoded
with open(sys.argv[1], "rb") as f:
    pdf = pdftotext.PDF(f)
# added by ashish folder name was incoorect 
with open('health_india/output1.txt', 'w') as f:
    f.write(" ".join(pdf))     
with open('health_india/output1.txt', 'r') as myfile:
    f = myfile.read()

# with open("HI-NIA-000887175-0-DAL.pdf", "rb") as f:
#   pdf = pdftotext.PDF(f)

# with open('hdfc/output1.txt', 'w') as f:
#   f.write(" ".join(pdf))     
# with open('hdfc/output1.txt', 'r') as myfile:
#   f = myfile.read()


try:
    hg=[]
    if f.find('Claim No')!=-1:
        w=f.find('Claim No')+len('Claim No')
        k=f[w:]
        u=k.find("\n")+w
        g=f[w:u]
        hg.append(g)
        print(hg[0])
    else:
        print("Claim No is not found")   
  
    if f.find('Patient Name')!=-1:
        w=f.find('Patient Name')+len('Patient Name')
        k=f[w:]
        u=k.find("\n")+w
        g=f[w:u]
        hg.append(g)
        print(hg[1])
    else:
        print("Patient Name is not found")
    
    if f.find('Policy No.')!=-1:
        w=f.find('Policy No')+len('Policy No.')
        k=f[w:]
        u=k.find("\n")+w
        g=f[w:u]
        hg.append(g)
        print(hg[2])
    else:
        print("Policy No is not found")

    if f.find('Requested')!=-1:     
        w=f.find('Requested')+len('Requested')
        k=f[w:]
        u=k.find("\n")+w
        g=f[w:u]
        hg.append(g)
        print(hg[3])
    else:
        print("Requested is not found")

    if f.find('HI Id No.')!=-1:
        w=f.find('HI Id No.')+len('HI Id No.')
        k=f[w:]
        u=k.find("\n")+w
        g=f[w:u]
        hg.append(g)
        print(hg[4])
    else:
        print("HI Id No is not found")   

    hg=[sub.replace(':','') for sub in hg]
    hg=[sub.replace('/-','') for sub in hg]      
    hg=[sub.replace('  ','') for sub in hg]
    hg=[sub.replace('Rs.','') for sub in hg]
    try:
        subprocess.run(["python", "test_api.py",hg[0],hg[3],hg[2],'','Denial',sys.argv[6],sys.argv[1],'',hg[4],hg[1]])
        '''wbk= openpyxl.load_workbook(wbkName)
        s2=wbk.worksheets[1]
        s2.cell(row_count_1+1, column=11).value='YES'
        '''
        subprocess.run(["python", "updation.py","1","max","11",'Yes'])  
    except Exception as e:
        #s2.cell(row_count_1+1, column=11).value='NO'
        subprocess.run(["python", "updation.py","1","max","11",'No'])
except Exception as e:
    #s2.cell(row_count_1+1, column=9).value='No'
    #s2.cell(row_count_1+1, column=11).value='NO'
    subprocess.run(["python", "updation.py","1","max","9",'Yes'])
    subprocess.run(["python", "updation.py","1","max","11",'No'])
now = datetime.datetime.now()
#s2.cell(row_count_1+1, column=6).value=now
#wbk.save(wbkName)
subprocess.run(["python", "updation.py","1","max","6",str(now)])





