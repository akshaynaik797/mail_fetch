import pandas as pd
import sys
from email.mime.text import MIMEText
import smtplib
import time
import imaplib
import email
import camelot
import PyPDF2
import csv
import xlsxwriter
from xlrd import open_workbook
import pdftotext
import xlwt
import xlrd
import os
import glob
import os.path
import json
from os import listdir
from os import path
from os.path import isfile, join
from html.parser import HTMLParser
import pdfkit
import subprocess
import tabula
import pandas as pd
import html2text

from decode_error import check_subject, read_from_delete
from make_log import log_exceptions
from app import config
import openpyxl

try:
    # path_wkhtmltopdf = r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe'
    # config = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf)
    subprocess.run(["python", "updation.py", "1", "max", "9", "X"])

    if path.exists(r'vipul/vipul' + str(sys.argv[6]) + '.xlsx'):
        os.remove(r'vipul/vipul' + str(sys.argv[6]) + '.xlsx')

    po = []
    fg = []
    repeat = []
    wbkName = 'vipul/vipul' + str(sys.argv[6]) + '.xlsx'
    wbk = openpyxl.Workbook()
    wbk.create_sheet('1')
    # wbk.create_sheet('Sheet3')
    s1 = wbk.worksheets[0]
    s2 = wbk.worksheets[1]
    wq = 0


    def read_email_from_gmail():
        SMTP_SERVER = str(sys.argv[5])
        mail = imaplib.IMAP4_SSL(SMTP_SERVER)
        # mail.login(user = 'tpa@psri.net', password = '@Dawakhana123#')
        # mail = smtplib.SMTP_SSL(host = 'smtp.gmail.com', port = 465)
        e_id = str(sys.argv[1])
        pswd = str(sys.argv[2])
        srt = str(sys.argv[3])
        stp = str(sys.argv[4])
        mail.login(user=e_id, password=pswd)
        # stp =str(input("Enter your end date : "))
        # mail = smtplib.SMTP_SSL(host = 'smtp.gmail.com', port = 465)
        mail.select("inbox", readonly=True)
        ###############################################<
        mail_uid = str(sys.argv[7])
        if mail_uid == -1:
            type, data = mail.search(None, '(SUBJECT "Payment of Your Claim File" since ' + srt + ' before ' + stp + ')')
            ids = data[0]
            id_list = ids.split()
        else:
            ids = mail_uid  # data is a list.
            # accept id from outside and put in id_list akshay var name = id

            id_list = []  # ids is a space separated string
            id_list.append(ids)
        ###############################################>
        # type, data = mail.search(None, '(SUBJECT "Payment of Your Claim File" since ' + srt + ' before ' + stp + ')')
        # ids = data[0]  # data is a list.
        # id_list = ids.split()  # ids is a space separated string
        for i in range(0, len(id_list)):
            latest_email_id = id_list[i]  # get the latest
            result, data = mail.fetch(latest_email_id, "(RFC822)")
            ##################################################ak
            try:
                # raw_email = data[0][1].decode('utf-8')
                try:
                    raw_email = data[0][1].decode('utf-8')
                except UnicodeDecodeError:
                    try:
                        raw_email = data[0][1].decode('ISO-8859-1')
                    except UnicodeDecodeError:
                        try:
                            raw_email = data[0][1].decode('ascii')
                        except UnicodeDecodeError:
                            pass
                email_message = email.message_from_string(raw_email)
                subject = email_message['Subject']
                result, sys.argv[8] = check_subject(subject, sys.argv[8], mail)
                if result == 'Changed':
                    # raise Exception('subject not matched')
                    raise Exception('subject not matched', )
            except:
                log_exceptions(syssubject=sys.argv[8], subject=subject, msg='subject not matched')
                if result != 'OK':
                    data = {'server': SMTP_SERVER,
                            'hospmail': e_id,
                            'pass': pswd,
                            'subject': sys.argv[8]}
                    try:
                        data = read_from_delete(data)
                        if data == None:
                            raise Exception("Not found")
                    except:
                        log_exceptions(msg='not found in deleted', subject=sys.argv[8])
            ##################################################akend

            raw_email = data[0][1].decode('utf-8')
            email_message = email.message_from_string(raw_email)
            if email_message['Subject'] not in fg:
                for part in email_message.walk():
                    # print(part.get_content_type())
                    if part.get_content_type() == "text/html":
                        # print('hi')
                        body = part.get_payload(decode=True)
                        file_name = 'vipul/email.html'  # "attachments_e/email"+str(i)+".html"
                        output_file = open(file_name, 'w')
                        output_file.write("Body: %s" % (body.decode('utf-8')))
                        output_file.close()
                        pdfkit.from_file('vipul/email.html',
                                         'vipul/attachments_' + str(sys.argv[6]) + '/' + str(i) + '.pdf',configuration=config)
            else:
                repeat.append(email_message['Subject'])

            fg.append(email_message['Subject'])


    mypath = os.getcwd() + '/vipul'
    if not path.exists(mypath):
        os.mkdir(mypath)
    if not path.exists(mypath + '/attachments_' + str(sys.argv[6])):
        os.mkdir(mypath + '/attachments_' + str(sys.argv[6]))
    mypath = os.getcwd() + '/vipul/attachments_' + str(sys.argv[6]) + '/'
    for filename in os.listdir(mypath):
        file_path = os.path.join(mypath, filename)
        if os.path.isfile(file_path) or os.path.islink(file_path):
            os.unlink(file_path)
    read_email_from_gmail()

    onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
    print(len(onlyfiles))
    for t in range(0, len(onlyfiles)):
        try:
            sh1 = ['Sr. No.', 'File No', 'Patient Name', 'Admin Date', 'Emp Code', 'Dis Date', 'Claim Amt',
                   'Total Bill Amt', 'Deduction', 'Co-Pay Deduction', 'Excess Pay Deduction', 'Approved Amount', 'Discount',
                   'Settled Amt', 'TDS Amount', 'Net Amount']
            sh2 = ['S.no', 'File No.', 'Deducted Amount', 'Deduction Reason']
            for i in range(0, len(sh2)):
                s2.cell(row=1, column=i + 1).value = sh2[i]
            for i in range(0, len(sh1)):
                s1.cell(row=1, column=i + 1).value = sh1[i]
            tables = camelot.read_pdf(mypath + onlyfiles[t], pages='all', line_scale=90)
            tables.export('vipul/foo1.xlsx', f='excel')
            loc = ("vipul/foo1.xlsx")
            wb = xlrd.open_workbook(loc)
            with open(mypath + onlyfiles[t], "rb") as f:
                pdf = pdftotext.PDF(f)

            with open('vipul/output.txt', 'w') as f:
                f.write(" ".join(pdf))
            with open('vipul/output.txt', 'r') as myfile:
                f = myfile.read()

            gh = []
            # print(f)
            w4 = f.find('Patient') + 12
            u2 = f.find('Admin Date')
            gh.append(f[w4:u2])
            if (gh[-1] == ' '):
                print(f)
            w5 = f.find('Admin Date') + 11
            u3 = f.find('Emp Code')
            gh.append(f[w5:u3])

            w5 = f.find('Emp Code') + 8
            g = f[w5:]
            u3 = g.find('\n') + w5
            gh.append(f[w5:u3])

            w1 = f.find('Dis') + 9
            u1 = f.find('File No')
            gh.append(f[w1:u1])

            w1 = f.find('File No') + 8
            g = f[w1:]
            u1 = g.find('\n') + w1
            ccn = f[w1:u1]
            ccn = ccn.replace('  ', '')
            # print(gh)

            gh = [sub.replace('  ', '') for sub in gh]
            if (gh[0] == ' ' or gh[0] == ''):
                w4 = f.find('information.') + 12
                u2 = f.find('Patient')
                gh[0] = f[w4:u2]
            # print(gh[0])
            gh = [sub.replace('\n', '') for sub in gh]
            gh = [sub.replace('  ', '') for sub in gh]

            sheet_1 = wb.sheet_by_index(1)
            sheet_1.cell_value(0, 0)
            max_row = sheet_1.nrows
            # print(max_row)
            max_col = sheet_1.ncols
            hg = []
            b = []
            p = []
            np = []
            r = []
            rt = []
            mh = []
            ro = []
            bo = []
            po = []
            no = []
            for i in range(1, max_row):
                ro.append(sheet_1.cell_value(i, 1))
                hg.append(sheet_1.cell_value(i, 2))
                b.append(sheet_1.cell_value(i, 3))
                p.append(sheet_1.cell_value(i, 4))
                np.append(sheet_1.cell_value(i, 5))
                r.append(sheet_1.cell_value(i, 6))
                rt.append(sheet_1.cell_value(i, 7))
                mh.append(sheet_1.cell_value(i, 8))
                bo.append(sheet_1.cell_value(i, 9))
                po.append(sheet_1.cell_value(i, 10))
                no.append(sheet_1.cell_value(i, 11))
                pass

            pass
            hg = [sub.replace('  ', '') for sub in hg]
            b = [sub.replace('  ', '') for sub in b]
            p = [sub.replace('  ', '') for sub in p]
            np = [sub.replace('  ', '') for sub in np]
            r = [sub.replace('  ', '') for sub in r]
            rt = [sub.replace('  ', '') for sub in rt]
            mh = [sub.replace('  ', '') for sub in mh]
            ro = [sub.replace('  ', '') for sub in ro]
            bo = [sub.replace('  ', '') for sub in bo]
            po = [sub.replace('  ', '') for sub in po]
            no = [sub.replace('\n', ' ') for sub in no]
            no = [sub.replace('\t', ' ') for sub in no]
            # print(no)
            xt = no[-1]
            op = xt.split(';')
            # print(op)
            s1.cell(row=t + 2, column=1).value = t + 1
            s1.cell(row=t + 2, column=2).value = ccn
            ph = ro[-1] + ' ' + hg[-1] + ' ' + b[-1] + ' ' + p[-1] + ' ' + np[-1] + ' ' + r[-1] + ' ' + rt[-1] + ' ' + mh[
                -1] + ' ' + bo[-1] + ' ' + po[-1]
            temp = [int(s) for s in ph.split() if s.isdigit()]
            # print(temp)
            for i in range(0, len(gh)):
                s1.cell(row=t + 2, column=2).value = ccn
                s1.cell(row=t + 2, column=i + 3).value = gh[i]
            for i in range(0, len(temp)):
                s1.cell(row=t + 2, column=len(gh) + 3 + i).value = temp[i]
            for i in op:
                row_num = s2.max_row + 1
                wq += 1
                s2.cell(row=row_num, column=1).value = wq
                s2.cell(row=row_num, column=2).value = ccn
                w5 = i.find('Rs.') + 3
                g = i[w5:]
                u3 = g.find(' ') + w5
                # print(i[w5:u3])
                s2.cell(row=row_num, column=3).value = (i[w5:u3])
                s2.cell(row=row_num, column=4).value = (i[u3:])
            tr = []
            w5 = f.find('UTR') + 14
            g = f[w5:]
            u3 = g.find('dated') + w5
            tr.append(f[w5:u3])
            u1 = g.find('.') + w5
            tr.append(f[u3 + 5:u1])
            tr = [sub.replace('NEFT-', '') for sub in tr]
            for i in range(0, len(tr)):
                s1.cell(row=t + 2, column=len(gh) + len(temp) + 3 + i).value = tr[i]
            os.rename(os.getcwd() + '/vipul/attachments_' + str(sys.argv[6]) + '/' + onlyfiles[t],
                      os.getcwd() + '/vipul/attachments_' + str(sys.argv[6]) + '/' + ccn + '.pdf')

        except Exception as e:
            log_exceptions()
            s1.cell(row=t + 2, column=1).value = 'error'
            os.rename(os.getcwd() + '/vipul/attachments_' + str(sys.argv[6]) + '/' + onlyfiles[t],
                      os.getcwd() + '/vipul/attachments_' + str(sys.argv[6]) + '/' + ccn + '.pdf')

    print("Done")
    wbk.save(wbkName)
    wbk.close
    wbkName = 'count/count.xlsx'
    wbk = openpyxl.load_workbook(wbkName)
    s1 = wbk.worksheets[0]

    row_ = s1.max_row + 1
    s1.cell(row=row_, column=1).value = 'vipul'
    s1.cell(row=row_, column=2).value = str(sys.argv[6])
    s1.cell(row=row_, column=3).value = len(fg)
    s1.cell(row=row_, column=4).value = len(onlyfiles)
    s1.cell(row=row_, column=5).value = len(repeat)
    wbk.save(wbkName)
    wbk.close()
    pass
    subprocess.run(["python", "updation.py", "1", "max", "9", " "])
except:
	log_exceptions()

