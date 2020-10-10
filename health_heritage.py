import pandas as pd
from email.mime.text import MIMEText
import smtplib
import time
import imaplib
import email
import camelot
import PyPDF2
import sys
import csv
import xlsxwriter
from xlrd import open_workbook
import xlwt
import os
import glob
import os.path
import xlrd
from os import listdir
from os import path
from os.path import isfile, join
from html.parser import HTMLParser
import pdfkit
import pandas as pd
import pdftotext
import html2text

from decode_error import check_subject, read_from_delete
from make_log import log_exceptions

if path.exists(r'health_heritage/health_heritage' + str(sys.argv[6]) + '.xlsx'):
    os.remove(r'health_heritage/health_heritage' + str(sys.argv[6]) + '.xlsx')
import openpyxl
import subprocess
try:
    subprocess.run(["python", "updation.py", "1", "max", "9", "X"])

    fg = []
    eu = []
    repeat = []
    from openpyxl.styles import Color, PatternFill, Font, Border

    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')


    def read_email_from_gmail():
        SMTP_SERVER = str(sys.argv[5])
        mail = imaplib.IMAP4_SSL(SMTP_SERVER)
        # mail.login(user = 'Mediclaim@therisingmedicare.com', password = 'cef@2018')
        b = 0
        e_id = str(sys.argv[1])
        pswd = str(sys.argv[2])
        srt = str(sys.argv[3])
        stp = str(sys.argv[4])
        mail.login(user=e_id, password=pswd)
        mail.select("inbox", readonly=True)
        ###############################################<
        mail_uid = str(sys.argv[7])
        if mail_uid == -1:
            type, data = mail.search(None,
                                     '(FROM "info@heritagehealthtpa.info" SUBJECT "Claim Settlement Advice" since ' + srt + ' before ' + stp + ')')
            ids = data[0]
            id_list = ids.split()
        else:
            ids = mail_uid  # data is a list.
            # accept id from outside and put in id_list akshay var name = id

            id_list = []  # ids is a space separated string
            id_list.append(ids)
        ###############################################>
        # type, data = mail.search(None,
        #                          '(FROM "info@heritagehealthtpa.info" SUBJECT "Claim Settlement Advice" since ' + srt + ' before ' + stp + ')')
        # ids = data[0]  # data is a list.
        # id_list = ids.split()  # ids is a space separated string
        # # print(id_list)
        for i in range(0, len(id_list)):
            latest_email_id = id_list[i]  # get the latest
            result, data = mail.fetch(latest_email_id, "(RFC822)")
            ##################################################ak
            try:
                # raw_email = data[0][1].decode('utf-8')
                try:
                    raw_email = data[0][1].decode('utf-8')
                except UnicodeDecodeError:
                    try:
                        raw_email = data[0][1].decode('ISO-8859-1')
                    except UnicodeDecodeError:
                        try:
                            raw_email = data[0][1].decode('ascii')
                        except UnicodeDecodeError:
                            pass
                email_message = email.message_from_string(raw_email)
                subject = email_message['Subject']
                result, sys.argv[8] = check_subject(subject, sys.argv[8], mail)
                if result == 'Changed':
                    # raise Exception('subject not matched')
                    raise Exception('subject not matched', )
            except:
                try:
                    log_exceptions(syssubject=sys.argv[8], subject=subject, error='subject not matched')
                except:
                    pass
                if result != 'OK':
                    data = {'server': SMTP_SERVER,
                            'hospmail': e_id,
                            'pass': pswd,
                            'subject': sys.argv[8]}
                    try:
                        data = read_from_delete(data)
                        if data == None:
                            raise Exception("Not found")
                    except:
                        log_exceptions(msg='not found in deleted', subject=sys.argv[8])
            ##################################################akend
            raw_email = data[0][1].decode('utf-8')
            email_message = email.message_from_string(raw_email)
            # if path.exists(r'/home/shivam/Desktop/vnu_scripts/Paramount/email.html'):
            # os.remove(r'email.html')
            # Body details
            if email_message['Subject'] not in fg:
                b += 1
                for part in email_message.walk():
                    if part.get_content_type() == "text/html":
                        body = part.get_payload(decode=True)
                        file_name = 'health_heritage/attachments_' + str(sys.argv[6]) + '/email' + str(b) + '.html'
                        output_file = open(file_name, 'w')
                        output_file.write("Body: %s" % (body.decode('utf-8')))
                        output_file.close()
            else:
                repeat.append(email_message['Subject'])
            fg.append(email_message['Subject'])


    mypath = os.getcwd() + '/health_heritage'
    if not path.exists(mypath):
        os.mkdir(mypath)
    if not path.exists(mypath + '/attachments_' + str(sys.argv[6])):
        os.mkdir(mypath + '/attachments_' + str(sys.argv[6]))
    mypath = os.getcwd() + '/health_heritage/attachments_' + str(sys.argv[6]) + '/'

    for filename in os.listdir(mypath):
        file_path = os.path.join(mypath, filename)
        if os.path.isfile(file_path) or os.path.islink(file_path):
            os.unlink(file_path)
    read_email_from_gmail()

    onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
    wq = 0
    wbkName = 'health_heritage/health_heritage' + str(sys.argv[6]) + '.xlsx'
    wbk = openpyxl.Workbook()
    wbk.create_sheet('1')
    s1 = wbk.worksheets[0]
    s2 = wbk.worksheets[1]
    for t in range(0, len(onlyfiles)):
        try:
            sh1 = ['Sr No.', 'Claim ID', 'Patient Name', 'I-Card No.', 'Policy No.', 'DOA', 'DOD', 'Illness',
                   'Amount Claimed', 'Net Amount', 'NEFT No.', 'Date of Transfer', 'TDS Deducted', 'insurance company',
                   ' primary beneficiary', 'bank', 'settle amt']
            sh2 = ['Sr No.', 'Claim ID', 'deductions amount', 'deductions Reasons']

            for i in range(0, len(sh1)):
                s1.cell(row=1, column=i + 1).value = sh1[i]
            for i in range(0, len(sh2)):
                s2.cell(row=1, column=i + 1).value = sh2[i]

            html = open('health_heritage/attachments_' + str(sys.argv[6]) + '/email' + str(t + 1) + ".html")
            f = str(html.read())
            w = open("health_heritage/out.txt", "w")
            w.write(html2text.html2text(f))
            html.close()
            w.close()
            with open('health_heritage/out.txt', 'r') as myfile:
                f = myfile.read()

            w = f.find('Details of deductions:') + 22
            u = f.find('Sincerely')
            g = f[w:u]
            g = g.replace('\n', '$$')
            g = g.replace('$$Rs.', '$$\n')
            g = g.replace('$$', ' ')
            g = g.replace('   ', '')
            g = g.replace('**', '')
            sy = g.split('\n')
            sy.remove('')

            hg = []
            w = f.find('Patient Name') + 13
            g = f[w:]
            u = g.find('|') + w
            hg.append(f[w:u])

            w = f.find('I-Card No.') + 12
            g = f[w:]
            u = g.find('|') + w
            hg.append(f[w:u])

            w1 = f.find('Policy No.') + 12
            g = f[w1:]
            u1 = g.find('|') + w1
            hg.append(f[w1:u1])

            w2 = f.find('DOA') + 6
            g = f[w2:]
            u2 = g.find('|') + w2
            hg.append(f[w2:u2])

            w9 = f.find('DOD') + 6
            g = f[w9:]
            u9 = g.find('\n') + w9
            hg.append(f[w9:u9])

            w2 = f.find('Illness') + 10
            g = f[w2:]
            u2 = g.find('|') + w2
            hg.append(f[w2:u2])

            w9 = f.find('Amount Claimed') + 18
            g = f[w9:]
            u9 = g.find('|') + w9
            hg.append(f[w9:u9])

            w9 = f.find('Net Amount') + 11
            g = f[w9:]
            u9 = g.find('\n') + w9
            hg.append(f[w9:u9])

            w9 = f.find('NEFT No.:') + 9
            g = f[w9:]
            u9 = g.find('Date of Transfer') + w9
            hg.append(f[w9:u9])

            w2 = f.find('Date of Transfer') + 16
            g = f[w2:]
            u2 = g.find('Settled Amount') + w2
            hg.append(f[w2:u2])

            w9 = f.find('TDS Deducted :') + 15
            g = f[w9:]
            u9 = g.find('Deducted') + w9
            hg.append(f[w9:u9])

            w9 = f.find('insurer') + 8
            g = f[w9:]
            u9 = g.find(',') + w9
            hg.append(f[w9:u9])

            w9 = f.find('Employee Name') + 14
            g = f[w9:]
            u9 = g.find('\n') + w9
            hg.append(f[w9:u9])
            if f.find('Drawn On') != -1:
                w9 = f.find('Drawn On') + 9
                g = f[w9:]
                u9 = g.find('\n') + w9
                hg.append(f[w9:u9])
            else:
                hg.append(' ')
            w9 = f.find('settled for') + 16
            g = f[w9:]
            u9 = g.find('on') + w9
            hg.append(f[w9:u9])

            w9 = f.find('CCN :') + 5
            g = f[w9:]
            u9 = g.find(')') + w9
            ccn = f[w9:u9]
            if ccn.find(' ') != -1:
                u9 = g.find(' ') + w9
                ccn = f[w9:u9]
            hg = [sub.replace('  ', '') for sub in hg]
            hg = [sub.replace(':', '') for sub in hg]
            hg = [sub.replace('*', '') for sub in hg]
            hg = [sub.replace('\n', ' ') for sub in hg]
            s1.cell(row=t + 2, column=1).value = t + 1
            s1.cell(row=t + 2, column=2).value = ccn
            for i in range(0, len(hg)):
                s1.cell(row=t + 2, column=i + 3).value = hg[i]
            # print(hg)
            for i in sy:
                max_row = s2.max_row + 1
                k = i.find('.')
                jk = i[:k + 3]
                kj = i[k + 3:]
                wq += 1
                s2.cell(row=max_row, column=1).value = wq
                s2.cell(row=max_row, column=2).value = ccn
                s2.cell(row=max_row, column=3).value = jk
                s2.cell(row=max_row, column=4).value = kj
            os.rename(os.getcwd() + '/health_heritage/attachments_' + str(sys.argv[6]) + '/email' + str(t + 1) + ".html",
                      os.getcwd() + '/health_heritage/attachments_' + str(sys.argv[6]) + '/' + ccn + ".html")
        except Exception as e:
            s1.cell(row=t + 2, column=1).value = 'error'
            os.rename(os.getcwd() + '/health_heritage/attachments_' + str(sys.argv[6]) + '/email' + str(t + 1) + ".html",
                      os.getcwd() + '/health_heritage/attachments_' + str(sys.argv[6]) + '/' + ccn + ".html")

    print("Done")
    wbk.save(wbkName)
    wbk.close
    wbkName = 'count/count.xlsx'
    wbk = openpyxl.load_workbook(wbkName)
    s1 = wbk.worksheets[0]

    row_ = s1.max_row + 1
    s1.cell(row=row_, column=1).value = 'health hertige'
    s1.cell(row=row_, column=2).value = str(sys.argv[6])
    s1.cell(row=row_, column=3).value = len(fg)
    s1.cell(row=row_, column=4).value = len(onlyfiles)
    s1.cell(row=row_, column=5).value = len(repeat)
    wbk.save(wbkName)
    subprocess.run(["python", "updation.py", "1", "max", "9", " "])
except:
    log_exceptions()