import os
import re
import sys
import struct, time
import subprocess
from datetime import date
import datetime
import openpyxl
import pdftotext
import time
import requests
import camelot
now = datetime.datetime.now()

subprocess.run(["python", "updation.py","1","max1","1",sys.argv[2]])
subprocess.run(["python", "updation.py","1","max","2",sys.argv[3]])
subprocess.run(["python", "updation.py","1","max","3",sys.argv[4]])
subprocess.run(["python", "updation.py","1","max","5",str(now)])
subprocess.run(["python", "updation.py","1","max","7",sys.argv[5]])
subprocess.run(["python", "updation.py","1","max","8",sys.argv[6]])

with open(sys.argv[1], "rb") as f:
	pdf = pdftotext.PDF(f)

with open('Cholamandalam/output.txt', 'w') as f:
	f.write(" ".join(pdf))     
with open('Cholamandalam/output.txt', 'r') as myfile:
 	f = myfile.read()
tables = camelot.read_pdf(sys.argv[1])#,line_scale=10)
tables.export('Cholamandalam/foo1.xlsx', f='excel')
loc = ("Cholamandalam/foo1.xlsx")
wb= openpyxl.load_workbook(loc)
s1_t=wb.worksheets[0]
s2_t=wb.worksheets[1]
row_count_t = s2_t.max_row
try:		
	hg=[]
	
	w=f.find('Claim No')+9
	g=f[w:]
	u=g.find('\n')+w
	hg.append(f[w:u])

	status='Approved'
	hg.append(s2_t.cell(row_count_t-1, column=4).value)
			
	
	hg.append(' ')
	
	w=f.find('AL No')+7
	g=f[w:]
	u=g.find('\n')+w
	hg.append(f[w:u])

	temp = re.compile(r"(?<=Patient).*(?=Membership)").search(f)
	if temp is None:
		pname = ""
	else:
		pname = temp.group()

	hg=[sub.replace(',','') for sub in hg]
	hg=[sub.replace(':','') for sub in hg]
	hg=[sub.replace('  ','') for sub in hg]
	hg=[sub.replace('Rs.','') for sub in hg]
	#s2.cell(row_count_1+1, column=9).value='Yes'
	#s2.cell(row_count_1+1, column=10).value='NA'
	#wbk.save(wbkName)
	
	subprocess.run(["python", "updation.py","1","max","9",'Yes'])
	subprocess.run(["python", "updation.py","1","max","10",'NA'])
	
	try:
		subprocess.run(["python", "test_api.py",hg[0],hg[1],hg[2],' ',status,sys.argv[6],sys.argv[1],' ',hg[3],pname])
		'''wbk= openpyxl.load_workbook(wbkName)
		s2=wbk.worksheets[1]
		s2.cell(row_count_1+1, column=11).value='YES'
		'''
		subprocess.run(["python", "updation.py","1","max","11",'Yes'])	
	except Exception as e:
		#s2.cell(row_count_1+1, column=11).value='NO'
		subprocess.run(["python", "updation.py","1","max","11",'No'])
except Exception as e:
	#s2.cell(row_count_1+1, column=9).value='No'
	#s2.cell(row_count_1+1, column=11).value='NO'
	subprocess.run(["python", "updation.py","1","max","9",'Yes'])
	subprocess.run(["python", "updation.py","1","max","11",'No'])
now = datetime.datetime.now()
#s2.cell(row_count_1+1, column=6).value=now
#wbk.save(wbkName)
subprocess.run(["python", "updation.py","1","max","6",str(now)])
