import os
import sys
import struct, time
import subprocess
from datetime import date
import datetime
import openpyxl
import pdftotext
import time
import requests
now = datetime.datetime.now()
'''
wbkName = 'log file.xlsx'
wbk= openpyxl.load_workbook(wbkName)
s1=wbk.worksheets[0]
s2=wbk.worksheets[1]
row_count_1 = s2.max_row
s2.cell(row_count_1+1, column=1).value=sys.argv[2]
s2.cell(row_count_1+1, column=2).value=sys.argv[3]
s2.cell(row_count_1+1, column=3).value=sys.argv[4]
s2.cell(row_count_1+1, column=5).value=now
s2.cell(row_count_1+1, column=7).value=sys.argv[5]
s2.cell(row_count_1+1, column=8).value=sys.argv[6]
'''

# subprocess.run(["python", "updation.py","1","max1","1",sys.argv[2]])
# subprocess.run(["python", "updation.py","1","max","2",sys.argv[3]])
# subprocess.run(["python", "updation.py","1","max","3",sys.argv[4]])
# subprocess.run(["python", "updation.py","1","max","5",str(now)])
# subprocess.run(["python", "updation.py","1","max","7",sys.argv[5]])
# subprocess.run(["python", "updation.py","1","max","8",sys.argv[6]])

with open("C:\\Users\\91798\\Desktop\\trial_shikha-master2\\genins\\genins.pdf", "rb") as f:
        pdf = pdftotext.PDF(f)

# with open('hdfc/output1.txt', 'w') as f:
#       f.write(" ".join(pdf))     
# with open('hdfc/output1.txt', 'r') as myfile:
#       f = myfile.read()

# Added by Ashish (To fix the encoding Exception)
with open('genins/output1.txt', 'w',encoding="utf-8") as f:
        f.write(" ".join(pdf))     
# with open('genins/output1.txt', 'r') as myfile:
#         f = myfile.read()
# try:            
#         hg=[]
#         if f.find('CCN No') != -1:
#                 w=f.find('CCN No')+len('CCN No')
#                 k=f[w:]
#                 u=k.find("\n")+w
#                 g=f[w:u]
#                 hg.append(g)
#         elif f.find('CCN:')!= -1:
#                 w=f.find('CCN:')+len(CCN:)
#                 k=f[w:]
#                 u=k.find("\n")+w
#                 g=f[w:u]
#                 hg.append(g)

#         else:
#                 print("CCN not found") 

#         if f.find('facility for ')!=-1:
#                         temp='facility for ' 
#                         temp1='at'
#         elif f.find('Patient Name :')!=-1:
#                         temp='Patient Name :'
#                         temp1='\n'
#         w=f.find(temp)+len(temp)
#         k=f[w:]
#         u=k.find(temp1)+w
#         g=f[w:u]
#         hg.append(g)

#         if f.find('HDFC ERGO ID')!= -1: 
#                 w=f.find('HDFC ERGO ID')+12
#                 k=f[w:]
#                 u=k.find("\n")+w
#                 g=f[w:u]
#                 hg.append(g)
#         else:
#                 print("HDFC ERGO ID not found")
#         hg=[sub.replace(':','') for sub in hg]  
#         hg=[sub.replace('  ','') for sub in hg]
#         hg=[sub.replace('Rs.','') for sub in hg]
#         #s2.cell(row_count_1+1, column=9).value='Yes'
#         #s2.cell(row_count_1+1, column=10).value='NA'
#         #wbk.save(wbkName)
        
#         subprocess.run(["python", "updation.py","1","max","9",'Yes'])
#         subprocess.run(["python", "updation.py","1","max","10",'NA'])
        
#         try:
#                 subprocess.run(["python", "test_api.py",hg[0],'','','','Denial',sys.argv[6],sys.argv[1],'',hg[2],hg[1]])
#                 '''wbk= openpyxl.load_workbook(wbkName)
#                 s2=wbk.worksheets[1]
#                 s2.cell(row_count_1+1, column=11).value='YES'
#                 '''
#                 subprocess.run(["python", "updation.py","1","max","11",'Yes'])  
#         except Exception as e:
                #s2.cell(row_count_1+1, column=11).value='NO'
#                 subprocess.run(["python", "updation.py","1","max","11",'No'])
# except Exception as e:
#         #s2.cell(row_count_1+1, column=9).value='No'
#         #s2.cell(row_count_1+1, column=11).value='NO'
#         subprocess.run(["python", "updation.py","1","max","9",'Yes'])
#         subprocess.run(["python", "updation.py","1","max","11",'No'])
# now = datetime.datetime.now()
#s2.cell(row_count_1+1, column=6).value=now
#wbk.save(wbkName)
#subprocess.run(["python", "updation.py","1","max","6",str(now)])
