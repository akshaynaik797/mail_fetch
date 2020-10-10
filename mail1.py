from email.mime.text import MIMEText
import smtplib
import time
import imaplib
import sys
import email
import os
import struct, time
import subprocess
from datetime import date
import datetime
import openpyxl
from itertools import chain
now = datetime.datetime.now()
today = datetime.date.today()
today = today.strftime('%d-%b-%Y')
f= open("unread_mail.txt","r")
fo= open("defualt_time_read.txt","r")
t=fo.read()
tg=t.split()
c=f.read()
fg=c.split()
#print(today)
f=open("unread_mail.txt", "a+")

wbkName = 'log file.xlsx'
wbk= openpyxl.load_workbook(wbkName)
s1=wbk.worksheets[0]	
row_count_1 = s1.max_row
wbk.save(wbkName)
'''
s1.cell(row_count_1+1, column=1).value=row_count_1
s1.cell(row_count_1+1, column=2).value=today
s1.cell(row_count_1+1, column=3).value=now
wbk.save(wbkName)
'''

subprocess.run(["python", "updation.py","0","max1","1",str(row_count_1)])
subprocess.run(["python", "updation.py","0","max","2",str(today)])
subprocess.run(["python", "updation.py","0","max","3",str(now)])
class Mail():
	import email
	import re
	import subprocess
	import pdfkit
	import openpyxl
	from datetime import datetime
	path_wkhtmltopdf = r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe'
	config = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf)
	b=0
	s_r=0
	def __init__(self):
		self.wbkName_config = 'email ids subjects.xlsx'
		self.wb= self.openpyxl.load_workbook(self.wbkName_config)
		s_c1=self.wb.worksheets[2]
		s_c2=self.wb.worksheets[1]
		self.subj=[]
		self.mail_id=[]
		for i in range(2,34):
			subj_t=[]
			mail_t=[]
			for j in range(2,9):
				temp_subj=s_c1.cell(row=i, column=j).value
				if(temp_subj!=None):
					#print(temp_subj)
                                        if (s_c1.cell(row=i, column=12).value==str(sys.argv[3])):
                                                subj_t.append(temp_subj.split(','))
                                        
                                        else:
                                                t=temp_subj.split(',')
                                                k=[]
                                                for h in t:
                                                        k.append('Do not need to process at all')
                                                subj_t.append(k)
                                                
				else:
					subj_t.append([''])
				
			self.subj.append(subj_t)
			if (s_c1.cell(row=i, column=9).value!=None):
                                if (s_c1.cell(row=i, column=12).value==str(sys.argv[3])):
                                        self.mail_id_1=s_c1.cell(row=i, column=9).value.split(',')
                                mail_t=s_c1.cell(row=i, column=9).value.split(',')
			else:
				mail_t=''
			self.mail_id.append(mail_t)
		for i in self.subj:
			for j in i:
				for kk,ss in enumerate(j):
					j[kk] = ss.replace("(","")
		for i in self.subj:
			for j in i:
				for kk,ss in enumerate(j):
					j[kk] = ss.replace(")","")
		#print(self.subj,self.mail_id)
		try:
			#SMTP_SERVER ='imap.gmail.com'
			SMTP_SERVER ='outlook.office365.com'
			self.mail = imaplib.IMAP4_SSL(SMTP_SERVER)	
			#self.mail.e_id ='mediclaim@inamdarhospital.org'
			#self.mail.pswd='Mediclaim@2019'
			self.mail.e_id ='Tpappg@maxhealthcare.com'
			self.mail.pswd='May@2020'
			self.mail.login(user =self.mail.e_id, password =self.mail.pswd)
			'''wbk= self.openpyxl.load_workbook(wbkName)
			s1=wbk.worksheets[0]
			s1.cell(row_count_1+1, column=6).value='YES'
			wbk.save(wbkName)
			'''
			subprocess.run(["python", "updation.py","0","max","6",'YES'])
		except Exception as e:
			'''wbk= self.openpyxl.load_workbook(wbkName)
			s1=wbk.worksheets[0]
			s1.cell(row_count_1+1, column=6).value='NO'
			wbk.save(wbkName)
			'''
			subprocess.run(["python", "updation.py","0","max","6",'NO'])
	def download_html(self,ins,ct):
		now = datetime.datetime.now()
		#wbkName = 'log file.xlsx'
		dirFiles = os.listdir(ins+'/attachments_'+ct)
		dirFiles.sort(key=lambda l: int(self.re.sub('\D', '', l)))
		#print(dirFiles[-1])
		if(len(dirFiles)>0):
			m=dirFiles[-1].find('.')
			self.b=int(dirFiles[-1][:m])
		else:
			self.b=0
		self.b+=1
		#print(self.b)
		for self.mail.part in self.mail.email_message.walk():
				if self.mail.part.get_content_type() == "text/html":
					self.mail.body = self.mail.part.get_payload(decode=True)
					self.mail.file_name = ins+'/email.html'
					self.mail.output_file = open(self.mail.file_name, 'w')
					self.mail.output_file.write("Body: %s" %(self.mail.body.decode('utf-8')))
					self.mail.output_file.close()
					
					self.pdfkit.from_file(ins+'/email.html', ins+'/attachments_'+ct+'/'+str(self.b)+'.pdf',configuration=self.config)
		'''wbk= self.openpyxl.load_workbook(wbkName)
		s1=wbk.worksheets[0]
		s1.cell(row_count_1+1, column=11).value=row_count_1
		wbk.save(wbkName)
		'''
		subprocess.run(["python", "updation.py","0","max","11",str(row_count_1)])
		#print(self.mail.email_message['Date'])
		self.l_time=self.mail.email_message['Date']
		w=self.l_time.find(',')+1
		g=self.l_time[w:]
		u=g.find('+')+w
		s=self.l_time[w:u]
		s=s.split(' ')
		while("" in s) : 
		    s.remove("") 
		#print(s)
		if(len(s)==4):
			t=s[0]+' '+s[1]+' '+s[2]+' '+s[3]

			d = self.datetime.strptime(t, '%d %b %Y %H:%M:%S')
			self.l_time=d.strftime('%d/%m/%Y %H:%M:%S')
		else:
			self.l_time=self.mail.email_message['Date']
		self.subprocess.run(["python", ins+"_"+ct+".py", ins+'/attachments_'+ct+'/'+str(self.b)+'.pdf',str(row_count_1),ins,ct,self.mail.email_message['Subject'],self.l_time])
		self.s_r+=1
		'''
		wbk= self.openpyxl.load_workbook(wbkName)
		s2=wbk.worksheets[1]
		row_count_2 = s2.max_row	
		s2.cell(row_count_2, column=4).value=now
		wbk.save(wbkName)
		'''
		subprocess.run(["python", "updation.py","1","max","4",str(now)])
	def download_pdf(self,ins,ct):
		import os
		print("pdf downloading")
		now = datetime.datetime.now()
		wbkName = 'log file.xlsx'
		dol=0
		#try:
		try:
			t_p=0
			for self.mail.part in self.mail.email_message.walk():
				if self.mail.part.get_content_maintype() == 'multipart':
				# print part.as_string()
					continue
				#if self.mail.part.get('Content-Disposition') is None:
				# print part.as_string()
				#	continue
				self.mail.fileName = self.mail.part.get_filename()
				self.mail.detach_dir=(os.getcwd()+'/'+ins+'/attachments_pdf_'+ct+'/')
				if bool(self.mail.fileName):
					if(ins=='MDINDIA' and self.mail.fileName.find('MDI')==-1):
						continue
					if(ins=='hdfc' and self.mail.fileName.find('KYCForm')!=-1):
						continue
					if(ins=='Raksha' and self.mail.fileName.find('DECLARATION')!=-1):
						continue
					else:
						t_p=1
						self.mail.filePath = os.path.join(self.mail.detach_dir, self.mail.fileName)
						#if not os.path.isfile(self.mail.filePath) :
						fp = open(self.mail.filePath, 'wb')
						fp.write(self.mail.part.get_payload(decode=True))
						fp.close()
			dol=1
			
		except Exception as e:
			'''wbk= self.openpyxl.load_workbook(wbkName)
			s1=wbk.worksheets[0]
			s1.cell(row_count_1+1, column=11).value=row_count_1
			wbk.save(wbkName)
			'''
			subprocess.run(["python", "updation.py","0","max","11",str(row_count_1)])
			wbk.save(wbkName)
			self.s_r+=1
			'''
			wbk= self.openpyxl.load_workbook(wbkName)
			s2=wbk.worksheets[1]
			row_count_2 = s2.max_row	
			#s2.cell(row_count_2, column=4).value=now
			s2.cell(row_count_2+1, column=1).value=row_count_1
			s2.cell(row_count_2+1, column=2).value=ins
			s2.cell(row_count_2+1, column=3).value=ct
			s2.cell(row_count_2+1, column=15).value='error while downloading'
			wbk.save(wbkName)
			'''
			subprocess.run(["python", "updation.py","1","max","1",str(row_count_1)])
			subprocess.run(["python", "updation.py","1","max","2",str(ins)])
			subprocess.run(["python", "updation.py","1","max","3",str(ct)])
			subprocess.run(["python", "updation.py","1","max","15",'error while downloading'])
			#print(e)		
		if (t_p==0):
			#print(self.mail.email_message['Subject'])
			self.download_html(ins,ct)
			return 0
		if(dol==1):
			'''wbk= self.openpyxl.load_workbook(wbkName)
			s1=wbk.worksheets[0]
			s1.cell(row_count_1+1, column=11).value=row_count_1
			wbk.save(wbkName)
			'''
			subprocess.run(["python", "updation.py","0","max","11",str(row_count_1)])
			#print(self.mail.email_message['Date'])
			self.l_time=self.mail.email_message['Date']
			w=self.l_time.find(',')+1
			g=self.l_time[w:]
			u=g.find('+')+w
			s=self.l_time[w:u]
			s=s.split(' ')
			while("" in s) : 
			    s.remove("") 
			#print(s)
			if(len(s)==4):
				t=s[0]+' '+s[1]+' '+s[2]+' '+s[3]

				d = self.datetime.strptime(t, '%d %b %Y %H:%M:%S')
				self.l_time=d.strftime('%d/%m/%Y %H:%M:%S')
			else:
				self.l_time=self.mail.email_message['Date']
			self.subprocess.run(["python", ins+"_"+ct+".py", self.mail.filePath,str(row_count_1),ins,ct,self.mail.email_message['Subject'],self.l_time])
			self.s_r+=1
			'''
			wbk= self.openpyxl.load_workbook(wbkName)
			s2=wbk.worksheets[1]
			row_count_2 = s2.max_row	
			s2.cell(row_count_2, column=4).value=now
			wbk.save(wbkName)
			'''
			subprocess.run(["python", "updation.py","1","max","4",str(now)])
	def process(self):
		for i in range (0,len(self.mail.id_list)):
			if len(fg)==0 or str(self.mail.id_list[i])>fg[-1]:
				try:
					self.mail.latest_email_id = self.mail.id_list[i] # get the latest
					self.mail.result, self.mail.data1 = self.mail.fetch(self.mail.latest_email_id, "(RFC822)") 
					self.mail.raw_email = self.mail.data1[0][1].decode('utf-8')
					self.mail.email_message = self.email.message_from_string(self.mail.raw_email)
					#print(self.mail.email_message['Subject'])
					#if(self.mail.email_message['Subject'] == self.subj[0][0][0]):
						#print('hohoho',[self.mail.email_message['From']])
					#print(self.mail.email_message['Subject'])
					if(self.mail.email_message['Subject']==None):
						continue
					elif(self.mail.email_message['Subject'].find(self.subj[9][0][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[9][0])!=-1):
						self.download_pdf('health_india','preauth')

					elif(self.mail.email_message['Subject'].find(self.subj[9][1][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[9][0])!=-1):
						self.download_pdf('health_india','query')

					elif(self.mail.email_message['Subject'].find(self.subj[9][3][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[9][0])!=-1):
						self.download_pdf('health_india','final')


					elif(self.mail.email_message['Subject'].find(self.subj[17][0][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[17][0])!=-1):
						self.download_pdf('Raksha','preauth')

					elif(self.mail.email_message['Subject'].find(self.subj[17][1][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[17][0])!=-1):
						self.download_pdf('Raksha','query')

					elif(self.mail.email_message['Subject'].find(self.subj[17][3][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[17][0])!=-1):
						self.download_pdf('Raksha','final')

					elif(self.mail.email_message['Subject'].find(self.subj[18][0][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[18][0])!=-1):
						self.download_pdf('safeway','preauth')

					elif(self.mail.email_message['Subject'].find(self.subj[18][1][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[18][0])!=-1):
						self.download_pdf('safeway','query')

					elif(self.mail.email_message['Subject'].find(self.subj[18][3][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[18][0])!=-1):
						self.download_pdf('safeway','final')

					elif(self.mail.email_message['Subject'].find(self.subj[19][0][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[19][0])!=-1):
						self.download_pdf('united','preauth')

					elif(self.mail.email_message['Subject'].find(self.subj[19][1][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[19][0])!=-1):
						self.download_pdf('united','query')

					elif(self.mail.email_message['Subject'].find(self.subj[19][3][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[19][0])!=-1):
						self.download_pdf('united','final')


					elif(self.mail.email_message['Subject'].find(self.subj[20][0][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[20][0])!=-1):
						self.download_pdf('vidal','preauth')

					elif(self.mail.email_message['Subject'].find(self.subj[20][1][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[20][0])!=-1):
						self.download_pdf('vidal','query')
						
					elif(self.mail.email_message['Subject'].find(self.subj[20][3][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[20][0])!=-1):
						self.download_pdf('vidal','final')

					elif(self.mail.email_message['Subject'].find(self.subj[21][0][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[21][0])!=-1):
						self.download_pdf('vipul','preauth')

					elif(self.mail.email_message['Subject'].find(self.subj[21][1][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[21][0])!=-1):
						self.download_pdf('vipul','query')

					elif(self.mail.email_message['Subject'].find(self.subj[22][0][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[22][0])!=-1):
						self.download_pdf('hdfc','preauth')

					elif(self.mail.email_message['Subject'].find(self.subj[22][1][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[22][0])!=-1):
						self.download_pdf('hdfc','query')

					elif(self.mail.email_message['Subject'].find(self.subj[22][3][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[22][0])!=-1):
						self.download_pdf('hdfc','final')


					elif(self.mail.email_message['Subject'].find(self.subj[25][0][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[25][0])!=-1):
						self.download_pdf('IFFCO','preauth')

					elif(self.mail.email_message['Subject'].find(self.subj[25][1][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[25][0])!=-1):
						self.download_pdf('IFFCO','query')

					elif(self.mail.email_message['Subject'].find(self.subj[25][3][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[25][0])!=-1):
						self.download_pdf('IFFCO','final')



					elif(self.mail.email_message['Subject'].find(self.subj[26][0][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[26][0])!=-1):
						self.download_pdf('Max_Bupa','preauth')

					elif(self.mail.email_message['Subject'].find(self.subj[26][1][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[26][0])!=-1):
						self.download_pdf('Max_Bupa','query')

					elif(self.mail.email_message['Subject'].find(self.subj[26][3][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[26][0])!=-1):
						self.download_pdf('Max_Bupa','final')	


					elif(self.mail.email_message['Subject'].find(self.subj[27][0][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[27][0])!=-1):
						self.download_pdf('reliance','preauth')

					elif(self.mail.email_message['Subject'].find(self.subj[27][1][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[27][0])!=-1):
						self.download_pdf('reliance','query')

					elif(self.mail.email_message['Subject'].find(self.subj[27][3][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[27][0])!=-1):
						self.download_pdf('reliance','final')

					elif(self.mail.email_message['Subject'].find(self.subj[28][0][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[28][0])!=-1):
						self.download_pdf('religare','preauth')

					elif(self.mail.email_message['Subject'].find(self.subj[28][1][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[28][0])!=-1):
						self.download_pdf('religare','query')

					elif(self.mail.email_message['Subject'].find(self.subj[28][3][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[28][0])!=-1):
						self.download_pdf('religare','final')

					elif(self.mail.email_message['Subject'].find(self.subj[29][0][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[29][1])!=-1):
						'''
						wbkName= 'log file.xlsx'
						wbk= self.openpyxl.load_workbook(wbkName)
						s2=wbk.worksheets[1]
						row_count_2 = s2.max_row	
						s2.cell(row_count_2+1, column=4).value=now
						s2.cell(row_count_2+1, column=1).value=row_count_1
						s2.cell(row_count_2+1, column=2).value='star'
						s2.cell(row_count_2+1, column=3).value='preauth'
						s2.cell(row_count_2+1, column=15).value='error while downloading'
						wbk.save(wbkName)
						'''
						subprocess.run(["python", "updation.py","1","max1","4",str(now)])
						subprocess.run(["python", "updation.py","1","max","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","1","max","2",'star'])
						subprocess.run(["python", "updation.py","1","max","3",'preauth'])
						subprocess.run(["python", "updation.py","1","max","7",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","1","max","15",'error while downloading'])

						self.s_r+=1
						star_subject=self.mail.email_message['Subject']
						w=star_subject.find('-')
						ccn=star_subject[:w-1]
					
						self.l_time=self.mail.email_message['Date']
						w=self.l_time.find(',')+1
						g=self.l_time[w:]
						u=g.find('+')+w
						s=self.l_time[w:u]
						s=s.split(' ')
						while("" in s) : 
						    s.remove("") 
						#print(s)
						if(len(s)==4):
							t=s[0]+' '+s[1]+' '+s[2]+' '+s[3]

							d = self.datetime.strptime(t, '%d %b %Y %H:%M:%S')
							self.l_time=d.strftime('%d/%m/%Y %H:%M:%S')
						else:
							self.l_time=self.mail.email_message['Date']
						
						self.subprocess.run(["python", "star_download.py",ccn,self.mail.email_message['Subject'],self.l_time,str(row_count_1),'preauth'])

					elif(self.mail.email_message['Subject'].find(self.subj[29][1][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[29][2])!=-1):
						self.download_pdf('star','query')

					elif(self.mail.email_message['Subject'].find(self.subj[29][3][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[29][1])!=-1):
						'''
						wbkName= 'log file.xlsx'
						wbk= self.openpyxl.load_workbook(wbkName)
						s2=wbk.worksheets[1]
						row_count_2 = s2.max_row	
						s2.cell(row_count_2+1, column=4).value=now
						s2.cell(row_count_2+1, column=1).value=row_count_1
						s2.cell(row_count_2+1, column=2).value='star'
						s2.cell(row_count_2+1, column=3).value='final'
						s2.cell(row_count_2+1, column=15).value='error while downloading'
						wbk.save(wbkName)
						'''
						subprocess.run(["python", "updation.py","1","max1","4",str(now)])
						subprocess.run(["python", "updation.py","1","max","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","1","max","2",'star'])
						subprocess.run(["python", "updation.py","1","max","3",'final'])
						subprocess.run(["python", "updation.py","1","max","7",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","1","max","15",'error while downloading'])
						self.s_r+=1
						star_subject=self.mail.email_message['Subject']
						w=star_subject.find('-')
						ccn=star_subject[:w-1]

						self.l_time=self.mail.email_message['Date']
						w=self.l_time.find(',')+1
						g=self.l_time[w:]
						u=g.find('+')+w
						s=self.l_time[w:u]
						s=s.split(' ')
						while("" in s) : 
						    s.remove("") 
						#print(s)
						if(len(s)==4):
							t=s[0]+' '+s[1]+' '+s[2]+' '+s[3]

							d = self.datetime.strptime(t, '%d %b %Y %H:%M:%S')
							self.l_time=d.strftime('%d/%m/%Y %H:%M:%S')
						else:
							self.l_time=self.mail.email_message['Date']
						self.subprocess.run(["python", "star_download.py",ccn,self.mail.email_message['Subject'],self.l_time,str(row_count_1),'final'])

					elif(self.mail.email_message['Subject'].find(self.subj[30][0][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[30][0])!=-1):
						self.download_pdf('Universal_Sompo','preauth')

					elif(self.mail.email_message['Subject'].find(self.subj[30][1][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[30][0])!=-1):
						self.download_pdf('Universal_Sompo','query')

					elif(self.mail.email_message['Subject'].find(self.subj[30][3][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[30][0])!=-1):
						self.download_pdf('Universal_Sompo','final')


					elif(self.mail.email_message['Subject'] in self.subj[0][0] and self.mail.email_message['From'].find(self.mail_id[0][0])!=-1):
						self.download_pdf('aditya','preauth')

					elif(self.mail.email_message['Subject'].find(self.subj[0][1][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[0][0])!=-1):
						self.download_pdf('aditya','query')

					elif(self.mail.email_message['Subject'].find(self.subj[0][2][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[0][0])!=-1):
						self.download_pdf('aditya','interim')

					elif(self.mail.email_message['Subject'] in self.subj[0][3] and self.mail.email_message['From'].find(self.mail_id[0][0])!=-1):
						self.download_pdf('aditya','final')

					elif(self.mail.email_message['Subject'] in self.subj[0][6] and self.mail.email_message['From'].find(self.mail_id[0][0])!=-1):
						self.download_pdf('aditya','settlement')

					elif(self.mail.email_message['Subject'].find(self.subj[1][0][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[1][0])!=-1):
						self.download_pdf('apollo','preauth')

					elif(self.mail.email_message['Subject'].find(self.subj[1][1][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[1][0])!=-1):
						self.download_pdf('apollo','query')

					#elif(self.mail.email_message['Subject'] in self.subj[1][2] and self.mail.email_message['From'].find(self.mail_id[0][0])!=-1):
						#self.download_pdf('apollo','final')
					#elif(self.mail.email_message['Subject'] in self.subj[1][5] and self.mail.email_message['From'].find(self.mail_id[0][0])!=-1):
						#self.download_pdf('aditya','settlement')

					elif(self.mail.email_message['Subject'].find(self.subj[2][0][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[2][0])!=-1):
						self.download_pdf('Cholamandalam','preauth')

					elif(self.mail.email_message['Subject'].find(self.subj[2][1][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[2][0])!=-1):
						self.download_pdf('Cholamandalam','query')

					elif(self.mail.email_message['Subject'].find(self.subj[2][3][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[2][0])!=-1):
						self.download_pdf('Cholamandalam','final')
					
					elif(self.mail.email_message['Subject'].find(self.subj[3][0][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[3][0])!=-1):
						self.download_pdf('fgh','preauth')

					elif(self.mail.email_message['Subject'].find(self.subj[3][1][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[3][0])!=-1):
						self.download_pdf('fgh','query')

					elif(self.mail.email_message['Subject'].find(self.subj[3][3][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[3][0])!=-1):
						self.download_pdf('fgh','final')
				
					elif(self.mail.email_message['Subject'].find(self.subj[4][0][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[4][0])!=-1):
						self.download_pdf('east_west','preauth')

					elif(self.mail.email_message['Subject'].find(self.subj[4][1][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[4][0])!=-1):
						self.download_pdf('east_west','query')

					elif(self.mail.email_message['Subject'].find(self.subj[4][3][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[4][0])!=-1):
						self.download_pdf('east_west','final')

					elif(self.mail.email_message['Subject'].find(self.subj[6][0][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[6][0])!=-1):
						self.download_pdf('fhpl','preauth')

					elif(self.mail.email_message['Subject'].find(self.subj[6][1][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[6][0])!=-1):
						self.download_pdf('fhpl','query')

					elif(self.mail.email_message['Subject'].find(self.subj[6][3][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[6][0])!=-1):
						self.download_pdf('fhpl','final')


					elif(self.mail.email_message['Subject'].find(self.subj[10][0][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[10][0])!=-1):
						self.download_pdf('health_insurance','preauth')


					elif(self.mail.email_message['Subject'].find(self.subj[11][0][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[11][0])!=-1):
						self.download_pdf('health_heritage','preauth')

					elif(self.mail.email_message['Subject'].find(self.subj[11][1][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[11][0])!=-1):
						self.download_pdf('health_heritage','query')

					elif(self.mail.email_message['Subject'].find(self.subj[11][3][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[11][0])!=-1):
						self.download_pdf('health_heritage','final')

					elif(self.mail.email_message['Subject'].find(self.subj[12][0][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[12][0])!=-1 and self.mail.email_message['Subject'].find('Denial')==-1):
						self.download_pdf('MDINDIA','preauth')

					elif(self.mail.email_message['Subject'].find(self.subj[12][1][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[12][0])!=-1):
						self.download_pdf('MDINDIA','query')

					#elif(self.mail.email_message['Subject'].find(self.subj[12][3][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[12][0])!=-1):
						#self.download_pdf('MDINDIA','final')


					elif(self.mail.email_message['Subject'].find(self.subj[13][0][0])!=-1 and self.mail.email_message['Subject'].find('Denial')==-1 and self.mail.email_message['From'].find(self.mail_id[13][0])!=-1):
						self.download_pdf('Medi_Assist','preauth')

					elif(self.mail.email_message['Subject'].find(self.subj[13][1][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[13][0])!=-1):
						self.download_pdf('Medi_Assist','query')

					elif(self.mail.email_message['Subject'].find(self.subj[13][3][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[13][0])!=-1):
						self.download_pdf('Medi_Assist','final')

					elif(self.mail.email_message['Subject'].find(self.subj[14][0][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[14][0])!=-1):
						self.download_pdf('Medsave','preauth')

					elif(self.mail.email_message['Subject'].find(self.subj[14][1][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[14][0])!=-1):
						self.download_pdf('Medsave','query')

					elif(self.mail.email_message['Subject'].find(self.subj[14][3][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[14][0])!=-1):
						self.download_pdf('Medsave','final')

					elif(self.mail.email_message['Subject'].find(self.subj[15][0][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[15][0])!=-1):
						self.download_pdf('Paramount','preauth')

					elif(self.mail.email_message['Subject'].find(self.subj[15][1][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[15][0])!=-1):
						self.download_pdf('Paramount','query')

					elif(self.mail.email_message['Subject'].find(self.subj[15][3][0])!=-1 and self.mail.email_message['From'].find(self.mail_id[15][0])!=-1):
						self.download_pdf('Paramount','final')

					elif(self.mail.email_message['Subject'] in self.subj[1][6][0] and self.mail.email_message['From'].find(self.mail_id[1][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('settlement')])

					elif(self.mail.email_message['Subject'] in self.subj[2][6][0] and self.mail.email_message['From'].find(self.mail_id[2][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('settlement')])

					elif(self.mail.email_message['Subject'] in self.subj[3][6][0] and self.mail.email_message['From'].find(self.mail_id[3][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('settlement')])

					elif(self.mail.email_message['Subject'] in self.subj[4][6][0] and self.mail.email_message['From'].find(self.mail_id[4][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('settlement')])

					elif(self.mail.email_message['Subject'] in self.subj[5][6][0] and self.mail.email_message['From'].find(self.mail_id[5][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('settlement')])

					elif(self.mail.email_message['Subject'] in self.subj[6][6][0] and self.mail.email_message['From'].find(self.mail_id[6][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('settlement')])

					elif(self.mail.email_message['Subject'] in self.subj[7][6][0] and self.mail.email_message['From'].find(self.mail_id[7][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('settlement')])

					elif(self.mail.email_message['Subject'] in self.subj[8][6][0] and self.mail.email_message['From'].find(self.mail_id[8][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('settlement')])


					elif(self.mail.email_message['Subject'] in self.subj[9][6][0] and self.mail.email_message['From'].find(self.mail_id[9][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('settlement')])

					elif(self.mail.email_message['Subject'] in self.subj[10][6][0] and self.mail.email_message['From'].find(self.mail_id[10][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('settlement')])

					elif(self.mail.email_message['Subject'] in self.subj[11][6][0] and self.mail.email_message['From'].find(self.mail_id[11][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('settlement')])

					elif(self.mail.email_message['Subject'] in self.subj[12][6][0] and self.mail.email_message['From'].find(self.mail_id[12][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('settlement')])

					elif(self.mail.email_message['Subject'] in self.subj[13][6][0] and self.mail.email_message['From'].find(self.mail_id[13][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('settlement')])

					elif(self.mail.email_message['Subject'] in self.subj[14][6][0] and self.mail.email_message['From'].find(self.mail_id[14][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('settlement')])

					elif(self.mail.email_message['Subject'] in self.subj[15][6][0] and self.mail.email_message['From'].find(self.mail_id[15][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('settlement')])

					elif(self.mail.email_message['Subject'] in self.subj[16][6][0] and self.mail.email_message['From'].find(self.mail_id[16][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('settlement')])


					elif(self.mail.email_message['Subject'] in self.subj[17][6][0] and self.mail.email_message['From'].find(self.mail_id[17][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('settlement')])

					elif(self.mail.email_message['Subject'] in self.subj[18][6][0] and self.mail.email_message['From'].find(self.mail_id[18][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('settlement')])

					elif(self.mail.email_message['Subject'] in self.subj[19][6][0] and self.mail.email_message['From'].find(self.mail_id[19][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('settlement')])

					elif(self.mail.email_message['Subject'] in self.subj[20][6][0] and self.mail.email_message['From'].find(self.mail_id[20][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('settlement')])

					elif(self.mail.email_message['Subject'] in self.subj[21][6][0] and self.mail.email_message['From'].find(self.mail_id[21][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('settlement')])

					elif(self.mail.email_message['Subject'] in self.subj[22][6][0] and self.mail.email_message['From'].find(self.mail_id[22][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('settlement')])

					elif(self.mail.email_message['Subject'] in self.subj[23][6][0] and self.mail.email_message['From'].find(self.mail_id[23][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('settlement')])

					elif(self.mail.email_message['Subject'] in self.subj[24][6][0] and self.mail.email_message['From'].find(self.mail_id[24][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('settlement')])


					elif(self.mail.email_message['Subject'] in self.subj[25][6][0] and self.mail.email_message['From'].find(self.mail_id[25][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('settlement')])

					elif(self.mail.email_message['Subject'] in self.subj[26][6][0] and self.mail.email_message['From'].find(self.mail_id[26][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('settlement')])

					elif(self.mail.email_message['Subject'] in self.subj[27][6][0] and self.mail.email_message['From'].find(self.mail_id[27][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('settlement')])

					elif(self.mail.email_message['Subject'] in self.subj[28][6][0] and self.mail.email_message['From'].find(self.mail_id[28][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('settlement')])

					elif(self.mail.email_message['Subject'] in self.subj[29][6][0] and self.mail.email_message['From'].find(self.mail_id[29][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('settlement')])

					elif(self.mail.email_message['Subject'] in self.subj[30][6][0] and self.mail.email_message['From'].find(self.mail_id[30][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('settlement')])


					#rejection mails
					elif(self.mail.email_message['Subject'] in self.subj[0][5][0] and self.mail.email_message['From'].find(self.mail_id[0][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])

					elif(self.mail.email_message['Subject'] in self.subj[1][5][0] and self.mail.email_message['From'].find(self.mail_id[1][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])

					elif(self.mail.email_message['Subject'] in self.subj[2][5][0] and self.mail.email_message['From'].find(self.mail_id[2][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])

					elif(self.mail.email_message['Subject'] in self.subj[3][5][0] and self.mail.email_message['From'].find(self.mail_id[3][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])

					elif(self.mail.email_message['Subject'] in self.subj[4][5][0] and self.mail.email_message['From'].find(self.mail_id[4][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])

					elif(self.mail.email_message['Subject'] in self.subj[5][5][0] and self.mail.email_message['From'].find(self.mail_id[5][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])

					elif(self.mail.email_message['Subject'] in self.subj[6][5][0] and self.mail.email_message['From'].find(self.mail_id[6][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])

					elif(self.mail.email_message['Subject'] in self.subj[7][5][0] and self.mail.email_message['From'].find(self.mail_id[7][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])

					elif(self.mail.email_message['Subject'] in self.subj[8][5][0] and self.mail.email_message['From'].find(self.mail_id[8][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])


					elif(self.mail.email_message['Subject'] in self.subj[9][5][0] and self.mail.email_message['From'].find(self.mail_id[9][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])

					elif(self.mail.email_message['Subject'] in self.subj[10][5][0] and self.mail.email_message['From'].find(self.mail_id[10][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])

					elif(self.mail.email_message['Subject'] in self.subj[11][5][0] and self.mail.email_message['From'].find(self.mail_id[11][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])

					elif(self.mail.email_message['Subject'] in self.subj[12][5][0] and self.mail.email_message['From'].find(self.mail_id[12][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])

					elif(self.mail.email_message['Subject'] in self.subj[13][5][0] and self.mail.email_message['From'].find(self.mail_id[13][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])

					elif(self.mail.email_message['Subject'] in self.subj[14][5][0] and self.mail.email_message['From'].find(self.mail_id[14][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])

					elif(self.mail.email_message['Subject'] in self.subj[15][5][0] and self.mail.email_message['From'].find(self.mail_id[15][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])

					elif(self.mail.email_message['Subject'] in self.subj[16][5][0] and self.mail.email_message['From'].find(self.mail_id[16][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])


					elif(self.mail.email_message['Subject'] in self.subj[17][5][0] and self.mail.email_message['From'].find(self.mail_id[17][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])

					elif(self.mail.email_message['Subject'] in self.subj[18][5][0] and self.mail.email_message['From'].find(self.mail_id[18][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])

					elif(self.mail.email_message['Subject'] in self.subj[19][5][0] and self.mail.email_message['From'].find(self.mail_id[19][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])

					elif(self.mail.email_message['Subject'] in self.subj[20][5][0] and self.mail.email_message['From'].find(self.mail_id[20][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])

					elif(self.mail.email_message['Subject'] in self.subj[21][5][0] and self.mail.email_message['From'].find(self.mail_id[21][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])

					elif(self.mail.email_message['Subject'] in self.subj[22][5][0] and self.mail.email_message['From'].find(self.mail_id[22][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])

					elif(self.mail.email_message['Subject'] in self.subj[23][5][0] and self.mail.email_message['From'].find(self.mail_id[23][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])

					elif(self.mail.email_message['Subject'] in self.subj[24][5][0] and self.mail.email_message['From'].find(self.mail_id[24][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])


					elif(self.mail.email_message['Subject'] in self.subj[25][5][0] and self.mail.email_message['From'].find(self.mail_id[25][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])

					elif(self.mail.email_message['Subject'] in self.subj[26][5][0] and self.mail.email_message['From'].find(self.mail_id[26][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])

					elif(self.mail.email_message['Subject'] in self.subj[27][5][0] and self.mail.email_message['From'].find(self.mail_id[27][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])

					elif(self.mail.email_message['Subject'] in self.subj[28][5][0] and self.mail.email_message['From'].find(self.mail_id[28][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])

					elif(self.mail.email_message['Subject'] in self.subj[29][5][0] and self.mail.email_message['From'].find(self.mail_id[29][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])

					elif(self.mail.email_message['Subject'] in self.subj[30][5][0] and self.mail.email_message['From'].find(self.mail_id[30][0])!=-1):
						subprocess.run(["python", "updation.py","4","max1","1",str(row_count_1)])
						subprocess.run(["python", "updation.py","4","max","2",str(self.mail.email_message['Date'])])
						subprocess.run(["python", "updation.py","4","max","3",str(self.mail.email_message['Subject'])])
						subprocess.run(["python", "updation.py","4","max","4",str(self.mail.email_message['From'])])
						subprocess.run(["python", "updation.py","4","max","5",str('Rejection')])					

				

					elif (1):
						temp_v=0
						for j in self.mail_id:
							for kk,ss in enumerate(j):
								if(self.mail.email_message['From'].find(j[kk])!=-1):
									temp_v=1
						temp_p=0
						for i in self.subj:
							for j in i:
								for kk,ss in enumerate(j):
									if(self.mail.email_message['Subject'].find(j[kk])!=-1 and j[kk]!=''):
										temp_p=1
						if temp_v==0:	
							'''		
							wbkName= 'log file.xlsx'
							wbk= self.openpyxl.load_workbook(wbkName)
							s3=wbk.worksheets[2]
							row_count_3 = s3.max_row
							s3.cell(row_count_3+1, column=1).value=row_count_1
							s3.cell(row_count_3+1, column=2).value=self.mail.email_message['Date']
							s3.cell(row_count_3+1, column=3).value=self.mail.email_message['Subject']
							s3.cell(row_count_3+1, column=4).value=self.mail.email_message['From']
							s3.cell(row_count_3+1, column=5).value='Invalid email_id'
							wbk.save(wbkName)
							'''
							subprocess.run(["python", "updation.py","2","max1","1",str(row_count_1)])
							subprocess.run(["python", "updation.py","2","max","2",str(self.mail.email_message['Date'])])
							subprocess.run(["python", "updation.py","2","max","3",str(self.mail.email_message['Subject'])])
							subprocess.run(["python", "updation.py","2","max","4",str(self.mail.email_message['From'])])
							subprocess.run(["python", "updation.py","2","max","5",str('Invalid email_id')])
						

						
						elif temp_p==0 and temp_v==1:
							'''
							wbk= self.openpyxl.load_workbook(wbkName)
							s3=wbk.worksheets[2]
							row_count_3 = s3.max_row
							s3.cell(row_count_3+1, column=1).value=row_count_1
							s3.cell(row_count_3+1, column=2).value=self.mail.email_message['Date']
							s3.cell(row_count_3+1, column=3).value=self.mail.email_message['Subject']
							s3.cell(row_count_3+1, column=4).value=self.mail.email_message['From']
							s3.cell(row_count_3+1, column=5).value='subject not known'
							wbk.save(wbkName)
							'''
							subprocess.run(["python", "updation.py","2","max1","1",str(row_count_1)])
							subprocess.run(["python", "updation.py","2","max","2",str(self.mail.email_message['Date'])])
							subprocess.run(["python", "updation.py","2","max","3",str(self.mail.email_message['Subject'])])
							subprocess.run(["python", "updation.py","2","max","4",str(self.mail.email_message['From'])])
							subprocess.run(["python", "updation.py","2","max","5",str('subject not known')])
						else:
							subprocess.run(["python", "updation.py","3","max1","1",str(row_count_1)])
							subprocess.run(["python", "updation.py","3","max","2",str(self.mail.email_message['Date'])])
							subprocess.run(["python", "updation.py","3","max","3",str(self.mail.email_message['Subject'])])
							subprocess.run(["python", "updation.py","3","max","4",str(self.mail.email_message['From'])])
							subprocess.run(["python", "updation.py","3","max","5",str('reason not known')])
							subprocess.run(["python", "updation.py","3","max","6",str(temp_v)])
							subprocess.run(["python", "updation.py","3","max","7",str(temp_p)])
					
					

				except Exception as e:
					print(self.mail.id_list[i],"can't process",e)
			
	def checkMail(self):
		self.mail.select('inbox',readonly=True)
		#type, self.mail.data = self.mail.search(None,'(since '+str(tg[-1])+')' )#+' before 5-Apr-2020'+')' )#
		self.mail.id_list_1=[]
		for i in self.mail_id_1:
                        #print(i)
                        type, self.mail.data = self.mail.search(None,'(FROM '+ i+' since '+str(sys.argv[1])+' before '+str(sys.argv[2])+')' )#+' before 5-Apr-2020'+')' )#
                        self.mail.ids = self.mail.data[0]# data is a list.
                        self.mail.id_list_1.append((self.mail.ids.split())) # ids is a space separated string

		
		self.mail.id_list = list(chain.from_iterable(self.mail.id_list_1)) 
		#type, self.mail.data = self.mail.search(None,'(since '+str(tg[-1])+')' )#+' before 5-Apr-2020'+')' )#
		if len(fg)==0 or str(self.mail.id_list[-1])>fg[-1]:
			if(len(fg)==0):
				'''
				wbk= self.openpyxl.load_workbook(wbkName)
				s1=wbk.worksheets[0]
				s1.cell(row_count_1+1, column=7).value=len(self.mail.id_list)
				wbk.save(wbkName)
				'''
				subprocess.run(["python", "updation.py","0","max","7",str(len(self.mail.id_list))])
			else:
				temp = self.re.findall(r'\d+', fg[-1])
				m_c=int(self.mail.id_list[-1].decode())-int(temp[0])
				'''
				wbk= self.openpyxl.load_workbook(wbkName)
				s1=wbk.worksheets[0]
				s1.cell(row_count_1+1, column=7).value=m_c
				wbk.save(wbkName)
				'''
				subprocess.run(["python", "updation.py","0","max","7",str(m_c)])
			print('new_mail')
			#f.write(str(self.mail.id_list[-1])+'\n')
			#f.close()
			
			self.process()
		else:
			'''
			wbk= self.openpyxl.load_workbook(wbkName)
			s1=wbk.worksheets[0]
			s1.cell(row_count_1+1, column=7).value=0
			wbk.save(wbkName)
			'''
			subprocess.run(["python", "updation.py","0","max","7",'0'])
		'''
		wbk= self.openpyxl.load_workbook(wbkName)
		s1=wbk.worksheets[0]
		s1.cell(row_count_1+1, column=8).value=self.s_r
		wbk.save(wbkName)
		'''
		subprocess.run(["python", "updation.py","0","max","8",str(self.s_r)])
			
		
email = Mail()
#while 1:
print ('Sending')
email.checkMail()
fo=open("defualt_time_read.txt", "a+")
if(str(today)!=str(tg[-1])):
	fo.write(str(today)+'\n')
now = datetime.datetime.now()
today = datetime.date.today()
today = today.strftime('%d-%b-%Y')
'''
wbk= openpyxl.load_workbook(wbkName)
s1=wbk.worksheets[0]
s1.cell(row_count_1+1, column=4).value=today
s1.cell(row_count_1+1, column=5).value=now
wbk.save(wbkName)
'''
subprocess.run(["python", "updation.py","0","max","4",str(today)])
subprocess.run(["python", "updation.py","0","max","5",str(now)])
print ('done1')
