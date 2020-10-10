import os
import sys
import struct, time
import subprocess
from datetime import date
import datetime
import openpyxl
import pdftotext
import time
import requests

from make_log import log_exceptions

now = datetime.datetime.now()


subprocess.run(["python", "updation.py","1","max1","1",sys.argv[2]])
subprocess.run(["python", "updation.py","1","max","2",sys.argv[3]])
subprocess.run(["python", "updation.py","1","max","3",sys.argv[4]])
subprocess.run(["python", "updation.py","1","max","5",str(now)])
subprocess.run(["python", "updation.py","1","max","7",sys.argv[5]])
subprocess.run(["python", "updation.py","1","max","8",sys.argv[6]])

with open(sys.argv[1], "rb") as f:
	pdf = pdftotext.PDF(f)

with open('medsave/output.txt', 'w') as f:
	f.write(" ".join(pdf))     
with open('medsave/output.txt', 'r') as myfile:
 	f = myfile.read()
try:
	hg=[]
	

	w=f.find('File No.')+len('File No.')
	k=f[w:]
	u=k.find("\n")+w 
	t=f[w:u]
	hg.append(t)

	w=f.find('Policy No.')+len('Policy No')
	k=f[w:]
	u=k.find("\n")+w 
	r=f[w:u]
	hg.append(r)

	w=f.find('Patient Name')+len('Patient Name')
	k=f[w:]
	u=k.find("\n")+w 
	n=f[w:u]
	hg.append(n)

	w=f.find('Card No. :')+ len("Card No. :") 
	k=f[w:] 
	u=k.find("\n")+w  
	hg.append(f[w:u])

	w=f.find('Claim Amount')+len('Claim Amount')
	k=f[w:]
	u=k.find("\n")+w 
	r=f[w:u]
	hg.append(r)


	# Policy No is also present in output.txt but not searched   Added by Ashish
	# w=f.find('Policy No.')+len('Policy No.')
	# k=f[w:]
	# u=k.find("\n")+w 
	# n=f[w:u]
	# hg.append(n)


	hg=[sub.replace(':','') for sub in hg]	
	hg=[sub.replace('  ','') for sub in hg]
	hg=[sub.replace('Rs.','') for sub in hg]
	#s2.cell(row_count_1+1, column=9).value='Yes'
	#s2.cell(row_count_1+1, column=10).value='NA'
	#wbk.save(wbkName)
	
	subprocess.run(["python", "updation.py","1","max","9",'Yes'])
	subprocess.run(["python", "updation.py","1","max","10",'NA'])
	
	try:
		subprocess.run(["python", "test_api.py",hg[0],hg[4],hg[1],'','Acknowledgement',sys.argv[6],sys.argv[1],'',hg[3],hg[2]])
	
		subprocess.run(["python", "updation.py","1","max","11",'Yes'])	
	except Exception as e:
		log_exceptions()
		#s2.cell(row_count_1+1, column=11).value='NO'
		subprocess.run(["python", "updation.py","1","max","11",'No'])
except Exception as e:
	log_exceptions()
	#s2.cell(row_count_1+1, column=9).value='No'
	#s2.cell(row_count_1+1, column=11).value='NO'
	subprocess.run(["python", "updation.py","1","max","9",'Yes'])
	subprocess.run(["python", "updation.py","1","max","11",'No'])
now = datetime.datetime.now()
#s2.cell(row_count_1+1, column=6).value=now
#wbk.save(wbkName)
subprocess.run(["python", "updation.py","1","max","6",str(now)])
