import openpyxl
import sys
wbkName_config = 'email ids subjects.xlsx'
wb= openpyxl.load_workbook(wbkName_config)
s_c1=wb.worksheets[2]
s_c2=wb.worksheets[1]
subj=[]
mail_id=[]
for i in range(2,34):
	subj_t=[]
	mail_t=[]
	for j in range(2,9):
		temp_subj=s_c1.cell(row=i, column=j).value
		
		if(temp_subj!=None):
			#print(temp_subj)
                        if (s_c1.cell(row=i, column=12).value==sys.argv[1]):
                                subj_t.append(temp_subj.split(','))
                        else:
                                t=temp_subj.split(',')
                                k=[]
                                for h in t:
                                        k.append('Do not need to process at all')
                                subj_t.append(k)
		else:
			subj_t.append([''])
	subj.append(subj_t)
	if (s_c1.cell(row=i, column=9).value!=None):
		mail_t=s_c1.cell(row=i, column=9).value.split(',')
	else:
		mail_t=''
	mail_id.append(mail_t)
for i in subj:
	for j in i:
		for kk,ss in enumerate(j):
			j[kk] = ss.replace("(","")
for i in subj:
	for j in i:
		for kk,ss in enumerate(j):
			j[kk] = ss.replace(")","")
print(subj)
