import os
import sys, re
import struct, time
import subprocess
from datetime import date
import datetime
import openpyxl
import pdftotext
import time
import requests

from patient_name_fun import pname_fun

now = datetime.datetime.now()

subprocess.run(["python", "updation.py","1","max1","1",sys.argv[2]])
subprocess.run(["python", "updation.py","1","max","2",sys.argv[3]])
subprocess.run(["python", "updation.py","1","max","3",sys.argv[4]])
subprocess.run(["python", "updation.py","1","max","5",str(now)])
subprocess.run(["python", "updation.py","1","max","7",sys.argv[5]])
subprocess.run(["python", "updation.py","1","max","8",sys.argv[6]])

with open(sys.argv[1], "rb") as f:
	pdf = pdftotext.PDF(f)
with open('east_west/output.txt', 'w',encoding='utf-8') as f:
	f.write(" ".join(pdf))     
with open('east_west/output.txt', 'r', encoding='utf-8') as myfile:
 	f = myfile.read()

try:		
	hg=[]

	if f.find('CCN. No.') != -1:
		x = re.search(r"CCN. No. +\S+", f)
		x1 = x.group()
		x2 = re.search(r"\d+", x1).group()
		hg.append(x2)
	else:
		hg.append('')

	status='Information Awaiting'
	hg.append('')

	#polici no
	if f.find('Policy No.') != -1:
		x = re.search(r"Policy No. +\S+", f)
		x1 = x.group()
		x2 = re.search(r"\d+", x1).group()
		hg.append(x2)
	else:
		hg.append('')

	hg.append('')



	#card no in memberid

	if f.find('Card No.') != -1:
		x = re.search(r"Card No. +\S+", f)
		x1 = x.group()
		x2 = x1.split()[-1]
		hg.append(x2)
	else:
		hg.append('')

	#name in last 

	if f.find('Name Of Beneficiary/Patient ') != -1:
		x = re.search(r"Name Of Beneficiary/Patient +\S+ +\S+", f)
		x1 = x.group()
		x2 = x1.split()[-2]+' '+x1.split()[-1]
		hg.append(x2)
	else:
		hg.append('')


	hg=[sub.replace(':','') for sub in hg]	
	hg=[sub.replace('  ','') for sub in hg]
	hg=[sub.replace('Rs.','') for sub in hg]

	
	subprocess.run(["python", "updation.py","1","max","9",'Yes'])
	subprocess.run(["python", "updation.py","1","max","10",'NA'])
	regex_list = [r"(?<=Name Of Beneficiary/Patient).*"]
	pname = pname_fun(f, regex_list)
	
	try:
		subprocess.run(["python", "test_api.py",hg[0],hg[1],' ',' ',status,sys.argv[6],sys.argv[1],' ',hg[2],pname])
		'''wbk= openpyxl.load_workbook(wbkName)
		s2=wbk.worksheets[1]
		s2.cell(row_count_1+1, column=11).value='YES'
		'''
		subprocess.run(["python", "updation.py","1","max","11",'Yes'])	
	except Exception as e:
		#s2.cell(row_count_1+1, column=11).value='NO'
		subprocess.run(["python", "updation.py","1","max","11",'No'])
except Exception as e:
	#s2.cell(row_count_1+1, column=9).value='No'
	#s2.cell(row_count_1+1, column=11).value='NO'
	subprocess.run(["python", "updation.py","1","max","9",'Yes'])
	subprocess.run(["python", "updation.py","1","max","11",'No'])
now = datetime.datetime.now()
#s2.cell(row_count_1+1, column=6).value=now
#wbk.save(wbkName)
subprocess.run(["python", "updation.py","1","max","6",str(now)])
