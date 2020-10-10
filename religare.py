from email.mime.text import MIMEText
import smtplib
import time
import imaplib
import email
import camelot
import PyPDF2
import sys
import xlsxwriter
from xlrd import open_workbook
from openpyxl import load_workbook
# from tabula import read_pdf
import xlwt
import json
import os
import os.path
import xlrd
from os import listdir
from os import path
from os.path import isfile, join
import pdftotext
import tabula, re
import pandas as pd
# from tabula import read_pdf
from decode_error import check_subject, read_from_delete
from make_log import log_exceptions
import subprocess

# sys.argv = ['religare.py', 'Tpappg@maxhealthcare.com', 'Sept@2020', '16-Sep-2020', '16-Sep-2020', 'outlook.office365.com', 'Max', '38522', 'Payment against Claim Reference Number:91237660-02 Policy No :\r\n 10953321  Proposer Name :VINOD KUMAR JAIN Patient Name :REENA JAIN']
try:
    subprocess.run(["python", "updation.py", "1", "max", "9", "X"])
    if path.exists(r'religare/religare' + str(sys.argv[6]) + '.xlsx'):
        os.remove(r'religare/religare' + str(sys.argv[6]) + '.xlsx')
    import openpyxl

    wq = 0
    eu = []
    fg = []
    repeat = []
    from openpyxl.styles import Color, PatternFill, Font, Border

    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')


    def read_email_from_gmail():
        SMTP_SERVER = str(sys.argv[5])
        mail = imaplib.IMAP4_SSL(SMTP_SERVER)
        e_id = str(sys.argv[1])
        pswd = str(sys.argv[2])
        srt = str(sys.argv[3])
        stp = str(sys.argv[4])
        mail.login(user=e_id, password=pswd)
        mail.select("inbox", readonly=True)
        ###############################################<
        mail_uid = str(sys.argv[7])
        if mail_uid == -1:
            type, data = mail.search(None,
                                     '(SUBJECT "Payment against Claim Reference Number:" since ' + srt + ' before ' + stp + ')')
            ids = data[0]
            id_list = ids.split()
        else:
            ids = mail_uid  # data is a list.
            # accept id from outside and put in id_list akshay var name = id

            id_list = []  # ids is a space separated string
            id_list.append(ids)
        ###############################################>
        # type, data = mail.search(None,
        #                          '(SUBJECT "Payment against Claim Reference Number:" since ' + srt + ' before ' + stp + ')')
        # ids = data[0]  # data is a list.
        # id_list = ids.split()  # ids is a space separated string
        # # print(id_list)
        for i in range(0, len(id_list)):
            latest_email_id = id_list[i]  # get the latest
            result, data = mail.fetch(latest_email_id,
                                      "(RFC822)")  # fetch the email body (RFC822)             for the given ID

            ##################################################ak
            try:
                # raw_email = data[0][1].decode('utf-8')
                try:
                    raw_email = data[0][1].decode('utf-8')
                except UnicodeDecodeError:
                    try:
                        raw_email = data[0][1].decode('ISO-8859-1')
                    except UnicodeDecodeError:
                        try:
                            raw_email = data[0][1].decode('ascii')
                        except UnicodeDecodeError:
                            pass
                email_message = email.message_from_string(raw_email)
                subject = email_message['Subject']
                result, sys.argv[8] = check_subject(subject, sys.argv[8], mail)
                if result == 'Changed':
                    # raise Exception('subject not matched')
                    raise Exception('subject not matched', )
            except:
                try:
                    log_exceptions(syssubject=sys.argv[8], subject=subject, error='subject not matched')
                except:
                    pass
                if result != 'OK':
                    data = {'server': SMTP_SERVER,
                            'hospmail': e_id,
                            'pass': pswd,
                            'subject': sys.argv[8]}
                    try:
                        data = read_from_delete(data)
                        if data == None:
                            raise Exception("Not found")
                    except:
                        log_exceptions(msg='not found in deleted', subject=sys.argv[8])
            ##################################################akend

            raw_email = data[0][1].decode('utf-8')
            email_message = email.message_from_string(raw_email)
            if email_message['Subject'] not in fg:
                for part in email_message.walk():

                    if part.get_content_maintype() == 'multipart':
                        # print part.as_string()
                        continue
                    if part.get('Content-Disposition') is None:
                        # print part.as_string()
                        continue
                    fileName = str(id_list[i]) + '.pdf'  # part.get_filename()
                    detach_dir = (os.getcwd() + '/religare/attachments_' + str(sys.argv[6]) + '/')
                    if bool(fileName):
                        # print (email_message['Subject'])
                        filePath = os.path.join(detach_dir, fileName)
                        if not os.path.isfile(filePath):
                            from reportlab.pdfgen import canvas
                            c = canvas.Canvas(fileName)
                            # print (fileName)
                            fp = open(filePath, 'wb')
                            fp.write(part.get_payload(decode=True))
                            fp.close()
            else:
                repeat.append(email_message['Subject'])
            fg.append(email_message['Subject'])


    mypath = os.getcwd() + '/religare'
    if not path.exists(mypath):
        os.mkdir(mypath)
    if not path.exists(mypath + '/attachments_' + str(sys.argv[6])):
        os.mkdir(mypath + '/attachments_' + str(sys.argv[6]))
    mypath = os.getcwd() + '/religare/attachments_' + str(sys.argv[6]) + '/'
    for filename in os.listdir(mypath):
        file_path = os.path.join(mypath, filename)
        if os.path.isfile(file_path) or os.path.islink(file_path):
            os.unlink(file_path)
    read_email_from_gmail()

    wbkName = 'religare/religare' + str(sys.argv[6]) + '.xlsx'
    wbk = openpyxl.Workbook()
    wbk.create_sheet('Sheet1')
    s1 = wbk.worksheets[0]
    s2 = wbk.worksheets[1]

    onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
    for t in range(0, len(onlyfiles)):
        try:
            sh1 = ['S.no', 'Claim No.', 'Policy No.', 'Bank Name', 'Proposer Name', 'Name of Patient',
                   'Instrument/ NEFT No', 'Instrument/ NEFT Date', 'Date of admission', 'Date of Discharge', 'bill amt',
                   'net amt', 'co-pay', 'deductible', 'Employee ID', 'Employee Name', 'hospital discount', 'Al no.',
                   'total deduction']
            sh2 = ['S.no', 'Claim No.', 'Description', 'Rate per day/Quantity', 'No.of Day/Visits/Quantity', 'Bill Amount',
                   'Admissible Amount', 'Deducted Amount', 'Deduction Reason']
            for i in range(0, len(sh2)):
                s2.cell(row=1, column=i + 1).value = sh2[i]
            for i in range(0, len(sh1)):
                s1.cell(row=1, column=i + 1).value = sh1[i]

            print(onlyfiles[t])
            with open(mypath + onlyfiles[t], "rb") as f:
                pdf = pdftotext.PDF(f)

            with open('religare/output.txt', 'w') as f:
                f.write(" ".join(pdf))
            with open('religare/output.txt', 'r') as myfile:
                f = myfile.read()

            hg = []
            w = f.find('Policy No.') + 10
            g = f[w:]
            u = g.find('\n') + w
            hg.append(f[w:u])

            w1 = f.find('Bank Name') + 10
            g = f[w1:]
            u1 = g.find('Successful') + w1
            hg.append(f[w1:u1])

            w2 = f.find('Proposer Name') + 14
            g = f[w2:]
            u2 = g.find('Policy No.') + w2
            u3 = g.find('\n') + w2
            u4 = g.find('Claimed') + w2
            hg.append(f[w2:u2] + f[u3:u4])

            w2 = f.find('Name of Patient') + 16
            g = f[w2:]
            u2 = g.find('\n') + w2
            u3 = g.find('Date of admission') + w2
            hg.append(f[w2:u2] + f[u2 + 60:u3])

            w3 = f.find('Instrument/ NEFT No') + 19
            g = f[w3:]
            u3 = g.find('\n') + w3
            hg.append(f[w3:u3])

            w4 = g.find('Instrument/ NEFT') + w3 + 17
            g = f[w4:]
            u4 = g.find('\n') + w4
            hg.append(f[w4:u4])

            w5 = f.find('Date of admission') + 18
            g = f[w5:]
            u5 = g.find('Date of Discharge') + w5
            hg.append(f[w5:u5])

            w6 = f.find('Date of Discharge') + 18
            g = f[w6:]
            u6 = g.find('\n') + w6
            hg.append(f[w6:u6])

            w6 = f.find('Bill Amount') + 18
            g = f[w6:]
            u6 = g.find('Instrument') + w6
            hg.append(f[w6:u6])

            w6 = f.find('Amount Paid') + 18
            g = f[w6:]
            u6 = g.find('Bank') + w6
            hg.append(f[w6:u6])

            if f.find('Co pay') != -1:
                w6 = f.find('Co pay') + 18
                g = f[w6:]
                u6 = g.find('Deductible') + w6
                hg.append(f[w6:u6])
            else:
                hg.append(' ')
            w6 = f.find('Deductible') + 18
            g = f[w6:]
            u6 = g.find('\n') + w6
            hg.append(f[w6:u6])

            w2 = f.find('Employee ID') + 14
            g = f[w2:]
            u3 = g.find('Employee Name') + w2
            hg.append(f[w2:u3])

            w2 = f.find('Employee Name') + 14
            g = f[w2:]
            u3 = g.find('Name of Proposer') + w2
            hg.append(f[w2:u3])

            w2 = f.find('Hospital Discount') + 25
            g = f[w2:]
            u3 = g.find('AL approved') + w2
            hg.append(f[w2:u3])

            w9 = f.find('AL No.') + 6
            g = f[w9:]
            u9 = g.find('\n') + w9 + 3
            ccn = (f[w9:u9])

            w9 = f.find('Claim No') + 15
            g = f[w9:]
            u9 = g.find('-') + w9 + 3
            # hg.append(f[w9:u9])
            regex = r"(?<=Claim No)[ \d]+"
            temp = re.compile(regex).search(f)
            if temp is not None:
                cno = temp.group().strip()
            else:
                cno = ""
            hg.append(cno)
            ccn = ccn.replace('\n', '')
            ccn = ccn.replace('.', '')
            ccn = ccn.replace(' ', '')
            if ccn == None:
                ccn = hg[-1]
            hg = [sub.replace('  ', '') for sub in hg]
            hg = [sub.replace(':', '') for sub in hg]
            hg = [sub.replace('\n', ' ') for sub in hg]
            s1.cell(row=t + 2, column=1).value = t + 1
            s1.cell(row=t + 2, column=2).value = ccn
            for i in range(0, len(hg)):
                s1.cell(row=t + 2, column=i + 3).value = hg[i]
            # df=tabula.read_pdf(mypath+onlyfiles[t],pages=2,multiple_tables=True,line_space=40)
            tabula.convert_into(mypath + onlyfiles[t], 'religare/out.xls', output_format='excel', pages=2)
            tabula.convert_into(mypath + onlyfiles[t], 'religare/out' + str(t) + '.json', output_format='json', pages='all')

            with open('religare/out' + str(t) + '.json') as f:
                data = json.load(f)
            k = 0
            l = 0
            u = 0
            w = 0
            mo = 0
            me = 0
            we = 0
            yt = 0
            p = []
            r = []
            ro = []
            po = []
            e = []
            eo = []
            kl = []
            lk = []
            for x in range(2, len(data)):
                d = data[x]["data"]
                # print(d)
                for i in d:
                    m = [0, 0, 0, 0, 0, 0, 0, 0]
                    for j in i:
                        for x, y in j.items():
                            h = str(y)
                            if (h.find('62.') != -1 and x == 'left'):
                                k = 1
                            if k == 1 and x == 'text':
                                k = 0
                                p.append(y)
                                m[0] = 1

                            if h.find('92.') != -1 and x == 'left':
                                u = 1
                            if u == 1 and x == 'text':
                                u = 0
                                r.append(y)
                                m[1] = 1

                            if h.find('152.') != -1 and x == 'left':
                                w = 1
                            if w == 1 and x == 'text':
                                w = 0
                                po.append(y)
                                m[2] = 1

                            if h.find('222.') != -1 and x == 'left':
                                l = 1
                            if l == 1 and x == 'text':
                                l = 0
                                ro.append(y)
                                m[3] = 1

                            if h.find('282.') != -1 and x == 'left':
                                mo = 1
                            if mo == 1 and x == 'text':
                                mo = 0
                                e.append(y)
                                m[4] = 1

                            if h.find('332.') != -1 and x == 'left':
                                me = 1
                            if me == 1 and x == 'text':
                                me = 0
                                eo.append(y)
                                m[5] = 1

                            if h.find('382.') != -1 and x == 'left':
                                we = 1
                            if we == 1 and x == 'text':
                                we = 0
                                kl.append(y)
                                m[6] = 1

                            if h.find('422.') != -1 and x == 'left':
                                yt = 1
                            if yt == 1 and x == 'text':
                                yt = 0
                                lk.append(y)
                                m[7] = 1

                    if m[0] == 0:
                        p.append('')
                    if m[1] == 0:
                        r.append(r[-1])
                    if m[2] == 0:
                        po.append('')
                    if m[3] == 0:
                        ro.append('')
                    if m[4] == 0:
                        e.append('')
                    if m[5] == 0:
                        eo.append('')
                    if m[6] == 0:
                        kl.append('')
                    if m[7] == 0:
                        lk.append('')
                # print(i)

            p = [sub.replace('\r', ' ') for sub in p]
            r = [sub.replace('\r', ' ') for sub in r]
            po = [sub.replace('\r', ' ') for sub in po]
            ro = [sub.replace('\r', ' ') for sub in ro]
            e = [sub.replace('\r', ' ') for sub in e]
            eo = [sub.replace('\r', ' ') for sub in eo]
            kl = [sub.replace('\r', ' ') for sub in kl]
            lk = [sub.replace('\r', ' ') for sub in lk]

            for i in range(0, len(p) - 1):
                if i == 0:
                    continue
                if r[i] == 'Description':
                    continue
                row_num = s2.max_row
                wq = wq + 1
                s2.cell(row=row_num + 1, column=1).value = wq
                s2.cell(row=row_num + 1, column=2).value = ccn
                s2.cell(row=row_num + 1, column=3).value = r[i]
                s2.cell(row=row_num + 1, column=4).value = ro[i]
                s2.cell(row=row_num + 1, column=5).value = po[i]
                s2.cell(row=row_num + 1, column=6).value = e[i]
                s2.cell(row=row_num + 1, column=7).value = eo[i]
                s2.cell(row=row_num + 1, column=8).value = kl[i]
                s2.cell(row=row_num + 1, column=9).value = lk[i]
            ded = kl[-1]
            s1.cell(row=t + 2, column=19).value = ded
            ccn = ccn.replace(' ', '')
            if ccn == ' ':
                ccn = cno
            os.rename(os.getcwd() + '/religare/attachments_' + str(sys.argv[6]) + '/' + onlyfiles[t],
                      os.getcwd() + '/religare/attachments_' + str(sys.argv[6]) + '/' + ccn + '.pdf')
        except Exception as e:
            if ccn == ' ':
                ccn = cno
            s1.cell(row=t + 2, column=1).value = 'error'
            os.rename(os.getcwd() + '/religare/attachments_' + str(sys.argv[6]) + '/' + onlyfiles[t],
                      os.getcwd() + '/religare/attachments_' + str(sys.argv[6]) + '/' + cno + '.pdf')

    for t in range(0, len(onlyfiles)):
        if t in eu:
            s1.cell(row=t + 2, column=1).fill = redFill
            s1.cell(row=t + 2, column=1).value = 'error'
    # print(po)
    print("Done")
    wbk.save(wbkName)
    wbk.close()
    wbkName = 'count/count.xlsx'
    wbk = openpyxl.load_workbook(wbkName)
    s1 = wbk.worksheets[0]

    row_ = s1.max_row + 1
    s1.cell(row=row_, column=1).value = 'religare'
    s1.cell(row=row_, column=2).value = str(sys.argv[6])
    s1.cell(row=row_, column=3).value = len(fg)
    s1.cell(row=row_, column=4).value = len(onlyfiles)
    s1.cell(row=row_, column=5).value = len(repeat)
    wbk.save(wbkName)
    subprocess.run(["python", "updation.py", "1", "max", "9", " "])
except:
	log_exceptions()
