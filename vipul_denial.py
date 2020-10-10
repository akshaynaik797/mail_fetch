import os
import sys
import struct, time
import subprocess
from datetime import date
import datetime
import openpyxl
import pdftotext
import time
import requests
now = datetime.datetime.now()

subprocess.run(["python", "updation.py","1","max1","1",sys.argv[2]])
subprocess.run(["python", "updation.py","1","max","2",sys.argv[3]])
subprocess.run(["python", "updation.py","1","max","3",sys.argv[4]])
subprocess.run(["python", "updation.py","1","max","5",str(now)])
subprocess.run(["python", "updation.py","1","max","7",sys.argv[5]])
subprocess.run(["python", "updation.py","1","max","8",sys.argv[6]])

with open(sys.argv[1], "rb") as f:
  pdf = pdftotext.PDF(f)

with open('vipul/output.txt', 'w') as f:
  f.write(" ".join(pdf))      
with open ('vipul/output.txt' , 'r' ) as myfile :
  f = myfile.read()


try:
    hg=[]
    if f.find('Authorisation No :') != -1:
      w=f.find('Authorisation No :')+len('Authorisation No :')
      k=f[w:]
      u=k.find("File No :")+w
      g=f[w:u]
      g=g.replace(' ','')
      mlist=g.split("\n\n")
      hg.append(mlist[2])
    else:
      print("Authorisation No : is not in output1.text")

    if f.find('File No :') != -1:
      w=f.find('File No :')+len('File No :')
      k=f[w:]
      u=k.find("Policy No :")+w
      g=f[w:u]
      # g=g.replace(' ','')
      mlist=g.split('\n\n')
      g=mlist[1]
      g=g.split()
      hg.append(g[-1])
    else:
      print("Policy No : not found in output1.text")

    if f.find('Policy No :') != -1:
      w=f.find('Policy No :')+len('Policy No :')
      k=f[w:]
      u=k.find("WITHOUT PREJUDICE")+w
      g=f[w:u]
      g=g.replace(' ','')
      mlist=g.split("\n\n")
      hg.append(mlist[1])
    else:
      print("Policy No : not found in output1.text")

    if f.find('Facility for') != -1:
      w=f.find('Facility for')+len('Facility for')
      k=f[w:]
      u=k.find("\n")+w
      g=f[w:u]
      hg.append(g)
    else:
      print("Facility for not found in output1.text")
    hg=[sub.replace(':','') for sub in hg]
    hg=[sub.replace('  ','') for sub in hg]
    hg=[sub.replace('Rs.','') for sub in hg]
    print(hg)

    subprocess.run(["python", "updation.py","1","max","9",'Yes'])
    subprocess.run(["python", "updation.py","1","max","10",'NA'])
    
    try:
     subprocess.run(["python", "test_api.py",hg[0],'',hg[2],'','Denial',sys.argv[6],sys.argv[1],'',hg[1],hg[3]])
     '''wbk= openpyxl.load_workbook(wbkName)
        s2=wbk.worksheets[1]
        s2.cell(row_count_1+1, column=11).value='YES'
        '''
     subprocess.run(["python", "updation.py","1","max","11",'Yes'])    
    except Exception as e:
      #s2.cell(row_count_1+1, column=11).value='NO'
      subprocess.run(["python", "updation.py","1","max","11",'No'])
except Exception as e:
  subprocess.run(["python", "updation.py","1","max","9",'Yes'])
  subprocess.run(["python", "updation.py","1","max","11",'No'])
now = datetime.datetime.now()
#s2.cell(row_count_1+1, column=6).value=now
#wbk.save(wbkName)
subprocess.run(["python", "updation.py","1","max","6",str(now)])
