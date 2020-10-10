from email.mime.text import MIMEText
import smtplib
import time
import imaplib
import sys
import email
import camelot
import PyPDF2
import xlsxwriter
from xlrd import open_workbook
import xlwt
import os
import os.path
import xlrd
from os import listdir
from os import path
from os.path import isfile, join

repeat = []


def read_email_from_gmail():
    SMTP_SERVER = 'imap.gmail.com'#str(sys.argv[5])
    mail = imaplib.IMAP4_SSL(SMTP_SERVER)
    e_id = 'mediclaim@inamdarhospital.org' #str(sys.argv[1])
    pswd = 'Mediclaim@2019'#str(sys.argv[2])
    srt = ''#str(sys.argv[3])
    stp = ''#str(sys.argv[4])
    mail.login(user=e_id, password=pswd)
    mail.select("inbox", readonly=True)
    ###############################################<
    mail_uid = '23130' #str(sys.argv[7])
    if mail_uid == -1:
        type, data = mail.search(None, '(SUBJECT "STAR HEALTH AND ALLIED INSUR04239 - 00040350005154" since ' + srt + ' before ' + stp + ')')
        ids = data[0]
        id_list = ids.split()
    else:
        ids = mail_uid  # data is a list.
        # accept id from outside and put in id_list akshay var name = id

        id_list = []  # ids is a space separated string
        id_list.append(ids)
    ###############################################>
    for i in range(0, len(id_list)):
        latest_email_id = id_list[i]  # get the latest
        result, data = mail.fetch(latest_email_id,
                                  "(RFC822)")  # fetch the email body (RFC822)             for the given ID

        raw_email = data[0][1].decode('utf-8')
        email_message = email.message_from_string(raw_email)

        for part in email_message.walk():
            if part.get_content_maintype() == 'multipart':
                # print part.as_string()
                continue
            if part.get('Content-Disposition') is None:
                # print part.as_string()
                continue
            fileName = part.get_filename()
            detach_dir = (os.getcwd() + '/star/attachments_1_' + str('inamdar'))
            if bool(fileName):
                filePath = os.path.join(detach_dir, fileName)
                if not os.path.isfile(filePath):
                    from reportlab.pdfgen import canvas
                    c = canvas.Canvas(fileName + '.pdf')
                    print(fileName)
                    fp = open(filePath, 'wb')
                    fp.write(part.get_payload(decode=True))
                    fp.close()


mypath = os.getcwd() + '/star'
if not path.exists(mypath):
    os.mkdir(mypath)
if not path.exists(mypath + '/attachments_1_' + str('inamdar')):
    os.mkdir(mypath + '/attachments_1_' + str('inamdar'))
mypath = os.getcwd() + '/star/attachments_1_' + str('inamdar') + '/'
for filename in os.listdir(mypath):
    file_path = os.path.join(mypath, filename)
    if os.path.isfile(file_path) or os.path.islink(file_path):
        os.unlink(file_path)
read_email_from_gmail()

if path.exists(r'star/star_small' + str('inamdar') + '.xlsx'):
    os.remove(r'star/star_small' + str('inamdar') + '.xlsx')
import openpyxl

wbkName = 'star/star_small' + str('inamdar') + '.xlsx'
wbk = openpyxl.Workbook()

onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
for t in range(0, len(onlyfiles)):
    try:
        s1 = wbk.worksheets[0]
        sh2 = ['Chq /DD/Ft No', 'Amount', 'IFSC Code', 'Credit A/c No.', 'transaction date']
        for i in range(0, len(sh2)):
            s1.cell(row=1, column=i + 3).value = sh2[i]
        pdfFileObj = open(mypath + onlyfiles[t], 'rb')
        pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
        
        pageObj = pdfReader.getPage(0)
        f = pageObj.extractText()
        f = f.replace('\n', '$$ ')
        # print(f)
        text_file = open('star/mail1.txt', "w")
        n = text_file.write(f)
        text_file.close()
        pdfFileObj.close()

        gh = []
        x1 = f.find('INTIMATION NO') + 14
        g = f[x1:]
        y1 = g.find('_') + x1
        cli = (f[x1:y1])
        cli = cli.replace(' ', '')
        x2 = f.find('Chq /DD/Ft No') + 13
        g = f[x2:]
        x3 = g.find(':') + x2 + 1
        y2 = g.find('Value') + x2
        gh.append(f[x3:y2])

        x3 = f.find('Amount') + 7
        g = f[x3:]
        x4 = g.find(':') + x3 + 1
        y3 = g.find('Amount') + x3
        gh.append(f[x4:y3])

        x4 = f.find('IFSC Code') + 10
        g = f[x4:]
        y4 = g.find('through') + x4
        gh.append(f[x4:y4])

        x5 = f.find('Account Number') + 14
        g = f[x5:]
        y5 = g.find('with') + x5
        gh.append(f[x5:y5])

        x5 = f.find('Value Date :') + 12
        g = f[x5:]
        y5 = g.find('Amount') + x5
        gh.append(f[x5:y5])

        gh = [sub.replace('  ', '') for sub in gh]
        s1.cell(row=1, column=1).value = 'Sr. No.'
        s1.cell(row=1, column=2).value = 'INTIMATION NO'
        for i in range(0, len(gh)):
            s1.cell(row=t + 2, column=1).value = t + 1
            s1.cell(row=t + 2, column=2).value = cli
            s1.cell(row=t + 2, column=i + 3).value = gh[i]
        m = cli
        m = m.replace('/', '')
        os.rename(os.getcwd() + '/star/attachments_1_' + str('inamdar') + '/' + onlyfiles[t],
                  os.getcwd() + '/star/attachments_1_' + str('inamdar') + '/' + m + '(1).pdf')
    except Exception as e:
        os.close(os.getcwd() + '/star/attachments_1_' + str('inamdar') + '/' + onlyfiles[t])
        s1.cell(row=t + 2, column=1).value = 'error'
        os.rename(os.getcwd() + '/star/attachments_1_' + str('inamdar') + '/' + onlyfiles[t],
                  os.getcwd() + '/star/attachments_1_' + str('inamdar') + '/' + m + '(1).pdf')

print("Done")
wbk.save(wbkName)
wbk.close
wbkName = 'count/count.xlsx'
wbk = openpyxl.load_workbook(wbkName)
s1 = wbk.worksheets[1]
row_ = s1.max_row + 1
s1.cell(row=row_, column=1).value = 'star_small'
s1.cell(row=row_, column=2).value = str('inamdar')
s1.cell(row=row_, column=3).value = len(fg)
s1.cell(row=row_, column=4).value = len(onlyfiles)
s1.cell(row=row_, column=5).value = len(repeat)
wbk.save(wbkName)
wbk.close
