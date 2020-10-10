import os
import re
import sys
import struct, time
import subprocess
from datetime import date
import datetime
import openpyxl
import pdftotext
import time
import requests

import camelot

from patient_name_fun import pname_fun

now = datetime.datetime.now()

subprocess.run(["python", "updation.py","1","max1","1",sys.argv[2]])
subprocess.run(["python", "updation.py","1","max","2",sys.argv[3]])
subprocess.run(["python", "updation.py","1","max","3",sys.argv[4]])
subprocess.run(["python", "updation.py","1","max","5",str(now)])
subprocess.run(["python", "updation.py","1","max","7",sys.argv[5]])
subprocess.run(["python", "updation.py","1","max","8",sys.argv[6]])
'''
tables = camelot.read_pdf(sys.argv[1])#,line_scale=10)
tables.export('Max_Bupa/foo1.xlsx', f='excel')
loc = ("Max_Bupa/foo1.xlsx")
wb= openpyxl.load_workbook(loc)
s1_t=wb.worksheets[0]
s2_t=wb.worksheets[1]
row_count_t = s2_t.max_row
'''
with open(sys.argv[1], "rb") as f:
	pdf = pdftotext.PDF(f)

with open('Max_Bupa/output.txt', 'w') as f:
	f.write(" ".join(pdf))     
with open('Max_Bupa/output.txt', 'r') as myfile:
 	f = myfile.read()

try:		
	hg=[]
	if f.find('Cashless Claim Number:') != -1:
		w=f.find('Cashless Claim Number:')+22
		g=f[w:]
		u=g.find('\n')+w
		hg.append(f[w:u])
	else:
		x = re.search(r"number +\d+", f)
		x1 = x.group()
		x2 = re.search(r"\d+", x1)
		bearing_no = x2.group()
		hg.append(bearing_no)


	status='Approved'
	if f.find('Total Authorised Amount (INR)') != -1:
		w=f.find('Total Authorised Amount (INR)')+29
		g=f[w:]
		u=g.find('\n')+w
		hg.append(f[w:u])
	else:
		x = re.search(r"approved for Rs.\d+", f)
		x1 = x.group()
		x2 = re.search(r"\d+", x1)
		aprvamt = x2.group()
		hg.append(aprvamt)

	if f.find('Policy No.:') != -1:
		w=f.find('Policy No.:')+11
		g=f[w:]
		u=g.find('Expected')+w
		hg.append(f[w:u])
	else:
		hg.append('')

	if f.find('Id of the Patient') != -1:
		w=f.find('Id of the Patient')+17
		g=f[w:]
		u=g.find('\n')+w
		hg.append(f[w:u])
	else:
		hg.append('')

	# w = f.find('Cashless Claim Number:') + 22
	# g = f[w:]
	# u = g.find('\n') + w
	# hg.append(f[w:u])
	#
	# status = 'Approved'
	# w = f.find('Total Authorised Amount (INR)') + 29
	# g = f[w:]
	# u = g.find('\n') + w
	# hg.append(f[w:u])
	#
	# w = f.find('Policy No.:') + 11
	# g = f[w:]
	# u = g.find('Expected') + w
	# hg.append(f[w:u])
	#
	# w = f.find('Id of the Patient') + 17
	# g = f[w:]
	# u = g.find('\n') + w
	# hg.append(f[w:u])

	hg=[sub.replace(':','') for sub in hg]
	hg=[sub.replace('  ','') for sub in hg]
	hg=[sub.replace('(INR)','') for sub in hg]
	hg = [sub.replace(',', '') for sub in hg]

	#s2.cell(row_count_1+1, column=9).value='Yes'
	#s2.cell(row_count_1+1, column=10).value='NA'
	#wbk.save(wbkName)
	
	subprocess.run(["python", "updation.py","1","max","9",'Yes'])
	subprocess.run(["python", "updation.py","1","max","10",'NA'])
	
	try:
		regex_list = [r"(?<=respect of).*(?=we are)"]
		pname = pname_fun(f, regex_list)
		subprocess.run(["python", "test_api.py",hg[0],hg[1],hg[2],'',status,sys.argv[6],sys.argv[1],'',hg[3], pname])
		'''wbk= openpyxl.load_workbook(wbkName)
		s2=wbk.worksheets[1]
		s2.cell(row_count_1+1, column=11).value='YES'
		'''
		subprocess.run(["python", "updation.py","1","max","11",'Yes'])	
	except Exception as e:
		print(e)
		#s2.cell(row_count_1+1, column=11).value='NO'
		subprocess.run(["python", "updation.py","1","max","11",'No'])
except Exception as e:
	print(e)
	#s2.cell(row_count_1+1, column=9).value='No'
	#s2.cell(row_count_1+1, column=11).value='NO'
	subprocess.run(["python", "updation.py","1","max","9",'Yes'])
	subprocess.run(["python", "updation.py","1","max","11",'No'])
now = datetime.datetime.now()
#s2.cell(row_count_1+1, column=6).value=now
#wbk.save(wbkName)
subprocess.run(["python", "updation.py","1","max","6",str(now)])
