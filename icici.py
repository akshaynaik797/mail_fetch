import pandas as pd
from email.mime.text import MIMEText
import smtplib
import time
import imaplib
import email
import camelot
import PyPDF2
import sys
import csv
import xlsxwriter
from xlrd import open_workbook
import xlwt
import os
import glob
import os.path
import xlrd
from os import listdir
from os import path
from os.path import isfile, join
from html.parser import HTMLParser
import pdfkit
import pandas as pd
import html2text
import pdftotext
import openpyxl

from decode_error import check_subject, read_from_delete
from make_log import log_exceptions
import subprocess
try:

    fg = []
    subprocess.run(["python", "updation.py", "1", "max", "9", "X"])
    eu = []
    repeat = []
    path_wkhtmltopdf = r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe'
    config = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf)
    from openpyxl.styles import Color, PatternFill, Font, Border

    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')
    if path.exists(r'icici_lombard/icici_lombard' + str(sys.argv[6]) + '.xlsx'):
        os.remove(r'icici_lombard/icici_lombard' + str(sys.argv[6]) + '.xlsx')


    def read_email_from_gmail():
        SMTP_SERVER = str(sys.argv[5])
        mail = imaplib.IMAP4_SSL(SMTP_SERVER)
        # mail.login(user = 'Mediclaim@therisingmedicare.com', password = 'cef@2018')

        e_id = str(sys.argv[1])
        pswd = str(sys.argv[2])
        srt = str(sys.argv[3])
        stp = str(sys.argv[4])
        mail.login(user=e_id, password=pswd)
        mail.select("inbox", readonly=True)
        ###############################################<
        mail_uid = str(sys.argv[7])
        if mail_uid == -1:
            type, data = mail.search(None, '(SUBJECT "payment processed for claim no" since ' + srt + ' before ' + stp + ')')
            ids = data[0]
            id_list = ids.split()
        else:
            ids = mail_uid  # data is a list.
            # accept id from outside and put in id_list akshay var name = id

            id_list = []  # ids is a space separated string
            id_list.append(ids)
        ###############################################>
        # type, data = mail.search(None, '(SUBJECT "payment processed for claim no" since ' + srt + ' before ' + stp + ')')
        # ids = data[0]  # data is a list.
        # id_list = ids.split()  # ids is a space separated string
        # # print(id_list)
        # b = 0
        for i in range(0, len(id_list)):
            latest_email_id = id_list[i]  # get the latest
            result, data = mail.fetch(latest_email_id, "(RFC822)")

            ##################################################ak
            try:
                subject = ''
                # raw_email = data[0][1].decode('utf-8')
                try:
                    raw_email = data[0][1].decode('utf-8')
                except UnicodeDecodeError:
                    try:
                        raw_email = data[0][1].decode('ISO-8859-1')
                    except UnicodeDecodeError:
                        try:
                            raw_email = data[0][1].decode('ascii')
                        except UnicodeDecodeError:
                            pass
                email_message = email.message_from_string(raw_email)
                subject = email_message['Subject']
                result, sys.argv[8] = check_subject(subject, sys.argv[8], mail)
                if result == 'Changed':
                    # raise Exception('subject not matched')
                    raise Exception('subject not matched', )
            except:
                try:
                    log_exceptions(syssubject=sys.argv[8], subject=subject, error='subject not matched')
                except:
                    pass
                if result != 'OK':
                    data = {'server': SMTP_SERVER,
                            'hospmail': e_id,
                            'pass': pswd,
                            'subject': sys.argv[8]}
                    try:
                        data = read_from_delete(data)
                        if data == None:
                            raise Exception("Not found")
                    except:
                        log_exceptions(msg='not found in deleted', subject=sys.argv[8])
            ##################################################akend
            # print(data)
            raw_email = data[0][1].decode('utf-8')

            email_message = email.message_from_string(raw_email)
            # if path.exists(r'/home/shivam/Desktop/vnu_scripts/Paramount/email.html'):
            # os.remove(r'email.html')
            # Body details
            if email_message['Subject'] not in fg:

                b = 1
                for part in email_message.walk():
                    # print(part.get_content_type())

                    if part.get_content_type() == "text/html":
                        # print('hi')
                        body = part.get_payload(decode=True)
                        # body=body[:6740]
                        file_name = "icici_lombard/email.html"
                        output_file = open(file_name, 'w')
                        try:
                            output_file.write("Body: %s" % (body.decode('utf-8')))
                        except Exception as e:
                            pof = e.isdigit()
                            print(e, pof)
                            body1 = body[:pof]
                            body2 = body[pof + 1:]
                        # output_file.write("Body: %s %s" %(body1.decode('utf-8'),%(body2.decode('utf-8')))
                        output_file.close()
                        try:
                            pdfkit.from_file('icici_lombard/email.html',
                                         'icici_lombard/attachments_pdf_' + str(sys.argv[6]) + '/' + str(b) + '.pdf',configuration=config)
                        except Exception as e:
                            pass

                    else:
                        continue
                for t, df in enumerate(pd.read_html(r'icici_lombard/email.html')):
                    writer = pd.ExcelWriter('icici_lombard/attachments_' + str(sys.argv[6]) + '/' + '%s' % (b) + '.xlsx',
                                            engine='xlsxwriter')
                    if (t == 5):
                        df_5 = df
                    if (t == 6):
                        df_6 = df
                    if (t == 14):
                        df_14 = df
                    if (t == 16):
                        df_16 = df

                df_17 = df
                df_5.to_excel(writer, '1')
                df_6.to_excel(writer, '2')
                df_14.to_excel(writer, '3')
                df_16.to_excel(writer, '4')
                df_17.to_excel(writer, '5')
                writer.save()

            else:
                repeat.append(email_message['Subject'])
            fg.append(email_message['Subject'])
        # df.to_csv('attachments/'+'%s'%i+'myfile_%s.csv'% t)


    # if path.exists(r'attachments'):
    # os.rmdir(r'attachments')

    mypath = os.getcwd() + '/icici_lombard'

    if not path.exists(mypath):
        os.mkdir(mypath)
    if not path.exists(mypath + '/attachments_' + str(sys.argv[6])):
        os.mkdir(mypath + '/attachments_' + str(sys.argv[6]))
    if not path.exists(mypath + '/attachments_pdf_' + str(sys.argv[6])):
        os.mkdir(mypath + '/attachments_pdf_' + str(sys.argv[6]))

    mypath = os.getcwd() + '/icici_lombard/attachments_' + str(sys.argv[6]) + '/'
    for filename in os.listdir(mypath):
        file_path = os.path.join(mypath, filename)
        if os.path.isfile(file_path) or os.path.islink(file_path):
            os.unlink(file_path)

    pdfpath = os.getcwd() + '/icici_lombard/attachments_pdf_' + str(sys.argv[6]) + '/'
    pdffiles = [f for f in listdir(pdfpath) if isfile(join(pdfpath, f))]

    for filename in os.listdir(pdfpath):
        file_path = os.path.join(pdfpath, filename)
        if os.path.isfile(file_path) or os.path.islink(file_path):
            os.unlink(file_path)
    read_email_from_gmail()

    wbkName = 'icici_lombard/icici_lombard' + str(sys.argv[6]) + '.xlsx'
    wbk = openpyxl.Workbook()
    wbk.create_sheet('1')
    # wbk.create_sheet('Sheet3')
    s1 = wbk.worksheets[0]
    s2 = wbk.worksheets[1]
    onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
    t = 0
    yu = []
    # print(onlyfiles)
    for i in onlyfiles:
        if i[-1] == '#':
            yu.append(t)
        t += 1
    # print(yu)
    t = 0
    for i in yu:
        onlyfiles.pop(i - t)
        t = t + 1
    pdfpath = os.getcwd() + '/icici_lombard/attachments_pdf_' + str(sys.argv[6]) + '/'
    pdffiles = [f for f in listdir(pdfpath) if isfile(join(pdfpath, f))]
    pdffiles.sort()
    # print(pdffiles)
    for t in range(0, len(pdffiles)):
        with open(pdfpath + pdffiles[t], "rb") as f:
            pdf = pdftotext.PDF(f)

        with open('icici_lombard/output.txt', 'w') as f:
            f.write(" ".join(pdf))
        with open('icici_lombard/output.txt', 'r') as myfile:
            f = myfile.read()
        eo = []
        w = f.find('amount of Rs.') + 13
        g = f[w:]
        u = u2 = g.find('.') + w
        eo.append(f[w:u])

        w1 = f.find('ref.') + 3
        g = f[w1:]
        x1 = g.find('no.') + 3 + w1
        u1 = g.find('dated') + w1
        eo.append(f[x1:u1])

        w = f.find('dated') + 5
        g = f[w:]
        u = u2 = g.find('towards') + w - 1
        eo.append(f[w:u])

        w1 = f.find('TDS is') + 7
        g = f[w1:]
        u1 = g.find('.') + w1
        eo.append(f[w1:u1])
        # print(eo)
        for i in range(0, len(eo)):
            s1.cell(row=t + 2, column=15 + i).value = eo[i]
    onlyfiles.sort()
    print(onlyfiles)
    for t in range(0, len(onlyfiles)):
        try:
            def select_sheet(wks, sheet_name):

                if not sheet_name in wks.sheetnames:
                    wks.create_sheet(sheet_name)
                # print(sheet_name)

                wks.save('icici_lombard.xlsx')


            def select_column(wks, s, ro):
                sheet = wks.worksheets[s]

                max_col = sheet.max_column
                s = []
                for i in range(1, max_col + 1):
                    cell_obj = sheet.cell(row=1, column=i)
                    s.append(cell_obj.value)


            # print(s)
            loc = (mypath + onlyfiles[t])
            wb = xlrd.open_workbook(loc)
            wb1 = openpyxl.load_workbook(loc)
            sh1 = ['Claim No', 'UHID NO', 'Name of the Patient', 'Policy Name', 'Requested Amount', 'Final Amount Settled',
                   'Diagnosis', 'Date of Admission', 'Date Of Discharge', 'CO-PAYMENT AMOUNT', 'DISALLOWED AMOUNT',
                   'DISALLOWED REASONS', 'settled amount', 'cheque/EFT vide ref. no.', 'date of payment', 'TDS']
            for i in range(0, len(sh1)):
                s1.cell(row=1, column=i + 3).value = sh1[i]
            sh2 = ['Charges Details', 'Claimed', 'Deductions', 'Paid', 'Reason for Deductions']
            for i in range(0, len(sh2)):
                s2.cell(row=1, column=i + 3).value = sh2[i]
            sr2 = wb1.worksheets[2]
            sheet_4 = wb.sheet_by_index(2)
            sheet_4.cell_value(0, 0)
            jf = []
            kl = []
            q = []
            w = []
            # row_num=sheet_4.nrows
            # column_num=sheet_4.ncols
            for i in range(1, sheet_4.nrows):
                jf.append(str(sheet_4.cell_value(i, 1)))
                kl.append(str(sheet_4.cell_value(i, 2)))
                q.append(str(sheet_4.cell_value(i, 3)))
                w.append(str(sheet_4.cell_value(i, 4)))
            # print(jf)
            res = [sub.replace('Â\xa0', '') for sub in jf]
            res1 = [sub.replace('Â', '') for sub in res]
            res5 = [sub.replace('Â\xa0', '') for sub in kl]
            res2 = [sub.replace('Â', '') for sub in res5]
            res6 = [sub.replace('Â\xa0', '') for sub in q]
            res3 = [sub.replace('Â', '') for sub in res6]
            res7 = [sub.replace('Â\xa0', '') for sub in w]
            res4 = [sub.replace('Â', '') for sub in res7]
            # print(res1,t)
            for i in range(0, len(res1)):
                sr2.cell(row=i + 2, column=2).value = res1[i]
                sr2.cell(row=i + 2, column=3).value = res2[i]
                sr2.cell(row=i + 2, column=4).value = res3[i]
                sr2.cell(row=i + 2, column=5).value = res4[i]
            # print(t,i)
            sr2 = wb1.worksheets[4]
            sheet_4 = wb.sheet_by_index(4)
            sheet_4.cell_value(0, 0)
            jf = []
            kl = []
            q = []
            w = []
            # row_num=sheet_4.nrows
            # column_num=sheet_4.ncols
            for i in range(1, sheet_4.nrows):
                jf.append(str(sheet_4.cell_value(i, 1)))
                kl.append(str(sheet_4.cell_value(i, 2)))
                q.append(str(sheet_4.cell_value(i, 3)))
                w.append(str(sheet_4.cell_value(i, 4)))
            # print(jf)
            res = [sub.replace('Â\xa0', '') for sub in jf]
            res1 = [sub.replace('Â', '') for sub in res]
            res5 = [sub.replace('Â\xa0', '') for sub in kl]
            res2 = [sub.replace('Â', '') for sub in res5]
            res6 = [sub.replace('Â\xa0', '') for sub in q]
            res3 = [sub.replace('Â', '') for sub in res6]
            res7 = [sub.replace('Â\xa0', '') for sub in w]
            res4 = [sub.replace('Â', '') for sub in res7]
            # print(res1,t)
            for i in range(0, len(res1)):
                sr2.cell(row=i + 2, column=2).value = res1[i]
                sr2.cell(row=i + 2, column=3).value = res2[i]
                sr2.cell(row=i + 2, column=4).value = res3[i]
                sr2.cell(row=i + 2, column=5).value = res4[i]
            # print(t,i)
            wb1.save(loc)
            wb = xlrd.open_workbook(loc)
            sheet_2 = wb.sheet_by_index(2)
            sheet_2.cell_value(0, 0)
            sheet_3 = wb.sheet_by_index(3)
            sheet_3.cell_value(0, 0)
            sheet_4 = wb.sheet_by_index(4)
            sheet_4.cell_value(0, 0)
            jf = []
            jf.append(sheet_2.cell_value(1, 2))
            jf.append(sheet_2.cell_value(2, 2))
            jf.append(sheet_2.cell_value(3, 2))
            jf.append(sheet_2.cell_value(8, 2))
            ccn = sheet_2.cell_value(1, 4)
            jf.append(sheet_2.cell_value(5, 4))
            jf.append(sheet_4.cell_value(2, 4))
            jf.append(sheet_4.cell_value(1, 4))
            jf.append(sheet_2.cell_value(7, 2))
            jf.append(sheet_2.cell_value(7, 4))
            # print(jf)

            m = []
            gh = []
            hg = []
            r = []
            rt = []
            for i in range(2, sheet_3.nrows):
                m.append(sheet_3.cell_value(i, 2))
                gh.append(sheet_3.cell_value(i, 3))
                hg.append(sheet_3.cell_value(i, 4))
                r.append(sheet_3.cell_value(i, 7))
                rt.append(sheet_3.cell_value(i, 8))
            hg = [str(sub).replace('!  td>', '') for sub in hg]
            for wd in wbk.worksheets[:1]:
                wd.cell(row=1, column=1).value = 'Sr. No.'
                wd.cell(row=1, column=2).value = 'AL NO'
                wd.cell(row=t + 2, column=1).value = t + 1
                wd.cell(row=t + 2, column=2).value = ccn
            s2.cell(row=1, column=1).value = 'Sr. No.'
            s2.cell(row=1, column=2).value = 'claim NO'
            for i in range(0, len(m)):
                row_num = s2.max_row
                s2.cell(row=row_num + 1, column=1).value = t + 1
                s2.cell(row=row_num + 1, column=2).value = jf[0]
                s2.cell(row=row_num + 1, column=3).value = m[i]
                s2.cell(row=row_num + 1, column=4).value = gh[i]
                s2.cell(row=row_num + 1, column=5).value = hg[i]
                s2.cell(row=row_num + 1, column=6).value = r[i]
                s2.cell(row=row_num + 1, column=7).value = rt[i]
            sheet_3 = wb.sheet_by_index(1)
            sheet_3.cell_value(0, 0)
            gh = []
            hg = []
            r = []
            rt = []
            for i in range(0, len(jf)):
                s1.cell(row=t + 2, column=i + 3).value = jf[i]
            # for i in range(2,sheet_3.nrows):
            hg.append(sheet_3.cell_value(2, 5))
            r.append(sheet_3.cell_value(2, 6))
            rt.append(sheet_3.cell_value(2, 7))
            sheet_1 = wb.sheet_by_index(0)
            sheet_1.cell_value(0, 0)
            e_id = sheet_1.cell_value(2, 3)
            e_name = sheet_1.cell_value(2, 4)
            for i in range(0, len(hg)):
                s1.cell(row=t + 2, column=len(jf) + 3).value = hg[i]
                s1.cell(row=t + 2, column=len(jf) + 4).value = r[i]
                s1.cell(row=t + 2, column=len(jf) + 5).value = rt[i]
            s1.cell(row=t + 2, column=19).value = e_id
            s1.cell(row=t + 2, column=20).value = e_name
            if (ccn == ''):
                ccn = 'al_no_not_found'
            os.rename(os.getcwd() + '/icici_lombard/attachments_pdf_' + str(sys.argv[6]) + '/' + pdffiles[t],
                      os.getcwd() + '/icici_lombard/attachments_pdf_' + str(sys.argv[6]) + '/' + ccn + '.pdf')

        except Exception as e:
            s1.cell(row=t + 2, column=1).value = 'error'
            os.rename(os.getcwd() + '/icici_lombard/attachments_pdf_' + str(sys.argv[6]) + '/' + pdffiles[t],
                      os.getcwd() + '/icici_lombard/attachments_pdf_' + str(sys.argv[6]) + '/' + ccn + '.pdf')

    # print(gh,hg,r,rt)
    print("Done")
    wbk.save(wbkName)
    wbk.close
    wbkName = 'count/count.xlsx'
    wbk = openpyxl.load_workbook(wbkName)
    s1 = wbk.worksheets[0]

    row_ = s1.max_row + 1
    s1.cell(row=row_, column=1).value = 'icici_lombard'
    s1.cell(row=row_, column=2).value = str(sys.argv[6])
    s1.cell(row=row_, column=3).value = len(fg)
    s1.cell(row=row_, column=4).value = len(onlyfiles)
    s1.cell(row=row_, column=5).value = len(repeat)
    wbk.save(wbkName)
    subprocess.run(["python", "updation.py", "1", "max", "9", " "])

except:
    log_exceptions()