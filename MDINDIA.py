import smtplib
import time
import imaplib
import email
import camelot
import PyPDF2
import xlsxwriter
from xlrd import open_workbook
import xlwt
import sys
import os
import os.path
import xlrd
from os import listdir
from os import path
from os.path import isfile, join
import pandas as pd
import tabula

from decode_error import check_subject, read_from_delete

if path.exists(r'MDINDIA' + str(sys.argv[6]) + '.xlsx'):
    os.remove(r'MDINDIA' + str(sys.argv[6]) + '.xlsx')
import openpyxl
from make_log import log_exceptions
import subprocess

try:


    po = []
    subprocess.run(["python", "updation.py", "1", "max", "9", "X"])

    eu = []
    wbkName = 'MDINDIA/MDINDIA' + str(sys.argv[6]) + '.xlsx'
    wbk = openpyxl.Workbook()
    wbk.create_sheet('Sheet1')
    wbk.create_sheet('DEDUCTION DETAILS')
    wbk.create_sheet('DISCOUNT DETAILS')
    z = 0
    c = 0
    fg = []
    repeat = []
    from openpyxl.styles import Color, PatternFill, Font, Border

    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')


    def read_email_from_gmail():
        SMTP_SERVER = str(sys.argv[5])
        mail = imaplib.IMAP4_SSL(SMTP_SERVER)
        # mail.login(user = 'Mediclaim@therisingmedicare.com', password = 'cef@2018')

        e_id = str(sys.argv[1])
        pswd = str(sys.argv[2])
        srt = str(sys.argv[3])
        stp = str(sys.argv[4])
        mail.login(user=e_id, password=pswd)
        mail.select("inbox", readonly=True)
        ###############################################<
        mail_uid = str(sys.argv[7])
        if mail_uid == -1:
            type, data = mail.search(None, '(SUBJECT "ECS PAYMENT DONE FOR" since ' + srt + ' before ' + stp + ')')
            ids = data[0]
            id_list = ids.split()
        else:
            ids = mail_uid  # data is a list.
            # accept id from outside and put in id_list akshay var name = id

            id_list = []  # ids is a space separated string
            id_list.append(ids)
        ###############################################>
        # type, data = mail.search(None, '(SUBJECT "ECS PAYMENT DONE FOR" since ' + srt + ' before ' + stp + ')')
        # ids = data[0]  # data is a list.
        # id_list = ids.split()  # ids is a space separated string
        # # print(id_list)
        for i in range(0, len(id_list)):
            latest_email_id = id_list[i]  # get the latest
            result, data = mail.fetch(latest_email_id,
                                      "(RFC822)")  # fetch the email body (RFC822)             for the given ID
            ##################################################ak
            try:
                # raw_email = data[0][1].decode('utf-8')
                try:
                    raw_email = data[0][1].decode('utf-8')
                except UnicodeDecodeError:
                    try:
                        raw_email = data[0][1].decode('ISO-8859-1')
                    except UnicodeDecodeError:
                        try:
                            raw_email = data[0][1].decode('ascii')
                        except UnicodeDecodeError:
                            pass
                email_message = email.message_from_string(raw_email)
                subject = email_message['Subject']
                result, sys.argv[8] = check_subject(subject, sys.argv[8], mail)
                if result == 'Changed':
                    # raise Exception('subject not matched')
                    raise Exception('subject not matched', )
            except:
                try:
                    log_exceptions(syssubject=sys.argv[8], subject=subject, error='subject not matched')
                except:
                    pass
                if result != 'OK':
                    data = {'server': SMTP_SERVER,
                            'hospmail': e_id,
                            'pass': pswd,
                            'subject': sys.argv[8]}
                    try:
                        data = read_from_delete(data)
                        if data == None:
                            raise Exception("Not found")
                    except:
                        log_exceptions(msg='not found in deleted', subject=sys.argv[8])
            ##################################################akend
            raw_email = data[0][1].decode('utf-8')
            email_message = email.message_from_string(raw_email)
            if email_message['Subject'] not in fg:
                for part in email_message.walk():

                    if part.get_content_maintype() == 'multipart':
                        # print part.as_string()
                        continue
                    # if part.get('Content-Disposition') is None:
                    # print part.as_string()
                    # continue
                    # print('hi')
                    fileName = part.get_filename()  # str(id_list[i])+'.pdf'
                    detach_dir = (os.getcwd() + '/MDINDIA/attachments_' + str(sys.argv[6]))
                    if bool(fileName):
                        # print (email_message['Subject'])
                        filePath = os.path.join(detach_dir, fileName)
                        if not os.path.isfile(filePath):
                            from reportlab.pdfgen import canvas
                            c = canvas.Canvas(fileName)
                            print(fileName)
                            fp = open(filePath, 'wb')
                            fp.write(part.get_payload(decode=True))
                            fp.close()
            else:
                repeat.append(email_message['Subject'])
            fg.append(email_message['Subject'])


    mypath = os.getcwd() + '/MDINDIA'
    if not path.exists(mypath):
        os.mkdir(mypath)
    if not path.exists(mypath + '/attachments_' + str(sys.argv[6])):
        os.mkdir(mypath + '/attachments_' + str(sys.argv[6]))
    mypath = mypath + '/attachments_' + str(sys.argv[6]) + '/'
    for filename in os.listdir(mypath):
        file_path = os.path.join(mypath, filename)
        if os.path.isfile(file_path) or os.path.islink(file_path):
            os.unlink(file_path)
    read_email_from_gmail()
    onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
    for t in range(0, len(onlyfiles)):
        try:
            # wbk.create_sheet('Sheet3')
            s1 = wbk.worksheets[0]
            s2 = wbk.worksheets[1]
            s3 = wbk.worksheets[2]
            s4 = wbk.worksheets[3]
            sh1 = ['Policy No.', 'Insurance Co.', 'Medi Assist ID', 'Patient Name', 'EMP.No.', 'Primary Beneficiary',
                   'Diagnosis', 'Beneficiary Bank Name', 'Beneficiary A/C Number', 'Beneficiary Name', 'utr', 'trans date']
            for i in range(0, len(sh1)):
                s1.cell(row=1, column=i + 3).value = sh1[i]
            sh2 = ['Balance Sum Insured Before Claim', 'Lodge Amt', 'Deduction Amt', 'Discount Amt', 'Authorized Amt',
                   'Settled Amt', 'NetPayable', 'Balance Sum Insured After Claim', 'TDS in %', 'TDS Amt', 'Final Payable',
                   'Remarks']
            for i in range(0, len(sh2)):
                s2.cell(row=1, column=i + 3).value = sh2[i]
            sh3 = ['Deducted Amount', 'Remarks']
            for i in range(0, len(sh3)):
                s3.cell(row=1, column=i + 3).value = sh3[i]
            sh4 = ['PARTICULARS', 'Discount Amount', 'Remarks']
            for i in range(0, len(sh4)):
                s4.cell(row=1, column=i + 3).value = sh4[i]
            tables = camelot.read_pdf(mypath + onlyfiles[t], line_scale=40, pages='all')
            tables.export('MDINDIA/foo1.xls', f='excel')
            loc = ("MDINDIA/foo1.xls")
            wb = xlrd.open_workbook(loc)
            sheet_1 = wb.sheet_by_index(0)
            sheet_1.cell_value(0, 0)
            f = (sheet_1.cell_value(1, 1))
            f = f.replace('\n', '$$ ')
            text_file = open("mail.txt", "w")
            n = text_file.write(f)
            text_file.close()

            gh = []
            x1 = f.find('Policy No.') + 13
            g = f[x1:]
            y1 = g.find('$$') + x1
            gh.append(f[x1:y1])

            x2 = f.find('Insurance Co.') + 16
            g = f[x2:]
            y2 = g.find('$$') + x2
            gh.append(f[x2:y2])

            x3 = f.find('CCN') + 6
            g = f[x3:]
            y3 = g.find('$$') + x3
            ccn = (f[x3:y3])

            x4 = f.find('MD ID No.') + 12
            g = f[x4:]
            y4 = g.find('$$') + x4
            gh.append(f[x4:y4])

            x5 = f.find('Patient Name') + 15
            g = f[x5:]
            y5 = g.find('$$') + x5
            gh.append(f[x5:y5])

            x6 = f.find('EMP.No.') + 10
            g = f[x6:]
            y6 = g.find('$$') + x6
            gh.append(f[x6:y6])

            x7 = y6 + 8
            g = f[x7:]
            y7 = g.find('$$') + x7
            gh.append(f[x7:y7])

            x8 = f.find('Diagnosis') + 16
            g = f[x8:]
            y8 = g.find('$$') + x8
            gh.append(f[x8:y8])

            sheet_2 = wb.sheet_by_index(1)
            f = ''
            sheet_2.cell_value(0, 0)
            for i in range(2, 8):
                k = sheet_2.cell_value(i, 1)
                f = f + '\n' + k
            f = f.replace('\n', '$$ ')
            text_file = open("mail.txt", "w")
            n = text_file.write(f)
            text_file.close()

            x9 = f.find(' ') + 1
            g = f[x9:]
            y9 = g.find('$$') + x9
            gh.append(f[x9:y9])

            x10 = f.find('Beneficiary Bank Name') + 24
            g = f[x10:]
            y10 = g.find('$$') + x10
            gh.append(f[x10:y10])

            x11 = f.find('Beneficiary A/C Number') + 25
            g = f[x11:]
            y11 = g.find('$$') + x11
            gh.append(f[x11:y11])

            if f.find('ECS Tran No.') != -1:

                x10 = f.find('ECS Tran No.') + 15
                g = f[x10:]
                y10 = g.find('$$') + x10
                gh.append(f[x10:y10])
            elif f.find('Cheque No') != -1:
                x11 = f.find('Cheque No') + 12
                g = f[x11:]
                y11 = g.find('$$') + x11
                gh.append(f[x11:y11])
            # print(gh[-1])

            if f.find('ECS Tran Date') != -1:
                x11 = f.find('ECS Tran Date') + 16
                g = f[x11:]
                y11 = g.find('$$') + x11
                gh.append(f[x11:y11])
            elif f.find('Cheque Date') != -1:
                x11 = f.find('Cheque Date') + 14
                g = f[x11:]
                y11 = g.find('$$') + x11
                gh.append(f[x11:y11])
            # print(gh[-1])
            x10 = f.find('period from') + 12
            g = f[x10:]
            y10 = g.find('to') + x10
            gh.append(f[x10:y10])

            g = f[y10:]
            y11 = g.find('.') + y10
            gh.append(f[y10 + 3:y11])
            # print(gh[-1],gh[-2])
            gh = [sub.replace('$', '') for sub in gh]
            for i in range(0, len(gh)):
                s1.cell(row=t + 2, column=i + 3).value = gh[i]
            f = ''
            n = sheet_1.nrows - 2

            for i in range(0, 2):
                k = sheet_1.cell_value(n, 1)
                f = f + '\n' + k
                n += 1
            f = f.replace('\n', '$$ ')
            text_file = open("MDINDIA/mail.txt", "w")
            n = text_file.write(f)
            text_file.close()

            hg = []
            w = f.find('$$') + 3
            g = f[w:]
            u = g.find('$$') + w
            hg.append(f[w:u])

            w1 = f.find('Balance Sum Insured Before Claim') + 35
            g = f[w1:]
            u1 = g.find('$$') + w1
            hg.append(f[w1:u1])

            w2 = u1 + 3
            g = f[w2:]
            u2 = g.find('$$') + w2
            hg.append(f[w2:u2])

            w3 = u2 + 3
            g = f[w3:]
            u3 = g.find('$$') + w3
            hg.append(f[w3:u3])

            w4 = f.find('Discount Amt') + 15
            g = f[w4:]
            u4 = g.find('$$') + w4
            hg.append(f[w4:u4])

            w5 = u4 + 3
            g = f[w5:]
            u5 = g.find('$$') + w5
            hg.append(f[w5:u5])

            w6 = u5 + 3
            g = f[w6:]
            u6 = g.find('$$') + w6
            hg.append(f[w6:u6])

            w7 = f.find('NetPayable') + 13
            g = f[w7:]
            u7 = g.find('$$') + w7
            hg.append(f[w7:u7])

            w8 = f.find('Balance Sum Insured After Claim') + 33
            g = f[w8:]
            u8 = g.find('$$') + w8
            hg.append(f[w8:u8])

            w9 = u8 + 3
            g = f[w9:]
            u9 = g.find('$$') + w9
            hg.append(f[w9:u9])

            w10 = u9 + 3
            g = f[w10:]
            u10 = g.find('$$') + w10
            hg.append(f[w10:u10])

            w11 = f.find('Remarks') + 10

            hg.append(f[w11:])
            # print(hg)
            for i in range(0, len(hg)):
                s2.cell(row=t + 2, column=i + 3).value = hg[i]
            sheet_1 = wb.sheet_by_index(0)
            sheet_1.cell_value(0, 0)
            b = []
            p = []
            np = []
            r = []
            sp = []
            n = sheet_1.nrows

            d = sheet_1.cell_value(2, 1)
            if (d == 'DEDUCTION DETAILS'):
                for i in range(4, n):
                    p.append(sheet_1.cell_value(i, 3))
                    r.append(sheet_1.cell_value(i, 4))
                    if p[i - 4] == '':
                        p.pop()
                        r.pop()
                        break
            # print(p)
            for i in range(0, len(p)):
                row_num = s3.max_row
                s3.cell(row=1, column=1).value = 'Sr. No.'
                s3.cell(row=1, column=2).value = 'CCN'
                s3.cell(row=row_num + 1, column=1).value = c + 1
                s3.cell(row=row_num + 1, column=2).value = ccn
                s3.cell(row=row_num + 1, column=3).value = p[i]
                s3.cell(row=row_num + 1, column=4).value = r[i]
                c = c + 1
            for i in range(1, n):
                l = sheet_1.cell_value(i, 1)
                if l == 'DISCOUNT DETAILS':
                    # print(t,i,n)
                    break
            m = i + 2
            for j in range(m, n - 2):
                b.append(sheet_1.cell_value(j, 1))
                np.append(sheet_1.cell_value(j, 3))
                sp.append(sheet_1.cell_value(j, 4))
            # print(b)
            for i in range(0, len(b)):
                row_num1 = s4.max_row
                s4.cell(row=1, column=1).value = 'Sr. No.'
                s4.cell(row=1, column=2).value = 'CCN'
                s4.cell(row=row_num1 + 1, column=1).value = z + 1
                s4.cell(row=row_num1 + 1, column=2).value = ccn
                s4.cell(row=row_num1 + 1, column=i + 3).value = b[i]
                s4.cell(row=row_num1 + 1, column=i + 4).value = np[i]
                s4.cell(row=row_num1 + 1, column=i + 5).value = sp[i]
                z = z + 1
            for wd in wbk.worksheets[:2]:
                wd.cell(row=1, column=1).value = 'Sr. No.'
                wd.cell(row=1, column=2).value = 'CCN'
                wd.cell(row=t + 2, column=1).value = t + 1
                wd.cell(row=t + 2, column=2).value = ccn
            os.rename(os.getcwd() + '/MDINDIA/attachments_' + str(sys.argv[6]) + '/' + onlyfiles[t],
                      os.getcwd() + '/MDINDIA/attachments_' + str(sys.argv[6]) + '/' + ccn + '.pdf')
        except Exception as e:
            s1.cell(row=t + 2, column=1).value = 'error'
            s1.cell(row=t + 2, column=2).value = ccn
            os.rename(os.getcwd() + '/MDINDIA/attachments_' + str(sys.argv[6]) + '/' + onlyfiles[t],
                      os.getcwd() + '/MDINDIA/attachments_' + str(sys.argv[6]) + '/' + ccn + '.pdf')

    print("Done")
    wbk.save(wbkName)
    wbk.close
    wbkName = 'count/count.xlsx'
    wbk = openpyxl.load_workbook(wbkName)
    s1 = wbk.worksheets[0]

    row_ = s1.max_row + 1
    s1.cell(row=row_, column=1).value = 'MDINDIA'
    s1.cell(row=row_, column=2).value = str(sys.argv[6])
    s1.cell(row=row_, column=3).value = len(fg)
    s1.cell(row=row_, column=4).value = len(onlyfiles)
    s1.cell(row=row_, column=5).value = len(repeat)
    wbk.save(wbkName)
    subprocess.run(["python", "updation.py", "1", "max", "9", " "])
except:
    log_exceptions()