import os
import sys
import struct, time
import subprocess
from datetime import date
import datetime
import openpyxl
import pdftotext
import time
import requests
now = datetime.datetime.now()
'''
wbkName = 'log file.xlsx'
wbk= openpyxl.load_workbook(wbkName)
s1=wbk.worksheets[0]
s2=wbk.worksheets[1]
row_count_1 = s2.max_row
s2.cell(row_count_1+1, column=1).value=sys.argv[2]
s2.cell(row_count_1+1, column=2).value=sys.argv[3]
s2.cell(row_count_1+1, column=3).value=sys.argv[4]
s2.cell(row_count_1+1, column=5).value=now
s2.cell(row_count_1+1, column=7).value=sys.argv[5]
s2.cell(row_count_1+1, column=8).value=sys.argv[6]
'''

subprocess.run(["python", "updation.py","1","max1","1",sys.argv[2]])
subprocess.run(["python", "updation.py","1","max","2",sys.argv[3]])
subprocess.run(["python", "updation.py","1","max","3",sys.argv[4]])
subprocess.run(["python", "updation.py","1","max","5",str(now)])
subprocess.run(["python", "updation.py","1","max","7",sys.argv[5]])
subprocess.run(["python", "updation.py","1","max","8",sys.argv[6]])

with open(sys.argv[1], "rb") as f:
	pdf = pdftotext.PDF(f)

with open('vidal/output.txt', 'w') as f:
	f.write(" ".join(pdf))     
with open('vidal/output.txt', 'r') as myfile:
 	f = myfile.read()

try:
    hg=[]
    # if f.find('Policy Number') != -1:
    #   w=f.find('Policy Number')+len('Policy Number')
    #   k=f[w:]
    #   u=k.find("\n")+w
    #   g=f[w:u]
    #   hg.append(g)
    # else:
    #       print("Policy No not found in output1.text")

    if f.find('Pre-Auth Number:') != -1:
      w=f.find('Pre-Auth Number:')+len('Pre-Auth Number:')
      k=f[w:]
      u=k.find("in your")+w
      g=f[w:u]
      hg.append(g)
    else:
      print("Pre-Auth Number not found in output1.text")

    if f.find('request for') != -1:
      w=f.find('request for')+len('request for')
      k=f[w:]
      u=k.find("and")+w
      g=f[w:u]
      hg.append(g)
    else:
      print("request for not found in output1.text")

    # if f.find('Claimed Amount') != -1:
    #   w=f.find('Claimed Amount')+len('Claimed Amount')
    #   k=f[w:]
    #   u=k.find("\n")+w
    #   g=f[w:u]
    #   hg.append(g.strip())
    # else:
    #    print("Claimed Amount not found in output1.text")   

       

    hg=[sub.replace(':','') for sub in hg]	
    hg=[sub.replace('  ','') for sub in hg]
    hg=[sub.replace('Rs.','') for sub in hg]
    hg=[sub.replace('\n','') for sub in hg]		
	
	#s2.cell(row_count_1+1, column=9).value='Yes'
	#s2.cell(row_count_1+1, column=10).value='NA'
	#wbk.save(wbkName)
	
    subprocess.run(["python", "updation.py","1","max","9",'Yes'])
    subprocess.run(["python", "updation.py","1","max","10",'NA'])
    try:  
     
      subprocess.run(["python", "test_api.py",hg[0],'','','','Acknowledgement',sys.argv[6],sys.argv[1],'','',hg[1]])

      # subprocess.run(["python", "test_api.py",hg[1],hg[3],hg[2],'','Denial',sys.argv[6],'',hg[0],'',hg[2]])

      subprocess.run(["python", "updation.py","1","max","11",'Yes'])
    except Exception as e:
		#s2.cell(row_count_1+1, column=11).value='NO'
        subprocess.run(["python", "updation.py","1","max","11",'No'])
except Exception as e:
	#s2.cell(row_count_1+1, column=9).value='No'
	#s2.cell(row_count_1+1, column=11).value='NO'
    subprocess.run(["python", "updation.py","1","max","9",'Yes'])
    subprocess.run(["python", "updation.py","1","max","11",'No'])
now = datetime.datetime.now()
#s2.cell(row_count_1+1, column=6).value=now
#wbk.save(wbkName)
subprocess.run(["python", "updation.py","1","max","6",str(now)])
