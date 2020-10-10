import os
import sys
import struct, time
import subprocess
from datetime import date
import datetime
import openpyxl
import pdftotext
import time
import requests
now = datetime.datetime.now()


subprocess.run(["python", "updation.py","1","max1","1",sys.argv[2]])
subprocess.run(["python", "updation.py","1","max","2",sys.argv[3]])
subprocess.run(["python", "updation.py","1","max","3",sys.argv[4]])
subprocess.run(["python", "updation.py","1","max","5",str(now)])
subprocess.run(["python", "updation.py","1","max","7",sys.argv[5]])
subprocess.run(["python", "updation.py","1","max","8",sys.argv[6]])

with open(sys.argv[1], "rb") as f:
	pdf = pdftotext.PDF(f)

with open('hdfc/output2.txt', 'w',encoding='utf-8') as f:
	f.write(" ".join(pdf))     
with open('hdfc/output2.txt', 'r',encoding='utf-8') as myfile:
 	f = myfile.read()

try:		
	hg=[]
	
	w=f.find('Claim Number')+12
	g=f[w:]
	u=g.find('(')+w
	if f[w:u].find('_')!=-1:
		u=g.find('_')+w
	hg.append(f[w:u])

	status='Approved'
	w=f.find('Total Authorised Amount')+23
	g=f[w:]
	u=g.find('(')+w
	hg.append(f[w:u])
	
	w=f.find('Policy Number')+13
	g=f[w:]
	u=g.find('Expected')+w
	hg.append(f[w:u])

	w=f.find('Member')+7
	g=f[w:]
	u=g.find('\n')+w
	hg.append(f[w:u])

	w=f.find('Patient Name')+len('Patient Name')
	g=f[w:]
	u=g.find('Age')+w
	hg.append(f[w:u])


	hg=[sub.replace(':','') for sub in hg]	
	hg=[sub.replace('  ','') for sub in hg]
	hg=[sub.replace('Rs.','') for sub in hg]
	#s2.cell(row_count_1+1, column=9).value='Yes'
	#s2.cell(row_count_1+1, column=10).value='NA'
	#wbk.save(wbkName)
	
	subprocess.run(["python", "updation.py","1","max","9",'Yes'])
	subprocess.run(["python", "updation.py","1","max","10",'NA'])
	
	try:
		subprocess.run(["python", "test_api.py",hg[0],hg[1],hg[2],'',status,sys.argv[6],sys.argv[1],'',hg[3],hg[4]])
		'''wbk= openpyxl.load_workbook(wbkName)
		s2=wbk.worksheets[1]
		s2.cell(row_count_1+1, column=11).value='YES'
		'''
		subprocess.run(["python", "updation.py","1","max","11",'Yes'])	
	except Exception as e:
		#s2.cell(row_count_1+1, column=11).value='NO'
		subprocess.run(["python", "updation.py","1","max","11",'No'])
except Exception as e:
	#s2.cell(row_count_1+1, column=9).value='No'
	#s2.cell(row_count_1+1, column=11).value='NO'
	subprocess.run(["python", "updation.py","1","max","9",'Yes'])
	subprocess.run(["python", "updation.py","1","max","11",'No'])
now = datetime.datetime.now()
#s2.cell(row_count_1+1, column=6).value=now
#wbk.save(wbkName)
subprocess.run(["python", "updation.py","1","max","6",str(now)])
