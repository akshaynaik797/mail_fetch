import os
import shutil
import sys
import subprocess
from datetime import date
import datetime
from os import listdir
from os import path
from os.path import isfile, join
from os import path
from openpyxl import load_workbook, Workbook
from clean_files import clean_f
from movemaster import move_master_to_master_insurer
from make_log import log_exceptions
try:
    st = str(sys.argv[1])  # input("Enter your start date :"))
    stp = str(sys.argv[2])  # input("Enter your end date :"))
    insurer = str(sys.argv[3])  # Accept insurer name
    hosp = str(sys.argv[4])  # Accept hospital name
    mail_uid = str(sys.argv[5])  # Accept mail uuid
    subject = str(sys.argv[6])  # Accept mail subject
    now = datetime.datetime.now()


    subprocess.run(["python", "updation.py","1","max","2",insurer])
    subprocess.run(["python", "updation.py","1","max","3","settlement"])
    subprocess.run(["python", "updation.py","1","max","20",mail_uid])
    subprocess.run(["python", "updation.py","1","max","21",hosp])
    subprocess.run(["python", "updation.py","1","max","5",str(now)])


    # delete the master if it exists
    if path.exists('master.xlsx'):
        os.remove('master.xlsx')

    with open('config.txt', 'r') as myfile:
        f = myfile.readlines()
        f = [i.strip('\n') for i in f]
        for i in f:
            if hosp in i:
                op = i
            else:
                print('Invalid insurer name')
                #exit()



    dest = 'count/count.xlsx'
    sh1 = ['insurance id',	'hospital',	'mail count',	'attachments count']

    if path.exists(dest):
        os.remove(dest)

    if not path.exists(dest):
        wb = Workbook()
        ws1 = wb.create_sheet("Sheet1", 1)
        ws1 = wb.create_sheet("Sheet2", 2)
        ws1 = wb.create_sheet("Sheet3", 3)
        ws1 = wb.create_sheet("Sheet4", 4)
        ws1 = wb.create_sheet("Sheet5", 5)
        ws1 = wb.create_sheet("Sheet6", 6)
        wb.save(dest)

    if path.exists(dest):
        wb = load_workbook(dest)
        main_s1 = wb.worksheets[0]
        for i in range(0, len(sh1)):
            main_s1.cell(row=1, column=i + 1).value = sh1[i]
        wb.save(dest)
        wb.close()
    # with open('config.txt', 'r') as myfile:
    # 	f = myfile.read()
    # op=f.split('\n')
    # op.pop(-1)

    # Python method getcwd() returns current working directory of a process.

    if insurer == 'park':
        for i in range(0,1):
            k = op.split(' ')
            # print(k)
            subprocess.run(["python", "Park_settlement.py",k[0],k[1],st,stp,k[2],k[3],mail_uid,subject])
            # os.system('python aditya_birla.py ' + k[0] + ' ' + k[1] + ' ' + st + ' ' + stp + ' ' + k[2] + ' ' + k[
            #     3] + ' ' + mail_uid)
            # mypath = os.getcwd() + '/Park/attachments_' + str(k[3]) + '/'
            # onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
            # for t in onlyfiles:
            #     shutil.copy(mypath + t, os.getcwd() + '/Attachments/')
            subprocess.run(["python", "master.py", insurer,op,mail_uid])
            move_master_to_master_insurer(mail_uid)
            insurer = 'park'
            if clean_f(insurer, k[3]):
                print(f'deleted {insurer} files.')


    if insurer == 'aditya':
        for i in range(0,1):
            k = op.split(' ')
            # print(k)
            subprocess.run(["python", "aditya_birla.py",k[0],k[1],st,stp,k[2],k[3],mail_uid,subject])
            # os.system('python aditya_birla.py ' + k[0] + ' ' + k[1] + ' ' + st + ' ' + stp + ' ' + k[2] + ' ' + k[
            #     3] + ' ' + mail_uid)
            mypath = os.getcwd() + '/aditya_birla/attachments_' + str(k[3]) + '/'
            onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
            if '.pdf' in onlyfiles:
                onlyfiles.remove('.pdf')
            for t in onlyfiles:
                shutil.copy(mypath + t, os.getcwd() + '/Attachments/')
            subprocess.run(["python", "master.py", insurer,op,mail_uid])
            move_master_to_master_insurer(mail_uid)
            if clean_f(insurer, k[3]):
                print(f'deleted {insurer} files.')

    elif insurer == 'apollo':
        for i in range(0,1):
            k = op.split(' ')
            subprocess.run(["python", "apollo_munich.py",k[0],k[1],st,stp,k[2],k[3],mail_uid,subject])
            # os.system('python apollo_munich.py ' + k[0] + ' ' + k[1] + ' ' + st + ' ' + stp + ' ' + k[2] + ' ' + k[
            #     3] + ' ' + mail_uid)
            mypath = os.getcwd() + '/apollo_munich/attachments_' + str(k[3]) + '/'
            onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
            if '.pdf' in onlyfiles:
                onlyfiles.remove('.pdf')
            for t in onlyfiles:
                shutil.copy(mypath + t, os.getcwd() + '/Attachments/')
            subprocess.run(["python", "master.py", insurer,op,mail_uid])
            move_master_to_master_insurer(mail_uid)
            if clean_f(insurer, k[3]):
                print(f'deleted {insurer} files.')

    elif insurer == 'big':
        for i in range(0,1):
            k = op.split(' ')
            subprocess.run(["python", "big.py",k[0],k[1],st,stp,k[2],k[3],mail_uid,subject])
            # os.system(
            #     'python big.py ' + k[0] + ' ' + k[1] + ' ' + st + ' ' + stp + ' ' + k[2] + ' ' + k[3] + ' ' + mail_uid)
            insurer = 'star'
            mypath = os.getcwd() + '/star/attachments_2_' + str(k[3]) + '/'
            onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
            if '.pdf' in onlyfiles:
                onlyfiles.remove('.pdf')
            for t in onlyfiles:
                shutil.copy(mypath + t, os.getcwd() + '/Attachments/')
            subprocess.run(["python", "master.py", insurer,op,mail_uid])
            move_master_to_master_insurer(mail_uid)
            if clean_f(insurer, k[3]):
                print(f'deleted {insurer} files.')

    elif insurer == 'small':
        for i in range(0,1):
            k=op.split(' ')
            subprocess.run(["python", "small.py",k[0],k[1],st,stp,k[2],k[3],mail_uid,subject])
            #os.system('python small.py '+k[0]+' '+k[1]+' '+st+' '+stp+' '+k[2]+' '+k[3]+' '+mail_uid)
            insurer = 'small_star'
            mypath = os.getcwd()+'/star/attachments_1_'+str(k[3])+'/'
            onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
            if '.pdf' in onlyfiles:
                onlyfiles.remove('.pdf')
            for t in onlyfiles:
                shutil.copy(mypath+t,os.getcwd()+'/Attachments/')
            subprocess.run(["python", "master.py", insurer,op,mail_uid])
            insurer = 'star'
            move_master_to_master_insurer(mail_uid)
            if clean_f(insurer, k[3]):
                print(f'deleted {insurer} files.')

    elif insurer == 'fgh':
        print(len(op))
        for i in range(0,1):
            k = op.split(' ')
            subprocess.run(["python", "fgh.py",k[0],k[1],st,stp,k[2],k[3],mail_uid,subject])
            # os.system(
            #     'python fgh.py ' + k[0] + ' ' + k[1] + ' ' + st + ' ' + stp + ' ' + k[2] + ' ' + k[3] + ' ' + mail_uid)
            mypath = os.getcwd() + '/FGH/attachments_' + str(k[3]) + '/'
            onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
            if '.pdf' in onlyfiles:
                onlyfiles.remove('.pdf')
            for t in onlyfiles:
                shutil.copy(mypath + t, os.getcwd() + '/Attachments/')
        subprocess.run(["python", "master.py", insurer,op,mail_uid])
        move_master_to_master_insurer(mail_uid)
        if clean_f(insurer, k[3]):
            print(f'deleted {insurer} files.')
        print("\nfgh")

    elif insurer == 'MDINDIA':
        for i in range(0,1):
            k = op.split(' ')
            subprocess.run(["python", "MDINDIA.py",k[0],k[1],st,stp,k[2],k[3],mail_uid,subject])
            # os.system(
            #     'python MDINDIA.py ' + k[0] + ' ' + k[1] + ' ' + st + ' ' + stp + ' ' + k[2] + ' ' + k[3] + ' ' + mail_uid)
            mypath = os.getcwd() + '/MDINDIA/attachments_' + str(k[3]) + '/'
            onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
            if '.pdf' in onlyfiles:
                onlyfiles.remove('.pdf')
            for t in onlyfiles:
                shutil.copy(mypath + t, os.getcwd() + '/Attachments/')
        subprocess.run(["python", "master.py", insurer,op,mail_uid])
        move_master_to_master_insurer(mail_uid)
        if clean_f(insurer, k[3]):
            print(f'deleted {insurer} files.')
        print("\nMDINDIA")

    elif insurer == 'fhpl':
        for i in range(0,1):
            k = op.split(' ')
            subprocess.run(["python", "fhpl.py",k[0],k[1],st,stp,k[2],k[3],mail_uid,subject])
            # os.system(
            #     'python fhpl.py ' + k[0] + ' ' + k[1] + ' ' + st + ' ' + stp + ' ' + k[2] + ' ' + k[3] + ' ' + mail_uid)
            mypath = os.getcwd() + '/fhpl/attachments_' + str(k[3]) + '/'
            onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
            if '.pdf' in onlyfiles:
                onlyfiles.remove('.pdf')
            for t in onlyfiles:
                shutil.copy(mypath + t, os.getcwd() + '/Attachments/')
            subprocess.run(["python", "master.py", insurer,op,mail_uid])
            move_master_to_master_insurer(mail_uid)
            if clean_f(insurer, k[3]):
                print(f'deleted {insurer} files.')

    elif insurer == 'hdfc':
        for i in range(0,1):
            k = op.split(' ')
            subprocess.run(["python", "hdfc.py",k[0],k[1],st,stp,k[2],k[3],mail_uid,subject])
            # os.system(
            #     'python hdfc.py ' + k[0] + ' ' + k[1] + ' ' + st + ' ' + stp + ' ' + k[2] + ' ' + k[3] + ' ' + mail_uid)
            mypath = os.getcwd() + '/hdfc/attachments_' + str(k[3]) + '/'
            onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
            if '.pdf' in onlyfiles:
                onlyfiles.remove('.pdf')
            for t in onlyfiles:
                shutil.copy(mypath + t, os.getcwd() + '/Attachments/')
            subprocess.run(["python", "master.py", insurer,op,mail_uid])
            move_master_to_master_insurer(mail_uid)
            if clean_f(insurer, k[3]):
                print(f'deleted {insurer} files.')

    elif insurer == 'health_heritage':
        for i in range(0,1):
            k = op.split(' ')
            subprocess.run(["python", "health_heritage.py",k[0],k[1],st,stp,k[2],k[3],mail_uid,subject])
            # os.system('python health_heritage.py ' + k[0] + ' ' + k[1] + ' ' + st + ' ' + stp + ' ' + k[2] + ' ' + k[
            #     3] + ' ' + mail_uid)
            mypath = os.getcwd() + '/health_heritage/attachments_' + str(k[3]) + '/'
            onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
            if '.pdf' in onlyfiles:
                onlyfiles.remove('.pdf')
            for t in onlyfiles:
                shutil.copy(mypath + t, os.getcwd() + '/Attachments/')
            subprocess.run(["python", "master.py", insurer,op,mail_uid])
            move_master_to_master_insurer(mail_uid)
            if clean_f(insurer, k[3]):
                print(f'deleted {insurer} files.')

    elif insurer == 'health_india':
        for i in range(0,1):
            k = op.split(' ')
            subprocess.run(["python", "health_india.py",k[0],k[1],st,stp,k[2],k[3],mail_uid,subject])
            # os.system('python health_india.py ' + k[0] + ' ' + k[1] + ' ' + st + ' ' + stp + ' ' + k[2] + ' ' + k[
            #     3] + ' ' + mail_uid)
            mypath = os.getcwd() + '/health_india/attachments_' + str(k[3]) + '/'
            onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
            if '.pdf' in onlyfiles:
                onlyfiles.remove('.pdf')
            for t in onlyfiles:
                shutil.copy(mypath + t, os.getcwd() + '/Attachments/')
            subprocess.run(["python", "master.py", insurer,op,mail_uid])
            move_master_to_master_insurer(mail_uid)
            if clean_f(insurer, k[3]):
                print(f'deleted {insurer} files.')

    elif insurer == 'icici_lombard':
        for i in range(0,1):
            k = op.split(' ')
            subprocess.run(["python", "icici.py",k[0],k[1],st,stp,k[2],k[3],mail_uid,subject])
            mypath = os.getcwd() + '/icici_lombard/attachments_pdf_' + str(k[3]) + '/'
            onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
            if '.pdf' in onlyfiles:
                onlyfiles.remove('.pdf')
            for t in onlyfiles:
                shutil.copy(mypath + t, os.getcwd() + '/Attachments/')
            subprocess.run(["python", "master.py", insurer,op,mail_uid])
            move_master_to_master_insurer(mail_uid)
            if clean_f(insurer, k[3]):
                print(f'deleted {insurer} files.')


    elif insurer == 'Medi_Assist':
        for i in range(0,1):
            k = op.split(' ')
            subprocess.run(["python", "Medi_Assist.py",k[0],k[1],st,stp,k[2],k[3],mail_uid,subject])
            #os.system('python Medi_Assist.py ' + k[0] + ' ' + k[1] + ' ' + st + ' ' + stp + ' ' + k[2] + ' ' + k[3] + ' ' + mail_uid)
            mypath = os.getcwd() + '/Medi_Assist/attachments_' + str(k[3]) + '/'
            onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
            if '.pdf' in onlyfiles:
                onlyfiles.remove('.pdf')
            for t in onlyfiles:
                shutil.copy(mypath + t, os.getcwd() + '/Attachments/')
            subprocess.run(["python", "master.py", insurer,op,mail_uid])
            move_master_to_master_insurer(mail_uid)
            if clean_f(insurer, k[3]):
                print(f'deleted {insurer} files.')

    elif insurer == 'Paramount':
        for i in range(0,1):
            k = op.split(' ')
            subprocess.run(["python", "Paramount.py",k[0],k[1],st,stp,k[2],k[3],mail_uid,subject])
            # os.system('python Paramount.py ' + k[0] + ' ' + k[1] + ' ' + st + ' ' + stp + ' ' + k[2] + ' ' + k[
            #     3] + ' ' + mail_uid)
            mypath = os.getcwd() + '/Paramount/attachments_mail_' + str(k[3]) + '/'
            onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
            if '.pdf' in onlyfiles:
                onlyfiles.remove('.pdf')
            for t in onlyfiles:
                shutil.copy(mypath + t, os.getcwd() + '/Attachments/')
            subprocess.run(["python", "master.py", insurer,op,mail_uid])
            move_master_to_master_insurer(mail_uid)
            if clean_f(insurer, k[3]):
                print(f'deleted {insurer} files.')

    elif insurer == 'religare':
        for i in range(0,1):
            k = op.split(' ')
            subprocess.run(["python", "religare.py",k[0],k[1],st,stp,k[2],k[3],mail_uid,subject])
            #
            # os.system(
            #     'python religare.py ' + k[0] + ' ' + k[1] + ' ' + st + ' ' + stp + ' ' + k[2] + ' ' + k[3] + ' ' + mail_uid)
            mypath = os.getcwd() + '/religare/attachments_' + str(k[3]) + '/'
            onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
            if '.pdf' in onlyfiles:
                onlyfiles.remove('.pdf')
            for t in onlyfiles:
                shutil.copy(mypath + t, os.getcwd() + '/Attachments/')
            subprocess.run(["python", "master.py", insurer,op,mail_uid])
            move_master_to_master_insurer(mail_uid)
            if clean_f(insurer, k[3]):
                print(f'deleted {insurer} files.')

    elif insurer == 'united':
        for i in range(0,1):
            k = op.split(' ')
            subprocess.run(["python", "united.py",k[0],k[1],st,stp,k[2],k[3],mail_uid,subject])
            # os.system(
            #     'python united.py ' + k[0] + ' ' + k[1] + ' ' + st + ' ' + stp + ' ' + k[2] + ' ' + k[3] + ' ' + mail_uid)
            mypath = os.getcwd() + '/united/attachments_pdf_' + str(k[3]) + '/'
            onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
            if '.pdf' in onlyfiles:
                onlyfiles.remove('.pdf')
            for t in onlyfiles:
                shutil.copy(mypath + t, os.getcwd() + '/Attachments/')
            subprocess.run(["python", "master.py", insurer,op,mail_uid])
            move_master_to_master_insurer(mail_uid)
            if clean_f(insurer, k[3]):
                print(f'deleted {insurer} files.')

    elif insurer == 'vidal':
        for i in range(0,1):
            k = op.split(' ')
            subprocess.run(["python", "vidal.py",k[0],k[1],st,stp,k[2],k[3],mail_uid,subject])
            # os.system(
            #     'python vidal.py ' + k[0] + ' ' + k[1] + ' ' + st + ' ' + stp + ' ' + k[2] + ' ' + k[3] + ' ' + mail_uid)
            mypath = os.getcwd() + '/vidal/attachments_' + str(k[3]) + '/'
            onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
            if '.pdf' in onlyfiles:
                onlyfiles.remove('.pdf')
            for t in onlyfiles:
                shutil.copy(mypath + t, os.getcwd() + '/Attachments/')
            subprocess.run(["python", "master.py", insurer,op,mail_uid])
            move_master_to_master_insurer(mail_uid)
            if clean_f(insurer, k[3]):
                print(f'deleted {insurer} files.')

    elif insurer == 'Universal_Sompo':
        for i in range(0,1):
            k = op.split(' ')
            subprocess.run(["python", "universal_sompo.py",k[0],k[1],st,stp,k[2],k[3],mail_uid,subject])
            # os.system('python universal_sompo.py ' + k[0] + ' ' + k[1] + ' ' + st + ' ' + stp + ' ' + k[2] + ' ' + k[
            #     3] + ' ' + mail_uid)
            mypath = os.getcwd() + '/universal_sompo/attachments_' + str(k[3]) + '/'
            onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
            if '.pdf' in onlyfiles:
                onlyfiles.remove('.pdf')
            for t in onlyfiles:
                shutil.copy(mypath + t, os.getcwd() + '/Attachments/')
            subprocess.run(["python", "master.py", insurer,op,mail_uid])
            move_master_to_master_insurer(mail_uid)
            if clean_f(insurer, k[3]):
                print(f'deleted {insurer} files.')

    elif insurer == 'vipul':
        for i in range(0,1):
            k = op.split(' ')
            subprocess.run(["python", "vipul.py",k[0],k[1],st,stp,k[2],k[3],mail_uid,subject])
            # os.system(
            #     'python vipul.py ' + k[0] + ' ' + k[1] + ' ' + st + ' ' + stp + ' ' + k[2] + ' ' + k[3] + ' ' + mail_uid)
            mypath = os.getcwd() + '/vipul/attachments_' + str(k[3]) + '/'
            onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
            if '.pdf' in onlyfiles:
                onlyfiles.remove('.pdf')
            for t in onlyfiles:
                shutil.copy(mypath + t, os.getcwd() + '/Attachments/')
            subprocess.run(["python", "master.py", insurer,op,mail_uid])
            move_master_to_master_insurer(mail_uid)
            if clean_f(insurer, k[3]):
                print(f'deleted {insurer} files.')

    elif insurer == 'east_west':
        for i in range(0,1):
            k = op.split(' ')
            subprocess.run(["python", "East_West.py",k[0],k[1],st,stp,k[2],k[3],mail_uid,subject])
            # os.system('python Ease_West.py ' + k[0] + ' ' + k[1] + ' ' + st + ' ' + stp + ' ' + k[2] + ' ' + k[
            #     3] + ' ' + mail_uid)
            mypath = os.getcwd() + '/East_West/attachments_' + str(k[3]) + '/'
            onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
            if '.pdf' in onlyfiles:
                onlyfiles.remove('.pdf')
            for t in onlyfiles:
                shutil.copy(mypath + t, os.getcwd() + '/Attachments/')
            subprocess.run(["python", "master.py", insurer,op,mail_uid])
            move_master_to_master_insurer(mail_uid)
            if clean_f(insurer, k[3]):
                print(f'deleted {insurer} files.')

    elif insurer == 'Good_health':
        for i in range(0,1):
            k = op.split(' ')
            subprocess.run(["python", "Good_health.py",k[0],k[1],st,stp,k[2],k[3],mail_uid,subject])
            os.system('python Good_health.py ' + k[0] + ' ' + k[1] + ' ' + st + ' ' + stp + ' ' + k[2] + ' ' + k[
                3] + ' ' + mail_uid)
            mypath = os.getcwd() + '/Good_health/attachments_' + str(k[3]) + '/'
            onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
            if '.pdf' in onlyfiles:
                onlyfiles.remove('.pdf')
            for t in onlyfiles:
                shutil.copy(mypath + t, os.getcwd() + '/Attachments/')
            subprocess.run(["python", "master.py", insurer,op,mail_uid])
            move_master_to_master_insurer(mail_uid)
            if clean_f(insurer, k[3]):
                print(f'deleted {insurer} files.')

    elif insurer == 'Medsave':
        for i in range(0,1):
            k = op.split(' ')
            subprocess.run(["python", "medsave.py",k[0],k[1],st,stp,k[2],k[3],mail_uid,subject])
            # os.system(
            #     'python medsave.py ' + k[0] + ' ' + k[1] + ' ' + st + ' ' + stp + ' ' + k[2] + ' ' + k[3] + ' ' + mail_uid)
            mypath = os.getcwd() + '/medsave/attachments_' + str(k[3]) + '/'
            onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
            if '.pdf' in onlyfiles:
                onlyfiles.remove('.pdf')
            for t in onlyfiles:
                shutil.copy(mypath + t, os.getcwd() + '/Attachments/')
            subprocess.run(["python", "master.py", insurer,op,mail_uid])
            move_master_to_master_insurer(mail_uid)
            if clean_f(insurer, k[3]):
                print(f'deleted {insurer} files.')

    elif insurer == 'Raksha':
        for i in range(0,1):
            k = op.split(' ')
            subprocess.run(["python", "raksha.py",k[0],k[1],st,stp,k[2],k[3],mail_uid,subject])
            # os.system(
            #     'python raksha.py ' + k[0] + ' ' + k[1] + ' ' + st + ' ' + stp + ' ' + k[2] + ' ' + k[3] + ' ' + mail_uid)
            mypath = os.getcwd() + '/raksha/attachments_' + str(k[3]) + '/'
            onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
            if '.pdf' in onlyfiles:
                onlyfiles.remove('.pdf')
            for t in onlyfiles:
                shutil.copy(mypath + t, os.getcwd() + '/Attachments/')
            subprocess.run(["python", "master.py", insurer,op,mail_uid])
            move_master_to_master_insurer(mail_uid)
            if clean_f(insurer, k[3]):
                print(f'deleted {insurer} files.')

    elif insurer == 'reliance':
        for i in range(0,1):
            k = op.split(' ')
            subprocess.run(["python", "reliance.py",k[0],k[1],st,stp,k[2],k[3],mail_uid,subject])
            # os.system(
            #     'python reliance.py ' + k[0] + ' ' + k[1] + ' ' + st + ' ' + stp + ' ' + k[2] + ' ' + k[3] + ' ' + mail_uid)
            mypath = os.getcwd() + '/reliance/attachments_' + str(k[3]) + '/'
            onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
            if '.pdf' in onlyfiles:
                onlyfiles.remove('.pdf')
            for t in onlyfiles:
                shutil.copy(mypath + t, os.getcwd() + '/Attachments/')
            subprocess.run(["python", "master.py", insurer,op,mail_uid])
            move_master_to_master_insurer(mail_uid)
            if clean_f(insurer, k[3]):
                print(f'deleted {insurer} files.')

    elif insurer == 'health_insurance':
        for i in range(0,1):
            k = op.split(' ')
            subprocess.run(["python", "health_insurance.py",k[0],k[1],st,stp,k[2],k[3],mail_uid,subject])
            # os.system('python health_insurance.py ' + k[0] + ' ' + k[1] + ' ' + st + ' ' + stp + ' ' + k[2] + ' ' + k[
            #     3] + ' ' + mail_uid)
            mypath = os.getcwd() + '/health_insurance/attachments_' + str(k[3]) + '/'
            onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
            if '.pdf' in onlyfiles:
                onlyfiles.remove('.pdf')
            for t in onlyfiles:
                shutil.copy(mypath + t, os.getcwd() + '/Attachments/')
            subprocess.run(["python", "master.py", insurer,op,mail_uid])
            move_master_to_master_insurer(mail_uid)
            if clean_f(insurer, k[3]):
                print(f'deleted {insurer} files.')

    elif insurer == 'bajaj':
        for i in range(0,1):
            k = op.split(' ')
            subprocess.run(["python", "bajaj.py",k[0],k[1],st,stp,k[2],k[3],mail_uid,subject])
            # os.system('python health_insurance.py ' + k[0] + ' ' + k[1] + ' ' + st + ' ' + stp + ' ' + k[2] + ' ' + k[
            #     3] + ' ' + mail_uid)
            mypath = os.getcwd() + '/bajaj/attachments_pdf_' + str(k[3]) + '/'
            onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
            if '.pdf' in onlyfiles:
                onlyfiles.remove('.pdf')
            for t in onlyfiles:
                shutil.copy(mypath + t, os.getcwd() + '/Attachments/')
            subprocess.run(["python", "master.py", insurer,op,mail_uid])
            move_master_to_master_insurer(mail_uid)
            if clean_f(insurer, k[3]):
                print(f'deleted {insurer} files.')
    now = datetime.datetime.now()
    subprocess.run(["python", "updation.py","1","max","6",str(now)])

except:
    log_exceptions()

###########################################################################
# for i in range(0,1):
# 	k=op[i].split(' ')
# 	#print(k)
# 	os.system('python aditya_birla.py '+k[0]+' '+k[1]+' '+st+' '+stp+' '+k[2]+' '+k[3])
# 	mypath = os.getcwd()+'/aditya_birla/attachments_'+str(k[3])+'/'
# 	onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
# 	for t in onlyfiles:
# 		shutil.copy(mypath+t,os.getcwd()+'/Attachments/')
#
# for i in range(0,1):
# 	k=op[i].split(' ')
# 	os.system('python apollo_munich.py '+k[0]+' '+k[1]+' '+st+' '+stp+' '+k[2]+' '+k[3])
# 	mypath = os.getcwd()+'/apollo_munich/attachments_'+str(k[3])+'/'
# 	onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
# 	for t in onlyfiles:
# 		shutil.copy(mypath+t,os.getcwd()+'/Attachments/')
#
# for i in range(0,1):
# 	k=op[i].split(' ')
# 	os.system('python big.py '+k[0]+' '+k[1]+' '+st+' '+stp+' '+k[2]+' '+k[3])
# 	mypath = os.getcwd()+'/star/attachments_2_'+str(k[3])+'/'
# 	onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
# 	for t in onlyfiles:
# 		shutil.copy(mypath+t,os.getcwd()+'/Attachments/')
#
# for i in range(0,1):
# 	k=op[i].split(' ')
# 	os.system('python small.py '+k[0]+' '+k[1]+' '+st+' '+stp+' '+k[2]+' '+k[3])
# 	mypath = os.getcwd()+'/star/attachments_1_'+str(k[3])+'/'
# 	onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
# 	for t in onlyfiles:
# 		shutil.copy(mypath+t,os.getcwd()+'/Attachments/')
#
# for i in range(0,1):
# 	k=op[i].split(' ')
# 	os.system('python fgh.py '+k[0]+' '+k[1]+' '+st+' '+stp+' '+k[2]+' '+k[3])
# 	mypath = os.getcwd()+'/FGH/attachments_'+str(k[3])+'/'
# 	onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
# 	for t in onlyfiles:
# 		shutil.copy(mypath+t,os.getcwd()+'/Attachments/')
# print("\nfgh")
#
# for i in range(0,1):
# 	k=op[i].split(' ')
# 	os.system('python MDINDIA.py '+k[0]+' '+k[1]+' '+st+' '+stp+' '+k[2]+' '+k[3])
# 	mypath = os.getcwd()+'/MDINDIA/attachments_'+str(k[3])+'/'
# 	onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
# 	for t in onlyfiles:
# 		shutil.copy(mypath+t,os.getcwd()+'/Attachments/')
# print("\nMDINDIA")
# for i in range(0,1):
# 	k=op[i].split(' ')
# 	os.system('python fhpl.py '+k[0]+' '+k[1]+' '+st+' '+stp+' '+k[2]+' '+k[3])
# 	mypath = os.getcwd()+'/fhpl/attachments_'+str(k[3])+'/'
# 	onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
# 	for t in onlyfiles:
# 		shutil.copy(mypath+t,os.getcwd()+'/Attachments/')
#
# for i in range(0,1):
# 	k=op[i].split(' ')
# 	os.system('python hdfc.py '+k[0]+' '+k[1]+' '+st+' '+stp+' '+k[2]+' '+k[3])
# 	mypath = os.getcwd()+'/hdfc/attachments_'+str(k[3])+'/'
# 	onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
# 	for t in onlyfiles:
# 		shutil.copy(mypath+t,os.getcwd()+'/Attachments/')
#
# for i in range(0,1):
# 	k=op[i].split(' ')
# 	os.system('python health_heritage.py '+k[0]+' '+k[1]+' '+st+' '+stp+' '+k[2]+' '+k[3])
# 	mypath = os.getcwd()+'/health_heritage/attachments_'+str(k[3])+'/'
# 	onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
# 	for t in onlyfiles:
# 		shutil.copy(mypath+t,os.getcwd()+'/Attachments/')
#
# for i in range(0,1):
# 	k=op[i].split(' ')
# 	os.system('python health_india.py '+k[0]+' '+k[1]+' '+st+' '+stp+' '+k[2]+' '+k[3])
# 	mypath = os.getcwd()+'/health_india/attachments_'+str(k[3])+'/'
# 	onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
# 	for t in onlyfiles:
# 		shutil.copy(mypath+t,os.getcwd()+'/Attachments/')
#
# for i in range(0,1):
# 	k=op[i].split(' ')
# 	os.system('python Medi_Assist.py '+k[0]+' '+k[1]+' '+st+' '+stp+' '+k[2]+' '+k[3])
# 	mypath = os.getcwd()+'/Medi_Assist/attachments_'+str(k[3])+'/'
# 	onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
# 	for t in onlyfiles:
# 		shutil.copy(mypath+t,os.getcwd()+'/Attachments/')
#
# for i in range(0,1):
# 	k=op[i].split(' ')
# 	os.system('python Paramount.py '+k[0]+' '+k[1]+' '+st+' '+stp+' '+k[2]+' '+k[3])
# 	mypath = os.getcwd()+'/Paramount/attachments_mail_'+str(k[3])+'/'
# 	onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
# 	for t in onlyfiles:
# 		shutil.copy(mypath+t,os.getcwd()+'/Attachments/')
# '''
# for i in range(0,1):
# 	k=op[i].split(' ')
# 	os.system('python icici.py '+k[0]+' '+k[1]+' '+st+' '+stp+' '+k[2]+' '+k[3])
# 	mypath = os.getcwd()+'/icici/attachments_pdf_'+str(k[3])+'/'
# 	onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
# 	for t in onlyfiles:
# 		shutil.copy(mypath+t,os.getcwd()+'/Attachments/')
# '''
# for i in range(0,1):
# 	k=op[i].split(' ')
# 	os.system('python religare.py '+k[0]+' '+k[1]+' '+st+' '+stp+' '+k[2]+' '+k[3])
# 	mypath = os.getcwd()+'/religare/attachments_'+str(k[3])+'/'
# 	onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
# 	for t in onlyfiles:
# 		shutil.copy(mypath+t,os.getcwd()+'/Attachments/')
#
# for i in range(0,1):
# 	k=op[i].split(' ')
# 	os.system('python united.py '+k[0]+' '+k[1]+' '+st+' '+stp+' '+k[2]+' '+k[3])
# 	mypath = os.getcwd()+'/united/attachments_pdf_'+str(k[3])+'/'
# 	onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
# 	for t in onlyfiles:
# 		shutil.copy(mypath+t,os.getcwd()+'/Attachments/')
#
# for i in range(0,1):
# 	k=op[i].split(' ')
# 	os.system('python vidal.py '+k[0]+' '+k[1]+' '+st+' '+stp+' '+k[2]+' '+k[3])
# 	mypath = os.getcwd()+'/vidal/attachments_'+str(k[3])+'/'
# 	onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
# 	for t in onlyfiles:
# 		shutil.copy(mypath+t,os.getcwd()+'/Attachments/')
#
# for i in range(0,1):
# 	k=op[i].split(' ')
# 	os.system('python universal_sompo.py '+k[0]+' '+k[1]+' '+st+' '+stp+' '+k[2]+' '+k[3])
# 	mypath = os.getcwd()+'/universal_sompo/attachments_'+str(k[3])+'/'
# 	onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
# 	for t in onlyfiles:
# 		shutil.copy(mypath+t,os.getcwd()+'/Attachments/')
#
# for i in range(0,1):
# 	k=op[i].split(' ')
# 	os.system('python vipul.py '+k[0]+' '+k[1]+' '+st+' '+stp+' '+k[2]+' '+k[3])
# 	mypath = os.getcwd()+'/vipul/attachments_'+str(k[3])+'/'
# 	onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
# 	for t in onlyfiles:
# 		shutil.copy(mypath+t,os.getcwd()+'/Attachments/')
#
# for i in range(0,1):
# 	k=op[i].split(' ')
# 	os.system('python Ease_West.py '+k[0]+' '+k[1]+' '+st+' '+stp+' '+k[2]+' '+k[3])
# 	mypath = os.getcwd()+'/Ease_West/attachments_'+str(k[3])+'/'
# 	onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
# 	for t in onlyfiles:
# 		shutil.copy(mypath+t,os.getcwd()+'/Attachments/')
#
# for i in range(0,1):
# 	k=op[i].split(' ')
# 	os.system('python Good_health.py '+k[0]+' '+k[1]+' '+st+' '+stp+' '+k[2]+' '+k[3])
# 	mypath = os.getcwd()+'/Good_health/attachments_'+str(k[3])+'/'
# 	onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
# 	for t in onlyfiles:
# 		shutil.copy(mypath+t,os.getcwd()+'/Attachments/')
#
# for i in range(0,1):
# 	k=op[i].split(' ')
# 	os.system('python medsave.py '+k[0]+' '+k[1]+' '+st+' '+stp+' '+k[2]+' '+k[3])
# 	mypath = os.getcwd()+'/medsave/attachments_'+str(k[3])+'/'
# 	onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
# 	for t in onlyfiles:
# 		shutil.copy(mypath+t,os.getcwd()+'/Attachments/')
#
# for i in range(0,1):
# 	k=op[i].split(' ')
# 	os.system('python raksha.py '+k[0]+' '+k[1]+' '+st+' '+stp+' '+k[2]+' '+k[3])
# 	mypath = os.getcwd()+'/raksha/attachments_'+str(k[3])+'/'
# 	onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
# 	for t in onlyfiles:
# 		shutil.copy(mypath+t,os.getcwd()+'/Attachments/')
#
# for i in range(0,1):
# 	k=op[i].split(' ')
# 	os.system('python reliance.py '+k[0]+' '+k[1]+' '+st+' '+stp+' '+k[2]+' '+k[3])
# 	mypath = os.getcwd()+'/reliance/attachments_'+str(k[3])+'/'
# 	onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
# 	for t in onlyfiles:
# 		shutil.copy(mypath+t,os.getcwd()+'/Attachments/')
#
# for i in range(0,1):
# 	k=op[i].split(' ')
# 	os.system('python health_insurance.py '+k[0]+' '+k[1]+' '+st+' '+stp+' '+k[2]+' '+k[3])
# 	mypath = os.getcwd()+'/health_insurance/attachments_'+str(k[3])+'/'
# 	onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
# 	for t in onlyfiles:
# 		shutil.copy(mypath+t,os.getcwd()+'/Attachments/')
