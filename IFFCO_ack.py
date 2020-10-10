import os
import sys
import struct, time
import subprocess
from datetime import date
import datetime
import openpyxl
import pdftotext
import time
import requests
now = datetime.datetime.now()


subprocess.run(["python", "updation.py","1","max1","1",sys.argv[2]])
subprocess.run(["python", "updation.py","1","max","2",sys.argv[3]])
subprocess.run(["python", "updation.py","1","max","3",sys.argv[4]])
subprocess.run(["python", "updation.py","1","max","5",str(now)])
subprocess.run(["python", "updation.py","1","max","7",sys.argv[5]])
subprocess.run(["python", "updation.py","1","max","8",sys.argv[6]])

with open(sys.argv[1], "rb") as f:
	pdf = pdftotext.PDF(f)

with open('IFFCO/output1.txt', 'w') as f:
	f.write(" ".join(pdf))     
with open('IFFCO//output1.txt', 'r') as myfile:
 	f = myfile.read()
    #print("d")

try:		
	hg=[]
		
	w=f.find('PATIENT NAME')+13
	g=f[w:]
	u=g.find("\n")+w
	hg.append(f[w:u])
	w=f.find('POLICY NO. ')+12
	g=f[w:]
	u=g.find("\n")+w
	hg.append(f[w:u])
	w=f.find('Incidence no')+13
	g=f[w:]
	u=g.find('.')+w
	hg.append(f[w:u])
	w=f.find('MEMBER ID. ')+11
	g=f[w:]
	u=g.find("\n")+w
	hg.append(f[w:u])
	hg=[sub.replace(':','') for sub in hg]	
	hg=[sub.replace('  ','') for sub in hg]
	hg=[sub.replace('Rs.','') for sub in hg]

	subprocess.run(["python", "updation.py","1","max","9",'Yes'])
	subprocess.run(["python", "updation.py","1","max","10",'NA'])
	
	try:
		subprocess.run(["python", "test_api.py",hg[2],'',hg[1],'','Acknowledgement',sys.argv[6],sys.argv[1],'',hg[3],hg[0]])
		'''wbk= openpyxl.load_workbook(wbkName)
		s2=wbk.worksheets[1]
		s2.cell(row_count_1+1, column=11).value='YES'
		'''
		subprocess.run(["python", "updation.py","1","max","11",'Yes'])	
	except Exception as e:
		#s2.cell(row_count_1+1, column=11).value='NO'
		subprocess.run(["python", "updation.py","1","max","11",'No'])
except Exception as e:
	#s2.cell(row_count_1+1, column=9).value='No'
	#s2.cell(row_count_1+1, column=11).value='NO'
	subprocess.run(["python", "updation.py","1","max","9",'Yes'])
	subprocess.run(["python", "updation.py","1","max","11",'No'])
now = datetime.datetime.now()
#s2.cell(row_count_1+1, column=6).value=now
#wbk.save(wbkName)
subprocess.run(["python", "updation.py","1","max","6",str(now)])

	