import os
import sys
import struct, time
import subprocess
from datetime import date
import datetime
import openpyxl
import pdftotext
import time
import requests

from patient_name_fun import pname_fun

now = datetime.datetime.now()
'''
wbkName = 'log file.xlsx'
wbk= openpyxl.load_workbook(wbkName)
s1=wbk.worksheets[0]
s2=wbk.worksheets[1]
row_count_1 = s2.max_row
s2.cell(row_count_1+1, column=1).value=sys.argv[2]
s2.cell(row_count_1+1, column=2).value=sys.argv[3]
s2.cell(row_count_1+1, column=3).value=sys.argv[4]
s2.cell(row_count_1+1, column=5).value=now
s2.cell(row_count_1+1, column=7).value=sys.argv[5]
s2.cell(row_count_1+1, column=8).value=sys.argv[6]
'''

subprocess.run(["python", "updation.py","1","max1","1",sys.argv[2]])
subprocess.run(["python", "updation.py","1","max","2",sys.argv[3]])
subprocess.run(["python", "updation.py","1","max","3",sys.argv[4]])
subprocess.run(["python", "updation.py","1","max","5",str(now)])
subprocess.run(["python", "updation.py","1","max","7",sys.argv[5]])
subprocess.run(["python", "updation.py","1","max","8",sys.argv[6]])

# Added by Ashish
with open(sys.argv[1], "rb") as f:
	pdf = pdftotext.PDF(f)

# with open("fhpl.pdf", "rb") as f:
# 	pdf = pdftotext.PDF(f)

# Added by Ashish
with open('fhpl/output1.txt', 'w') as f:
	f.write(" ".join(pdf))     
with open('fhpl/output1.txt', 'r') as myfile:
 	f = myfile.read()

# with open('hdfc/output1.txt', 'w') as f:
# 	f.write(" ".join(pdf))     
# with open('hdfc/output1.txt', 'r') as myfile:
#  	f = myfile.read()



try:
	w=f.find('Subject: Cashless Denial Of')+8
	k=f[w:]
	u=k.find(":")+w
	g=f[w:u]
	kk=g.strip()
	kk=str(kk)
	yy=str('Cashless Denial Of')
	if kk == yy:
		hg=[]
		w=f.find('Preauth ID:')+12
		k=f[w:]
		u=k.find("\n")+w
		g=f[w:u]
		kk=g.strip()
		hg.append(kk)

		w=f.find('UHID')+5
		k=f[w:]
		u=k.find("\n")+w
		g=f[w:u]
		kk=g.strip()
		kk=kk[5:13]
		hg.append(kk)

		w=f.find('Cashless Denial Of')+19
		k=f[w:]
		u=k.find(" with Card No")+w
		g=f[w:u]
		kk=g.strip()
		hg.append(kk)
	else:
		hg=[]
		w=f.find('PA.No ')+8
		k=f[w:]
		u=k.find("\n")+w
		g=f[w:u]
		kk=g.strip()
		ch=""
		for i in kk:
			if i is '/':
				break
			else:
				ch+=i
		hg.append(ch)

		w=f.find('Card No')+10
		k=f[w:]
		u=k.find("\n")+w
		g=f[w:u]
		pp=g.strip()
		pp=pp[5:]
		hg.append(pp)

		w=f.find('for the following reason.')+26
		k=f[w:]
		u=k.find("For any further queries")+w
		g=f[w:u]
		hg.append(g.strip())

		subprocess.run(["python", "updation.py","1","max","9",'Yes'])
	subprocess.run(["python", "updation.py","1","max","10",'NA'])
	
	try:
		regex_list = [r"(?<=Patient Name).*(?=Insurer Name)"]
		pname = pname_fun(f, regex_list)
		subprocess.run(["python", "test_api.py",hg[0],'','','','Denial',sys.argv[6],sys.argv[1],hg[1],hg[2], pname])
		'''wbk= openpyxl.load_workbook(wbkName)
		s2=wbk.worksheets[1]
		s2.cell(row_count_1+1, column=11).value='YES'
		'''
		subprocess.run(["python", "updation.py","1","max","11",'Yes'])	
	except Exception as e:
		#s2.cell(row_count_1+1, column=11).value='NO'
		subprocess.run(["python", "updation.py","1","max","11",'No'])
except Exception as e:
	#s2.cell(row_count_1+1, column=9).value='No'
	#s2.cell(row_count_1+1, column=11).value='NO'
	subprocess.run(["python", "updation.py","1","max","9",'Yes'])
	subprocess.run(["python", "updation.py","1","max","11",'No'])
now = datetime.datetime.now()
#s2.cell(row_count_1+1, column=6).value=now
#wbk.save(wbkName)
subprocess.run(["python", "updation.py","1","max","6",str(now)])
