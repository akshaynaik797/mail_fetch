import pandas as pd
from email.mime.text import MIMEText
import smtplib
import time
import sys
import imaplib
import email
import camelot
import PyPDF2
import csv
import xlsxwriter
from xlrd import open_workbook
import xlwt
import os
import glob
import os.path
import xlrd
from os import listdir
from os import path
from os.path import isfile, join
from html.parser import HTMLParser
import pdfkit
import pandas as pd
import pdftotext
import html2text
from openpyxl.styles import Color, PatternFill, Font, Border
import subprocess
from decode_error import check_subject, read_from_delete
from make_log import log_exceptions

try:
    subprocess.run(["python", "updation.py", "1", "max", "9", "X"])

    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')
    fg = []
    eu = []
    path_wkhtmltopdf = r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe'
    config = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf)
    # config = pdfkit.configuration(wkhtmltopdf='/usr/bin/wkhtmltopdf')



    def read_email_from_gmail():
        b = 0
        SMTP_SERVER = str(sys.argv[5])
        mail = imaplib.IMAP4_SSL(SMTP_SERVER)
        # mail.login(user = 'Mediclaim@therisingmedicare.com', password = 'cef@2018')

        e_id = str(sys.argv[1])
        pswd = str(sys.argv[2])
        srt = str(sys.argv[3])
        stp = str(sys.argv[4])
        mail.login(user=e_id, password=pswd)
        mail.select("inbox", readonly=True)
        ###############################################<
        mail_uid = str(sys.argv[7])
        if mail_uid == -1:
            type, data = mail.search(None, '(SUBJECT "Claims settlement-NEFT" since ' + srt + ' before ' + stp + ')')
            ids = data[0]
            id_list = ids.split()
        else:
            ids = mail_uid  # data is a list.
            # accept id from outside and put in id_list akshay var name = id

            id_list = []  # ids is a space separated string
            id_list.append(ids)
        ###############################################>
        # type, data = mail.search(None,
        #                          '(FROM "claim.support@apollomunichinsurance.com" SUBJECT "Claim Settlement for Claim ID" since ' + srt + ' before ' + stp + ')')
        # ids = data[0]  # data is a list.
        # id_list = ids.split()  # ids is a space separated string
        # # print(id_list)
        for i in range(0, len(id_list)):
            latest_email_id = id_list[i]  # get the latest
            result, data = mail.fetch(latest_email_id, "(RFC822)")
            ##################################################ak
            try:
                # raw_email = data[0][1].decode('utf-8')
                try:
                    raw_email = data[0][1].decode('utf-8')
                except UnicodeDecodeError:
                    try:
                        raw_email = data[0][1].decode('ISO-8859-1')
                    except UnicodeDecodeError:
                        try:
                            raw_email = data[0][1].decode('ascii')
                        except UnicodeDecodeError:
                            pass
                email_message = email.message_from_string(raw_email)
                subject = email_message['Subject']
                result, sys.argv[8] = check_subject(subject, sys.argv[8], mail)
                if result == 'Changed':
                    # raise Exception('subject not matched')
                    raise Exception('subject not matched', )
            except:
                try:
                    log_exceptions(syssubject=sys.argv[8], subject=subject, error='subject not matched')
                except:
                    pass
                if result != 'OK':
                    data = {'server': SMTP_SERVER,
                            'hospmail': e_id,
                            'pass': pswd,
                            'subject': sys.argv[8]}
                    try:
                        data = read_from_delete(data)
                        if data == None:
                            raise Exception("Not found")
                    except:
                        log_exceptions(msg='not found in deleted', subject=sys.argv[8])
            ##################################################akend
            raw_email = data[0][1].decode('utf-8')
            email_message = email.message_from_string(raw_email)
            # if path.exists(r'/home/shivam/Desktop/vnu_scripts/Paramount/email.html'):
            # os.remove(r'email.html')
            # Body details
            if email_message['Subject'] not in fg:
                b += 1
                for part in email_message.walk():
                    # print(part.get_content_type())
                    if part.get_content_type() == "text/html":
                        # print('hi')
                        body = part.get_payload(decode=True)
                        file_name = "apollo_munich/email.html"
                        output_file = open(file_name, 'w')
                        output_file.write("Body: %s" % (body.decode('utf-8')))
                        output_file.close()

                        pdfkit.from_file('apollo_munich/email.html',
                                         'apollo_munich/attachments_' + str(sys.argv[6]) + '/' + str(b) + '.pdf',
                                         configuration=config)
            fg.append(email_message['Subject'])


    mypath = os.getcwd() + '/apollo_munich'
    if not path.exists(mypath):
        os.mkdir(mypath)
    if not path.exists(mypath + '/attachments_' + str(sys.argv[6])):
        os.mkdir(mypath + '/attachments_' + str(sys.argv[6]))
    mypath = os.getcwd() + '/apollo_munich/attachments_' + str(sys.argv[6]) + '/'
    for filename in os.listdir(mypath):
        file_path = os.path.join(mypath, filename)
        if os.path.isfile(file_path) or os.path.islink(file_path):
            os.unlink(file_path)
    read_email_from_gmail()
    # print(fg)
    ccn = []
    name = []
    uhid = []
    for i in fg:
        d = i.find('Claim ID:') + 9
        k = i.find('of')
        ccn.append(i[d:k])
        d = i.find('covered')
        k = i.find('of') + 2
        name.append(i[k:d])
        d = i.find('UHID No:') + 9
        k = i.find('.')
        uhid.append(i[d:k])
    # print(ccn,uhid)
    onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
    onlyfiles.sort()
    # print(onlyfiles)
    if path.exists(r'apollo_munich/apollo' + str(sys.argv[6]) + '.xlsx'):
        os.remove(r'apollo_munich/apollo' + str(sys.argv[6]) + '.xlsx')
    import openpyxl

    po = []
    wq = 0
    wbkName = 'apollo_munich/apollo' + str(sys.argv[6]) + '.xlsx'
    wbk = openpyxl.Workbook()
    wbk.create_sheet('1')
    s1 = wbk.worksheets[0]
    s2 = wbk.worksheets[1]
    for t in range(0, len(onlyfiles)):
        sh1 = ['Sr No.', 'Preauth Id', 'Hospital Name', 'Claimed Amount', 'Diagnosis', 'Billed Amount', 'Date of Admission',
               'settled Amount', 'Date of Discharge', 'Cheque Amount', 'Cheque Number/NEFT reference', 'Disallowed Amount',
               'Cheque/NEFT date', 'Discount Amount', 'TDS Amount', 'IP Number', 'Bill No', 'ccn', 'uhid', 'patient name']
        sh2 = ['Sr No.', 'Claim ID', 'category', 'Disallowance amount', 'Disallowance Reasons']

        for i in range(0, len(sh1)):
            s1.cell(row=1, column=i + 1).value = sh1[i]
        for i in range(0, len(sh2)):
            s2.cell(row=1, column=i + 1).value = sh2[i]
        tables = camelot.read_pdf(mypath + onlyfiles[t], pages='all', Line_scale=10)
        tables.export('apollo_munich/foo1.xls', f='excel')
        loc = ("apollo_munich/foo1.xls")
        wb = xlrd.open_workbook(loc)
        s = []
        sheet_3 = wb.sheet_by_index(0)
        sheet_3.cell_value(0, 0)

        for i in range(1, sheet_3.nrows):
            s.append(sheet_3.cell_value(i, 2))
            s.append(sheet_3.cell_value(i, 4))
        mid = s[-1]
        s.pop(-1)
        s = [sub.replace('\t', ' ') for sub in s]
        s = [sub.replace('Rs.', '') for sub in s]
        # print(s)
        s1.cell(row=t + 2, column=1).value = t + 1
        s1.cell(row=t + 2, column=2).value = mid
        for i in range(0, len(s)):
            s1.cell(row=t + 2, column=i + 3).value = s[i]
        s1.cell(row=t + 2, column=i + 4).value = ccn[t]
        s1.cell(row=t + 2, column=i + 5).value = uhid[t]
        s1.cell(row=t + 2, column=i + 6).value = name[t]
        with open(mypath + onlyfiles[t], "rb") as f:
            pdf = pdftotext.PDF(f)

        with open('apollo_munich/output.txt', 'w') as f:
            f.write(" ".join(pdf))
        with open('apollo_munich/output.txt', 'r') as myfile:
            f = myfile.read()
        hg = []
        w = f.find('Disallowance Reasons :') + 22
        u = f.find('Please note')
        g = f[w:u]
        sy = g.split('\n')
        sy.pop(0)
        sy.pop(-1)
        for i in sy:
            # print(i)
            if (i.find(':') != -1):
                k = i
                k = k.replace(':', '')
                continue
            else:
                w1 = i.find('Rs.') + 3
                g = i[w1:]
                u1 = g.find('.') + w1 + 3
                m = i[w1:u1]
                h = i[u1:]
            row_num = s2.max_row + 1
            wq += 1
            s2.cell(row=row_num, column=1).value = wq
            s2.cell(row=row_num, column=2).value = mid
            s2.cell(row=row_num, column=3).value = k
            s2.cell(row=row_num, column=4).value = m
            s2.cell(row=row_num, column=5).value = h
        # print(s)
        os.rename(os.getcwd() + '/apollo_munich/attachments_' + str(sys.argv[6]) + '/' + onlyfiles[t],
                  os.getcwd() + '/apollo_munich/attachments_' + str(sys.argv[6]) + '/' + mid + '.pdf')
        '''w1=f.find('Reference/UTR No.')+17
        g=f[w1:]
        u1=g.find('\n')+w1
        hg.append(f[w1:u1])
    
        w2=f.find('Payment Date')+13
        g=f[w2:]
        u2=g.find('\n')+w2
        hg.append(f[w2:u2])
        
    
        w9=f.find('ID Card Number :')+17
        g=f[w9:]
        u9=g.find('\n')+w9
        hg.append(f[w9:u9])
    
        hg=[sub.replace('  ','') for sub in hg]
        hg=[sub.replace(':','') for sub in hg]
        hg=[sub.replace('\n',' ') for sub in hg]'''
        # print(sy[1].find(':'))
        '''except Exception as e:
            eu.append(t)
            print(onlyfiles[t],e)
    for t in range(0,len(onlyfiles)):
        if t in eu:
            s1.cell(row=t+1, column=1).fill = redFill'''
    print("Done")
    wbk.save(wbkName)
    wbk.close
    wbkName = 'count/count.xlsx'
    wbk = openpyxl.load_workbook(wbkName)
    s1 = wbk.worksheets[0]
    s1.cell(row=3, column=1).value = 'apollo munich'
    s1.cell(row=3, column=2).value = len(fg)
    s1.cell(row=3, column=3).value = len(onlyfiles)
    wbk.save(wbkName)
    wbk.close
    subprocess.run(["python", "updation.py", "1", "max", "9", " "])
except:
    log_exceptions()