import os
import sys
import struct, time
import subprocess
from datetime import date
import datetime
import openpyxl
import pdftotext
import time
import requests
now = datetime.datetime.now()

subprocess.run(["python", "updation.py","1","max1","1",sys.argv[2]])
subprocess.run(["python", "updation.py","1","max","2",sys.argv[3]])
subprocess.run(["python", "updation.py","1","max","3",sys.argv[4]])
subprocess.run(["python", "updation.py","1","max","5",str(now)])
subprocess.run(["python", "updation.py","1","max","7",sys.argv[5]])
subprocess.run(["python", "updation.py","1","max","8",sys.argv[6]])

with open(sys.argv[1], "rb") as f:
    pdf = pdftotext.PDF(f)
# added by ashish
with open('Max_Bupa/output1.txt', 'w') as f:
    f.write(" ".join(pdf))     
with open('Max_Bupa/output1.txt', 'r') as myfile:
    f = myfile.read()

# with open('max/output1.txt', 'w') as f:
#   f.write(" ".join(pdf))     
# with open('max/output1.txt', 'r') as myfile:
#   f = myfile.read()

try:
    hg=[]
    if f.find("Policy No.")!=-1:
        w=f.find('Policy No.')+10
        k=f[w:]
        hg.append(f[w:w+50])
    else:
        print("Policy No. is not found")
        hg.append('')

    if f.find("Cashless Claim Number")!=-1:   
        w=f.find('Cashless Claim Number')+21
        k=f[w:]
        u=k.find('(')+w
        hg.append(f[w:u])
    else:
        print("Cashless Claim Number is not found")
        hg.append('')

    if f.find("Patient Name")!=-1:
        w=f.find('Patient Name')+12
        k=f[w:]
        hg.append(f[w:w+50])
    else:
        print("Patient Name is not found")
        if f.find("in respect of")!=-1:
            startPos=f.find('in respect of')+len('in respect of')
            k=f[startPos:]
            endPos=k.find('we')
            p_name=k[0:endPos]
            hg.append(p_name) 


    if f.find("ID/TPA/Insurer Id of the Patient")!=-1:
        w=f.find('ID/TPA/Insurer Id of the Patient')+32
        k=f[w:]
        u=k.find('\n')+w
        hg.append(f[w:u])
    else:
        print("ID/TPA/Insurer Id of the Patient is not found")
        hg.append('')
  

    hg=[sub.replace('\n','') for sub in hg]
    hg=[sub.replace(':','') for sub in hg]  
    hg=[sub.replace('  ','') for sub in hg]
    hg=[sub.replace('Rs.','') for sub in hg]
    
    
    

    subprocess.run(["python", "updation.py","1","max","9",'Yes'])
    subprocess.run(["python", "updation.py","1","max","10",'NA'])

    try:
          subprocess.run(["python", "test_api.py",hg[1],'',hg[0],'','Denial',sys.argv[6],sys.argv[1],'',hg[3],hg[2]])
          subprocess.run(["python", "updation.py","1","max","11",'Yes'])    
    except Exception as e:
        #s2.cell(row_count_1+1, column=11).value='NO'
          subprocess.run(["python", "updation.py","1","max","11",'No'])
except Exception as e:
    #s2.cell(row_count_1+1, column=9).value='No'
    #s2.cell(row_count_1+1, column=11).value='NO'
    subprocess.run(["python", "updation.py","1","max","9",'Yes'])
    subprocess.run(["python", "updation.py","1","max","11",'No'])
now = datetime.datetime.now()
#s2.cell(row_count_1+1, column=6).value=now
#wbk.save(wbkName)
subprocess.run(["python", "updation.py","1","max","6",str(now)])





