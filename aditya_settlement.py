import pandas as pd
from email.mime.text import MIMEText
import smtplib
import time
import imaplib
import sys
import email
import camelot
import PyPDF2
import csv
import xlsxwriter
from xlrd import open_workbook
import xlwt
import os
import glob
import os.path
import xlrd
from os import listdir
from os import path
import pdftotext
from os.path import isfile, join
from html.parser import HTMLParser
import pdfkit
import pandas as pd
import html2text
if path.exists(r'aditya_birla/aditya'+str(sys.argv[6])+'.xlsx'):
	os.remove(r'aditya_birla/aditya'+str(sys.argv[6])+'.xlsx')
import openpyxl
po=[]
wbkName = 'aditya/aditya.xlsx'
wbk = openpyxl.Workbook()
wbk.create_sheet('1')
wbk.create_sheet('Sheet3')
s1=wbk.worksheets[0]
s2=wbk.worksheets[1]
s3=wbk.worksheets[2]
t=0
try:
	tables = camelot.read_pdf(sys.argv[1],line_scale=10)
	tables.export('aditya/foo1.xls', f='excel')
	loc = ("aditya/foo1.xls")
	with open(sys.argv[1], "rb") as f:
    		pdf = pdftotext.PDF(f)

	with open('aditya/output.txt', 'w') as f:
		f.write(" ".join(pdf))     
	with open('aditya/output.txt', 'r') as myfile:
	 	f = myfile.read()
	f=f.replace('\n', '$$ ')	
	#print(f)	

	wb = xlrd.open_workbook(loc)
	sh1=['Sum Insured','Claimed amount (Rs.)','AL Amount(in case of cashless)','Approved amount (Rs.)','Deduction amount (Rs.)','Hospital Amount','TDS','Discount Amount','Reason for deduction','Amount Utilised']
	for i in range (0,len(sh1)):
		s2.cell(row=1, column=i+3).value = sh1[i]
	sh2=['Policy Number','Member id','Patient Name','Policy Holder','Payee Bank name','Payee account number','Amount of transfer','UTR number','Diagnosis','doa','dod','Deduction amount','discount','transaction date']
	for i in range (0,len(sh2)):
		s1.cell(row=1, column=i+3).value = sh2[i]
	sh3=['Sr','CCN','Deduction amount','Deduction reason']
	for i in range (0,len(sh3)):
		s3.cell(row=1, column=i+1).value = sh3[i]
	hg=[]
	w=f.find('Policy Number')+15
	g=f[w:]
	u=g.find('$$')+w
	hg.append(f[w:u])

	w1=f.find('Member id')+11
	g=f[w1:]
	u1=g.find('$$')+w1
	hg.append(f[w1:u1])

	w2=f.find('Patient Name')+14
	g=f[w2:]
	u2=g.find('$$')+w2
	hg.append(f[w2:u2])

	w3=f.find('Policy Holder')+14
	g=f[w3:]
	u3=g.find('$$')+w3
	hg.append(f[w3:u3])

	w4=f.find('Payee Bank name')+17
	g=f[w4:]
	u4=g.find('$$')+w4
	hg.append(f[w4:u4])

	w5=f.find('Payee account number')+22
	g=f[w5:]
	u5=g.find('$$')+w5
	hg.append(f[w5:u5])

	w6=f.find('Amount of transfer')+25
	g=f[w6:]
	u6=g.find('$$')+w6
	hg.append(f[w6:u6])

	w7=f.find('UTR number')+12
	g=f[w7:]
	u7=g.find('$$')+w7
	hg.append(f[w7:u7])

	w8=f.find('Ailment Name')+14
	g=f[w8:]
	u8=g.find('Please note')+w8
	hg.append(f[w8:u8])

	w8=f.find('Date of Admission:')+19
	g=f[w8:]
	u8=g.find('$$')+w8
	hg.append(f[w8:u8])

	w8=f.find('Date of Discharge:')+19
	g=f[w8:]
	u8=g.find('$$')+w8
	hg.append(f[w8:u8])

	w8=f.find('Deduction amount')+23
	g=f[w8:]
	u8=g.find('$$')+w8
	hg.append(f[w8:u8])

	w8=f.find('Discount Amount')+26
	g=f[w8:]
	u8=g.find('$$')+w8
	hg.append(f[w8:u8])

	w8=f.find('Date of transfer :')+19
	g=f[w8:]
	u8=g.find('$$')+w8
	hg.append(f[w8:u8])

	w9=f.find('Claim registration number')+26
	g=f[w9:]
	u9=g.find('$$')+w9
	ccn=(f[w9:u9])
	hg=[sub.replace('  ','') for sub in hg]
	hg=[sub.replace('$$ ','') for sub in hg]
	#print(hg)
	for i in range (0,len(hg)):
		s1.cell(row=t+2, column=i+3).value = hg[i]

	for wd in wbk.worksheets[:2]:
		wd.cell(row=1, column=1).value = 'Sr. No.'
		wd.cell(row=1, column=2).value = 'CCN'
		wd.cell(row=t+2, column=1).value = t+1
		wd.cell(row=t+2, column=2).value = ccn

	sheet_2 = wb.sheet_by_index(0) 
	sheet_2.cell_value(0, 0) 
	b=[]

	for i in range(2,sheet_2.nrows): 
		b.append(sheet_2.cell_value(i,2))
	rd=b[8]
	rd=rd.replace('Rs','rs')
	rd=rd.replace(',RS','rs') 
	rd=rd.replace('\t',' ')
	rd=rd.replace('\n',' ')
	rd=rd.replace(',,',',')
	reason=rd.split('rs')
	deduct=[]
	deduct_res=[]
	for i in reason: 
		w8=i.find('rs')+2
		g=i[w8:]
		u8=g.find('/-')+w8
		deduct.append(i[w8:u8])
		w8=i.find('/-')+2
		deduct_res.append(i[w8:])
	temp_k=0
	for i in deduct_res:
		deduct_res[temp_k]=i[:-1]
		temp_k+=1
	#print(deduct,deduct_res,reason)
	for i in range(0,len(b)):
		s2.cell(row=t+2, column=i+3).value = b[i]
	for i in range(1,len(deduct)):			
		if(deduct[i]!=''):			
			row=s3.max_row+1
			s3.cell(row=row, column=1).value = row+1
			s3.cell(row=row, column=2).value = ccn
			s3.cell(row=row, column=3).value = deduct[i]
			s3.cell(row=row, column=4).value = deduct_res[i]
except Exception as e:
	s1.cell(row=t+2, column=1).value = 'error'
print("Done")
wbk.save(wbkName)
wbk.close
'''
Name = 'count/count.xlsx'
wbk = openpyxl.Workbook()
s1=wbk.worksheets[0]
s1.cell(row=1, column=1).value = 'insurance id'
s1.cell(row=1, column=2).value = 'hospital id'
s1.cell(row=1, column=3).value = 'mail count'
s1.cell(row=1, column=4).value = 'attachments count'
s1.cell(row=1, column=5).value = 'repeating mail'
row_=s1.max_row+1
s1.cell(row=row_, column=1).value = 'aditya birla'
s1.cell(row=row_, column=2).value = str(sys.argv[6])
s1.cell(row=row_, column=3).value = len(fg)
s1.cell(row=row_, column=4).value = len(onlyfiles)
s1.cell(row=row_, column=5).value = len(repeat)
wbk.save(Name)'''
wbk.close
