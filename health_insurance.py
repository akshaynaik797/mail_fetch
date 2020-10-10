import pandas as pd
from email.mime.text import MIMEText
import smtplib
import time
import reportlab
import imaplib
import sys
import email
import camelot
import PyPDF2
import csv
import xlsxwriter
from xlrd import open_workbook
import xlwt
import os
import glob
import os.path
import xlrd
from os import listdir
from os import path
import pdftotext
from os.path import isfile, join
from html.parser import HTMLParser
import pdfkit
import re
import subprocess
from decode_error import check_subject, read_from_delete
from make_log import log_exceptions

try:
    # from tabula import convert_into
    fg = []
    subprocess.run(["python", "updation.py", "1", "max", "9", "X"])
    repeat = []
    wq = 0


    def read_email_from_gmail():
        SMTP_SERVER = str(sys.argv[5])
        mail = imaplib.IMAP4_SSL(SMTP_SERVER)
        e_id = str(sys.argv[1])
        pswd = str(sys.argv[2])
        srt = str(sys.argv[3])
        stp = str(sys.argv[4])
        mail.login(user=e_id, password=pswd)
        mail.select("inbox", readonly=True)
        ###############################################<
        mail_uid = str(sys.argv[7])
        if mail_uid == -1:
            type, data = mail.search(None,
                                     '(FROM "customerservice@hitpa.co.in" SUBJECT "Claim Settlement" since ' + srt + ' before ' + stp + ')')
            ids = data[0]
            id_list = ids.split()
        else:
            ids = mail_uid  # data is a list.
            # accept id from outside and put in id_list akshay var name = id

            id_list = []  # ids is a space separated string
            id_list.append(ids)
        ###############################################>
        # type, data = mail.search(None,
        #                          '(FROM "customerservice@hitpa.co.in" SUBJECT "Claim Settlement" since ' + srt + ' before ' + stp + ')')
        # ids = data[0]  # data is a list.
        # id_list = ids.split()  # ids is a space separated string
        # # print(id_list)
        for i in range(0, len(id_list)):
            latest_email_id = id_list[i]  # get the latest
            result, data = mail.fetch(latest_email_id,
                                      "(RFC822)")  # fetch the email body (RFC822)             for the given ID

            ##################################################ak
            try:
                # raw_email = data[0][1].decode('utf-8')
                try:
                    raw_email = data[0][1].decode('utf-8')
                except UnicodeDecodeError:
                    try:
                        raw_email = data[0][1].decode('ISO-8859-1')
                    except UnicodeDecodeError:
                        try:
                            raw_email = data[0][1].decode('ascii')
                        except UnicodeDecodeError:
                            pass
                email_message = email.message_from_string(raw_email)
                subject = email_message['Subject']
                result, sys.argv[8] = check_subject(subject, sys.argv[8], mail)
                if result == 'Changed':
                    # raise Exception('subject not matched')
                    raise Exception('subject not matched', )
            except:
                try:
                    log_exceptions(syssubject=sys.argv[8], subject=subject, error='subject not matched')
                except:
                    pass
                if result != 'OK':
                    data = {'server': SMTP_SERVER,
                            'hospmail': e_id,
                            'pass': pswd,
                            'subject': sys.argv[8]}
                    try:
                        data = read_from_delete(data)
                        if data == None:
                            raise Exception("Not found")
                    except:
                        log_exceptions(msg='not found in deleted', subject=sys.argv[8])
            ##################################################akend

            raw_email = data[0][1].decode('utf-8')
            email_message = email.message_from_string(raw_email)
            if email_message['Subject'] not in fg:
                for part in email_message.walk():
                    if part.get_content_maintype() == 'multipart':
                        # print part.as_string()
                        continue
                    if part.get('Content-Disposition') is None:
                        # print part.as_string()
                        continue
                    fileName = part.get_filename()
                    detach_dir = (os.getcwd() + '/health_insurance/attachments_' + str(sys.argv[6]))
                    if bool(fileName):
                        filePath = os.path.join(detach_dir, fileName)
                        if not os.path.isfile(filePath):
                            from reportlab.pdfgen import canvas
                            c = canvas.Canvas(fileName + '.pdf')
                            print(fileName)
                            fp = open(filePath, 'wb')
                            fp.write(part.get_payload(decode=True))
                            fp.close()
            else:
                repeat.append(email_message['Subject'])

            fg.append(email_message['Subject'])


    mypath = os.getcwd() + '/health_insurance'
    if not path.exists(mypath):
        os.mkdir(mypath)
    if not path.exists(mypath + '/attachments_' + str(sys.argv[6])):
        os.mkdir(mypath + '/attachments_' + str(sys.argv[6]))
    mypath = os.getcwd() + '/health_insurance/attachments_' + str(sys.argv[6]) + '/'

    for filename in os.listdir(mypath):
        file_path = os.path.join(mypath, filename)
        if os.path.isfile(file_path) or os.path.islink(file_path):
            os.unlink(file_path)
    read_email_from_gmail()
    import sys

    for filename in os.listdir(mypath):
        file_path = os.path.join(mypath, filename)
        # print(filename)
        if filename.find('.pdf') == -1:
            # print(file_path)
            os.remove(file_path)
    onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
    if path.exists(r'health_insurance/health_insurance' + str(sys.argv[6]) + '.xlsx'):
        os.remove(r'health_insurance/health_insurance' + str(sys.argv[6]) + '.xlsx')
    import openpyxl

    po = []
    wbkName = 'health_insurance/health_insurance' + str(sys.argv[6]) + '.xlsx'
    wbk = openpyxl.Workbook()
    wbk.create_sheet('1')
    s1 = wbk.worksheets[0]
    s2 = wbk.worksheets[1]
    for t in range(0, len(onlyfiles)):
        try:
            sh1 = ['sr no', 'CCN', 'Insurer', 'Proposer Name', 'diagnosis', 'Policy Number', 'Patient Name', 'doa', 'dod',
                   'UHID', 'NEFT Date', 'NEFT No.', 'BilledAmount', 'disallowed', 'DiscountAmt', 'COPay', 'deduction',
                   'TDS', 'SettledAmount']
            for i in range(0, len(sh1)):
                s1.cell(row=1, column=i + 1).value = sh1[i]
            sh2 = ['sr no', 'CCN', 'Category', 'Sub Category', 'Requested Amount', 'Approved Amount', 'Deducted Amount',
                   'Reason']
            for i in range(0, len(sh2)):
                s2.cell(row=1, column=i + 1).value = sh2[i]
            # tables = camelot.read_pdf(mypath+onlyfiles[t],pages='all',line_scale=50)
            # tables.export('health_insurance/foo1.xls', f='excel')
            # loc = ("health_insurance/foo1.xls")
            with open(mypath + onlyfiles[t], "rb") as f:
                pdf = pdftotext.PDF(f)

            with open('health_insurance/output.txt', 'w') as f:
                f.write(" ".join(pdf))
            with open('health_insurance/output.txt', 'r') as myfile:
                f = myfile.read()
            hg = []
            w = f.find('claim number') + 12
            g = f[w:]
            x1 = g.find(':') + w
            u = g.find('\n') + w
            hg.append(f[x1 + 1:u])

            w = f.find('Insurer') + 7
            g = f[w:]
            x1 = g.find(':') + w
            u = g.find('Proposer Name') + w
            hg.append(f[x1 + 1:u])

            w = f.find('Proposer Name') + 13
            g = f[w:]
            x1 = g.find(':') + w
            u = g.find('Diagnosis') + w
            hg.append(f[x1 + 1:u])

            w = f.find('Diagnosis') + 9
            g = f[w:]
            x1 = g.find(':') + w
            u = g.find('Policy Number') + w
            hg.append(f[x1 + 1:u])

            w = f.find('Policy Number') + 13
            g = f[w:]
            x1 = g.find(':') + w
            u = g.find('Patient Name') + w
            hg.append(f[x1 + 1:u])

            w = f.find('Patient Name') + 12
            g = f[w:]
            x1 = g.find(':') + w
            u = g.find('Hospital Name') + w
            hg.append(f[x1 + 1:u])

            w = f.find('Period of Hospitalization') + 25
            g = f[w:]
            x1 = g.find(':') + w
            u = g.find('to') + w
            hg.append(f[x1 + 1:u])
            x1 = g.find('Patient’s Member UHID') + w
            hg.append(f[u + 2:x1])

            w = f.find('Patient’s Member UHID') + 22
            g = f[w:]
            x1 = g.find(':') + w
            u = g.find('\n') + w
            hg.append(f[x1 + 1:u])

            w = f.find('Date') + 5
            g = f[w:]
            u = g.find('\n') + w
            hg.append(f[w:u])

            w = f.find('Neft/Cheque number') + 18
            g = f[w:]
            u = g.find('\n') + w
            hg.append(f[w:u])

            w = f.find('Billed Amount') + 12
            g = f[w:]
            x1 = g.find(':') + w
            u = g.find('\n') + w
            hg.append(f[x1 + 1:u])

            w = f.find('Amount Disallowed') + 17
            g = f[w:]
            x1 = g.find(':') + w
            u = g.find('\n') + w
            hg.append(f[x1 + 1:u])

            w = f.find('Discount') + 8
            g = f[w:]
            x1 = g.find(':') + w
            u = g.find('\n') + w
            hg.append(f[x1 + 1:u])

            w = f.find('Co-payment') + 10
            g = f[w:]
            x1 = g.find(':') + w
            u = g.find('\n') + w
            hg.append(f[x1 + 1:u])

            w = f.find('Deductible') + 10
            g = f[w:]
            x1 = g.find(':') + w
            u = g.find('\n') + w
            hg.append(f[x1 + 1:u])

            w = f.find('TDS') + 3
            g = f[w:]
            x1 = g.find(':') + w
            u = g.find('\n') + w
            hg.append(f[x1 + 1:u])

            w = f.find('Amount Paid') + 11
            g = f[w:]
            x1 = g.find(':') + w
            u = g.find('(') + w
            hg.append(f[x1 + 1:u])

            hg = [sub.replace('\n', ' ') for sub in hg]
            hg = [sub.replace('  ', '') for sub in hg]
            hg = [sub.replace('Rs', '') for sub in hg]
            hg = [sub.replace(':', '') for sub in hg]

            # print(hg)

            for i in range(0, len(hg)):
                s1.cell(row=t + 2, column=1).value = t + 1
                s1.cell(row=t + 2, column=i + 2).value = hg[i]

            w = f.find('Reason') + 5
            g = f[w:]
            x1 = g.find('\n') + w
            u = g.find('Payment Summary') + w
            temp = f[x1:u]
            temp = temp.split('\n')
            so1 = []
            so2 = []
            so3 = []
            so4 = []
            so5 = []
            so6 = []
            for k in temp:
                if (k != ''):
                    temp2 = re.findall("\d+\.\d+", k)
                    if len(temp2) != 0:
                        k = k.replace('   ', '$')
                        while (k.find('$$') != -1):
                            k = k.replace('$$', '$')

                        temp1 = k.split('$')
                        so1.append(temp1[0])
                        so2.append(temp1[1])
                        so3.append(temp2[0])
                        so4.append(temp2[1])
                        so5.append(temp2[2])
                        if (len(temp1) == 6):
                            so6.append(temp1[5])
                        else:
                            so6.append(' ')
                    else:
                        continue
            so1 = [sub.replace('  ', '') for sub in so1]
            so2 = [sub.replace('  ', '') for sub in so2]
            so3 = [sub.replace('  ', '') for sub in so3]
            so4 = [sub.replace('  ', '') for sub in so4]
            so5 = [sub.replace('  ', '') for sub in so5]
            so6 = [sub.replace('  ', '') for sub in so6]

            for i in range(0, len(so1)):
                wq += 1
                row_num = s2.max_row
                s2.cell(row=row_num + 1, column=1).value = wq
                s2.cell(row=row_num + 1, column=2).value = hg[0]
                s2.cell(row=row_num + 1, column=3).value = so1[i]
                s2.cell(row=row_num + 1, column=4).value = so2[i]
                s2.cell(row=row_num + 1, column=5).value = so3[i]
                s2.cell(row=row_num + 1, column=6).value = so4[i]
                s2.cell(row=row_num + 1, column=7).value = so5[i]
                s2.cell(row=row_num + 1, column=8).value = so6[i]
            os.rename(os.getcwd() + '/health_insurance/attachments_' + str(sys.argv[6]) + '/' + onlyfiles[t],
                      os.getcwd() + '/health_insurance/attachments_' + str(sys.argv[6]) + '/' + hg[0] + '.pdf')
        except Exception as e:
            s1.cell(row=t + 2, column=1).value = 'error'
            os.rename(os.getcwd() + '/health_insurance/attachments_' + str(sys.argv[6]) + '/' + onlyfiles[t],
                      os.getcwd() + '/health_insurance/attachments_' + str(sys.argv[6]) + '/' + hg[0] + '.pdf')

    # print(po)
    print("Done")
    wbk.save(wbkName)
    wbk.close
    wbkName = 'count/count.xlsx'
    wbk = openpyxl.load_workbook(wbkName)
    s1 = wbk.worksheets[0]

    row_ = s1.max_row + 1
    s1.cell(row=row_, column=1).value = 'health insurance'
    s1.cell(row=row_, column=2).value = str(sys.argv[6])
    s1.cell(row=row_, column=3).value = len(fg)
    s1.cell(row=row_, column=4).value = len(onlyfiles)
    s1.cell(row=row_, column=5).value = len(repeat)
    wbk.save(wbkName)
    wbk.close
    subprocess.run(["python", "updation.py", "1", "max", "9", " "])
except:
    log_exceptions()