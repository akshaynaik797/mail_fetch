import os
import re
import sys
import struct, time
import subprocess
from datetime import date
import datetime
import openpyxl
import pdftotext
from bad_pdf import text_from_pdf
import time
import requests

from patient_name_fun import pname_fun

now = datetime.datetime.now()

subprocess.run(["python", "updation.py","1","max1","1",sys.argv[2]])
subprocess.run(["python", "updation.py","1","max","2",sys.argv[3]])
subprocess.run(["python", "updation.py","1","max","3",sys.argv[4]])
subprocess.run(["python", "updation.py","1","max","5",str(now)])
subprocess.run(["python", "updation.py","1","max","7",sys.argv[5]])
subprocess.run(["python", "updation.py","1","max","8",sys.argv[6]])


with open(sys.argv[1], "rb") as f:
    pdf = pdftotext.PDF(f)

with open('star/output.txt', 'w') as f:
    f.write(" ".join(pdf))
num_lines = sum(1 for line in open('star/output.txt'))
if num_lines < 2:
    text_from_pdf(sys.argv[1], 'star/output.txt')

with open('star/output.txt', 'r') as myfile:
    f = myfile.read()

try:		
    hg=[]

    subject = sys.argv[5]
    temp = re.compile(r"[^-]*(?=-)").search(subject)
    if temp is not None:
        hg.append(temp.group().strip())
    else:
        hg.append("")

    status='Approved'
    if f.find('Initial')!=-1:
        w=f.find('Initial Approval Amount')+25
        g=f[w:]
        u=g.find('(')+w
        hg.append(f[w:u])
    else:
        w=f.find('Total Authorized amount')+25
        g=f[w:]
        u=g.find('(')+w
        hg.append(f[w:u])
    w=f.find('Policy Number')+13
    g=f[w:]
    u=g.find('Expected')+w
    hg.append(f[w:u])

    w=f.find('Id of the')+9
    g=f[w:]
    u=g.find('\n')+w
    hg.append(f[w:u])
    hg[1]=hg[1].replace('-','')
    hg=[sub.replace(':','') for sub in hg]
    hg=[sub.replace('Rs.','') for sub in hg]
    hg=[sub.replace('  ','') for sub in hg]
    hg = [sub.replace('-', '') for sub in hg]
    hg = [sub.replace('|', '') for sub in hg]

    #s2.cell(row_count_1+1, column=9).value='Yes'
    #s2.cell(row_count_1+1, column=10).value='NA'
    #wbk.save(wbkName)

    subprocess.run(["python", "updation.py","1","max","9",'Yes'])
    subprocess.run(["python", "updation.py","1","max","10",'NA'])

    try:
        regex_list = [r"(?<=Patient’s Member).*"]
        pname = pname_fun(f, regex_list)
        subprocess.run(["python", "test_api.py",hg[0],hg[1],hg[2],'Pa',status,sys.argv[6],sys.argv[1],'',hg[3], pname])
        '''wbk= openpyxl.load_workbook(wbkName)
        s2=wbk.worksheets[1]
        s2.cell(row_count_1+1, column=11).value='YES'
        '''
        subprocess.run(["python", "updation.py","1","max","11",'Yes'])
    except Exception as e:
        #s2.cell(row_count_1+1, column=11).value='NO'
        subprocess.run(["python", "updation.py","1","max","11",'No'])
except Exception as e:
    #s2.cell(row_count_1+1, column=9).value='No'
    #s2.cell(row_count_1+1, column=11).value='NO'
    subprocess.run(["python", "updation.py","1","max","9",'Yes'])
    subprocess.run(["python", "updation.py","1","max","11",'No'])

now = datetime.datetime.now()
#s2.cell(row_count_1+1, column=6).value=now
#wbk.save(wbkName)
subprocess.run(["python", "updation.py","1","max","6",str(now)])
