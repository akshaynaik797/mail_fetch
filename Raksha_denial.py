import os
import sys
import struct, time
import subprocess
from datetime import date
import datetime
import openpyxl
import pdftotext
import re
import time
import requests
now = datetime.datetime.now()

subprocess.run(["python", "updation.py","1","max1","1",sys.argv[2]])
subprocess.run(["python", "updation.py","1","max","2",sys.argv[3]])
subprocess.run(["python", "updation.py","1","max","3",sys.argv[4]])
subprocess.run(["python", "updation.py","1","max","5",str(now)])
subprocess.run(["python", "updation.py","1","max","7",sys.argv[5]])
subprocess.run(["python", "updation.py","1","max","8",sys.argv[6]])

with open(sys.argv[1], "rb") as f:
        pdf = pdftotext.PDF(f)
with open('Raksha/output1.txt', 'w') as f:
        f.write(" ".join(pdf))     
with open('Raksha/output1.txt', 'r') as myfile:
        f = myfile.read()
try:
    hg=[]
    w=f.find('Claim Number')+12
    k=f[w:]
    u=k.find("(")+w
    hg.append(f[w:u])

    
    w=f.find("Patient's Member UHID")+21
    k=f[w:]
    hg.append(f[w:w+16])


    w=f.find('Patient Name')+12
    k=f[w:]
    hg.append(f[w:w+35])


    w=f.find('Policy Number')+13
    k=f[w:]
    hg.append(f[w:w+35])
    if 'Admission Is Not Between Policy Start Date And End Date' in f:
        hg = []
        hg.append('')
        regex = r'(?<=Member Id).*'
        temp = re.compile(regex).search(f)
        if temp is not None:
            hg.append(temp.group())
        else:
            hg.append('')

        regex = r'(?<=Patient Name).*'
        temp = re.compile(regex).search(f)
        if temp is not None:
            hg.append(temp.group())
        else:
            hg.append('')
        hg.append('')

    temp = re.compile(r"(?<=Patient Name).*(?=Age)").search(f)

    if temp is not None:
        pname = temp.group().strip()
        pname = pname.replace(':', "")
        pname = pname.strip()
    else:
        pname = ''

    hg=[sub.replace(':','') for sub in hg]      
    hg=[sub.replace('  ','') for sub in hg]
    hg=[sub.replace('Rs.','') for sub in hg]

        
    subprocess.run(["python", "updation.py","1","max","9",'Yes'])
    subprocess.run(["python", "updation.py","1","max","10",'NA'])

    try:
        subprocess.run(["python", "test_api.py",hg[0],'',hg[3],'','Denial',sys.argv[6],sys.argv[1],'',hg[1],hg[2], pname])
        subprocess.run(["python", "updation.py","1","max","11",'Yes'])  
    except Exception as e:
        #s2.cell(row_count_1+1, column=11).value='NO'
        subprocess.run(["python", "updation.py","1","max","11",'No'])
except Exception as e:
        #s2.cell(row_count_1+1, column=9).value='No'
        #s2.cell(row_count_1+1, column=11).value='NO'
        subprocess.run(["python", "updation.py","1","max","9",'Yes'])
        subprocess.run(["python", "updation.py","1","max","11",'No'])
now = datetime.datetime.now()
#s2.cell(row_count_1+1, column=6).value=now
#wbk.save(wbkName)
subprocess.run(["python", "updation.py","1","max","6",str(now)])


