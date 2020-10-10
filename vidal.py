import pandas as pd
from email.mime.text import MIMEText
import smtplib
import sys
import time
import imaplib
import email
import camelot
import PyPDF2
import csv
import xlsxwriter
from xlrd import open_workbook
import pdftotext
import xlwt
import os
import glob
import os.path
import xlrd
from os import listdir
from os import path
from os.path import isfile, join
from html.parser import HTMLParser
import pdfkit
import pandas as pd
import html2text
import subprocess

from decode_error import check_subject, read_from_delete
from make_log import log_exceptions

try:

    eu = []
    subprocess.run(["python", "updation.py", "1", "max", "9", "X"])

    repeat = []
    if path.exists(r'vidal/vidal' + str(sys.argv[6]) + '.xlsx'):
        os.remove(r'vidal/vidal' + str(sys.argv[6]) + '.xlsx')
    import openpyxl

    fg = []
    from openpyxl.styles import Color, PatternFill, Font, Border

    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')


    def read_email_from_gmail():
        SMTP_SERVER = str(sys.argv[5])
        mail = imaplib.IMAP4_SSL(SMTP_SERVER)
        # mail.login(user = 'Mediclaim@therisingmedicare.com', password = 'cef@2018')
        e_id = str(sys.argv[1])
        pswd = str(sys.argv[2])
        srt = str(sys.argv[3])
        stp = str(sys.argv[4])
        mail.login(user=e_id, password=pswd)
        mail.select("inbox", readonly=True)
        ###############################################<
        mail_uid = str(sys.argv[7])
        if mail_uid == -1:
            type, data = mail.search(None, '(SUBJECT "EFT/Cheque Details" since ' + srt + ' before ' + stp + ')')
            ids = data[0]
            id_list = ids.split()
        else:
            ids = mail_uid  # data is a list.
            # accept id from outside and put in id_list akshay var name = id

            id_list = []  # ids is a space separated string
            id_list.append(ids)
        ###############################################>
        # type, data = mail.search(None, '(SUBJECT "EFT/Cheque Details" since ' + srt + ' before ' + stp + ')')
        # ids = data[0]  # data is a list.
        # id_list = ids.split()  # ids is a space separated string
        # # print(id_list)
        for i in range(0, len(id_list)):
            latest_email_id = id_list[i]  # get the latest
            result, data = mail.fetch(latest_email_id, "(RFC822)")
            ##################################################ak
            try:
                # raw_email = data[0][1].decode('utf-8')
                try:
                    raw_email = data[0][1].decode('utf-8')
                except UnicodeDecodeError:
                    try:
                        raw_email = data[0][1].decode('ISO-8859-1')
                    except UnicodeDecodeError:
                        try:
                            raw_email = data[0][1].decode('ascii')
                        except UnicodeDecodeError:
                            pass
                email_message = email.message_from_string(raw_email)
                subject = email_message['Subject']
                result, sys.argv[8] = check_subject(subject, sys.argv[8], mail)
                if result == 'Changed':
                    # raise Exception('subject not matched')
                    raise Exception('subject not matched', )
            except:
                try:
                    log_exceptions(syssubject=sys.argv[8], subject=subject, error='subject not matched')
                except:
                    pass
                if result != 'OK':
                    data = {'server': SMTP_SERVER,
                            'hospmail': e_id,
                            'pass': pswd,
                            'subject': sys.argv[8]}
                    try:
                        data = read_from_delete(data)
                        if data == None:
                            raise Exception("Not found")
                    except:
                        log_exceptions(msg='not found in deleted', subject=sys.argv[8])
            ##################################################akend
            raw_email = data[0][1].decode('utf-8')
            email_message = email.message_from_string(raw_email)
            if email_message['Subject'] not in fg:
                for part in email_message.walk():

                    if part.get_content_maintype() == 'multipart':
                        # print part.as_string()
                        continue
                    if part.get('Content-Disposition') is None:
                        # print part.as_string()
                        continue
                    fileName = part.get_filename()
                    detach_dir = (os.getcwd() + '/vidal/attachments_' + str(sys.argv[6]))
                    if bool(fileName):
                        filePath = os.path.join(detach_dir, fileName)
                        if not os.path.isfile(filePath):
                            from reportlab.pdfgen import canvas
                            c = canvas.Canvas(fileName + '.pdf')
                            # print (fileName)
                            fp = open(filePath, 'wb')
                            fp.write(part.get_payload(decode=True))
                            fp.close()
            else:
                repeat.append(email_message['Subject'])

            fg.append(email_message['Subject'])


    mypath = os.getcwd() + '/vidal'
    if not path.exists(mypath):
        os.mkdir(mypath)
    if not path.exists(mypath + '/attachments_' + str(sys.argv[6])):
        os.mkdir(mypath + '/attachments_' + str(sys.argv[6]))
    mypath = os.getcwd() + '/vidal/attachments_' + str(sys.argv[6]) + '/'
    for filename in os.listdir(mypath):
        file_path = os.path.join(mypath, filename)
        if os.path.isfile(file_path) or os.path.islink(file_path):
            os.unlink(file_path)
    read_email_from_gmail()

    wbkName = 'vidal/vidal' + str(sys.argv[6]) + '.xlsx'
    wbk = openpyxl.Workbook()
    wbk.create_sheet('1')
    s1 = wbk.worksheets[0]
    s2 = wbk.worksheets[1]
    wq = 0
    onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
    for t in range(0, len(onlyfiles)):
        try:
            def select_sheet(wks, sheet_name):

                if not sheet_name in wks.sheetnames:
                    wks.create_sheet(sheet_name)
                # print(sheet_name)

                wks.save('vidal.xlsx')


            def select_column(wks, s, ro):
                sheet = wks.worksheets[s]

                max_col = sheet.max_column
                s = []
                for i in range(1, max_col + 1):
                    cell_obj = sheet.cell(row=1, column=i)
                    s.append(cell_obj.value)


            # print(s)

            # print(mypath+'attachments/'+onlyfiles[t])
            CCN = onlyfiles[t].replace('.pdf', '')
            # print(CCN)

            tables = camelot.read_pdf(mypath + onlyfiles[t], pages='all', line_scale=100)
            tables.export('vidal/foo1.xlsx', f='excel')
            loc = ("vidal/foo1.xlsx")
            wb1 = openpyxl.load_workbook(loc)
            wb = xlrd.open_workbook(loc)

            with open(mypath + onlyfiles[t], "rb") as f:
                pdf = pdftotext.PDF(f)

            with open('vidal/output.txt', 'w') as f:
                f.write(" ".join(pdf))
            with open('vidal/output.txt', 'r') as myfile:
                f = myfile.read()

            sh1 = ['Sr No.', 'ccn', 'Claim No.', 'Claimant Name', 'Insured Person', 'Policy No.', 'Corporate Name',
                   'Insurance Company', 'Emp no./Ref-no.', 'Diagnosis', 'TDS Amount', 'Co-pay Amt', 'Deductible Amt',
                   'Discount allowed', 'Settled Amt', 'EFT No', 'Transaction date', 'DOA', 'DOD', 'IP No.', 'Auth.Amt',
                   'Claim File', 'insurance company', 'Deductible Amt', 'net payable', 'settled amt']
            for i in range(0, len(sh1)):
                s1.cell(row=1, column=i + 1).value = sh1[i]
            sh2 = ['Sr No.', 'ccn', 'Bill No.', 'Bill Date', 'Nature of  Expenditure', 'Amt Claimed',
                   'Disallowed / Non Medical Expenses Rs.', 'Amount Settled Rs.', 'Remarks']
            for i in range(0, len(sh2)):
                s2.cell(row=1, column=i + 1).value = sh2[i]

            sheet_1 = wb.sheet_by_index(0)
            sheet_1.cell_value(0, 0)
            max_row = sheet_1.nrows
            max_col = sheet_1.ncols
            wd = wb1.worksheets[0]
            for i in range(2, max_col):
                if sheet_1.cell_value(1, i) == '':
                    # print(i)
                    wd.delete_cols(i + 1)
            wb1.save(loc)
            wb1.close
            wb = xlrd.open_workbook(loc)
            sheet_1 = wb.sheet_by_index(0)
            sheet_1.cell_value(0, 0)
            max_row = sheet_1.nrows
            max_col = sheet_1.ncols
            hg = []
            b = []
            p = []
            np = []
            r = []
            rt = []
            mh = []
            m = 0
            for i in range(2, max_row):
                if sheet_1.cell_value(i, 1) == 'Total :':
                    m = 1
                    break
                hg.append(sheet_1.cell_value(i, 2))
                b.append(sheet_1.cell_value(i, 3))
                p.append(sheet_1.cell_value(i, 4))
                np.append(sheet_1.cell_value(i, 5))
                r.append(sheet_1.cell_value(i, 6))
                k = sheet_1.cell_value(i, 7)
                if k != '':
                    x = k.find('.00')
                    op = k[:x + 3]
                    oy = k[x + 3:]
                    rt.append(op)
                    mh.append(oy)
                else:
                    rt.append(sheet_1.cell_value(i, 7))
                    mh.append(sheet_1.cell_value(i, 8))

                # hg=[sub.replace('a.ii)', '') for sub in hg]

            if (m == 0):
                sheet_2 = wb.sheet_by_index(1)
                sheet_2.cell_value(0, 0)
                max_row = sheet_2.nrows
                for i in range(2, max_row):
                    hg.append(sheet_2.cell_value(i, 2))
                    b.append(sheet_2.cell_value(i, 3))
                    p.append(sheet_2.cell_value(i, 4))
                    np.append(sheet_2.cell_value(i, 5))
                    r.append(sheet_2.cell_value(i, 6))
                    rt.append(sheet_2.cell_value(i, 7))
                    mh.append(sheet_1.cell_value(i, 8))
                    if sheet_2.cell_value(i, 4) == 'Total':
                        break

            # print(hg)

            gh = []

            w5 = f.find('Claim No. :') + 11
            g = f[w5:]
            u3 = g.find('Claim File') + w5
            gh.append(f[w5:u3])

            w4 = f.find('Claimant Name :') + 15
            g = f[w4:]
            u2 = g.find('\n') + w4
            gh.append(f[w4:u2])

            w5 = f.find('Insured Name :') + 15
            g = f[w5:]
            u3 = g.find('Claimant Name :') + w5
            gh.append(f[w5:u3])

            w4 = f.find('Policy No. :') + 12
            g = f[w4:]
            u2 = g.find('Policy Start Date') + w4
            gh.append(f[w4:u2])

            w5 = f.find('Claim Settlement No.') + 21
            g = f[w5:]
            u3 = g.find('Settlement Date:') + w5
            ccn = f[w5:u3]

            w4 = f.find('Corporate Name') + 14
            g = f[w4:]
            u2 = g.find('Payee Name') + w4
            gh.append(f[w4:u2])

            w5 = f.find('Insurance Company :') + 20
            g = f[w5:]
            u3 = g.find('\n') + w5
            gh.append(f[w5:u3])

            w4 = f.find('Emp no./Ref-no.') + 17
            g = f[w4:]
            u2 = g.find('IP No.') + w4
            gh.append(f[w4:u2])

            w5 = f.find('Diagnosis :') + 11
            g = f[w5:]
            u3 = g.find('\n') + w5
            gh.append(f[w5:u3])

            w4 = f.find('TDS Amount') + 17
            g = f[w4:]
            u2 = g.find('\n') + w4
            gh.append(f[w4:u2])

            w5 = f.find('Co-pay Amt.') + 18
            g = f[w5:]
            u3 = g.find('\n') + w5
            gh.append(f[w5:u3])

            w5 = f.find('Deductible Amt') + 20
            g = f[w5:]
            u3 = g.find('\n') + w5
            gh.append(f[w5:u3])

            w5 = f.find('Discount allowed') + 24
            g = f[w5:]
            u3 = g.find('\n') + w5
            gh.append(f[w5:u3])

            w4 = f.find('payment of Rs.') + 14
            g = f[w4:]
            u2 = g.find('vide') + w4
            gh.append(f[w4:u2])

            w5 = f.find('EFT No.') + 7
            g = f[w5:]
            u3 = g.find('dated') + w5
            gh.append(f[w5:u3])

            g = f[u3:]
            u2 = g.find('to the') + u3
            gh.append(f[u3 + 6:u2])
            gh = [sub.replace(':', '') for sub in gh]
            w5 = f.find('DOA :') + 5
            g = f[w5:]
            u3 = g.find('Hospital :') + w5
            gh.append(f[w5:u3])

            w5 = f.find('DOD :') + 5
            g = f[w5:]
            u3 = g.find('\n') + w5
            gh.append(f[w5:u3])

            w5 = f.find('IP No.') + 8
            g = f[w5:]
            u3 = g.find('\n') + w5
            gh.append(f[w5:u3])

            w5 = f.find('Auth.Amt (Rs.)') + 17
            g = f[w5:]
            u3 = g.find('\n') + w5
            gh.append(f[w5:u3])

            w5 = f.find('Claim File No. :') + 17
            g = f[w5:]
            u3 = g.find('\n') + w5
            gh.append(f[w5:u3])

            w5 = f.find('Insurance Company :') + 20
            g = f[w5:]
            u3 = g.find('\n') + w5
            gh.append(f[w5:u3])

            w5 = f.find('Deductible Amt (Rs.) :') + 23
            g = f[w5:]
            u3 = g.find('\n') + w5
            gh.append(f[w5:u3])

            w5 = f.find('Total Approved (Rs.) :') + 23
            g = f[w5:]
            u3 = g.find('\n') + w5
            gh.append(f[w5:u3])

            w5 = f.find('Approval Amount (Rs) :') + 23
            g = f[w5:]
            u3 = g.find('\n') + w5
            gh.append(f[w5:u3])

            ccn = ccn.replace(':', '')
            gh = [sub.replace('  ', '') for sub in gh]
            # print(gh,ccn)

            for i in range(0, len(gh)):
                s1.cell(row=t + 2, column=1).value = t + 1
                s1.cell(row=t + 2, column=2).value = ccn
                s1.cell(row=t + 2, column=i + 3).value = gh[i]
            for i in range(0, len(hg)):
                row_num = s2.max_row + 1
                wq += 1
                s2.cell(row=row_num, column=1).value = wq
                s2.cell(row=row_num, column=2).value = ccn
                s2.cell(row=row_num, column=3).value = hg[i]
                s2.cell(row=row_num, column=4).value = b[i]
                s2.cell(row=row_num, column=5).value = p[i]
                s2.cell(row=row_num, column=6).value = np[i]
                s2.cell(row=row_num, column=7).value = r[i]
                s2.cell(row=row_num, column=8).value = rt[i]
                s2.cell(row=row_num, column=9).value = mh[i]
        except Exception as e:
            s1.cell(row=t + 2, column=1).value = 'error'

    print("Done")
    wbk.save(wbkName)
    wbk.close
    wbkName = 'count/count.xlsx'
    wbk = openpyxl.load_workbook(wbkName)
    s1 = wbk.worksheets[0]

    row_ = s1.max_row + 1
    s1.cell(row=row_, column=1).value = 'vidal'
    s1.cell(row=row_, column=2).value = str(sys.argv[6])
    s1.cell(row=row_, column=3).value = len(fg)
    s1.cell(row=row_, column=4).value = len(onlyfiles)
    s1.cell(row=row_, column=5).value = len(repeat)
    wbk.save(wbkName)
    subprocess.run(["python", "updation.py", "1", "max", "9", " "])
except:
	log_exceptions()