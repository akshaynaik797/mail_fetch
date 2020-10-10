import os
import re
import sys
import struct, time
import subprocess
from datetime import date
import datetime
import openpyxl
import pdftotext
import time
import requests
now = datetime.datetime.now()

subprocess.run(["python", "updation.py","1","max1","1",sys.argv[2]])
subprocess.run(["python", "updation.py","1","max","2",sys.argv[3]])
subprocess.run(["python", "updation.py","1","max","3",sys.argv[4]])
subprocess.run(["python", "updation.py","1","max","5",str(now)])
subprocess.run(["python", "updation.py","1","max","7",sys.argv[5]])
subprocess.run(["python", "updation.py","1","max","8",sys.argv[6]])

with open(sys.argv[1], "rb") as f:
	pdf = pdftotext.PDF(f)

with open('Raksha/output.txt', 'w') as f:
	f.write(" ".join(pdf))     
with open('Raksha/output.txt', 'r') as myfile:
 	f = myfile.read()
try:
	if (f.find("Cashless Query Letter")!=-1):
		hg=[]
		w=f.find('Claim Number:')+13
		g=f[w:]
		u=g.find('(')+w
		hg.append(f[w:u])

		status='Information Awaiting'
		hg.append('')

		w=f.find('Policy Number :')+15
		g=f[w:]
		u=g.find('Expected')+w
		hg.append(f[w:u])
	
		w=f.find('Member UHID:')+12
		g=f[w:]
		u=g.find('\n')+w
		hg.append(f[w:u])

		temp = re.compile(r"(?<=Patient Name).*(?=Age)").search(f)

		if temp is not None:
			pname = temp.group().strip()
			pname = pname.replace(':', "")
			pname = pname.strip()
		else:
			pname = ''
		
		hg=[sub.replace(':','') for sub in hg]	
		hg=[sub.replace('  ','') for sub in hg]
		hg=[sub.replace('Rs.','') for sub in hg]
		#s2.cell(row_count_1+1, column=9).value='Yes'
		#s2.cell(row_count_1+1, column=10).value='NA'
		#wbk.save(wbkName)
		
		subprocess.run(["python", "updation.py","1","max","9",'Yes'])
		subprocess.run(["python", "updation.py","1","max","10",'NA'])
		
		try:
			subprocess.run(["python", "test_api.py",hg[0],hg[1],hg[2],'',status,sys.argv[6],sys.argv[1],'',hg[3], pname])
			'''wbk= openpyxl.load_workbook(wbkName)
			s2=wbk.worksheets[1]
			s2.cell(row_count_1+1, column=11).value='YES'
			'''
			subprocess.run(["python", "updation.py","1","max","11",'Yes'])	
		except Exception as e:
			#s2.cell(row_count_1+1, column=11).value='NO'
			subprocess.run(["python", "updation.py","1","max","11",'No'])
	elif(f.find("Ref No.")!=-1):
		hg=[]
		w=f.find('Claim No:')+9
		g=f[w:]
		u=g.find('\n')+w
		hg.append(f[w:u])

		status='Information Awaiting'
		w=f.find('MemberId :')+10
		g=f[w:]
		u=g.find('\n')+w
		hg.append(f[w:u])#policy no
		w=f.find('Policy No :')+11
		g=f[w:]
		u=g.find('\n')+w
		hg.append(f[w:u])
		#hg.append('')
		
		hg.append('')#member id
		hg=[sub.replace(':','') for sub in hg]	
		hg=[sub.replace('  ','') for sub in hg]
		hg=[sub.replace('Rs.','') for sub in hg]
		#s2.cell(row_count_1+1, column=9).value='Yes'
		#s2.cell(row_count_1+1, column=10).value='NA'
		#wbk.save(wbkName)
		
		subprocess.run(["python", "updation.py","1","max","9",'Yes'])
		subprocess.run(["python", "updation.py","1","max","10",'NA'])
		
		try:
			subprocess.run(["python", "test_api.py",hg[0],'',hg[2],'',status,sys.argv[6],sys.argv[1],'',hg[1]])
			subprocess.run(["python", "updation.py","1","max","11",'Yes'])	
		except Exception as e:
			#s2.cell(row_count_1+1, column=11).value='NO'
			subprocess.run(["python", "updation.py","1","max","11",'No'])
	elif f.find("Sub: First Reminder for Claim") != -1:
		hg = []
		badchars = ('/', ',', ':', '-')
		datadict = {}
		regexdict = {'preid': [r"(?<=Claim No).*"],
					 'amount': [r"(?<=Claim Amount).*"],
					 'memid': [r"(?<=MemberId).*"],
					 'polno': [r"(?<=Policy No).*"],
					 'pname': [r"(?<=Claim in respect of).*"],
					 }

		for i in regexdict:
			for j in regexdict[i]:
				data = re.compile(j).search(f)
				if data is not None:
					temp = data.group().strip()
					for k in badchars:
						temp = temp.replace(k, "")
					datadict[i] = temp.strip()
					break
				datadict[i] = ""
		a = 1
		hg = datadict
		status = 'Information Awaiting'

		subprocess.run(["python", "updation.py", "1", "max", "9", 'Yes'])
		subprocess.run(["python", "updation.py", "1", "max", "10", 'NA'])

		try:
			subprocess.run(
				["python", "test_api.py", hg["preid"], hg["amount"], hg["polno"], " ", status, sys.argv[6], sys.argv[1], '', hg["memid"],hg["pname"]])
			'''wbk= openpyxl.load_workbook(wbkName)
			s2=wbk.worksheets[1]
			s2.cell(row_count_1+1, column=11).value='YES'
			'''
			subprocess.run(["python", "updation.py", "1", "max", "11", 'Yes'])
		except Exception as e:
			# s2.cell(row_count_1+1, column=11).value='NO'
			subprocess.run(["python", "updation.py", "1", "max", "11", 'No'])

except Exception as e:
	subprocess.run(["python", "updation.py","1","max","9",'Yes'])
	subprocess.run(["python", "updation.py","1","max","11",'No'])
	
	
now = datetime.datetime.now()
subprocess.run(["python", "updation.py","1","max","6",str(now)])

