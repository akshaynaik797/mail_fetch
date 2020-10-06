import requests
import sys
import openpyxl
import subprocess
#files={'doc':open('/home/shivam/Desktop/vnu_scripts/automation/aditya/attachments_preauth/1.pdf', 'rb')}
files = {'doc': open(sys.argv[7],  'rb')}

subprocess.run(["python", "updation.py","1","max","22",sys.argv[1]])
subprocess.run(["python", "updation.py","1","max","23",sys.argv[2]])
subprocess.run(["python", "updation.py","1","max","24",sys.argv[5]])
subprocess.run(["python", "updation.py","1","max","25",sys.argv[6]])
subprocess.run(["python", "updation.py","1","max","26",sys.argv[3]])
subprocess.run(["python", "updation.py","1","max","27",sys.argv[9]])



API_ENDPOINT ="https://vnusoftware.com/iclaimtest/api/preauth"
if len(sys.argv) == 10:
        sys.argv.insert(10,"test")
subprocess.run(["python", "updation.py","1","max","28",sys.argv[10]])
data = {
	'preauthid':sys.argv[1],
	#'pname': sys.argv[10],
        'amount':sys.argv[2], 
        'status':sys.argv[5],
	'process':sys.argv[4],
	'lettertime':sys.argv[6],
	'policyno':sys.argv[3],
	'memberid':sys.argv[9],
	'comment':sys.argv[10]
	}

#r = requests.post(url = API_ENDPOINT, data = data , files = files)
pastebin_url = 'Not called.'#r.text
if pastebin_url!="Data Update Success":
        print("API commented out")
	#subprocess.run(["python", "sms_api.py","api error"])


s_temp='NA'
print(pastebin_url)
for i in range(1,10):
	if sys.argv[i]=='':
                if s_temp=='NA':
                        s_temp=''
                if i==1:
                        subprocess.run(["python", "updation.py","1","max","9",'No'])
			#subprocess.run(["python", "updation.py","1","max","10",'preauthid,'])
                        s_temp=s_temp+'preauthid,'
                if i==2 and sys.argv[5]!='Information Awaiting':
                        subprocess.run(["python", "updation.py","1","max","9",'No'])
			#subprocess.run(["python", "updation.py","1","max","10",'amount,'])
                        s_temp=s_temp+'amount,'
                if i==3:
                        subprocess.run(["python", "updation.py","1","max","9",'No'])
			#subprocess.run(["python", "updation.py","1","max","10",'policyno,'])
                        s_temp=s_temp+'policyno,'
                if i==4:
                        subprocess.run(["python", "updation.py","1","max","9",'No'])
			#subprocess.run(["python", "updation.py","1","max","10",'process,'])
                        s_temp=s_temp+'process,'
                if i==5:
                        subprocess.run(["python", "updation.py","1","max","9",'No'])
			#subprocess.run(["python", "updation.py","1","max","10",'status,'])
                        s_temp=s_temp+'status,'
                if i==6:
                        subprocess.run(["python", "updation.py","1","max","9",'No'])
			#subprocess.run(["python", "updation.py","1","max","10",'lettertime,'])
                        s_temp=s_temp+'lettertime,'
                if i==9:
                        subprocess.run(["python", "updation.py","1","max","9",'No'])
			#subprocess.run(["python", "updation.py","1","max","10",'memberid,'])
                        s_temp=s_temp+'memberid,'
subprocess.run(["python", "updation.py","1","max","10",str(s_temp)])
#s2.cell(row_count_1, column=12).value=str(data)
subprocess.run(["python", "updation.py","1","max","12",str(data)])
#s2.cell(row_count_1, column=13).value=str(pastebin_url)
subprocess.run(["python", "updation.py","1","max","13",str(pastebin_url)])
#print(s2.cell(row_count_1, column=13).value)
