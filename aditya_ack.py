#######################
import os
import re
import sys
import struct, time
import subprocess
from datetime import date
import datetime
import openpyxl
import pdftotext
import time
import requests

now = datetime.datetime.now()
subprocess.run(["python", "updation.py","1","max1","1",sys.argv[2]])
subprocess.run(["python", "updation.py","1","max","2",sys.argv[3]])
subprocess.run(["python", "updation.py","1","max","3",sys.argv[4]])
subprocess.run(["python", "updation.py","1","max","5",str(now)])
subprocess.run(["python", "updation.py","1","max","7",sys.argv[5]])
subprocess.run(["python", "updation.py","1","max","8",sys.argv[6]])


with open(sys.argv[1], "rb") as f:
    pdf = pdftotext.PDF(f)

with open('aditya/output1.txt', 'w') as f:
    f.write(" ".join(pdf))     
with open('aditya/output1.txt', 'r') as myfile:
     f = myfile.read()

try:
    badchars = ('/', ',', ':', '-')
    datadict = {}
    regexdict = {'preid': [r"(?<=Preauth Number:).*"],
                 'polno': [r"(?<=Policy Number:).*"],
                 'pname': [r"(?<=Patient name:).*"],
                 'amount': [r"(?<=Claim Amount:).*"], }

    for i in regexdict:
        for j in regexdict[i]:
            data = re.compile(j).search(f)
            if data is not None:
                temp = data.group().strip()
                for k in badchars:
                    temp = temp.replace(k, "")
                datadict[i] = temp.strip()
                break
            datadict[i] = ""
    a = 1

    try:
        subprocess.run(["python", "test_api.py",datadict["preid"],datadict["amount"],datadict["polno"],'','Acknowledgement',sys.argv[6],sys.argv[1],'','',datadict["pname"]])
        '''wbk= openpyxl.load_workbook(wbkName)
        s2=wbk.worksheets[1]
        s2.cell(row_count_1+1, column=11).value='YES'
        '''
        subprocess.run(["python", "updation.py","1","max","11",'Yes'])    
    except Exception as e:
        #s2.cell(row_count_1+1, column=11).value='NO'
        subprocess.run(["python", "updation.py","1","max","11",'No'])
except Exception as e:
    #s2.cell(row_count_1+1, column=9).value='No'
    #s2.cell(row_count_1+1, column=11).value='NO'
    subprocess.run(["python", "updation.py","1","max","9",'Yes'])
    subprocess.run(["python", "updation.py","1","max","11",'No'])
now = datetime.datetime.now()
#s2.cell(row_count_1+1, column=6).value=now
#wbk.save(wbkName)
subprocess.run(["python", "updation.py","1","max","6",str(now)])


