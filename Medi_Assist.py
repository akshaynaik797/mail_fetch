from email.mime.text import MIMEText
import pdftotext
import smtplib
import time
import imaplib
import email
import camelot
import PyPDF2
import xlsxwriter
from xlrd import open_workbook
from decode_error import read_from_delete, check_subject
from email.header import decode_header

import subprocess
import xlwt
import sys
import os
import os.path
import xlrd
from os import listdir
from os import path
from os.path import isfile, join
from make_log import log_exceptions

try:
    subprocess.run(["python", "updation.py", "1", "max", "9", "X"])
    # sys.argv = ['Medi_Assist.py', 'Tpappg@maxhealthcare.com', 'May@2020', '25-Aug-2020', '25-Aug-2020', 'outlook.office365.com', 'Max', '27271']
    if path.exists(r'Medi_Assist/Medi_Assist' + str(sys.argv[6]) + '.xlsx'):
        os.remove(r'Medi_Assist/Medi_Assist' + str(sys.argv[6]) + '.xlsx')
    import openpyxl

    po = []
    repeat = []
    fg = []
    from openpyxl.styles import Color, PatternFill, Font, Border

    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')


    def read_email_from_gmail():
        SMTP_SERVER = str(sys.argv[5])
        mail = imaplib.IMAP4_SSL(SMTP_SERVER)
        e_id = str(sys.argv[1])
        pswd = str(sys.argv[2])
        srt = str(sys.argv[3])
        stp = str(sys.argv[4])
        mail.login(user=e_id, password=pswd)
        mail.select("inbox", readonly=True)
        ###############################################<
        mail_uid = str(sys.argv[7])
        if mail_uid == -1:
            type, data = mail.search(None,
                                     '(SUBJECT "Settlement of your Claim Reference" since ' + srt + ' before ' + stp + ')')
            ids = data[0]
            id_list = ids.split()
        else:
            ids = mail_uid  # data is a list.
            # accept id from outside and put in id_list akshay var name = id

            id_list = []  # ids is a space separated string
            id_list.append(ids)
        ###############################################>
        # type, data = mail.search(None,
        #                          '(SUBJECT "Settlement of your Claim Reference" since ' + srt + ' before ' + stp + ')')
        # ids = data[0]  # data is a list.
        # id_list = ids.split()  # ids is a space separated string
        # # print(id_list)
        for i in range(0, len(id_list)):
            latest_email_id = id_list[i]  # get the latest
            result, data = mail.fetch(latest_email_id,
                                      "(RFC822)")  # fetch the email body (RFC822)             for the given ID
            ##################################################ak
            try:
                # raw_email = data[0][1].decode('utf-8')
                try:
                    raw_email = data[0][1].decode('utf-8')
                except UnicodeDecodeError:
                    try:
                        raw_email = data[0][1].decode('ISO-8859-1')
                    except UnicodeDecodeError:
                        try:
                            raw_email = data[0][1].decode('ascii')
                        except UnicodeDecodeError:
                            pass
                email_message = email.message_from_string(raw_email)
                subject = email_message['Subject']
                result, sys.argv[8] = check_subject(subject, sys.argv[8], mail)
                if result == 'Changed':
                    # raise Exception('subject not matched')
                    raise Exception('subject not matched', )
            except:
                try:
                    log_exceptions(syssubject=sys.argv[8], subject=subject, error='subject not matched')
                except:
                    pass
                if result != 'OK':
                    data = {'server': SMTP_SERVER,
                            'hospmail': e_id,
                            'pass': pswd,
                            'subject': sys.argv[8]}
                    try:
                        data = read_from_delete(data)
                        if data == None:
                            raise Exception("Not found")
                    except:
                        log_exceptions(msg='not found in deleted', subject=sys.argv[8])
            ##################################################akend
                raw_email = data[0][1].decode('utf-8')
                email_message = email.message_from_string(raw_email)
            if email_message['Subject'] not in fg:
                for part in email_message.walk():
                    if part.get_content_maintype() == 'multipart':
                        # print part.as_string()
                        continue
                    if part.get('Content-Disposition') is None:
                        # print part.as_string()
                        continue
                    fileName = part.get_filename()
                    detach_dir = (os.getcwd() + '/Medi_Assist/attachments_' + str(sys.argv[6]))
                    if bool(fileName):
                        filePath = os.path.join(detach_dir, fileName)
                        if not os.path.isfile(filePath):
                            from reportlab.pdfgen import canvas
                            c = canvas.Canvas(fileName + '.pdf')
                            # print (fileName)
                            fp = open(filePath, 'wb')
                            fp.write(part.get_payload(decode=True))
                            fp.close()
            else:
                repeat.append(email_message['Subject'])
            fg.append(email_message['Subject'])


    mypath = os.getcwd() + '/Medi_Assist'
    if not path.exists(mypath):
        os.mkdir(mypath)
    if not path.exists(mypath + '/attachments_' + str(sys.argv[6])):
        os.mkdir(mypath + '/attachments_' + str(sys.argv[6]))
    mypath = os.getcwd() + '/Medi_Assist/attachments_' + str(sys.argv[6]) + '/'
    for filename in os.listdir(mypath):
        file_path = os.path.join(mypath, filename)
        if os.path.isfile(file_path) or os.path.islink(file_path):
            os.unlink(file_path)
    read_email_from_gmail()

    onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
    wbkName = 'Medi_Assist/Medi_Assist' + str(sys.argv[6]) + '.xlsx'
    wbk = openpyxl.Workbook()
    wbk.create_sheet('Sheet1')
    wbk.create_sheet('Sheet2')
    # wbk.create_sheet('Sheet3')
    s1 = wbk.worksheets[0]
    s2 = wbk.worksheets[1]
    s3 = wbk.worksheets[2]
    eu = []
    goble_total = []
    for t in range(0, len(onlyfiles)):
        try:
            with open(mypath + onlyfiles[t], "rb") as f:
                pdf = pdftotext.PDF(f)

            with open('Medi_Assist/output.txt', 'w') as f:
                f.write(" ".join(pdf))
            with open('Medi_Assist/output.txt', 'r') as myfile:
                f = myfile.read()


            def select_sheet(wks, sheet_name):

                if not sheet_name in wks.sheetnames:
                    wks.create_sheet(sheet_name)
                # print(sheet_name)

                wks.save('Medi_Assist/Medi_Assist' + str(sys.argv[6]) + '.xlsx')


            def select_column(wks, s, ro):
                sheet = wks.worksheets[s]

                max_col = sheet.max_column
                s = []
                for i in range(1, max_col + 1):
                    cell_obj = sheet.cell(row=1, column=i)
                    s.append(cell_obj.value)


            # print(s)

            # print(mypath+'attachments/'+onlyfiles[t])
            CCN = onlyfiles[t].replace('.pdf', '')
            # print(CCN)

            tables = camelot.read_pdf(mypath + onlyfiles[t], pages='1-end')
            a = tables.n

            # print(tables.n)
            # for 2 table case
            if tables.n == 2:
                sh1 = ['Patient Name', 'Insurance Company', 'Medi Assist ID', 'Policy Holder', 'IP No.', 'Policy No.',
                       'Primary Beneficiary', 'Employee ID', 'Insurer Claim No', 'Insurer Member ID', 'diagnosis', 'doa',
                       'dod']
                sh2 = ['Settled Amount (INR)', 'Settlement Date', 'UTR Number', 'Account Holder Name', 'Bank Name',
                       'Account Number']

                tables.export('Medi_Assist/foo1.xls', f='excel')
                loc = ("Medi_Assist/foo1.xls")
                wb = xlrd.open_workbook(loc)
                pdfFileObj = open(mypath + onlyfiles[t], 'rb')
                pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
                pageObj = pdfReader.getPage(0)
                f = pageObj.extractText()
                w = f.find('recommended')
                g = f[w:]
                u = g.find('€') + w + 1
                total = f[u:]
                if total.find(' ') != -1:
                    uop = g.find('The') + w
                    total = f[u:uop]

                if total[0] < '0' or total[0] > '9':
                    pageObj = pdfReader.getPage(1)
                    df = pageObj.extractText()
                    w = df.find('payment')
                    g = df[w:]
                    u = df.find('€') + 1
                    uop = df.find('The')
                    total = df[u:uop]
                goble_total.append(total)
                # print(f)
                '''text_file = open("mail.txt", "w")
                n = text_file.write(f)
                text_file.close()'''

                for i in range(0, len(sh1)):
                    s1.cell(row=1, column=i + 3).value = sh1[i]

                x1 = f.find('treatment of') + 12
                g = f[x1:]
                # print(g)
                y1 = g.find(' at ') + x1

                x = f.find('Claimant Name :')
                x3 = f.find('issued by') + 9
                g = f[x3:]
                y3 = g.find(',') + x3
                y = f.find('MAID:')
                x2 = f.find('Relationship')
                x4 = f.find('Policy No:') + 10
                y4 = f.find('Period of Insurance')
                x5 = f.find('Primary Member') + 16
                g = f[x5:]
                y5 = g.find('(') + x5
                y6 = g.find(')') + x5
                x7 = f.find('Member Id') + 10
                y7 = f.find('Address')
                gh = []

                gh.append(f[x + 15:y])
                gh.append(f[x3:y3])
                gh.append(f[y + 5:x2])
                gh.append(' ')  # policy holder not exracted
                gh.append(' ')
                gh.append(f[x4:y4])
                gh.append(f[x5:y5])
                gh.append(f[y5 + 1:y6])
                gh.append(' ')
                gh.append(f[x7:y7])
                gh.append(f[x1:y1])

                w = f.find('from') + 5
                g = f[w:]
                u = g.find('to') + w
                gh.append(f[w:u])

                u1 = g.find('.') + w
                gh.append(f[u + 2:u1])

                for i in range(0, len(gh)):
                    s1.cell(row=t + 2, column=i + 3).value = gh[i]

                for i in range(0, len(sh2)):
                    s2.cell(row=1, column=i + 3).value = sh2[i]

                hg = []
                x8 = f.find('Amount Settled') + 18
                y8 = f.find('Category Break')
                x9 = f.find('Settlement Date') + 16
                y9 = f.find('Insurer')
                x10 = f.find('Transaction Id') + 15
                y10 = f.find('Account Holder Name')
                x11 = f.find('Bank name')
                y11 = f.find('Branch')
                y10 = f.find('Acc No')
                hg.append(f[x8:y8])
                hg.append(f[x9:y9])
                hg.append(f[x10:y10])
                hg.append(' ')
                hg.append(f[x11 + 10:y11])
                hg.append(f[y10 + 7:x11])

                for i in range(0, len(hg)):
                    s2.cell(row=t + 2, column=i + 3).value = hg[i]

                s = []
                d = []
                sheet_3 = wb.sheet_by_index(1)
                sheet_3.cell_value(0, 0)

                for i in range(1, sheet_3.nrows):
                    s.append(sheet_3.cell_value(i, 1))
                    d.append(sheet_3.cell_value(i, 2))
                s = [sub.replace('\n', ' ') for sub in s]
                # print(s,d)
                if (t == 0):
                    for i in range(0, len(s) - 1):
                        s3.cell(row=1, column=i + 3).value = s[i]
                        po.append(s[i])
                        length = i + 3
                        s3.cell(row=t + 2, column=i + 3).value = d[i]
                else:

                    for i in range(0, len(s) - 1):
                        if s[i] not in po:
                            po.append(s[i])
                            s3.cell(row=1, column=length + 1).value = s[i]
                            s3.cell(row=t + 2, column=length + 1).value = d[i]
                            length = length + 1
                        if s[i] in po:
                            # print('hi')
                            u = po.index(s[i])
                            s3.cell(row=t + 2, column=u + 3).value = d[i]
                s_v = []

                sheet_2 = wb.sheet_by_index(0)
                sheet_2.cell_value(0, 0)
                b = []
                p = []
                np = []
                r = []
                for i in range(3, sheet_2.nrows):
                    s_v.append(sheet_2.cell_value(i, 1))
                    b.append(sheet_2.cell_value(i, 2))
                    p.append(sheet_2.cell_value(i, 3))
                    np.append(sheet_2.cell_value(i, 4))
                    r.append(sheet_2.cell_value(i, 5))
                for i in range(len(s_v)):
                    select_sheet(wbk, s_v[i])
                ro = []
                ro.append(sheet_2.row_values(2))
                ro = ro[0][2:]
                # print(ro)
                xls = xlrd.open_workbook('Medi_Assist/Medi_Assist' + str(sys.argv[6]) + '.xlsx', on_demand=True)
                sheet_list = xls.sheet_names()
                # print(sheet_list)
                for i in range(len(s_v)):
                    sheet_name = s_v[i]
                    for y in range(len(sheet_list)):
                        if sheet_name in sheet_list[y]:
                            for k in range(0, len(ro)):
                                sheet = wbk.worksheets[y]
                                # print(ro[k],k,sheet)
                                sheet.cell(row=1, column=k + 3).value = ro[k]
                            sheet.cell(row=t + 2, column=3).value = b[i]
                            sheet.cell(row=t + 2, column=4).value = p[i]
                            sheet.cell(row=t + 2, column=5).value = np[i]
                            sheet.cell(row=t + 2, column=6).value = r[i]
                        # select_column(wbk,y,ro)

            if tables.n == 4 or tables.n == 3:
                if tables.n == 4:
                    tables.export('Medi_Assist/foo1.xls', f='excel')
                    loc = ("Medi_Assist/foo1.xls")
                if tables.n == 3:
                    tables = camelot.read_pdf(mypath + onlyfiles[t], line_scale=100)
                    tables.export('Medi_Assist/foo1.xls', f='excel')
                    loc = ("Medi_Assist/foo1.xls")

                wb = xlrd.open_workbook(loc)
                sheet_0 = wb.sheet_by_index(0)
                sheet_0.cell_value(0, 0)
                k = []
                v = []
                for i in range(1, sheet_0.nrows):
                    k.append(sheet_0.cell_value(i, 1))
                    v.append(sheet_0.cell_value(i, 2))

                pdfFileObj = open(mypath + onlyfiles[t], 'rb')
                pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
                pageObj = pdfReader.getPage(0)
                f = pageObj.extractText()
                w = f.find('recommended')
                g = f[w:]
                u = g.find('€') + w + 1
                total = f[u:]
                if total.find(' ') != -1:
                    uop = g.find('The') + w
                    total = f[u:uop]

                if total[0] < '0' or total[0] > '9':
                    pageObj = pdfReader.getPage(1)
                    df = pageObj.extractText()
                    w = df.find('payment')
                    g = df[w:]
                    u = df.find('€') + 1
                    uop = df.find('The')
                    total = df[u:uop]
                goble_total.append(total)
                x1 = f.find('treatment of') + 12
                g = f[x1:]

                # print(g)
                y1 = g.find(' at ') + x1
                k.append('diagnosis')
                k.append('doa')
                k.append('dod')
                v.append(f[x1:y1])

                w = f.find('from') + 5
                g = f[w:]
                u = g.find('to') + w
                v.append(f[w:u])

                u1 = g.find('.') + w
                v.append(f[u + 2:u1])
                m = []
                l = []
                sheet_1 = wb.sheet_by_index(1)
                sheet_1.cell_value(0, 0)
                for i in range(0, len(k)):
                    s1.cell(row=1, column=i + 3).value = k[i]
                for i in range(1, sheet_1.nrows):
                    m.append(sheet_1.cell_value(i, 1))
                    l.append(sheet_1.cell_value(i, 2))
                for i in range(0, len(m)):
                    s2.cell(row=1, column=i + 3).value = m[i]
                s = []
                d = []
                sheet_3 = wb.sheet_by_index(3)
                sheet_3.cell_value(0, 0)

                for i in range(1, sheet_3.nrows):
                    s.append(sheet_3.cell_value(i, 1))
                    d.append(sheet_3.cell_value(i, 2))
                s = [sub.replace('\n', ' ') for sub in s]
                if (t == 0):
                    for i in range(0, len(s) - 1):
                        s3.cell(row=1, column=i + 3).value = s[i]
                        po.append(s[i])
                        length = i + 3
                        s3.cell(row=t + 2, column=i + 3).value = d[i]
                    # print(po)
                else:

                    for i in range(0, len(s) - 1):
                        for j in range(0, len(po)):
                            if s[i] not in po:
                                # print(s[i])
                                po.append(s[i])
                                s3.cell(row=1, column=length + 1).value = s[i]
                                s3.cell(row=t + 2, column=length + 1).value = d[i]
                                length = length + 1
                            elif s[i] in po:
                                u = po.index(s[i])
                                s3.cell(row=t + 2, column=u + 3).value = d[i]

                # print(len(k))
                # res = {k[i]: v[i] for i in range(len(k))}
                # print (res)

                # if(wbk.worksheets=='<Worksheet "Sheet1">'):

                for i in range(0, len(v)):
                    s1.cell(row=t + 2, column=i + 3).value = v[i]
                for i in range(0, len(l)):
                    s2.cell(row=t + 2, column=i + 3).value = l[i]

                # dynamic method table 3
                s_v = []

                sheet_2 = wb.sheet_by_index(2)
                sheet_2.cell_value(0, 0)
                b = []
                p = []
                np = []
                r = []
                for i in range(2, sheet_2.nrows):
                    s_v.append(sheet_2.cell_value(i, 1))
                    b.append(sheet_2.cell_value(i, 2))
                    p.append(sheet_2.cell_value(i, 3))
                    np.append(sheet_2.cell_value(i, 4))
                    r.append(sheet_2.cell_value(i, 5))
                for i in range(len(s_v)):
                    select_sheet(wbk, s_v[i])
                ro = []
                ro.append(sheet_2.row_values(1))
                ro = ro[0][2:]
                # print(s_v)
                xls = xlrd.open_workbook('Medi_Assist/Medi_Assist' + str(sys.argv[6]) + '.xlsx', on_demand=True)
                sheet_list = xls.sheet_names()
                # print(sheet_list)
                for i in range(len(s_v)):
                    sheet_name = s_v[i]
                    for y in range(len(sheet_list)):
                        if sheet_name in sheet_list[y]:
                            for k in range(0, len(ro)):
                                sheet = wbk.worksheets[y]
                                # print(sheet)
                                sheet.cell(row=1, column=k + 3).value = ro[k]
                            sheet.cell(row=t + 2, column=3).value = b[i]
                            sheet.cell(row=t + 2, column=4).value = p[i]
                            sheet.cell(row=t + 2, column=5).value = np[i]
                            sheet.cell(row=t + 2, column=6).value = r[i]
                        # select_column(wbk,y,ro)


        except Exception as e:
            log_exceptions()
            s1.cell(row=t + 2, column=1).value = 'error'
    for t in range(0, len(onlyfiles)):
        CCN = onlyfiles[t].replace('.pdf', '')
        for wd in wbk.worksheets:
            wd.cell(row=1, column=1).value = 'Sr. No.'
            wd.cell(row=1, column=2).value = 'CCN'
            wd.cell(row=t + 2, column=1).value = t + 1
            wd.cell(row=t + 2, column=2).value = CCN
        if t in eu:
            s1.cell(row=t + 2, column=1).fill = redFill
            s1.cell(row=t + 2, column=1).value = 'error'

    # static method
    '''
            sheet_2 = wb.sheet_by_index(2) 
            sheet_2.cell_value(0, 0) 
            s_v=[]
            b=[]
            p=[]
            np=[] 
            r=[]
            for i in range(2,sheet_2.nrows): 
                s_v.append(sheet_2.cell_value(i,1))
                b.append(sheet_2.cell_value(i, 2))	
                p.append(sheet_2.cell_value(i,3))
                np.append(sheet_2.cell_value(i,4))
                r.append(sheet_2.cell_value(i, 5))
            ro=[]
            ro.append(sheet_2.row_values(1))
            ro=ro[0][2:7]
    
            g=[s4,s5,s6,s7,s8,s9,s10,s11]
            for m in g:
                for i in range (0,len(ro)):
                    m.cell(row=1, column=i+3).value = ro[i]
    
            for i in range(len(s_v)):
                if s_v[i] == 'Pharmacy & Medicine Charges':
                    
                    s4.cell(row=t+2, column=3).value = b[i]
                    s4.cell(row=t+2, column=4).value = p[i]
                    s4.cell(row=t+2, column=5).value = np[i]
                    s4.cell(row=t+2, column=6).value = r[i]
                if s_v[i] == 'Consultant Charges':
                    s5.cell(row=t+2, column=3).value = b[i]
                    s5.cell(row=t+2, column=4).value = p[i]
                    s5.cell(row=t+2, column=5).value = np[i]
                    s5.cell(row=t+2, column=6).value = r[i]
                if s_v[i] == 'Miscellaneous Charges':
                    s6.cell(row=t+2, column=3).value = b[i]
                    s6.cell(row=t+2, column=4).value = p[i]
                    s6.cell(row=t+2, column=5).value = np[i]
                    s6.cell(row=t+2, column=6).value = r[i]
                if s_v[i] == 'Investigation & Lab Charges':
                    s7.cell(row=t+2, column=3).value = b[i]
                    s7.cell(row=t+2, column=4).value = p[i]
                    s7.cell(row=t+2, column=5).value = np[i]
                    s7.cell(row=t+2, column=6).value = r[i]
                if s_v[i] == 'Hospital Charges':
                    s8.cell(row=t+2, column=3).value = b[i]
                    s8.cell(row=t+2, column=4).value = p[i]
                    s8.cell(row=t+2, column=5).value = np[i]
                    s8.cell(row=t+2, column=6).value = r[i]
                if s_v[i] == 'Package':
                    print('hi')
                    s9.cell(row=t+2, column=3).value = b[i]
                    s9.cell(row=t+2, column=4).value = p[i]
                    s9.cell(row=t+2, column=5).value = np[i]
                    s9.cell(row=t+2, column=6).value = r[i]
                if s_v[i] == 'Surgery Charges':
                    print('hi')
                    s11.cell(row=t+2, column=3).value = b[i]
                    s11.cell(row=t+2, column=4).value = p[i]
                    s11.cell(row=t+2, column=5).value = np[i]
                    s11.cell(row=t+2, column=6).value = r[i]
                if s_v[i] == 'Total':
                    s10.cell(row=t+2, column=3).value = b[i]
                    s10.cell(row=t+2, column=4).value = p[i]
                    s10.cell(row=t+2, column=5).value = np[i]
                    s10.cell(row=t+2, column=6).value = r[i]
        
    
    '''
    if (t != 0):
        u = length + 1
        s3.cell(row=1, column=u).value = 'Net amount recommended for payment'
    for t in range(0, len(onlyfiles)):
        s3.cell(row=t + 2, column=u).value = goble_total[t]

    # print(po)
    print("Done")
    wbk.save(wbkName)
    wbk.close
    wbkName = 'count/count.xlsx'
    wbk = openpyxl.load_workbook(wbkName)
    s1 = wbk.worksheets[0]

    row_ = s1.max_row + 1
    s1.cell(row=row_, column=1).value = 'Medi Assist'
    s1.cell(row=row_, column=2).value = str(sys.argv[6])
    s1.cell(row=row_, column=3).value = len(fg)
    s1.cell(row=row_, column=4).value = len(onlyfiles)
    s1.cell(row=row_, column=5).value = len(repeat)
    wbk.save(wbkName)
    subprocess.run(["python", "updation.py", "1", "max", "9", " "])
except:
    log_exceptions()
    pass